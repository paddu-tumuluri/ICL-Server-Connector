import time
import threading
import socket
import configparser
import datetime
import os
import re
import shutil
import sys
import select
import json
import queue
from turtle import width
from xml.etree.ElementTree import tostring
from datetime import datetime, timedelta  
from datetime import datetime
from pathlib import Path
import ctypes
import subprocess
import glob

import openpyxl
from openpyxl import Workbook

import tkinter as tk
import tkinter.messagebox as messagebox
from tkinter import scrolledtext, filedialog, simpledialog

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap import Combobox
from PIL import Image, ImageTk
import configparser
import pandas as pd
import traceback

# Default values for IP and Port
DEFAULT_IP = "192.168.254.62"
DEFAULT_PORT = 12005
DEFAULT_COMMANDS = ["busmap()", "2+2", "Cat()", 'Help("cos")']
# Default values for DSLAM drivers and ports
DEFAULT_DSLAM_DRIVERS = []
DEFAULT_PORT2 = "0"

response3 = None
response3_tp = None
devices_listbox = None
dslam_profile_dropdown = None
dut_profile_dropdown = None
dpu_profile_dropdown = None
dslam_name_dropdown= None
dslam_busmap_dropdown = None
dslam_host_dropdown = None
operator_dropdown = None

# Initialize configuration for saving preferences
config = configparser.ConfigParser()
# Add with other global variables at the top
global timestamp_start
timestamp_start = None
# Queue for thread-safe communication
message_queue = queue.Queue()

# Placeholder for connection object
sock = None
_startup_complete = False
# Global variable to track test plan file between pages 2 and 3
current_tp_file_2_3 = None

# Add with other global variables
global current_tp_file
current_tp_file = None
global current_log_file
current_log_file = None
global current_results_csv
current_results_csv = None  # Global variable to store current Results.csv path

page3_functions = {}  # Dictionary to store page 3 functions
page4_functions = {}  # Dictionary to store page 4 functions    
page2_functions = {}  # Dictionary to store page 2 functions
  # Global variable to store current test plan file path
test_plan_entry = None
test_plan_tree = None
notebook = None
dslam_driver_var = None
is_shutting_down = False
# Command history
command_history = ["Execute all"] + DEFAULT_COMMANDS.copy()  # Add "Execute all" as the first option

# Create the main window
root = ttk.Window(themename="flatly")  # Initialize ttkbootstrap window with a theme
root.geometry("1500x900")  # Increase the window size
root.state('zoomed')

def on_closing():
    global sock, is_shutting_down
    
    # Set a flag to indicate we're shutting down
    is_shutting_down = True
    
    try:
        if sock:
            # Show disconnecting message (safely)
            try:
                status_label.config(text="Disconnecting...")
                root.update()
            except:
                # Ignore errors if widgets can't be updated
                pass
            
            try:
                # Close the socket
                sock.close()
                sock = None
                print("Socket connection closed")
            except Exception as e:
                print(f"Error closing socket: {e}")
            
            # Final status update (safely)
            try:
                status_label.config(text="Disconnected")
                root.update()
            except:
                # Ignore errors if widgets can't be updated
                pass
            
    except Exception as e:
        print(f"Error during closing: {e}")
    finally:
        # Add a slight delay to let threads notice we're shutting down
        time.sleep(0.5)
        
        # Actually quit and destroy
        root.quit()
        root.destroy()

# Make sure this line is after the function definition
root.protocol("WM_DELETE_WINDOW", on_closing)

# Create main container
main_container = ttk.Frame(root)
main_container.pack(fill="both", expand=True)

# Create and pack the logo frame at the top
logo_frame = ttk.Frame(main_container)
logo_frame.pack(fill="x", padx=5, pady=2)

try:
    # Get the absolute path to the logo file
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        base_path = sys._MEIPASS
    else:
        # Running as script
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    logo_path = os.path.join(base_path, 'sparnex-logo.jpg')
    
    if os.path.exists(logo_path):
        # Load and resize logo using PIL
        logo_img = Image.open(logo_path)
        logo_img = logo_img.resize((450, 100), Image.Resampling.LANCZOS)
        
        # Convert to PhotoImage after theme initialization
        logo_photo = ImageTk.PhotoImage(logo_img)
        
        # Create logo label with explicit background color from theme
        logo_label = ttk.Label(logo_frame, image=logo_photo, style='TLabel')
        logo_label.image = logo_photo  # Keep a reference
        
        # Configure label background to match theme
        style = ttk.Style()
        bg_color = style.lookup('TLabel', 'background')
        if bg_color:
            logo_label.configure(background=bg_color)
        
        logo_label.pack(side="left", padx=(5, 0), pady=5)
        
        # Create a Close button at the right side of the logo frame
        style.configure('Close.TButton', font=("Helvetica", 10, "bold"), foreground='white')
        close_button = ttk.Button(
            logo_frame, 
            text="✕ Close Application", 
            style='Close.TButton',
            cursor="hand2",
            command=on_closing  # Use your existing on_closing function
            )
        close_button.pack(side="right", padx=(0, 20), pady=5)

    else:
        print(f"Logo file not found at: {logo_path}")
        ttk.Label(logo_frame, text="Sparnex Instruments", 
                 font=("Helvetica", 16, "bold")).pack(side="left", padx=(20, 0), pady=5)
    
except Exception as e:
    print(f"Error loading logo: {e}")
    ttk.Label(logo_frame, text="Sparnex Instruments", 
             font=("Helvetica", 16, "bold")).pack(side="left", padx=(20, 0), pady=5)

root.bind('<Control-q>', lambda e: on_closing())
# Initialize the style
style = ttk.Style()

# Define a default style for all widgets
style.configure('.', font=("Helvetica", 10, "bold"), foreground='black')

# Define a default style for all buttons
style.configure('TButton', font=("Helvetica", 10, "bold"), cursor="hand2")
#style.map('TButton', background=[('active', '#1abc9c'), ('!active', '#16a085'), ('hover', '#2ecc71')])

# Define a default style for all checkbuttons
style.configure('TCheckbutton', font=("Helvetica", 10, "bold"))

# Define a default style for all text fields
style.configure('TEntry', font=("Helvetica", 10, "bold"), foreground='black')
#style.map('TEntry', fieldbackground=[('readonly', '#3e3d3b')])

# Define a default style for all combo boxes
style.configure('TCombobox', font=("Helvetica", 10, "bold"), foreground='black')
#style.map('TCombobox', fieldbackground=[('readonly', '#3e3d3b')], background=[('readonly', '#3e3d3b')])



APP_VERSION = "13.10"
# Function to update the window title with status
def update_window_title(status):
    root.title(f"ServerConnector \t - \t Version:{APP_VERSION} \t - \t {status}")

# Initial window title
update_window_title("Ready to connect")



# # Load user preferences from config file
# def load_preferences():
#     config.read('user_config.ini')
#     if 'SERVER' in config:
#         ip_entry.delete(0, tk.END)  # Clear the entry first
#         ip_entry.insert(0, config['SERVER']['IP'])
#         port_entry.delete(0, tk.END)  # Clear the entry first
#         port_entry.insert(0, config['SERVER']['Port'])

# Save user preferences (IP, Port) to config file

toggle_var = tk.BooleanVar()
# Global variables to store parsed details
global bcm_address, FN_profile, bcm_file,profile_filename
# Update bcm6400x_profile for the session
global bcm_profile
bcm_profile = "BCM6400x_Default_VDSL.txt"
bcm_address = "DMS4_192_168_001_001"
FN_profile = r"..\DSLAM\DMS4\Profiles\BCM6400x_Default_VDSL.txt"
bcm_file = r"..\DSLAM\DMS4\BCM6400x_100.bin"
profile_filename = None
global selected_port
selected_port = 0 #default
profile_listbox = None
preview_tree = None  # Initialize as None
global dslam_folder_var
global address_tp
address_tp = {}
global libname_var
global dut_libname_var
# Global variables at the top of file
status_label = None  # Main status label in header
  # Page 3 specific status label
status_var_3 = tk.StringVar()  # Page 3 status variable
status_label_3 = None
status_label_4 = None  # Page 4 specific status label
status_var_4 = tk.StringVar()  # Page 4 status variable

netstat_content_frame = None
netstat_canvas = None
# Initialize globals
# Add these global variables at the top of your file
global connection_content_frame, connection_canvas
connection_content_frame = None
connection_canvas = None
global root_dir, workspace_dir, out_dir, dslam_profiles_dir, icl_scripts_dir

global test_search_var
# Global variables for Page 4
page4_frame = None
config_inner_frame = None
test_inner_frame = None
config_canvas = None
test_canvas = None
response_text_4 = None
status_label_4 = None
status_var_4 = tk.StringVar()
config_search_var = tk.StringVar()
test_search_var = tk.StringVar()
execute_window_instance = None
operator_var = tk.StringVar()
page4_functions = {}
# Define variables for DSLAM/DPU related fields globally
dslam_driver_entry = tk.StringVar()
dslam_port_var = tk.StringVar()
dslam_slot_var = tk.IntVar()
dslam_frame_var = tk.IntVar()
dslam_host_var = tk.StringVar()
dslam_folder_var = tk.StringVar(value="")
dslam_profile_var = tk.StringVar()
layout_var = tk.StringVar(value="default")
output_folder_var = tk.StringVar(value='..\\..\\Myworkspace\\Out\\') # Default value
output_filename_var = tk.StringVar(value='') # Default value
libname_var = tk.StringVar(value="")
dut_libname_var = tk.StringVar(value="")
startup_time_var = tk.IntVar(value=90) # Default value 90 seconds
error_free_time = tk.IntVar(value=60)
stabilize_time = tk.IntVar(value=0)

min_stabilize_time_var = tk.IntVar()  # Set default min value
max_stabilize_time_var = tk.IntVar()  # Set default max value

min_error_free_time_var = tk.IntVar()  # Set default min value
max_error_free_time_var = tk.IntVar()  # Set default max value


max_startup_time_var = tk.IntVar()  # Set default max value
min_startup_time_var = tk.IntVar()  # Set default min value


dslam_name_var = tk.StringVar()
dslam_serialnumber_var = tk.StringVar()
dslam_driver_var = tk.StringVar()
dslam_busmap_var = tk.StringVar()
dslam_port_var = tk.StringVar()
dslam_slot_var = tk.IntVar()
dslam_frame_var = tk.IntVar()
dslam_host_var =tk.StringVar()
dslam_profile_var = tk.StringVar()
# Define variables for DUT related fields globally
dut_name_var = tk.StringVar()
dut_busmap_var = tk.StringVar()
dut_driver_var = tk.StringVar()
dut_serialnumber_var = tk.StringVar()
dut_port_var = tk.IntVar(value=1)
dut_slot_var = tk.StringVar()
#dut_frame_var = tk.StringVar()
dut_folder_var = tk.StringVar(value="")
dut_profile_var = tk.StringVar()
dut_host_var = tk.StringVar()
dut_testplan_var = tk.StringVar()
testscenario_var = tk.StringVar(value="FS")  # Default value
platform_name_var = tk.StringVar()
platform_noise_generator_var = tk.StringVar()
platform_linesimulator_var = tk.StringVar()
platform_switch_var = tk.StringVar()
platform_traffic_generator_var = tk.StringVar()
database_file_var = tk.StringVar(value=r"\\hermes\4_Engineering\7-test campaign\database.measurement.xlsx")
# At the top with other global variables
dut_folder_entry = None
dslam_folder_entry = None
global df
# Add at top with other globals
dslam_drivers_dict = {}  # Global dictionary to store all driver details
results_excel_var = tk.StringVar()
results_folder_var = tk.StringVar(value = r"\\hermes\4_Engineering\7-test campaign\results")
# Add this global variable at the top
page3_instance = None  # Will store the instance of page 3 class
content_frame = None
canvas = None
global dropdowns
# Global variable to track if execution window is open
execute_window_instance = None
operator_var = tk.StringVar()
# Add the global declaration at the top of the file with your other global variables
global results_path_var

# Initialize it with other variables
results_path_var = None


canvas = None
current_tp_file = None
test_plan_entry = None
status_label_3 = None
status_var_3 = None
listbox1 = None
listbox2 = None
preview_tree = None
refresh_button = None
all_items = []
page3_functions = {}  # If not already defined

def save_preferences():
    global driver_path_changed
    try:
        # Start the progress bar
        #progress_bar.start()
       
        # Prompt the user for a file name
        file_path = filedialog.asksaveasfilename(
            title="Save Preferences As",
            defaultextension=".ini",
            filetypes=[("INI files", "*.ini"), ("All files", "*.*")]
        )
        if file_path:
            # Determine the driver path to save
            driver_path_to_save = driver_path_var.get() if driver_path_changed else bcm_file
           
            # Check if an item is selected in the profile_listbox
            selected_profile = profile_listbox.get(profile_listbox.curselection()) if profile_listbox.curselection() else ""
           
            # Ensure all values are strings
            config['PREFERENCES'] = {
                'IP': str(ip_entry.get()) if ip_entry.get() else "",
                'Port': str(port_entry.get()) if port_entry.get() else "",
                'Driver': str(driver_path_to_save),  # Save the appropriate driver path
                'ProfilePath': str(profile_path_var.get()) if profile_path_var.get() else "",  # Ensure this is a string
                'FileNameProfile': str(selected_profile),  # Use the selected profile
                'SelectedPort': str(port_var.get()) if port_var.get() else "",
                'BCMAddress': str(bcm_address) if bcm_address else ""  # Add bcm_address to preferences
            }
            with open(file_path, 'w') as configfile:
                config.write(configfile)
            display_message("Preferences saved successfully.", "black")
        else:
            display_message("Save operation cancelled.", "red")
    except Exception as e:
        display_message(f"Failed to save preferences: {str(e)}", "red")
    finally:
        # Stop the progress bar
        #progress_bar.stop()
        display_message(f" ", "red")



# Function to load settings from a file
def load_settings(file_path):
    config = configparser.ConfigParser()
    if file_path.endswith('.tp'):
        with open(file_path, 'r') as file:
            lines = file.readlines()
       
        config_section = False
        config_data = ['[config]\n']  # Initialize with section header
       
        for line in lines:
            if line.strip() == '[config]':
                config_section = True
                continue
            elif line.strip() == '[\\config]':
                config_section = False
                continue
            
            if config_section and '=' in line:
                # Convert config.dpu.name='value' format to INI format
                key, value = line.strip().split('=', 1)
                # Remove 'config.' prefix and clean up value
                key = key.replace('config.', '')
                value = value.strip("'\"")
                # Add to config data in INI format
                config_data.append(f"{key}={value}\n")
       
        config_data.append('\n')  # End the section
        config.read_string(''.join(config_data))
    elif file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path)
        for index, row in df.iterrows():
            section = row['Section']
            option = row['Option']
            value = row['Value']
            if not config.has_section(section):
                config.add_section(section)
            config.set(section, option, value)
    return config


def validate_startup_time(*args):
    min_value = min_startup_time_var.get()
    max_value = max_startup_time_var.get()
    try:
        current_value = int(startup_time_var.get())
    except (tk.TclError, ValueError):
        current_value = min_value  # or set to a default value

    if current_value < min_value or current_value > max_value:
        messagebox.showerror("Invalid Input", f"Startup Time must be between {min_value} and {max_value}")
        startup_time_var.set(min_value)  # Reset to min value if out of range


def validate_stabilize_time(*args):
    min_value = min_stabilize_time_var.get()
    max_value = max_stabilize_time_var.get()
    try:
        current_value = int(stabilize_time.get())
    except (tk.TclError, ValueError):
        current_value = min_value  # or set to a default value

    if current_value < min_value or current_value > max_value:
        messagebox.showerror("Invalid Input", f"Stabilize Time must be between {min_value} and {max_value}")
        stabilize_time.set(min_value)  # Reset to min value if out of range


def validate_error_free_time(*args):
    min_value = min_error_free_time_var.get()
    max_value = max_error_free_time_var.get()
    try:
        current_value = int(error_free_time.get())
    except (tk.TclError, ValueError):
        current_value = min_value  # or set to a default value

    if current_value < min_value or current_value > max_value:
        messagebox.showerror("Invalid Input", f"Error Free Time must be between {min_value} and {max_value}")
        error_free_time.set(min_value)  # Reset to min value if out of range


def setup_validation_traces():
    startup_time_var.trace_add("write", validate_startup_time)
    stabilize_time.trace_add("write", validate_stabilize_time)
    error_free_time.trace_add("write", validate_error_free_time)

def sync_profile_and_folder(profile_path, is_dslam=True):
    try:
        if not profile_path:
            print("No profile path provided")
            return
        remote_host = ip_entry.get()
        if not remote_host:
            print("No remote host provided")
            return

        # Get parent folder path and profile name directly
        folder_path = os.path.dirname(profile_path)
        profile_name = os.path.basename(profile_path)
        
        print(f"Syncing {'DSLAM' if is_dslam else 'DUT'} profile:")
        print(f"Profile path: {profile_path}")
        print(f"Folder path: {folder_path}")
        print(f"Profile: {profile_name}")

        # Convert to network path if needed
        if folder_path.startswith('C:'):
            # Remove C: and convert to network path format
            path_without_colon = folder_path.replace('C:', 'C')
            network_path = f"//{remote_host}/{path_without_colon}"  # Use forward slashes

            folder_path = network_path.replace('/', '\\')  # Convert to backslashes at the end
            print(f"Using network path: {folder_path}")
        elif folder_path.startswith('..'):
            # Get the base path from dpu.driver
            driver_path = dslam_driver_var.get()  # "C:\Sparnex.testplan.DLA351\ICL_Server\DSLAM\BCM_UDP\BCM_udp_100.bin"
            if driver_path:
                # Extract the base part up to ICL_Server
                base_parts = driver_path.split('ICL_Server')[0]  # "C:\Sparnex.testplan.DLA351\"
                base_path = base_parts.replace('C:', f'//{remote_host}/C')
                
                # Remove the ..\..\
                clean_path = folder_path.replace('..\\..\\', '')  # "ICL_Server\DSLAM\BCM_UDP\Profiles"
                network_path = f"{base_path}\\{clean_path}"
                folder_path = network_path.replace('/', '\\')
        # List files directly using os.listdir
        try:
            profile_files = [f for f in os.listdir(folder_path) if f.endswith('.txt')]
            print(f"Found {len(profile_files)} profile files in {folder_path}")
            
            # Update UI with found files
            if is_dslam:
                if 'DSLAM_Profile' in dropdowns and dropdowns['DSLAM_Profile']['widget']:
                    dropdown = dropdowns['DSLAM_Profile']['widget']
                    if profile_name not in profile_files:
                        profile_files.append(profile_name)
                    dropdown['values'] = [profile_name] + ["Select other Profiles"] + [
                        f for f in profile_files if f != profile_name
                    ]
                    dropdown.set(profile_name)
                    
                # Update folder entry
                dslam_folder_var.set(ensure_trailing_backslash(folder_path))
                if dslam_folder_entry:
                    dslam_folder_entry.config(state='normal')
                    dslam_folder_entry.delete(0, tk.END)
                    dslam_folder_entry.insert(0, ensure_trailing_backslash(folder_path))
                    dslam_folder_entry.config(state='readonly')
            else:
                if 'DUT_Profile' in dropdowns and dropdowns['DUT_Profile']['widget']:
                    dropdown = dropdowns['DUT_Profile']['widget']
                    if profile_name not in profile_files:
                        profile_files.append(profile_name)
                    dropdown['values'] = [profile_name] + ["Select other Profiles"] + [
                        f for f in profile_files if f != profile_name
                    ]
                    dropdown.set(profile_name)
                    
                # Update folder entry
                dut_folder_var.set(ensure_trailing_backslash(folder_path))
                if dut_folder_entry:
                    dut_folder_entry.config(state='normal')
                    dut_folder_entry.delete(0, tk.END)
                    dut_folder_entry.insert(0, ensure_trailing_backslash(folder_path))
                    dut_folder_entry.config(state='readonly')
                    
        except Exception as e:
            print(f"Error accessing directory {folder_path}: {str(e)}")
            traceback.print_exc()
            
    except Exception as e:
        print(f"Error in sync_profile_and_folder: {str(e)}")
        traceback.print_exc()

def convert_to_relative_path(full_path):
    """Convert absolute path to relative path starting with ../../"""
    try:
        if not full_path:
            return full_path
            
        # Normalize path separators and handle both forward/backward slashes
        normalized_path = os.path.normpath(full_path.replace('/', os.sep))
        parts = normalized_path.split(os.sep)
        
        # Find 'ICL_Server' in the path (case insensitive)
        icl_index = -1
        for i, part in enumerate(parts):
            if part.lower() == 'icl_server':
                icl_index = i
                break
                
        if icl_index != -1:
            # Take the path from 'ICL_Server' onwards
            relative_parts = parts[icl_index:]
            # Add '../../' prefix
            relative_path = os.path.join('..', '..', *relative_parts)
            # Ensure consistent path separators
            relative_path = relative_path.replace('/', os.sep)
            print(f"Converted path: {full_path} -> {relative_path}")
            return relative_path
            
        print(f"ICL_Server not found in path: {full_path}")
        return full_path  # Return original if ICL_Server not found
        
    except Exception as e:
        print(f"Error converting path: {str(e)}")
        return full_path

# Function to update GUI fields
def update_gui_fields(config):
    root.update_idletasks()
    if 'config' in config.sections():
        for key, value in config.items('config'):
            # Remove both 'config.' prefixes if present
            stripped_key = key.replace('config.', '')
            clean_value = value.strip("'\"")
            print(f"Setting {stripped_key} to {clean_value}")  # Debugging output
            
            # Map the keys to their corresponding variables
            if stripped_key == 'dpu.name':
                dslam_name_var.set(clean_value)
                dropdowns['DSLAM_Name']['widget'].set(clean_value)
            
            elif stripped_key == 'dpu.serialnumber':
                dslam_serialnumber_var.set(clean_value)
                dropdowns['DSLAM_Serialnumber']['widget'].set(clean_value)
            
            elif stripped_key == 'dpu.busmap':
                dslam_busmap_var.set(clean_value)
                dropdowns['DSLAM_Busmap']['widget'].set(clean_value)
            
            elif stripped_key == 'dpu.port':
                dslam_port_var.set(clean_value if clean_value else "0")
            
            elif stripped_key == 'dpu.slot':
                dslam_slot_var.set(clean_value)
            
            elif stripped_key == 'dpu.frame':
                dslam_frame_var.set(clean_value)

            elif stripped_key == 'dpu.libname':
                libname_var.set(clean_value)    

            elif stripped_key == 'dpu.host':
                dslam_host_var.set(clean_value)
                try:
                    if 'DSLAM_Host' in dropdowns and dropdowns['DSLAM_Host']['widget']:
                        current_values = list(dropdowns['DSLAM_Host']['widget']['values'])
                        if clean_value not in current_values:
                            current_values.append(clean_value)
                            dropdowns['DSLAM_Host']['widget']['values'] = current_values
                    dropdowns['DSLAM_Host']['widget'].set(clean_value)
                except Exception as e:
                    print(f"Error setting DSLAM Host dropdown: {str(e)}")    

            elif stripped_key == 'dpu.folder':
                #\elative_path = convert_to_relative_path(clean_value)
                folder_path = ensure_trailing_backslash(clean_value)
                dslam_folder_var.set(folder_path)
                print(f"DSLAM folder set to: {folder_path}")
            
            elif stripped_key == 'dpu.profile':
                dslam_profile_var.set(clean_value)
                folder_path = dslam_folder_var.get()
                if folder_path:
                    full_path = os.path.join(folder_path, clean_value)
                    print(f"Syncing dslam profile with : {full_path}")
                    sync_profile_and_folder(full_path, is_dslam=True)
                #dropdowns['DSLAM_Profile']['widget'].set(clean_value)
            
            elif stripped_key == 'dpu.driver':
                try:
                    
                    normalized_path = normalize_path(clean_value)
                    dslam_driver_var.set(normalized_path)
                    if ('DSLAM_Driver' in dropdowns and 
                        dropdowns['DSLAM_Driver']['widget'] and 
                        hasattr(dropdowns['DSLAM_Driver']['widget'], 'set')):
                        dropdowns['DSLAM_Driver']['widget'].set(normalized_path)
                    else:
                        print(f"Warning: DSLAM_Driver widget not available for update")
                except Exception as e:
                    print(f"Error updating DSLAM driver: {str(e)}")
            
            # DUT fields
            elif stripped_key == 'dut.name':
                dut_name_var.set(clean_value)
                dropdowns['DUT_Name']['widget'].set(clean_value)
            
            elif stripped_key == 'dut.busmap':
                dut_busmap_var.set(clean_value)
                dropdowns['DUT_Busmap']['widget'].set(clean_value)
            
            elif stripped_key == 'dut.driver':
                try:
                    normalized_path = normalize_path(clean_value)
                    dut_driver_var.set(normalized_path)
                    if ('DUT_Driver' in dropdowns and 
                        dropdowns['DUT_Driver']['widget'] and 
                        hasattr(dropdowns['DUT_Driver']['widget'], 'set')):
                        dropdowns['DUT_Driver']['widget'].set(normalized_path)
                    else:
                        print(f"Warning: DUT_Driver widget not available for update")
                except Exception as e:
                    print(f"Error updating DUT driver: {str(e)}")
            
            elif stripped_key == 'dut.serialnumber':
                dut_serialnumber_var.set(clean_value)
                dropdowns['DUT_Serialnumber']['widget'].set(clean_value)
            
            elif stripped_key == 'dut.testplan':
                dut_testplan_var.set(clean_value)
                dropdowns['DUT_Testplan']['widget'].set(clean_value)
           
            elif stripped_key == 'dut.libname':
                dut_libname_var.set(clean_value)
           
            elif stripped_key == 'dut.folder':
                relative_path = convert_to_relative_path(clean_value)   
                folder_path = ensure_trailing_backslash(relative_path)
                dut_folder_var.set(folder_path)
                print(f"DUT folder set to: {folder_path}")
            
            elif stripped_key == 'operator':
                operator_var.set(clean_value)
                dropdowns['Operator']['widget'].set(clean_value)

            elif stripped_key == 'dut.port':
                dut_port_var.set(clean_value if clean_value else "1")
            
            elif stripped_key == 'dut.host':
                dut_host_var.set(clean_value)
                try:
                    if 'DUT_Host' in dropdowns and dropdowns['DUT_Host']['widget']:
                        current_values = list(dropdowns['DUT_Host']['widget']['values'])
                        if clean_value not in current_values:
                          dropdowns['DUT_Host']['widget']['values'] = current_values  
                    dropdowns['DUT_Host']['widget'].set(clean_value)
                except Exception as e:
                    print(f"Error setting DUT host dropdown :{str(e)}")
            
            elif stripped_key == 'dut.profile':
                dut_profile_var.set(clean_value)
                folder_path = dut_folder_var.get()
                if folder_path:
                    full_path = os.path.join(folder_path, clean_value)
                    print(f"Syncing dut profile with : {full_path}")
                    sync_profile_and_folder(full_path, is_dslam=False)
                #dropdowns['DUT_Profile']['widget'].set(clean_value)
            
            # Validation fields
            elif stripped_key == 'validation.max_startup_time':
                startup_time_var.set(clean_value if clean_value else "90")
            elif stripped_key == 'validation.stabilize_time':
                stabilize_time.set(clean_value if clean_value else "0")
            elif stripped_key == 'validation.error_free_time':
                error_free_time.set(clean_value if clean_value else "60")
            
            # Other fields
            elif stripped_key == 'testscenario':
                testscenario_var.set(clean_value if clean_value else "FS")
            elif stripped_key == 'layout':
                layout_var.set(clean_value if clean_value else "default")
            elif stripped_key == 'output.folder':
                output_folder_var.set(ensure_trailing_backslash(clean_value) if clean_value else "..\\..\\Myworkspace\\Out\\")
            elif stripped_key == 'output.filename':
                output_filename_var.set(clean_value)
            elif stripped_key == 'database.file':
                database_file_var.set(clean_value)
            
            # Platform fields
            elif stripped_key == 'platform.name':
                platform_name_var.set(clean_value)
            elif stripped_key == 'platform.linesimulator':
                platform_linesimulator_var.set(clean_value)
            elif stripped_key == 'platform.noise_generator':
                platform_noise_generator_var.set(clean_value)
            elif stripped_key == 'platform.switch':
                platform_switch_var.set(clean_value)
            elif stripped_key == 'platform.traffic_generator':
                platform_traffic_generator_var.set(clean_value)



def ensure_trailing_backslash(path):
    """Ensure the path ends with a backslash and converts forward slashes to backslashes"""
    if not path:
        return path
    # Convert forward slashes to backslashes
    path = path.replace('/', '\\')
    # Add trailing backslash if not present
    if not path.endswith('\\'):
        path += '\\'
    return path

# Function to update folder entries with trailing backslash
def update_folder_entries():
    """Update folder entries to ensure they end with backslash"""
    try:
        # Update DSLAM folder
        dslam_folder = dslam_folder_var.get()
        if dslam_folder:
            dslam_folder_var.set(ensure_trailing_backslash(dslam_folder))
            
        # Update DUT folder
        dut_folder = dut_folder_var.get()
        if dut_folder:
            dut_folder_var.set(ensure_trailing_backslash(dut_folder))
    except Exception as e:
        print(f"Error updating folder entries: {str(e)}")

# Function to create a new test plan
def create_new_test_plan():
    # if not check_socket(silent=True):
    #     return False
    global current_tp_file
    current_tp_file = None
    file_path = filedialog.asksaveasfilename(defaultextension=".tp", filetypes=[("Test Plan Files", "*.tp"), ("All Files", "*.*")])
    if file_path:
        try:
            with open(file_path, 'w') as configfile:
                # Write the [config] section
                configfile.write("[config]\n")
                
                for field, var in dropdowns.items():
                    # Skip if the variable is None
                    if var['var'].get() is None:
                        continue
                        
                    value = str(var['var'].get())
                    
                    # Special handling for dpu.busmap to extract only address
                    if field == 'dpu.busmap':
                        # Split on space and take the last part (address)
                        value = value.split()[-1] if ' ' in value else value
                        configfile.write(f"config.{field}='{value}'\n")
                        
                    # Numeric fields without quotes
                    elif field in ['dpu.port', 'validation.max_startup_time']:
                        configfile.write(f"config.{field}={value}\n")
                        
                    # All other fields with quotes
                    else:
                        configfile.write(f"config.{field}='{value}'\n")
                
                # Write the closing tags
                configfile.write("[\\config]\n\n[test]\n\n[\\test]")
            
            current_tp_file = file_path
            # Set as current file
            save_test_plan.current_tp_file = file_path           
            messagebox.showinfo("Success", "New Test Plan created and saved successfully.")
            update_status_2(f"Editing: {os.path.basename(file_path)}")
            update_file_status(source_page='2')  # Update status after successful save
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create new test plan: {str(e)}")
            update_status_2(f"Error creating test plan: {str(e)}", True)
            update_file_status(str(e), True, source_page='2')
         

# Function to save a test plan
def save_test_plan():
    global current_tp_file
    
    try:
        if not sock:
            if not hasattr(save_test_plan, 'initialized'):
                save_test_plan.initialized = True
                return
                
        if not current_tp_file:
            return save_as_test_plan()

        
        # Prepare config data
        config_data = "[config]\n"
    
    # DSLAM Configuration
        config_data += f"config.dpu.name='{dslam_name_var.get()}'\n"
        config_data += f"config.dpu.serialnumber='{dslam_serialnumber_var.get()}'\n"  # Add serial number
        busmap_value = dslam_busmap_var.get().split()[-1] if ' ' in dslam_busmap_var.get() else dslam_busmap_var.get()
        config_data += f"config.dpu.busmap='{busmap_value}'\n"
        config_data += f"config.dpu.driver='{dslam_driver_var.get()}'\n"
        config_data += f"config.dpu.port={get_safe_int(dslam_port_var,0)}\n"
        config_data += f"config.dpu.slot={get_safe_int(dslam_slot_var,0)}\n"
        config_data += f"config.dpu.frame={get_safe_int(dslam_frame_var,0)}\n"
        config_data += f"config.dpu.host='{dslam_host_var.get()}'\n"
        config_data += f"config.dpu.folder='{ensure_trailing_backslash(dslam_folder_var.get())}'\n"
        config_data += f"config.dpu.profile='{dslam_profile_var.get()}'\n"
        config_data +=f"config.dpu.libname='{libname_var.get()}'\n"



    # DUT Configuration
        config_data += f"config.dut.name='{dut_name_var.get()}'\n"
        config_data += f"config.dut.busmap='{dut_busmap_var.get()}'\n"
        config_data += f"config.dut.driver='{dut_driver_var.get()}'\n"
        config_data += f"config.dut.serialnumber='{dut_serialnumber_var.get()}'\n"
        config_data += f"config.dut.testplan='{dut_testplan_var.get()}'\n"
        config_data += f"config.dut.folder='{ensure_trailing_backslash(dut_folder_var.get())}'\n"
        config_data += f"config.dut.port={get_safe_int(dut_port_var,0)}\n"
        config_data += f"config.dut.host='{dut_host_var.get()}'\n"
        config_data += f"config.dut.profile='{dut_profile_var.get()}'\n"
        config_data +=f"config.dut.libname='{dut_libname_var.get()}'\n"

    # Validation Settings
        config_data += f"config.validation.max_startup_time={get_safe_int(startup_time_var,0)}\n"
        config_data += f"config.validation.startup_time_min={get_safe_int(min_startup_time_var,0)}\n"
        config_data += f"config.validation.startup_time_max={get_safe_int(max_startup_time_var,0)}\n"
    
        config_data += f"config.validation.stabilize_time={get_safe_int(stabilize_time,0)}\n"
        config_data += f"config.validation.stabilize_time_min={get_safe_int(min_stabilize_time_var,0)}\n"
        config_data += f"config.validation.stabilize_time_max={get_safe_int(max_stabilize_time_var,0)}\n"
        
        config_data += f"config.validation.error_free_time={get_safe_int(error_free_time,0)}\n"
        config_data += f"config.validation.error_free_time_min={get_safe_int(min_error_free_time_var,0)}\n"
        config_data += f"config.validation.error_free_time_max={get_safe_int(max_error_free_time_var,0)}\n"

    # Test Scenario
        config_data += f"config.testscenario='{testscenario_var.get()}'\n"

    # Layout
        config_data += f"config.layout='{layout_var.get()}'\n"

        config_data += f"config.operator='{operator_var.get()}'\n"

    # Platform Configuration
        config_data += f"config.platform.name='{platform_name_var.get()}'\n"
        config_data += f"config.platform.linesimulator='{platform_linesimulator_var.get()}'\n"
        config_data += f"config.platform.noise_generator='{platform_noise_generator_var.get()}'\n"
        config_data += f"config.platform.switch='{platform_switch_var.get()}'\n"
        config_data += f"config.platform.traffic_generator='{platform_traffic_generator_var.get()}'\n"

    # Database Configuration
        config_data += f"config.database.file='{database_file_var.get()}'\n"
        
        folder_path = output_folder_var.get().rstrip('\\')
        config_data += f"config.output.folder='{folder_path}\\'\n"
    # Output Results
        #config_data += f"config.output.folder='{output_folder_var.get().rstrip('\\') + '\\'}'\n"
        config_data += f"config.output.filename='{output_filename_var.get()}'\n"

        config_data += "[\\config]\n"
        # Handle file operations
        if os.path.exists(current_tp_file):
            # Read existing file content
            with open(current_tp_file, 'r') as file:
                content = file.read()
            
            # Find config section boundaries
            config_start = content.find("[config]")
            config_end = content.find("[\\config]")
            
            if config_start != -1 and config_end != -1:
                # Replace only the config section while preserving everything else
                new_content = (
                    content[:config_start] +  # Keep everything before [config]
                    config_data +             # Insert new config section
                    content[config_end + len("[\\config]"):]  # Keep everything after [\\config]
                )
            else:
                # If no config section exists, add it at the beginning
                new_content = config_data + content
            
            # Write updated content
            with open(current_tp_file, 'w') as file:
                file.write(new_content)
        else:
            # Create new file with config section and empty test section
            with open(current_tp_file, 'w') as file:
                file.write(config_data)
                file.write("\n[test]\n[\\test]")

        #messagebox.showinfo("Success", f"Test Plan saved successfully to {os.path.basename(current_tp_file)}")
        thread_safe_display_message(f"\nTest Plan saved successfully to {os.path.basename(current_tp_file)}\n","black")
        update_status_2(f"Saved: {os.path.basename(current_tp_file)}")
        update_file_status(source_page='2')
        return True
        
    except Exception as e:
        update_status_2(f"Error saving test plan: {str(e)}", True)
        update_file_status(str(e), True, source_page='2')
        return False

def browse_file(entry_widget, var=None):
    """
    Enhanced file browse function that checks for existing path in entry widget
    and uses it as initial directory if valid.
    """
    try:
        # Get current path from entry widget or variable
        current_path = var.get() if var else entry_widget.get()
        print(f"DEBUG: Current path from {'variable' if var else 'entry'}: {current_path}")
        
        # Determine initial directory
        initial_dir = os.path.expanduser('~')  # Default to home directory
        print(f"DEBUG: Default initial directory: {initial_dir}")
        
        try:
            if current_path:
                print(f"DEBUG: Processing current path: {current_path}")
                if current_path.startswith('\\\\'): # Network path
                    # Convert network path to a format that tkinter can understand
                    initial_dir = os.path.dirname(current_path).replace('\\', '/')
                    print(f"DEBUG: Network path detected, converted to: {initial_dir}")
                elif current_path.startswith('..'): # Relative path
                    print("DEBUG: Relative path detected, resolving...")
                    pwd_process = subprocess.Popen('cd', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                    current_dir, _ = pwd_process.communicate()
                    if current_dir:
                        full_path = os.path.abspath(os.path.join(current_dir.strip(), current_path))
                        initial_dir = os.path.dirname(full_path)
                        print(f"DEBUG: Resolved relative path to: {initial_dir}")
                else: # Local path
                    if os.path.exists(os.path.dirname(current_path)):
                        initial_dir = os.path.dirname(current_path)
                        print(f"DEBUG: Using existing local path: {initial_dir}")
            elif os.path.exists('config.ini'):
                config.read('config.ini')
                saved_path = config.get('PREFERENCES', 'LastUsedPath', fallback=os.path.expanduser('~'))
                if os.path.exists(saved_path):
                    initial_dir = saved_path
                    print(f"DEBUG: Using saved path from config: {initial_dir}")
        except Exception as path_error:
            print(f"DEBUG: Path resolution error: {path_error}")
            print(f"DEBUG: Using default home directory: {initial_dir}")
        
        print(f"DEBUG: Opening file dialog at: {initial_dir}")
        
        file_path = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="Select File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            # Convert to Windows path format if needed
            if '/' in file_path:
                file_path = file_path.replace('/', '\\')
                print(f"DEBUG: Converted to Windows path: {file_path}")

            # Update entry widget and/or variable
            if var:
                print(f"DEBUG: Setting StringVar to: {file_path}")
                var.set(file_path)
                print(f"DEBUG: StringVar value after set: {var.get()}")
            else:
                print(f"DEBUG: Updating entry widget with: {file_path}")
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, file_path)
                
    except Exception as e:
        print(f"DEBUG: Browse error: {str(e)}")
        print(f"DEBUG: Full traceback: {traceback.format_exc()}")
        messagebox.showerror("Error", f"Failed to browse: {str(e)}")





def load_config_data_page2(file_path):
    try:
        def load_and_update(file_path):
            """Load config data from file path to page 2"""
            try:
                # Load settings and update GUI
                config = load_settings(file_path)
                root.after(0, lambda: update_gui_fields(config))
                
                # Update status
                root.after(0, lambda: status_label.config(text=f"Current file: {os.path.basename(file_path)}"))
                root.after(0, lambda: update_status_2(f"Config loaded: {os.path.basename(file_path)}"))
                
            except Exception as e:
                root.after(0, lambda: messagebox.showerror("Error", f"Failed to load config: {str(e)}"))
                root.after(0, lambda: update_status_2(f"Error loading config: {str(e)}", True))
        
        threading.Thread(target=load_and_update, args=(file_path,), daemon=True).start()
        
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("Error", f"Failed to start config loading: {str(e)}"))
        root.after(0, lambda: update_status_2(f"Error starting config load: {str(e)}", True))

def get_safe_int(var, default=0):
    """Safely get integer value from a tkinter variable"""
    try:
        value = var.get()
        return int(value) if value is not None else default
    except (ValueError, tk.TclError):
        return default

# Function to save the layout configuration
def save_layout_config():
    layout = layout_var.get()
    config['Config'] = {'layout': layout}
    with open('config.ini', 'w') as configfile:
        config.write(configfile)
    display_message(f"Layout configuration saved: {layout}", "black")

def edit_test_plan(tp_file):
    global current_tp_file
    
    try:
        # Check server connection only when performing operations
        if not sock:
            # During initialization, silently return
            if not hasattr(edit_test_plan, 'initialized'):
                edit_test_plan.initialized = True
                return
            # After initialization, show error
            #messagebox.showerror("Error", "Please connect to server first")
            #return

        # file_path = filedialog.askopenfilename(
        #     filetypes=[("Test Plan Files", "*.tp")],
        #     title="Open Test Plan"
        # )
        
        # if not file_path:  # User cancelled
        #     return
        file_path = tp_file
        # Read file content for validation
        with open(file_path, 'r') as f:
            content = f.read()
            
        # Load settings without requiring specific sections
        config = load_settings(file_path)
        if config:
            update_gui_fields(config)
        else:
            # Just log the info, don't show error
            print("No config section found or empty config")
        
        # Update current file and status AFTER successful load
        current_tp_file = file_path
        save_test_plan.current_tp_file = file_path
        status_label.config(text=f"Current file: {os.path.basename(file_path)}")

        #messagebox.showinfo("Success", "Test Plan loaded successfully for editing.")
        update_status_2(f"Editing: {os.path.basename(file_path)}")
        update_file_status(source_page='2')  # Update status after successful load
        
    except Exception as e:
        current_tp_file = None  # Reset on error
        messagebox.showerror("Error", f"Failed to load test plan: {str(e)}")
        update_status_2(f"Error loading test plan: {str(e)}", True)
        update_file_status(str(e), True, source_page='2')

# Add the new save_as_test_plan function:
def save_as_test_plan(source_page='2'):
    global current_tp_file
    """Save config values to a new file, overwriting any existing content"""
    try:
        # if not check_socket(silent=True):
        #     return False
                # Get test data from original file if it exists
        test_data = ""
        # Store the original file path and name FIRST
        original_file = current_tp_file
        original_filename = os.path.basename(original_file) if original_file else None

        config_data = "[config]\n"
    
        # DSLAM Configuration
        config_data += f"config.dpu.name='{dslam_name_var.get()}'\n"
        config_data += f"config.dpu.serialnumber='{dslam_serialnumber_var.get()}'\n"  # Add serial number
        busmap_value = dslam_busmap_var.get().split()[-1] if ' ' in dslam_busmap_var.get() else dslam_busmap_var.get()
        config_data += f"config.dpu.busmap='{busmap_value}'\n"
        config_data += f"config.dpu.driver='{dslam_driver_var.get()}'\n"
        config_data += f"config.dpu.port={get_safe_int(dslam_port_var,0)}\n"
        config_data += f"config.dpu.slot={get_safe_int(dslam_slot_var,0)}\n"
        config_data += f"config.dpu.frame={get_safe_int(dslam_frame_var,0)}\n"
        config_data += f"config.dpu.host='{dslam_host_var.get()}'\n"
        config_data += f"config.dpu.folder='{ensure_trailing_backslash(dslam_folder_var.get())}'\n"
        config_data += f"config.dpu.profile='{dslam_profile_var.get()}'\n"
        config_data +=f"config.dpu.libname='{libname_var.get()}'\n"

        # DUT Configuration
        config_data += f"config.dut.name='{dut_name_var.get()}'\n"
        config_data += f"config.dut.busmap='{dut_busmap_var.get()}'\n"
        config_data += f"config.dut.driver='{dut_driver_var.get()}'\n"
        config_data += f"config.dut.serialnumber='{dut_serialnumber_var.get()}'\n"
        config_data += f"config.dut.testplan='{dut_testplan_var.get()}'\n"
        config_data += f"config.dut.folder='{ensure_trailing_backslash(dut_folder_var.get())}'\n"
        config_data += f"config.dut.port={get_safe_int(dut_port_var,0)}\n"
        config_data += f"config.dut.host='{dut_host_var.get()}'\n"
        config_data += f"config.dut.profile='{dut_profile_var.get()}'\n"
        config_data +=f"config.dut.libname='{dut_libname_var.get()}'\n"

        # Validation Settings
        config_data += f"config.validation.max_startup_time={get_safe_int(startup_time_var,0)}\n"
        config_data += f"config.validation.startup_time_min={get_safe_int(min_startup_time_var,0)}\n"
        config_data += f"config.validation.startup_time_max={get_safe_int(max_startup_time_var,0)}\n"
    
        config_data += f"config.validation.stabilize_time={get_safe_int(stabilize_time,0)}\n"
        config_data += f"config.validation.stabilize_time_min={get_safe_int(min_stabilize_time_var,0)}\n"
        config_data += f"config.validation.stabilize_time_max={get_safe_int(max_stabilize_time_var,0)}\n"
    
        config_data += f"config.validation.error_free_time={get_safe_int(error_free_time,0)}\n"
        config_data += f"config.validation.error_free_time_min={get_safe_int(min_error_free_time_var,0)}\n"
        config_data += f"config.validation.error_free_time_max={get_safe_int(max_error_free_time_var,0)}\n"

        # Test Scenario
        config_data += f"config.testscenario='{testscenario_var.get()}'\n"

        # Layout
        config_data += f"config.layout='{layout_var.get()}'\n"
        config_data += f"config.operator='{operator_var.get()}'\n"

        # Platform Configuration
        config_data += f"config.platform.name='{platform_name_var.get()}'\n"
        config_data += f"config.platform.linesimulator='{platform_linesimulator_var.get()}'\n"
        config_data += f"config.platform.noise_generator='{platform_noise_generator_var.get()}'\n"
        config_data += f"config.platform.switch='{platform_switch_var.get()}'\n"
        config_data += f"config.platform.traffic_generator='{platform_traffic_generator_var.get()}'\n"

        # Database Configuration
        config_data += f"config.database.file='{database_file_var.get()}'\n"

        # Output Results
        folder_path = output_folder_var.get().rstrip('\\')
        config_data += f"config.output.folder='{folder_path}\\'\n"
        #config_data += f"config.output.folder='{output_folder_var.get().rstrip('\\') + '\\'}'\n"
        config_data += f"config.output.filename='{output_filename_var.get()}'\n"

        config_data += "[\\config]"
        # Show save dialog
        file_path = filedialog.asksaveasfilename(
            defaultextension=".tp",
            filetypes=[("Test Plan Files", "*.tp")],
            title="Save Test Plan As"
        )
        

        if not file_path:  # User cancelled
            return
        # Get test data from original file if it exists
        test_data = ""
        if original_file and os.path.exists(original_file):  # Check OLD file
            with open(original_file, 'r') as file:  # Read from OLD file
                content = file.read()
                test_start = content.find("[test]")
                test_end = content.find("[\\test]")
                if test_start != -1 and test_end != -1:
                    test_data = content[test_start:test_end + 7]

        # NOW update current_tp_file to new path
        current_tp_file = file_path
        

        # Write to new file
        with open(file_path, 'w') as file:
            file.write(config_data)  # Write config section
            if test_data:  # Write existing test data if available
                file.write("\n" + test_data)
            else:  # Write empty test section
                file.write("\n[test]\n[\\test]")
            
        # Show appropriate success message
        if original_filename:
            messagebox.showinfo("Success", 
                f"Test Plan saved to {os.path.basename(file_path)}\nTest data copied from {original_filename}")
        else:
            messagebox.showinfo("Success", 
                f"Test Plan saved to {os.path.basename(file_path)}")
        update_status_2(f"Saved: {os.path.basename(file_path)}")
        update_file_status(source_page='2')  # Update status after successful save
        return True
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save test plan: {str(e)}")
        update_status_2(f"Error saving test plan: {str(e)}", True)
        update_file_status(str(e), True, source_page='2')
        return False

page2_functions['save_as_test_plan'] = lambda: save_as_test_plan('2')  # Maps to wrapper that creates thread
#page2_functions['save_as_test_plan_o2'] = save_as_test_plan_o2_thread  # Maps to actual implementation

# def sync_save_as_test_plan(source_page='2'):
#     """Synchronized save-as function that coordinates between pages"""
#     save_event = threading.Event()
#     def on_save_complete():
#         save_event.set()
#     try:
#         if not sock:
#             if not hasattr(sync_save_as_test_plan, 'initialized'):
#                 sync_save_as_test_plan.initialized = True
#                 return False
#             root.after(0, lambda: messagebox.showerror("Error", "Please connect to server first"))
#             return False
#         if source_page == '3':
#             # First do page 3's save as
#             if not page3_functions['save_as_test_plan_o3_thread'](on_complete=on_save_complete):
#                 return False
#             save_event.wait()
                
#             # If config section is empty, save config from page 2
#             with open(current_tp_file, 'r') as f:
#                 content = f.read()
#                 if '[config]\n[\\config]' in content:
#                     # Save current config data from page 2
#                     test_data = content[content.find('[\\config]')+9:]
#                     page2_functions['save_as_test_plan_o2_thread'](on_complete=on_save_complete)
#                     save_event.wait()
#                     # Restore test data
#                     with open(current_tp_file, 'a') as f:
#                         f.write(test_data)
                        
#         else:  # source_page == '2'
#             # First do page 2's save as
#             if not page2_functions['save_as_test_plan_o2_thread'](on_complete=on_save_complete):
#                 return False
                
#             # If test section is empty, get and save test data from page 3
#             with open(current_tp_file, 'r') as f:
#                 content = f.read()
#                 if '[test]\n[\\test]' in content:
#                     test_rows = page3_functions['get_all_test_data']()
#                     if test_rows:
#                         with open(current_tp_file, 'a') as f:
#                             f.write('[test]\n')
#                             for row in test_rows:
#                                 f.write(f"{row}\n")
#                             f.write('[\\test]')
        
#         return True

#     except Exception as e:
#         error_msg = f"Error in sync save: {str(e)}"
#         messagebox.showerror("Error", error_msg)
#         if source_page == '2':
#             update_status_2(error_msg, True)
#         else:
#             update_status_3(error_msg, True)
#         return False

# # Update function references for both pages
# page2_functions['save_as_test_plan'] = lambda: sync_save_as_test_plan('2')
# page3_functions['save_as_test_plan'] = lambda: sync_save_as_test_plan('3')
            




# def sync_save_test_plan(source_page='2'):
#     """Synchronized save function that coordinates between page 2 and 3"""
#     try:
#         if not check_socket(silent=True):
#             return False

#         if not current_tp_file:
#             return save_as_test_plan('2')

#         success = False
        
#         # Save config data from page 2
#         if source_page == '2':
#             # First save config data
#             success = page2_functions['save_test_plan']()
#             if success and 'save_test_plan' in page3_functions:
#                 # Then save test data from page 3
#                 threading.Thread(
#                     target=lambda: page3_functions['save_test_plan'](),
#                     daemon=True
#                 ).start()
                
#         # Save test data from page 3
#         elif source_page == '3':
#             # First preserve config data
#             config_data = ""
#             if os.path.exists(current_tp_file):
#                 with open(current_tp_file, 'r') as file:
#                     content = file.read()
#                     config_start = content.find("[config]")
#                     config_end = content.find("[\\config]")
#                     if config_start != -1 and config_end != -1:
#                         config_data = content[config_start:config_end + len("[\\config]")]
            
#             # Then save test data
#             success = page3_functions['save_test_plan']()
            
#             # Finally restore config data if needed
#             if success and config_data:
#                 with open(current_tp_file, 'r') as file:
#                     content = file.read()
#                 config_start = content.find("[config]")
#                 config_end = content.find("[\\config]")
#                 if config_start != -1 and config_end != -1:
#                     new_content = (
#                         content[:config_start] +
#                         config_data +
#                         content[config_end + len("[\\config]"):]
#                     )
#                     with open(current_tp_file, 'w') as file:
#                         file.write(new_content)

#         # Update status and UI
#         if success:
#             thread_safe_display_message(
#                 f"\nTest Plan saved successfully to {os.path.basename(current_tp_file)}\n",
#                 "black"
#             )
#             update_status_2(f"Saved: {os.path.basename(current_tp_file)}")
#             update_status_3(f"Saved: {os.path.basename(current_tp_file)}")
#             update_file_status(source_page=source_page)
        
#         return success

#     except Exception as e:
#         error_msg = f"Error in synchronized save: {str(e)}"
#         print(error_msg)
#         if source_page == '2':
#             update_status_2(error_msg, True)
#         else:
#             update_status_3(error_msg, True)
#         update_file_status(str(e), True, source_page=source_page)
#         return False

def browse_save_dslam_folder():
    """Browse and select a DSLAM folder"""
    try:
        # Get current folder path if exists
        current_path = dslam_folder_var.get()
        initial_dir = None

        if current_path:
            if current_path.startswith('\\\\'): # Network path
                # Convert network path to a format that tkinter can understand
                initial_dir = current_path.replace('\\', '/')
            elif current_path.startswith('..'): # Relative path
                # Try to resolve relative path
                pwd_process = subprocess.Popen('cd', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                current_dir, _ = pwd_process.communicate()
                if current_dir:
                    initial_dir = os.path.abspath(os.path.join(current_dir.strip(), current_path))
            else: # Local path
                initial_dir = current_path

        # Open folder dialog with initial directory if available
        folder_path = filedialog.askdirectory(
            title="Select DSLAM Folder",
            initialdir=initial_dir
        )
        
        if folder_path:
            # Convert back to Windows path format if needed
            if '/' in folder_path:
                folder_path = folder_path.replace('/', '\\')
            
            folder_path = ensure_trailing_backslash(folder_path)
            
            # Update both the StringVar and Entry widget
            dslam_folder_var.set(folder_path)
            
            # Force update the entry widget
            dslam_folder_entry.delete(0, tk.END)
            dslam_folder_entry.insert(0, folder_path)
            
            # Automatically get DSLAM files after setting the folder
            get_dslam_folder_files()
            
    except Exception as e:
        messagebox.showerror("Error", f"Failed to browse folder: {str(e)}")
        print(f"Error in browse_save_dslam_folder: {traceback.format_exc()}")

def get_dslam_folder_files():
    """Get list of files from the selected DSLAM folder"""
    try:
        folder_path = dslam_folder_var.get()
        
        if not folder_path:
            messagebox.showwarning("Warning", "Please select a DSLAM folder first")
            return
            
        if os.path.exists(folder_path):
            # Get all files in the folder
            files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
            
            if files:
                # Update the dropdown
                if dropdowns['DSLAM_Profile']['widget']:
                    dropdowns['DSLAM_Profile']['widget']['values'] = files
                    dropdowns['DSLAM_Profile']['widget'].set('')  # Clear current selection
                messagebox.showinfo("Success", f"Found {len(files)} files in DSLAM folder")
            else:
                messagebox.showinfo("Info", "No files found in selected folder")
                
        else:
            messagebox.showwarning("Warning", f"Selected DSLAM folder not found: {folder_path}")
            
    except Exception as e:
        messagebox.showerror("Error", f"Failed to get DSLAM files list: {str(e)}")
        print(f"Error in get_dslam_folder_files: {traceback.format_exc()}")

def on_dslam_file_select(event):
    """Handle DSLAM file selection from listbox"""
    try:
        selection = dslam_profile_dropdown.curselection()
        if selection:
            file_name = dslam_profile_dropdown.get(selection[0])
            folder_path = dslam_folder_var.get()
            full_path = os.path.join(folder_path, file_name)
            
            # Update the profile variable with the selected file
            dslam_profile_var.set(file_name)
            
            # Update the dropdown if it exists
            if 'DSLAM_Profile' in dropdowns and dropdowns['DSLAM_Profile']['widget']:
                dropdowns['DSLAM_Profile']['widget'].set(file_name)
                
    except Exception as e:
        messagebox.showerror("Error", f"Failed to select DSLAM file: {str(e)}")
        print(f"Error in on_dslam_file_select: {traceback.format_exc()}")

def load_dslam_profile():
    """Load the selected DSLAM file"""
    try:
        selection = dslam_profile_dropdown.curselection()
        if selection:
            file_name = dslam_profile_dropdown.get(selection[0])
            folder_path = dslam_folder_var.get()
            full_path = os.path.join(folder_path, file_name)
            
            if os.path.exists(full_path):
                # Set the profile name in the variable and dropdown
                dslam_profile_var.set(file_name)
                if 'DSLAM_Profile' in dropdowns and dropdowns['DSLAM_Profile']['widget']:
                    dropdowns['DSLAM_Profile']['widget'].set(file_name)
            
                
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load DSLAM file: {str(e)}")
        print(f"Error in load_dslam_profile: {traceback.format_exc()}")

# Function to load and populate dropdowns
def load_and_populate_dropdowns(file_path, sheet_name, dropdowns, *vars):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    dropdown_data = {}  # Dictionary to store DataFrame for each dropdown
    print(f"Loading data from sheet: {sheet_name}")  # Debugging statement
    print(df.head())  # Print the first few rows of the DataFrame for debugging

    def process_values(values_list, column_data):
        """Helper function to process values and handle None/NaN cases"""
        # Convert all values to strings for case-insensitive comparison
        str_values = [str(v).lower() if v is not None else 'none' for v in values_list]
        # Check if any variation of None/NaN exists
        none_variants = {'none', 'nan', ''}
        if any(v in none_variants for v in str_values):
            if 'None' not in values_list:  # Add 'None' if not already present
                values_list.append('None')
        return list(dict.fromkeys(values_list))  # Remove duplicates while preserving order
    dropdown_data[sheet_name] = df  # Store the entire DataFrame for the sheet

    if sheet_name == 'DSLAM_serialnumber':
        try:
            # Find the first non-NaN values for Max_Port and Min_Port
            max_port = df['Max_Port'].dropna().iloc[0]
            min_port = df['Min_Port'].dropna().iloc[0]
            

            # Generate list of port values from min to max (inclusive)
            port_values = [str(int(i)) for i in range(int(min_port), int(max_port) + 1)]
                    
            # Update the dropdown with the port values
            dslam_port_var.set(port_values[0] if port_values else "None")
            dropdowns['DSLAM_Port']['widget']['values'] = port_values
            print(f"Port values loaded: {port_values}")  # Debug print
            dropdown_data['DSLAM_Port'] = df
        except Exception as e:
            print(f"Error loading DSLAM port values: {str(e)}")
            # Fallback to default values if there's an error
            port_values = ['None', '0', '1', '2', '3']
            dslam_port_var.set(port_values[0])
            dropdowns['DSLAM_Port']['widget']['values'] = port_values

        # Continue with existing DSLAM_serialnumber processing
        for column in df.columns:
            values = df[column].replace({pd.NA: None, pd.NaT: None}).dropna().tolist()
            
            if column == 'DSLAM_Name':
                values = process_values(values, df[column])
                dropdowns['DSLAM_Name']['widget']['values'] = values
                dslam_name_var.set(values[0] if values else "None")
                dropdown_data['DSLAM_Name'] = df
            
            elif column == 'DSLAM_busmap':
                busmap_values = process_values(values, df[column])
                dslam_busmap_var.set(busmap_values[0] if busmap_values else "None")
                dropdowns['DSLAM_Busmap']['widget']['values'] = busmap_values
                dropdown_data['DSLAM_Busmap'] = df
            elif column == 'DSLAM_Serialnumber':
                serialnumber_values = process_values(values, df[column])
                dslam_serialnumber_var.set(serialnumber_values[0] if serialnumber_values else "None")
                dropdowns['DSLAM_Serialnumber']['widget']['values'] = serialnumber_values
                dropdown_data['DSLAM_Serialnumber'] = df


    


    elif sheet_name == 'operator':
        first_col = df.iloc[:,0]
        operator_values = first_col.replace({pd.NA: None, pd.NaT: None}).dropna().tolist()
        operator_values = process_values(operator_values, first_col)

        default_value = operator_values[0] if operator_values else "Tester"
        operator_var.set(default_value)

        if 'Operator' in dropdowns and dropdowns['Operator']['widget']:
            dropdowns['Operator']['widget']['values'] = operator_values
            dropdowns['Operator']['widget'].set(default_value)

        dropdown_data['Operator'] = df
        print(f"Operator values loaded : {operator_values}")       


    elif sheet_name == 'DUT_serialnumber':
        try:
            # Find the first non-NaN values for Max_Port and Min_Port
            max_port = df['Max_Port'].dropna().iloc[0]
            min_port = df['Min_Port'].dropna().iloc[0]
            
            # Generate list of port values from min to max (inclusive)
            port_values = [str(int(i)) for i in range(int(min_port), int(max_port) + 1)]
            
            # Add 'None' as an option at the beginning
            port_values = ['None'] + port_values
            
            # Update the dropdown with the port values
            dut_port_var.set(port_values[0] if port_values else "None")
            dropdowns['DUT_Port']['widget']['values'] = port_values
            print(f"DUT Port values loaded: {port_values}")  # Debug print

            dropdown_data['DUT_Port'] = df
            
        except Exception as e:
            print(f"Error loading DUT port values: {str(e)}")
            # Fallback to default values if there's an error
            port_values = ['None', '0', '1', '2', '3']
            dut_port_var.set(port_values[0])
            dropdowns['DUT_Port']['widget']['values'] = port_values

        # Continue with existing DUT_serialnumber processing
        for column in df.columns:
            values = df[column].replace({pd.NA: None, pd.NaT: None}).dropna().tolist()
            
            if column == 'DUT_Name':
                values = process_values(values, df[column])
                dropdowns['DUT_Name']['widget']['values'] = values
                dut_name_var.set(values[0] if values else "None")
                dropdown_data['DUT_Name'] = df            
            elif column == 'DUT_busmap':
                busmap_values = process_values(values, df[column])
                dut_busmap_var.set(busmap_values[0] if busmap_values else "None")
                dropdowns['DUT_Busmap']['widget']['values'] = busmap_values
                dropdown_data['DUT_Busmap'] = df
            elif column == 'DUT_Serialnumber':
                serialnumber_values = process_values(values, df[column])
                dut_serialnumber_var.set(serialnumber_values[0] if serialnumber_values else "None")
                dropdowns['DUT_Serialnumber']['widget']['values'] = serialnumber_values
                dropdown_data['DUT_Serialnumber'] = df


    elif sheet_name == 'DUT_testplan':
        test_plan_values = df['Test Plan'].replace({pd.NA: None, pd.NaT: None}).dropna().tolist()
        test_plan_values = process_values(test_plan_values, df['Test Plan'])
        dut_testplan_var.set(test_plan_values[0] if test_plan_values else "None")
        dropdowns['DUT_Testplan']['widget']['values'] = test_plan_values
        dropdown_data['Test Plan'] = df

    elif sheet_name == 'DUT_host':

        try:
            column_name = 'Host' if 'Host' in df.columns else 'host'
            # Check if the 'host' column exists
            if column_name in df.columns:
                # Get all host values from the single column
                host_values = df[column_name].replace({pd.NA: 'None', pd.NaT: 'None'}).tolist()

                # Process values and ensure 'None' is at the beginning
                if 'None' not in host_values:
                    host_values = ['None'] + [v for v in host_values if v != 'None' and pd.notna(v)]
                else:
                    host_values = [v for v in host_values if pd.notna(v)]

                # Remove duplicates while preserving order
                host_values = list(dict.fromkeys(host_values))
            
                # Set default values - use the first actual value from the list
                default_value = host_values[1] if len(host_values) > 1 else host_values[0]
                dut_host_var.set(default_value)
                dslam_host_var.set(default_value)
            
                # Update both dropdowns with the same list of values
                if 'DUT_Host' in dropdowns and dropdowns['DUT_Host']['widget']:
                    dropdowns['DUT_Host']['widget']['values'] = host_values
                    if not dropdowns['DUT_Host']['widget'].get():
                         dropdowns['DUT_Host']['widget'].set(dut_host_var.get())
            
                if 'DSLAM_Host' in dropdowns and dropdowns['DSLAM_Host']['widget']:
                    dropdowns['DSLAM_Host']['widget']['values'] = host_values
                    if not dropdowns['DSLAM_Host']['widget'].get():
                         dropdowns['DSLAM_Host']['widget'].set(dslam_host_var.get())            
                # Store dataframe
                dropdown_data['DUT Host'] = df
                dropdown_data['DSLAM Host'] = df

                print(f"Host values loaded : {host_values}")
        except Exception as e:
            print(f"Error processing DUT_host sheet: {str(e)}")

    elif sheet_name == 'Test_Scenario':
        test_scenario_values = df['Test_Scenario'].replace({pd.NA: None, pd.NaT: None}).dropna().tolist()
        test_scenario_values = process_values(test_scenario_values, df['Test_Scenario'])
        testscenario_var.set(test_scenario_values[0] if test_scenario_values else "None")
        dropdowns['DUT_Testscenario']['widget']['values'] = test_scenario_values
        dropdown_data['Test_Scenario'] = df

    elif sheet_name == 'Validation_Fields':
        # Extract min and max values for Validation_startup
        min_value = df.loc[df['Validation_Fields'] == 'Validation_startup', 'Min'].values[0]
        max_value = df.loc[df['Validation_Fields'] == 'Validation_startup', 'Max'].values[0]
        min_startup_time_var.set(min_value)
        max_startup_time_var.set(max_value)
        # Extract min and max values for Validation_stabilize
        min_value = df.loc[df['Validation_Fields'] == 'Validation_stabilize', 'Min'].values[0]
        max_value = df.loc[df['Validation_Fields'] == 'Validation_stabilize', 'Max'].values[0]
        min_stabilize_time_var.set(min_value)
        max_stabilize_time_var.set(max_value)
        # Extract min and max values for Validation_error
        min_value = df.loc[df['Validation_Fields'] == 'Validation_error', 'Min'].values[0]
        max_value = df.loc[df['Validation_Fields'] == 'Validation_error', 'Max'].values[0]
        min_error_free_time_var.set(min_value)
        max_error_free_time_var.set(max_value)


    elif sheet_name == 'Platform_data':
        for column in df.columns:
            values = df[column].replace({pd.NA: None, pd.NaT: None}).dropna().tolist()
            
            if column == 'Platform_name':
                values = process_values(values, df[column])
                platform_name_var.set(values[0] if values else "None")
                dropdowns['Platform_name']['widget']['values'] = values
                dropdown_data['Platform_name'] = df
            
            elif column == 'Linesimulator':
                linesimulator_values = process_values(values, df[column])
                platform_linesimulator_var.set(linesimulator_values[0] if linesimulator_values else "None")
                dropdowns['Linesimulator']['widget']['values'] = linesimulator_values
                dropdown_data['Linesimulator'] = df
            
            elif column == 'NoiseGenerator':
                noise_generator_values = process_values(values, df[column])
                platform_noise_generator_var.set(noise_generator_values[0] if noise_generator_values else "None")
                dropdowns['NoiseGenerator']['widget']['values'] = noise_generator_values
                dropdown_data['NoiseGenerator'] = df
            
            elif column == 'Switch':
                switch_values = process_values(values, df[column])
                platform_switch_var.set(switch_values[0] if switch_values else "None")
                dropdowns['Switch']['widget']['values'] = switch_values
                dropdown_data['Switch'] = df
            
            elif column == 'trafficgenerator':
                traffic_generator_values = process_values(values, df[column])
                platform_traffic_generator_var.set(traffic_generator_values[0] if traffic_generator_values else "None")
                dropdowns['trafficgenerator']['widget']['values'] = traffic_generator_values
                dropdown_data['trafficgenerator'] = df
        
        dropdown_data['Platform_data'] = df

    elif sheet_name == 'Platform_Fields':
        for column in df.columns:
            values = df[column].replace({pd.NA: None, pd.NaT: None}).dropna().tolist()
            
            if column == 'Switch':
                switch_values = process_values(values, df[column])
                platform_switch_var.set(switch_values[0] if switch_values else "None")
                dropdowns['Switch']['widget']['values'] = switch_values
                dropdown_data['Switch'] = df
            
            elif column == 'NoiseGenerator':
                noise_generator_values = process_values(values, df[column])
                platform_noise_generator_var.set(noise_generator_values[0] if noise_generator_values else "None")
                dropdowns['NoiseGenerator']['widget']['values'] = noise_generator_values
                dropdown_data['NoiseGenerator'] = df
            
            elif column == 'trafficgenerator':
                traffic_generator_values = process_values(values, df[column])
                platform_traffic_generator_var.set(traffic_generator_values[0] if traffic_generator_values else "None")
                dropdowns['trafficgenerator']['widget']['values'] = traffic_generator_values
                dropdown_data['trafficgenerator'] = df
            
            elif column == 'Linesimulator':
                linesimulator_values = process_values(values, df[column])
                platform_linesimulator_var.set(linesimulator_values[0] if linesimulator_values else "None")
                dropdowns['Linesimulator']['widget']['values'] = linesimulator_values
                dropdown_data['Linesimulator'] = df
        
        dropdown_data['Platform_Fields'] = df

    return dropdown_data


# Function to Update GUI Fields for the Platform dropdowns
def update_platform_dropdowns(dropdown_data, selected_platform, platform_name_dropdown, platform_linesimulator_var, platform_noise_generator_var, platform_switch_var, platform_traffic_generator_var):
    try:
        df = dropdown_data.get('Platform_data')  # Retrieve the DataFrame from the dictionary
        if df is None:
            print("Platform_data not found in dropdown_data")
            return
            
        # Print available columns for debugging
        print(f"Available columns: {df.columns.tolist()}")
            
        filtered_df = df[df['Platform_name'] == selected_platform]
        if not filtered_df.empty:
            # Update the platform name in the dropdown
            platform_name_dropdown.set(selected_platform)
            
            # Try different possible column names for traffic generator with case-insensitive matching
            traffic_gen_col = None
            for col in df.columns:
                if col.lower() == 'trafficgenerator':
                    traffic_gen_col = col
                    break
                    
            if traffic_gen_col is None:
                print("Warning: TrafficGenerator column not found. Using 'TrafficGenerator' as fallback.")
                traffic_gen_col = 'TrafficGenerator'
            
            # Get values from DataFrame and handle NaN values
            linesimulator = filtered_df['Linesimulator'].values[0] if 'Linesimulator' in filtered_df else None
            noise_generator = filtered_df['NoiseGenerator'].values[0] if 'NoiseGenerator' in filtered_df else None
            switch = filtered_df['Switch'].values[0] if 'Switch' in filtered_df else None
            
            # Use the found column name or handle missing column
            try:
                traffic_generator = filtered_df[traffic_gen_col].values[0]
            except KeyError:
                print(f"Column '{traffic_gen_col}' not found. Setting empty value.")
                traffic_generator = None

            print(f"Linesimulator: {linesimulator}, NoiseGenerator: {noise_generator}, Switch: {switch}, trafficgenerator: {traffic_generator}")

            # Set values, replacing NaN with empty string
            
            platform_linesimulator_var.set("" if pd.isna(linesimulator) else linesimulator)
            platform_noise_generator_var.set("" if pd.isna(noise_generator) else noise_generator)
            platform_switch_var.set("" if pd.isna(switch) else switch)
            platform_traffic_generator_var.set("" if pd.isna(traffic_generator) else traffic_generator)
        else:
            platform_linesimulator_var.set("")
            platform_noise_generator_var.set("")
            platform_switch_var.set("")
            platform_traffic_generator_var.set("")
            
    except Exception as e:
        print(f"Error updating dropdowns: {str(e)}")
        import traceback
        traceback.print_exc()
        # Still set empty values to avoid UI inconsistency
        platform_linesimulator_var.set("")
        platform_noise_generator_var.set("")
        platform_switch_var.set("")
        platform_traffic_generator_var.set("")
 
def update_dropdowns_from_sheet(file_path, sheet_name, dropdowns, platform_vars):

    try:
        # Load the specified sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(f"Loading data from sheet: {sheet_name}")  # Debugging output

        # Iterate through the columns of the DataFrame
        for column in df.columns:
            values = df[column].dropna().tolist()  # Get non-null values

            # Update the corresponding dropdown if it exists
            if column in dropdowns:
                dropdowns[column]['widget']['values'] = values  # Update dropdown values
                if values:
                    platform_vars[column].set(values[0])  # Set the first value as the default

        print("Dropdowns updated successfully.")  # Debugging output
    except Exception as e:
        print(f"Error updating dropdowns: {str(e)}")  # Error handling            






# Bind the Event for the "Other" Platform Selection
def on_platform_selected(platform_name_dropdown, dropdown_data, platform_name_var, 
                        platform_linesimulator_var, platform_noise_generator_var,
                        platform_traffic_generator_dropdown, platform_switch_var,
                        platform_switch_dropdown, platform_traffic_generator_var,
                        platform_linesimulator_dropdown, platform_noise_generator_dropdown):
    
    try:
        selected_platform = platform_name_var.get()

        if selected_platform == "other":
            # Clear existing values
            platform_linesimulator_var.set("")
            platform_noise_generator_var.set("")
            platform_switch_var.set("")
            platform_traffic_generator_var.set("")

            # Enable the fields
            platform_linesimulator_dropdown.config(state='readonly')
            platform_noise_generator_dropdown.config(state='readonly')
            platform_switch_dropdown.config(state='readonly')
            platform_traffic_generator_dropdown.config(state='readonly')

            file_path = get_excel_file_path()

            # Call function to update dropdowns
            update_dropdowns_from_sheet(file_path, "Platform_Fields", dropdowns, {
                'Linesimulator': platform_linesimulator_var,
                'NoiseGenerator': platform_noise_generator_var,
                'Switch': platform_switch_var,
                'TrafficGenerator': platform_traffic_generator_var
            })
            
            # Add empty option to each dropdown
            for dropdown, var in [
                (platform_linesimulator_dropdown, platform_linesimulator_var),
                (platform_noise_generator_dropdown, platform_noise_generator_var),
                (platform_switch_dropdown, platform_switch_var),
                (platform_traffic_generator_dropdown, platform_traffic_generator_var)
            ]:
                # Get current values and add empty option if not already present
                current_values = list(dropdown['values']) if 'values' in dropdown.keys() else []
                
                # Add empty option at the beginning if not already there
                if '' not in current_values:
                    new_values = [''] + current_values
                    dropdown['values'] = new_values
                    print(f"DEBUG: Added empty option to dropdown. Values: {new_values}")
                
                # Set to empty value
                var.set('')
                
        else:
            # Handle other platform selections - existing code
            platform_linesimulator_var.set("")  # Clear the field
            platform_noise_generator_var.set("")  # Clear the field
            platform_switch_var.set("")  # Clear the field
            platform_traffic_generator_var.set("")  # Clear the field
       
            # Disable the fields
            platform_linesimulator_dropdown.config(state='disabled')
            platform_noise_generator_dropdown.config(state='disabled')
            platform_switch_dropdown.config(state='disabled')
            platform_traffic_generator_dropdown.config(state='disabled')

            file_path = get_excel_file_path()

            # Load values from 'Platform_data' for the selected platform
            dropdown_data.update(load_and_populate_dropdowns(file_path, "Platform_data", dropdowns,
                dut_name_var, dut_testplan_var, dut_host_var, dut_busmap_var,
                dut_serialnumber_var, testscenario_var,
                min_startup_time_var, max_startup_time_var,
                min_error_free_time_var, max_error_free_time_var,
                min_stabilize_time_var, max_stabilize_time_var,
                platform_name_var, platform_linesimulator_var,
                platform_noise_generator_var, platform_switch_var,
                platform_traffic_generator_var))
                
            # Handle other platform selections
            update_platform_dropdowns(dropdown_data, selected_platform, platform_name_dropdown, 
                                    platform_linesimulator_var, platform_noise_generator_var, 
                                    platform_switch_var, platform_traffic_generator_var)
    except FileNotFoundError as e:
        messagebox.showerror("Error", str(e))
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred:\n{str(e)}")
        print(f"DEBUG: Error in on_platform_selected: {str(e)}")
        print(f"DEBUG: Traceback: {traceback.format_exc()}") 

# Dictionary to store dropdowns
dropdowns = {}

def update_dut_serialnumber_dropdown(dropdown_data, selected_dut_serialnumber, dut_serialnumber_dropdown, dut_name_var, dut_busmap_var,dut_libname_var):
    try:
        # Safely get the DataFrame using .get()
        df = dropdown_data.get('DUT_serialnumber')  # Use sheet name as key
        if df is None:
            print("Warning: DUT_serialnumber DataFrame not found in dropdown_data")
            return

        # Filter for the selected serial number
        filtered_df = df[df['DUT_Serialnumber'] == selected_dut_serialnumber]
        print(f"Selected DUT Serial Number: {selected_dut_serialnumber}")  # Debug print
   
        # Update the serial number dropdown
        serial_numbers = df['DUT_Serialnumber'].dropna().tolist()
        dut_serialnumber_dropdown['values'] = serial_numbers
        if serial_numbers:
            dut_serialnumber_dropdown.set(selected_dut_serialnumber)
        else:
            dut_serialnumber_dropdown.set("")
            print("No serial numbers found in DataFrame")

        # Update the dut_name, dut_busmap, and dut_libname based on the selected serial number
        if not filtered_df.empty:
            dut_name = filtered_df['DUT_Name'].values[0]
            dut_busmap = filtered_df['DUT_busmap'].values[0]
            dut_name_var.set(dut_name)
            dut_busmap_var.set(dut_busmap)

            # Update libname 
            if 'libname' in filtered_df.columns:
                dut_libname = filtered_df['libname'].values[0]
                dut_libname_var.set(dut_libname)
                print(f"Updated DUT libname to: {dut_libname}")
            
            # Update port values if Max_Port and Min_Port exist
            if 'Max_Port' in filtered_df.columns and 'Min_Port' in filtered_df.columns:
                try:
                    max_port = filtered_df['Max_Port'].iloc[0]
                    min_port = filtered_df['Min_Port'].iloc[0]

                    if pd.notna(max_port) and pd.notna(min_port):
                        port_values = [str(i) for i in range(int(min_port), int(max_port) + 1)]
                        dropdowns['DUT_Port']['widget']['values'] = port_values
                        dropdowns['DUT_Port']['var'].set(port_values[0])
                        print(f"Updated port values for {selected_dut_serialnumber}: {port_values}")
                except Exception as e:
                    print(f"Error updating port values: {str(e)}")
        else:
            dut_name_var.set("")
            dut_busmap_var.set("")
            dut_libname_var.set("")
            print(f"No matching data found for serial number: {selected_dut_serialnumber}")

    except Exception as e:
        print(f"Error in update_dut_serialnumber_dropdown: {str(e)}")
        # Set default values in case of error
        dut_name_var.set("")
        dut_busmap_var.set("")
        dut_libname_var.set("")


def update_dslam_name_dropdown(data, selected_name, name_dropdown, serialnumber_dropdown, busmap_var, libname_var):
    """Update DSLAM dropdowns (serials, busmap, libname) when a DSLAM name is selected"""
    try:
        if not data or 'DSLAM_serialnumber' not in data:
            return
            
        df = data['DSLAM_serialnumber']
        if selected_name:
            # Filter rows matching the selected DSLAM name
            matching_rows = df[df['DSLAM_Name'] == selected_name]
            if not matching_rows.empty:
                # Update serial numbers
                matching_serials = matching_rows['DSLAM_Serialnumber'].dropna().tolist()
                serialnumber_dropdown['values'] = matching_serials
                
                # Update busmap (take first match)
                matching_busmap = matching_rows['DSLAM_busmap'].iloc[0]
                if matching_busmap:
                    busmap_var.set(matching_busmap)

                # Update libname (new requirement)
                if 'libname' in matching_rows.columns:
                    matching_libname = matching_rows['libname'].iloc[0]
                    if matching_libname:
                        libname_var.set(matching_libname)
                        print(f"Updated libname to: {matching_libname}")
                
    except Exception as e:
        print(f"Error updating DSLAM dropdowns: {e}")


# def update_dslam_name_dropdown(data, selected_name, name_dropdown, serialnumber_dropdown, busmap_var):
#     """Update DSLAM dropdowns when name is selected"""
#     try:
#         if not data or 'DSLAM_serialnumber' not in data:
#             return
            
#         df = data['DSLAM_serialnumber']
#         if selected_name:
#             # Filter serial numbers for selected name
#             matching_serials = df[df['DSLAM_Name'] == selected_name]['DSLAM_Serialnumber'].tolist()
#             serialnumber_dropdown['values'] = matching_serials
            
#             # Update busmap if there's only one match
#             matching_busmap = df[df['DSLAM_Name'] == selected_name]['DSLAM_busmap'].iloc[0]
#             if matching_busmap:
#                 busmap_var.set(matching_busmap)

#             # Update libname (new requirement)
#             if 'libname' in matching_rows.columns:
#                 matching_libname = matching_rows['libname'].iloc[0]
#                 if matching_libname:
#                     libname_var.set(matching_libname)
#                     print(f"Updated libname to: {matching_libname}")
                
#     except Exception as e:
#         print(f"Error updating DSLAM dropdowns: {e}")
        

def update_dslam_details_with_libname(data, selected_name, name_dropdown, serialnumber_dropdown, busmap_var, libname_var):
    """
    Update DSLAM dropdowns and libname when a DSLAM name is selected.
    - Fills serial numbers, busmap, and libname from the Excel data.
    """
    try:
        if not data or 'DSLAM_serialnumber' not in data:
            return
        
        df = data['DSLAM_serialnumber']
        if selected_name:
            # Filter rows matching the selected DSLAM name
            matching_rows = df[df['DSLAM_Name'] == selected_name]
            if not matching_rows.empty:
                # Update serial numbers
                matching_serials = matching_rows['DSLAM_Serialnumber'].dropna().tolist()
                serialnumber_dropdown['values'] = matching_serials

                # Update busmap (take the first match)
                matching_busmap = matching_rows['DSLAM_busmap'].iloc[0]
                if matching_busmap:
                    busmap_var.set(matching_busmap)

                # Update libname (new requirement)
                if 'libname' in matching_rows.columns:
                    matching_libname = matching_rows['libname'].iloc[0]
                    if matching_libname:
                        libname_var.set(matching_libname)
                        print(f"Updated libname to: {matching_libname}")
    except Exception as e:
        print(f"Error updating DSLAM details with libname: {e}")



def on_dslam_serialnumber_selected(event):
    """Update DSLAM details (name, busmap, ports, libname) when DSLAM serial number is selected."""
    try:
        selected_serial = dslam_serialnumber_var.get()
        print(f"\nDEBUG: Function called with selection: {selected_serial}")
        
        # Get the Excel file path
        file_path = get_excel_file_path()
        
        # Get the Excel data
        df = pd.read_excel(file_path, sheet_name='DSLAM_serialnumber')
        print("\nDEBUG: DataFrame content:")
        print(df[['DSLAM_Name', 'DSLAM_Serialnumber', 'DSLAM_busmap', 'Max_Port', 'Min_Port', 'libname']])
        
        # Find the row matching the selected serial number
        row = df[df['DSLAM_Serialnumber'].astype(str).str.contains(str(selected_serial), na=False)]
        print(f"\nDEBUG: Found matching row for {selected_serial}:")
        print(row)
        if not row.empty:
            try:
                on_dslam_serialnumber_port_update(event)
                
                # Update DSLAM name
                dslam_name = row['DSLAM_Name'].iloc[0]
                dslam_name_var.set(dslam_name)
                print(f"Updated DSLAM name to: {dslam_name}")

                # Update busmap
                dslam_busmap = row['DSLAM_busmap'].iloc[0]
                dslam_busmap_var.set(dslam_busmap)
                print(f"Updated DSLAM busmap to: {dslam_busmap}")

                # Update libname (NEW)
                if 'libname' in row.columns:
                    matching_libname = row['libname'].iloc[0]
                    if matching_libname:
                        libname_var.set(matching_libname)
                        print(f"Updated libname to: {matching_libname}")
                
                # Update port values
                max_port = int(float(row['Max_Port'].iloc[0]))
                min_port = int(float(row['Min_Port'].iloc[0]))
                print(f"\nDEBUG: Min Port: {min_port}, Max Port: {max_port}")
                
                port_values = [str(i) for i in range(min_port, max_port + 1)]
                dropdowns['DSLAM_Port']['widget']['values'] = port_values
                dropdowns['DSLAM_Port']['widget'].set(port_values[0])
                
                print(f"Updated port values for {selected_serial}: {port_values}")
                
            except ValueError as ve:
                print(f"Error converting ports to integers: {ve}")
                initialize_dslam_port_dropdown()
        else:
            print(f"\nDEBUG: No matching row found for serial number: {selected_serial}")
            initialize_dslam_port_dropdown()
            
    except FileNotFoundError as e:
        print(f"Error: Excel file not found - {str(e)}")
        messagebox.showerror("Error", f"Could not find Excel file:\n{str(e)}")
        initialize_dslam_port_dropdown()
    except Exception as e:
        print(f"Error in on_dslam_serialnumber_selected: {str(e)}")
        traceback.print_exc()
        messagebox.showerror("Error", f"An unexpected error occurred:\n{str(e)}")
        initialize_dslam_port_dropdown()


# def on_dslam_serialnumber_selected(event):
#     """Update DSLAM details and port dropdown when DSLAM serial number is selected."""


#     try:
#         selected_serial = dslam_serialnumber_var.get()
#         print(f"\nDEBUG: Function called with selection: {selected_serial}")
        
#         # Get the Excel file path
#         file_path = get_excel_file_path()
        
#         # Get the Excel data
#         df = pd.read_excel(file_path, sheet_name='DSLAM_serialnumber')
#         print("\nDEBUG: DataFrame content:")
#         print(df[['DSLAM_Name', 'DSLAM_Serialnumber', 'DSLAM_busmap', 'Max_Port', 'Min_Port']])
        
#         # Find the row matching the selected serial number
#         row = df[df['DSLAM_Serialnumber'].str.contains(selected_serial, na=False)]
#         print(f"\nDEBUG: Found matching row for {selected_serial}:")
#         print(row)
#         if not row.empty:
#             try:

#                 on_dslam_serialnumber_port_update(event)
#                 # Update DSLAM name and busmap
#                 dslam_name = row['DSLAM_Name'].iloc[0]
#                 dslam_name_var.set(dslam_name)
#                 print(f"Updated DSLAM name to: {dslam_name}")

#                 dslam_busmap = row['DSLAM_busmap'].iloc[0]
#                 dslam_busmap_var.set(dslam_busmap)
#                 print(f"Updated DSLAM busmap to: {dslam_busmap}")
                
#                 # Update port values
#                 max_port = int(float(row['Max_Port'].iloc[0]))
#                 min_port = int(float(row['Min_Port'].iloc[0]))
#                 print(f"\nDEBUG: Min Port: {min_port}, Max Port: {max_port}")
                
#                 # Generate port values from min to max (inclusive)
#                 port_values = [str(i) for i in range(min_port, max_port + 1)]
                
#                 # Update the dropdown with new values
#                 dropdowns['DSLAM_Port']['widget']['values'] = port_values
#                 dropdowns['DSLAM_Port']['widget'].set(port_values[0])
                
#                 print(f"Updated port values for {selected_serial}: {port_values}")
                
#             except ValueError as ve:
#                 print(f"Error converting ports to integers: {ve}")
#                 initialize_dslam_port_dropdown()
#         else:
#             print(f"\nDEBUG: No matching row found for serial number: {selected_serial}")
#             initialize_dslam_port_dropdown()
            
#     except FileNotFoundError as e:
#         print(f"Error: Excel file not found - {str(e)}")
#         messagebox.showerror("Error", f"Could not find Excel file:\n{str(e)}")
#         initialize_dslam_port_dropdown()
#     except Exception as e:
#         print(f"Error in on_dslam_serialnumber_selected: {str(e)}")
#         traceback.print_exc()
#         messagebox.showerror("Error", f"An unexpected error occurred:\n{str(e)}")
#         initialize_dslam_port_dropdown()


def update_dut_name_dropdown(dropdown_data, selected_dut_name, dut_name_dropdown, dut_serialnumber_dropdown, dut_busmap_var, dut_libname_var):
    df = dropdown_data['DUT_Serialnumber']  # Retrieve the DataFrame from the dictionary
    filtered_df = df[df['DUT_Name'] == selected_dut_name]
   
    # Update the name dropdown
    names = df['DUT_Name'].dropna().unique().tolist()
    dut_name_dropdown['values'] = names
    if names:
        dut_name_dropdown.set(selected_dut_name)
    else:
        dut_name_dropdown.set("")
   
    # Update the serial number dropdown based on the selected name
    serial_numbers = filtered_df['DUT_Serialnumber'].dropna().tolist()
    dut_serialnumber_dropdown['values'] = serial_numbers
    if serial_numbers:
        dut_serialnumber_dropdown.set(serial_numbers[0])
        # Automatically update the busmap and libname based on the first serial number
        update_dut_serialnumber_dropdown(
            dropdown_data,
            serial_numbers[0],
            dut_serialnumber_dropdown,
            dut_name_var,
            dut_busmap_var,
            dut_libname_var
        )
    else:
        dut_serialnumber_dropdown.set("")
        dut_busmap_var.set("")
        dut_libname_var.set("")

def browse_with_default_path(entry_widget, var=None, is_profile=False, profile_dropdown=None):
    """
    Generic browse function that checks for existing path in entry widget
    and uses it as initial directory if valid.
    """
    try:
        # Get current path from entry widget or variable
        current_path = var.get() if var else entry_widget.get()
        print(f"DEBUG: Current path from {'variable' if var else 'entry'}: {current_path}")
        # Determine initial directory
        initial_dir = os.path.expanduser('~')  # Default to home directory
        print(f"DEBUG: Initial directory: {initial_dir}")
        try:
            if current_path:
                print(f"DEBUG: Current path: {current_path}")
                if current_path.startswith('\\\\'): # Network path
                    # Convert network path to a format that tkinter can understand
                    initial_dir = current_path.replace('\\', '/')
                    print(f"DEBUG: Network path detected, converted to: {initial_dir}")
                elif current_path.startswith('..'): # Relative path
                    print("DEBUG: Relative path detected, resolving...")
                    # Try to resolve relative path
                    pwd_process = subprocess.Popen('cd', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                    current_dir, _ = pwd_process.communicate()
                    if current_dir:
                        initial_dir = os.path.abspath(os.path.join(current_dir.strip(), current_path))
                        print(f"DEBUG: Resolved relative path to: {initial_dir}")
                else: # Local path
                    if os.path.exists(os.path.dirname(current_path)):
                        initial_dir = os.path.dirname(current_path)
                        print(f"DEBUG: Local path detected, resolved to: {initial_dir}")
            elif os.path.exists('config.ini'):
                config.read('config.ini')
                saved_path = config.get('PREFERENCES', 'LastUsedPath', fallback=os.path.expanduser('~'))
                if os.path.exists(saved_path):
                    initial_dir = saved_path
        except Exception as path_error:
            print(f"Path resolution error: {path_error}")  # Debug print
            pass  # Use default home directory if any issues with paths
        
        print(f"Opening folder dialog at: {initial_dir}")  # Debug print
        
        new_path = filedialog.askdirectory(
            initialdir=initial_dir,
            title="Select Folder"
        )
            
        if new_path:
            # Convert back to Windows path format if needed
            if '/' in new_path:
                new_path = new_path.replace('/', '\\')
                
            new_path = ensure_trailing_backslash(new_path)

            # Update entry widget and/or variable
            if var:
                var.set(new_path)
                print(f"DEBUG: Updating StringVar with new path")
                print(f"DEBUG: StringVar value after set: {var.get()}")
            else:
                print(f"DEBUG: Updating entry widget with new path")
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, new_path)
            
            # If this is a profile folder browse, update the profile list
            if is_profile and profile_dropdown:
                print(f"Updating profiles for path: {new_path}")  # Debug print
                update_profile_list(new_path, profile_dropdown)
                
    except Exception as e:
        print(f"Browse error: {str(e)}")  # Debug print
        messagebox.showerror("Error", f"Failed to browse: {str(e)}")

def update_profile_list(folder_path, profile_dropdown):
    """Update profile dropdown with files from selected folder"""
    try:
        print(f"Checking folder: {folder_path}")  # Debug print
        if os.path.exists(folder_path):
            # Get all files in the folder
            files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
            
            print(f"Found files: {files}")  # Debug print
            
            if files:
                # Update the dropdown
                profile_dropdown['values'] = files
                profile_dropdown.set(files[0])  # Select first file by default
                print(f"Updated dropdown with {len(files)} files")  # Debug print
            else:
                print("No files found in folder")  # Debug print
                profile_dropdown['values'] = []
                profile_dropdown.set('')
                messagebox.showinfo("Info", "No files found in selected folder")
        else:
            print(f"Folder not found: {folder_path}")  # Debug print
            messagebox.showwarning("Warning", f"Selected profile folder not found: {folder_path}")
            
    except Exception as e:
        print(f"Update profile error: {str(e)}")  # Debug print
        messagebox.showerror("Error", f"Failed to update profile list: {str(e)}")
# Main status update function (for header status)
def update_status():
    """Updates main header status label for file operations"""
    try:
        global current_tp_file, status_label
        
        if current_tp_file and os.path.exists(current_tp_file):
            file_name = os.path.basename(current_tp_file)
            status_label.config(
                text=f"Current file: {file_name}",
                foreground='black'
            )
        else:
            status_label.config(
                text="No file opened",
                foreground='black'
            )
        status_label.update()
        
    except Exception as e:
        print(f"Error updating status label: {str(e)}")
        status_label.config(
            text="No file opened",
            foreground='black'
        
        )
def update_status_2(message, is_error=False):
    """Updates Page 2 specific status messages"""
    try:
        global status_var_2
        status_var_2.set(message)
        if is_error:
            status_label_2.configure(foreground="red")
        else:
            status_label_2.configure(foreground='black')
    except Exception as e:
        print(f"Error updating page 2 status: {str(e)}")
dropdowns = {
    # DSLAM Configuration
    'DSLAM_Name': {'var': dslam_name_var, 'widget': None},
    'DSLAM_Serialnumber': {'var': dslam_serialnumber_var, 'widget': None},
    'DSLAM_Busmap': {'var': dslam_busmap_var, 'widget': None},
    'DSLAM_Driver': {'var': dslam_driver_var, 'widget': None},
    'DSLAM_Port': {'var': dslam_port_var, 'widget': None},
    'DSLAM_Slot': {'var': dslam_slot_var, 'widget': None},
    'DSLAM_Frame': {'var': dslam_frame_var, 'widget': None},
    'DSLAM_Host': {'var': dslam_host_var, 'widget' :None},
    'DSLAM_Folder': {'var': dslam_folder_var, 'widget': None},
    'DSLAM_Profile': {'var': dslam_profile_var, 'widget': None},
    
    # DUT Configuration
    'DUT_Name': {'var': dut_name_var, 'widget': None},
    'DUT_Serialnumber': {'var': dut_serialnumber_var, 'widget': None},
    'DUT_Driver': {'var': dut_driver_var, 'widget': None},
    'DUT_Busmap': {'var': dut_busmap_var, 'widget': None},
    'DUT_Profile': {'var': dut_profile_var, 'widget': None},
    'DUT_Testplan': {'var': dut_testplan_var, 'widget': None},
    'DUT_Host': {'var': dut_host_var, 'widget': None},
    'DUT_Port': {'var': dut_port_var, 'widget': None},
    'DUT_Folder': {'var': dut_folder_var, 'widget': None},
    'DUT_Testscenario': {'var': testscenario_var, 'widget': None},
    
    # Platform Configuration
    'Platform_name': {'var': platform_name_var, 'widget': None},
    'Linesimulator': {'var': platform_noise_generator_var, 'widget': None},
    'NoiseGenerator': {'var': platform_linesimulator_var, 'widget': None},
    'Switch': {'var': platform_switch_var, 'widget': None},
    'trafficgenerator': {'var': platform_traffic_generator_var, 'widget': None},
    'Layout': {'var': layout_var, 'widget': None},
    'Operator': {'var': operator_var, 'widget': None}
}

def create_zoomed_dropdown(parent, **kwargs):
    """Create a dropdown that zooms when clicked"""
    dropdown = ttk.Combobox(parent, **kwargs)
    
    def on_dropdown_open(event):
        # Get current position
        x = dropdown.winfo_rootx()
        y = dropdown.winfo_rooty()
        
        # Create zoomed popup
        popup = tk.Toplevel(parent)
        popup.withdraw()  # Hide initially to prevent flickering
        popup.transient(parent)  # Make it float on top
        popup.overrideredirect(True)  # Remove window decorations
        
        # Create larger listbox with same values
        listbox = tk.Listbox(popup, font=("Helvetica", 12), 
                            width=60,  # Increased width
                            height=10)  # Show more items
        listbox.pack(fill=tk.BOTH, expand=True)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.configure(yscrollcommand=scrollbar.set)
        
        # Populate with current values
        values = dropdown['values']
        for value in values:
            listbox.insert(tk.END, value)
        
        def on_select(event):
            selection = listbox.get(listbox.curselection())
            dropdown.set(selection)
            popup.destroy()
            dropdown.event_generate('<<ComboboxSelected>>')
        
        listbox.bind('<Double-Button-1>', on_select)
        listbox.bind('<Return>', on_select)
        listbox.bind('<Escape>', lambda e: popup.destroy())
        
        # Position and show popup
        popup.update_idletasks()
        popup.geometry(f"+{x}+{y}")
        popup.deiconify()
        
        # Give focus to listbox
        listbox.focus_set()
        
    dropdown.bind('<Button-1>', on_dropdown_open)
    return dropdown

def initialize_dut_port_dropdown():
    """Initialize the DUT port dropdown with a full range of port values."""
    full_port_range = [str(i) for i in range(0, 12)]  # Assuming max is 11
    dropdowns['DUT_Port']['widget']['values'] = full_port_range
    dut_port_var.set(full_port_range[0])
    
def initialize_dslam_port_dropdown():
    """Initialize the DSLAM port dropdown with a full range of port values."""
    full_port_range = [str(i) for i in range(0, 12)]  # Assuming max is 11
    dropdowns['DSLAM_Port']['widget']['values'] = full_port_range
    dslam_port_var.set(full_port_range[0])

def on_dslam_serialnumber_port_update(event):
    """Update port dropdown when DSLAM serial number is selected."""
    def get_excel_file_path():
        # Network path (preferred)
        network_path = r"\\hermes\4_Engineering\Software\Projects\Python ServerConnector\Input Test Plan Configurator\test_plan_configurator.xlsx"
        
        # Determine if running as executable or script
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")
        
        # Local path as fallback
        local_path = os.path.join(base_path, 'test_plan_configurator.xlsx')
        
        # Try network path first
        if os.path.exists(network_path):
            print(f"Using network file: {network_path}")
            return network_path
        # Then try local path
        elif os.path.exists(local_path):
            print(f"Using local file: {local_path}")
            return local_path
        else:
            raise FileNotFoundError(
                f"Could not find test_plan_configurator.xlsx in either:\n"
                f"Network: {network_path}\n"
                f"Local: {local_path}"
            )

    try:
        selected_serial = dslam_serialnumber_var.get()
        print(f"\nDEBUG: Function called with selection: {selected_serial}")
        
        # Get the Excel file path
        file_path = get_excel_file_path()
        
        # Get the Excel data
        df = pd.read_excel(file_path, sheet_name='DSLAM_serialnumber')
        
        # Find the row matching the selected serial number
        row = df[df['DSLAM_Serialnumber'].str.contains(selected_serial, na=False)]
        print(f"\nDEBUG: Row found: {row}")
        
        if not row.empty:
            try:
                max_port = int(float(row['Max_Port'].iloc[0]))
                min_port = int(float(row['Min_Port'].iloc[0]))
                print(f"\nDEBUG: Min Port: {min_port}, Max Port: {max_port}")
                
                # Filter the full port range to the relevant range
                port_values = [str(i) for i in range(min_port, max_port + 1)]
                
                # Update the dropdown with new values
                dropdowns['DSLAM_Port']['widget']['values'] = port_values
                dropdowns['DSLAM_Port']['widget'].set(port_values[0])
                dslam_port_var.set(port_values[0])
                
                print(f"Updated port values for {selected_serial}: {port_values}")
                
            except ValueError as ve:
                print(f"Error converting ports to integers: {ve}")
                # Reset to full range on error
                initialize_dslam_port_dropdown()
        else:
            print(f"\nDEBUG: No matching row found for serial number: {selected_serial}")
            # Reset to full range if no match found
            initialize_dslam_port_dropdown()
            
    except FileNotFoundError as e:
        print(f"Error: Excel file not found - {str(e)}")
        messagebox.showerror("Error", f"Could not find Excel file:\n{str(e)}")
        initialize_dslam_port_dropdown()
    except pd.errors.EmptyDataError:
        print("Error: Excel file is empty or corrupted")
        messagebox.showerror("Error", "The Excel file appears to be empty or corrupted")
        initialize_dslam_port_dropdown()
    except Exception as e:
        print(f"Error in on_dslam_serialnumber_selected: {str(e)}")
        traceback.print_exc()
        messagebox.showerror("Error", f"An unexpected error occurred:\n{str(e)}")
        initialize_dslam_port_dropdown()



def normalize_path(path):
    """Convert forward slashes to backslashes and ensure proper double backslashes at start"""
    if path:
        # First replace all forward slashes with single backslashes
        path = path.replace('/', '\\')
        
        # Ensure network paths start with exactly two backslashes
        if path.startswith('\\'):
            # Remove all backslashes from the start
            path = path.lstrip('\\')
            # Add exactly two backslashes at the start
            path = '\\\\' + path
            
    return path 

def get_excel_file_path():
    """Get Excel file path from network location only"""
    network_path = r"\\hermes\4_Engineering\Software\Projects\Python ServerConnector\Input Test Plan Configurator\test_plan_configurator.xlsx"
    
    try:
        if os.path.exists(network_path):
            print(f"Using network file: {network_path}")
            return network_path
        else:
            # Log the issue
            print(f"Network file not found: {network_path}")
            print("Using default values instead")
            
            # Show a temporary message instead of an error
            show_temp_message("Network File Unavailable", 
                             "Using default values. Check network connection.")
            
            # Apply default values
            reset_to_defaults()
            
            # Return None to indicate we're using defaults instead
            return None
    except Exception as e:
        # Handle any other errors (like permission issues)
        print(f"Error accessing network file: {str(e)}")
        print("Using default values instead")
        
        # Show a temporary message instead of an error
        show_temp_message("Network Access Error", 
                         f"Using default values. Error: {str(e)}")
        
        # Apply default values
        reset_to_defaults()
        
        # Return None to indicate we're using defaults instead
        return None

def open_in_explorer(path_entry, var=None):
    """Opens the specified path in File Explorer"""
    try:
        # Get path from variable or entry
        path = var.get() if var else path_entry.get()
        if path and os.path.exists(path):
            # Use normalized path
            normalized_path = os.path.normpath(path)
            print(f"DEBUG: Opening explorer at: {normalized_path}")
            os.startfile(normalized_path)
        else:
            print(f"DEBUG: Invalid path: {path}")
            messagebox.showwarning("Warning", "Please select a valid folder first.")
    except Exception as e:
        print(f"DEBUG: Error opening explorer: {str(e)}")
        messagebox.showerror("Error", f"Failed to open folder: {str(e)}")

# def open_execute_window():
#     """Open the execution window with Test Plan execution UI"""
#     global execute_window_instance, current_tp_file, response_text_4, status_label_4
#     global config_inner_frame, test_inner_frame, config_canvas, test_canvas
    
    
#     try:
        
#         # If no test plan is currently loaded or if unsaved changes exist, 
#         # save both pages first
#         if current_tp_file:
#             # Save both page 2 and page 3 content before opening execution window
#             print("Saving current test plan before execution")
#             coordinated_save_test_plan()
#         else:
#             # If no file is open, prompt to save first
#             result = messagebox.askyesno(
#                 "No Test Plan", 
#                 "No test plan is currently open. Would you like to create or open one first?"
#             )
#             if result:
#                 # Let user create or open a file
#                 choice = messagebox.askquestion(
#                     "Test Plan", 
#                     "Do you want to create a new test plan? (No will open an existing one)"
#                 )
#                 if choice == 'yes':
#                     # Create new test plan
#                     if not create_coordinated_test_plan():
#                         return  # User cancelled
#                 else:
#                     # Open existing test plan
#                     if not coordinated_edit_test_plan():
#                         return  # User cancelled
#             else:
#                 return  # User cancelled
                
#         # Check if a test plan is now available
#         if not current_tp_file:
#             messagebox.showerror("Error", "No test plan selected for execution.")
#             return
        
#         # Check if page4 already exists
#         if page4_frame and page4_frame.winfo_exists():
#             # Page4 exists - just switch to it
#             print("Page4 exists - switching to it")
#             # Find the index of page4 in the notebook
#             for i in range(notebook.index('end')):
#                 if notebook.select(i) == str(page4_frame):
#                     notebook.select(i)
#                     break
#             return
            
#         # Create page4 in the main window
#         create_page4(notebook, as_tab=True)
        
#         # Find the index of the newly created page4 and select it
#         for i in range(notebook.index('end')):
#             if notebook.select(i) == str(page4_frame):
#                 notebook.select(i)
#                 break
        

        
        
#         #  # Check if window already exists
#         # if execute_window_instance and execute_window_instance.winfo_exists():
#         #     # Window already exists - restore it from minimized state if needed
#         #     print("Execution window exists - restoring from taskbar")
#         #     execute_window_instance.deiconify()  # Show if it was withdrawn
#         #     execute_window_instance.state('normal')  # Restore if minimized
#         #     execute_window_instance.focus_force()  # Force focus
#         #     execute_window_instance.lift()  # Bring to front
#         #     return
            
#         # # Create top-level window for execution
#         # execute_window = tk.Toplevel(root)
#         # execute_window_instance = execute_window
#         # execute_window.title(f"Test Plan Execution - {os.path.basename(current_tp_file)}")
#         # execute_window.geometry("1200x800")
        
#         # # Configure window properties
#         # execute_window.wm_transient('')  # Allow window to appear in the taskbar
#         # execute_window.resizable(True, True)
#         # execute_window.attributes('-toolwindow', False)  # Standard window with min/max buttons
        
#         # Add minimize handler
#         def on_minimize():
#             # Minimize to taskbar instead of closing
#             execute_window.withdraw()  # Hide the window
#             # Show a notification or update status to indicate it's minimized
#             print("Execution window minimized to taskbar")
            
#         # Add restore handler
#         def on_restore():
#             # Restore from taskbar
#             execute_window.deiconify()  # Show the window
#             execute_window.state('normal')  # Ensure it's not minimized
#             execute_window.focus_force()  # Force focus
#             execute_window.lift()  # Bring to front
#             print("Execution window restored from taskbar")
            
#         # Add close handler
#         def on_execute_window_close():
#             global execute_window_instance, response_text_4, status_label_4, config_inner_frame, test_inner_frame
#             global config_canvas, test_canvas  

            
#             # Ask user if they want to close or minimize
#             result = messagebox.askyesnocancel(
#                 "Close Execution Window",
#                 "Do you want to close the execution window?\n\n"
#                 "Yes: Close completely (logs will be lost)\n"
#                 "No: Minimize to taskbar (logs will be preserved)\n"
#                 "Cancel: Keep window open"
#             )
            
#             if result is None:  # Cancel
#                 return
#             elif result:  # Yes - close completely
#                 try:
#                     # Clean up all references
#                     execute_window_instance = None
#                     response_text_4 = None
#                     status_label_4 = None
#                     config_inner_frame = None
#                     test_inner_frame = None
#                     config_canvas = None
#                     test_canvas = None
#                     print("Execution window closed - all references cleaned up")
#                 except Exception as e:
#                     print(f"Error during window cleanup: {str(e)}")
#                 finally:
#                     execute_window.destroy()
#             else:  # No - minimize to taskbar
#                 on_minimize()
        
#         # # Bind window events
#         # execute_window.protocol("WM_DELETE_WINDOW", on_execute_window_close)
        
#         # # Add buttons frame at the bottom
#         # button_frame = ttk.Frame(execute_window)
#         # button_frame.pack(side="bottom", pady=10, fill="x")
        
#         # # Add minimize button
#         # ttk.Button(
#         #     button_frame, 
#         #     text="Minimize to Taskbar",
#         #     command=on_minimize
#         # ).pack(side="left", padx=5)
        
#         # # Add close button
#         # ttk.Button(
#         #     button_frame, 
#         #     text="Close Window",
#         #     command=on_execute_window_close
#         # ).pack(side="right", padx=5)
        
#         # # Create execution frame inside window
#         # create_page4(execute_window, as_tab=False)
        
        
#         # # Try to copy the main window's icon for consistency
#         # try:
#         #     execute_window.iconbitmap(root.iconbitmap())
#         # except:
#         #     pass  # Skip if icon setting fails
        
#         # # Give focus to the new window
#         # execute_window.focus_set()
        

        
#         # # Center the window on screen
#         # execute_window.update_idletasks()  # Ensure window dimensions are calculated
        
#         # # Use explicit window size rather than requested size
#         # width, height = 1200, 800
        
#         # # Get monitor info
#         # screen_width = root.winfo_screenwidth()
#         # screen_height = root.winfo_screenheight()
        
#         # # Calculate position to center relative to main window
#         # root_x = root.winfo_rootx()
#         # root_y = root.winfo_rooty()
#         # root_width = root.winfo_width()
#         # root_height = root.winfo_height()
        
#         # # Center on parent, but ensure it stays on screen
#         # x = max(0, min(screen_width - width, root_x + (root_width - width) // 2))
#         # y = max(0, min(screen_height - height, root_y + (root_height - height) // 2))
        
#         # execute_window.geometry(f"{width}x{height}+{x}+{y}")
        
#         # EXECUTE TEST PLAN IMMEDIATELY
#         print("Automatically executing test plan")

#         # Delay running the test plan to ensure UI is fully loaded
#         execute_window.after(1000, lambda: safely_run_test_plan(root))
#     except Exception as e:
#         print(f"Error opening execution window: {str(e)}")
#         messagebox.showerror("Error", f"Could not open execution window:\n{str(e)}")

# 
def open_execute_window():
    """Open the execution window with Test Plan execution UI"""
    print("open_execute_window called")
    global execute_window_instance, current_tp_file, response_text_4, status_label_4
    global config_inner_frame, test_inner_frame, config_canvas, test_canvas, page4_frame, notebook
    
    try:
        print(f"Current test plan file: {current_tp_file}")
        # First time run - save and execute
        if not current_tp_file:
            print("No test plan file - auto-saving current work")
            if coordinated_save_test_plan():
                print(f"Successfully saved and will use new test plan: {current_tp_file}")
            else:
                print("Save failed")
                return
        else:
            print("Saving current test plan before execution")
            coordinated_save_test_plan()

        if not current_tp_file:
            print("No test plan file available after save attempt")
            return

        # Now handle page4 - Check if it exists and is valid
        found = False
        if page4_frame and page4_frame.winfo_exists():
            print("Checking for existing page4 tab")
            for i in range(notebook.index('end')):
                if notebook.tab(i, "text") == "Test Plan Execution":
                    print(f"Found existing page4 at index {i}")
                    notebook.select(i)
                    found = True
                    if status_label_4 and status_label_4.winfo_exists():
                        status_label_4.config(text=f"Current Test Plan: {os.path.basename(current_tp_file)}")                           
                    # Don't run again, just switch to the tab
                    break
        
        # Only create new page4 and execute test plan if not found
        if not found:
            print("No existing page4, creating new one and executing test plan")
            page4_frame = create_page4(notebook, as_tab=True)
            
            if page4_frame is None:
                print("Error: page4_frame is None after creation")
                messagebox.showerror("Error", "Failed to create execution page")
                return
                
            print(f"Page4 created successfully: {page4_frame}")
            
            # Find and select the new tab
            for i in range(notebook.index('end')):
                if notebook.tab(i, "text") == "Test Plan Execution":
                    notebook.select(i)
                    print(f"Selected new page4 at index {i}")
                    if status_label_4 and status_label_4.winfo_exists():
                        status_label_4.config(text=f"Current Test Plan: {os.path.basename(current_tp_file)}")
                    # Only run test plan on first creation
                    root.after(500, run_current_or_new_tp)
                    break
                    
    except Exception as e:
        print(f"Error in open_execute_window: {str(e)}")
        traceback.print_exc()
        messagebox.showerror("Error", f"Could not open execution window:\n{str(e)}")

def safely_run_test_plan(window):
    """Safely run the test plan with error handling"""
    try:
        print("Attempting to run test plan")
        if 'run_current_or_new_tp' in globals() and callable(globals()['run_current_or_new_tp']):
            run_current_or_new_tp()
        elif 'run_current_or_new_tp' in page4_functions and callable(page4_functions['run_current_or_new_tp']):
            page4_functions['run_current_or_new_tp']()
        else:
            print("Warning: run_current_or_new_tp function not found")
    except Exception as e:
        print(f"Error running test plan: {str(e)}")
        traceback.print_exc()


def get_file_dialog_directory():
    """
    Get the appropriate directory to open file dialogs in.
    Returns:
        str: Path to use as initialdir for file dialogs
    """
    global workspace_dir, root_dir, current_tp_file

    # Use workspace_dir if available
    if workspace_dir:
        print(f"Using MyWorkspace directory: {workspace_dir}")
        return workspace_dir

    # Fall back to root_dir if available
    if root_dir:
        print(f"Using root directory: {root_dir}")
        return root_dir

    # Fall back to current test plan directory
    if current_tp_file:
        current_dir = os.path.dirname(current_tp_file)
        print(f"Using current file directory: {current_dir}")
        return current_dir

    # Last resort
    cwd = os.getcwd()
    print(f"Using current working directory: {cwd}")
    return cwd


def create_coordinated_test_plan(source_page='2'):
    """Create a new test plan with content from both pages
    
    Args:
        source_page: Which page initiated the creation ('2' or '3')
    """
    global current_tp_file
    
    try:
        # Get the appropriate initial directory
        initial_dir = get_file_dialog_directory()
        
        # Prompt user to save the file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".tp",
            filetypes=[("Test Plan Files", "*.tp"), ("All Files", "*.*")],
            title="Save New Test Plan As",
            initialdir=initial_dir
        )
    
        if not file_path:  # User cancelled
            return False
        
        # Update global reference
        current_tp_file = file_path
        print(f"Creating new test plan at: {file_path}")
        
        # Different handling based on source page
        if source_page == '2':
            # If called from page 2, create new test plan using page 2's function first
            create_new_test_plan()  # This creates the config section from page 2
            # Then save test section using page 3's function
            save_page3_test_plan()  # This adds the test section to the same file
        else:  # source_page == '3'
            # If called from page 3, create new test plan using page 3's function first
            create_test_plan()  # This creates the test section from page 3
            # Then save config section using page 2's function
            save_test_plan()  # This adds the config section to the same file
        
        # Update status on both pages
        update_status_2(f"Created: {os.path.basename(file_path)}")
        update_status_3(f"Created: {os.path.basename(file_path)}")
        update_file_status(source_page='both')
        
        # Show success message
        messagebox.showinfo("Success", 
            f"Created {os.path.basename(file_path)} file with " +
            f"config and test sections successfully.")
        
        return True
        
    except Exception as e:
        messagebox.showerror("Error", f"Error creating test plan: {str(e)}")
        print(f"Error in create_coordinated_test_plan: {str(e)}")
        return False

def coordinated_edit_test_plan():
    """Edit an existing test plan - Loading both config and test sections"""
    global current_tp_file
    
    try:
        # Get the appropriate initial directory
        initial_dir = get_file_dialog_directory()
        
        # Let user select a file
        tp_file = filedialog.askopenfilename(
            title="Select Test Plan File",
            filetypes=[("Test Plan Files", "*.tp"), ("All Files", "*.*")],
            initialdir=initial_dir
        )
        
        if not tp_file:
            return  # User cancelled
            
        # Update global reference
        current_tp_file = tp_file
        print(f"Opening test plan: {tp_file}")
        
        # Load config section for page 2
        config_loaded = False
        test_loaded = False
        try:
            print(f"Loading config section from: {tp_file}")
            edit_test_plan(tp_file)  # This loads the config section
            config_loaded = True
        except Exception as e:
            print(f"Warning: Error loading config section: {str(e)}")
            messagebox.showwarning("Warning", f"Error loading configuration section: {str(e)}")
        
        # Load test section for page 3
        test_loaded = False
        try:
            print(f"Loading test section from: {tp_file}")
            edit_page3_test_plan(tp_file)  # This loads the test section
            test_loaded = True
        except Exception as e:
            print(f"Warning: Error loading test section: {str(e)}")
            messagebox.showwarning("Warning", f"Error loading test section: {str(e)}")
            
        # Update status on both pages
        if config_loaded and test_loaded:
            update_status_2(f"Loaded config from: {os.path.basename(tp_file)}")
            update_status_3(f"Loaded tests from: {os.path.basename(tp_file)}")
            update_file_status(source_page='both')
            messagebox.showinfo("Success", 
                f"Test plan '{os.path.basename(tp_file)}' loaded successfully with both configuration and test sections.")
        elif config_loaded:
            update_status_2(f"Loaded config from: {os.path.basename(tp_file)}")
            update_file_status(source_page='2')
            messagebox.showinfo("Partial Success", 
                f"Only configuration section was loaded from '{os.path.basename(tp_file)}'.")
        elif test_loaded:
            update_status_3(f"Loaded tests from: {os.path.basename(tp_file)}")
            update_file_status(source_page='3')
            messagebox.showinfo("Partial Success", 
                f"Only test section was loaded from '{os.path.basename(tp_file)}'.")
        else:
            messagebox.showerror("Error", f"Failed to load any sections from '{os.path.basename(tp_file)}'.")
            
    except Exception as e:
        messagebox.showerror("Error", f"Error opening test plan: {str(e)}")
        print(f"Error in coordinated_edit_test_plan: {str(e)}")

def coordinated_save_test_plan():
    """Save current test plan - saving both config and test sections"""
    global current_tp_file
    
    try:
        # Check if we have a file to save to
        if not current_tp_file:
            return coordinated_save_as_test_plan()
            
        # Step 1: Save config section from page 2
        try:
            save_test_plan()  # Original page 2 save function
            print("Config section saved")
        except Exception as e:
            print(f"Warning: Error saving config section: {str(e)}")
            
        # Step 2: Save test section from page 3 - DIRECT CALL
        try:
            save_page3_test_plan()  # Direct call to the page3 save function
            print("Test section saved")
        except Exception as e:
            print(f"Warning: Error saving test section: {str(e)}")
            
        # Update status
        update_status_2(f"Saved to: {os.path.basename(current_tp_file)}")
        update_status_3(f"Saved to: {os.path.basename(current_tp_file)}")
        
        return True
        
    except Exception as e:
        messagebox.showerror("Error", f"Error saving test plan: {str(e)}")
        print(f"Error in coordinated_save_test_plan: {str(e)}")
        return False
        
def coordinated_save_as_test_plan():
    """Save test plan with a new name - saving both config and test sections"""
    global current_tp_file
    
    try:
        # Get the appropriate initial directory
        initial_dir = get_file_dialog_directory()
        
        # Prompt user for new file name
        file_path = filedialog.asksaveasfilename(
            defaultextension=".tp",
            filetypes=[("Test Plan Files", "*.tp"), ("All Files", "*.*")],
            title="Save Test Plan As",
            initialdir=initial_dir
        )
        
        if not file_path:  # User cancelled
            return False
            
        # Store original path and update global reference
        original_file = current_tp_file
        current_tp_file = file_path
        
        # Save to the new file using our coordinated save function
        success = coordinated_save_test_plan()
        
        if success:
            messagebox.showinfo("Success", f"Test plan saved as: {os.path.basename(file_path)}")
            
        return success
        
    except Exception as e:
        messagebox.showerror("Error", f"Error in Save As: {str(e)}")
        print(f"Error in coordinated_save_as_test_plan: {str(e)}")
        return False

def debug_run_button():
    print("Run button clicked")
    open_execute_window()     

def show_temp_message(title, message, duration=2000):
    """Show a temporary message that disappears after the specified duration in milliseconds"""
    popup = tk.Toplevel()
    popup.title(title)
    popup.geometry("300x100")  # Set size
    popup.resizable(False, False)  # Not resizable
    
    # Position in center of root window
    popup.geometry("+{}+{}".format(
        root.winfo_rootx() + (root.winfo_width() // 2) - 150,
        root.winfo_rooty() + (root.winfo_height() // 2) - 50
    ))
    
    # Make it stay on top and remove window decorations for cleaner look
    popup.attributes("-topmost", True)
    popup.overrideredirect(True)  # Remove title bar
    
    # Add a border
    popup.configure(bg="lightgrey")
    inner_frame = tk.Frame(popup, bg="white", padx=20, pady=20)
    inner_frame.pack(padx=2, pady=2, fill="both", expand=True)
    
    # Add message
    tk.Label(inner_frame, text=message, bg="white", font=("Helvetica", 10)).pack()
    
    # Schedule the popup to be destroyed after duration
    popup.after(duration, popup.destroy)
    
    # Allow clicking anywhere to dismiss early
    popup.bind("<Button-1>", lambda e: popup.destroy())
    
    return popup


def reset_to_defaults():
    """Reset all fields to default values after user confirmation"""
    # Show confirmation dialog
    if messagebox.askyesno("Reset Test Plan", 
                          "This will clear all current settings and reset to defaults. Continue?"):
        
        # Clear current test plan reference
        global current_tp_file
        current_tp_file = None
        
        # Reset all configuration fields to defaults
        default_values = {
        # DSLAM Configuration
        'DSLAM_Name': "DMS4",
        'DSLAM_Serialnumber': "DMS4 CO 65400 192.168.1.1",
        'DSLAM_Busmap': "DMS4_192_168_001_001",
        'DSLAM_Driver': "../DSLAM/BCM6400x/BCM6400x_100.bin",  # Based on driver_path_var default
        'DSLAM_Port': "0",                     # Based on port_var default
        'DSLAM_Slot': "0",
        'DSLAM_Frame': "0",
        'DSLAM_Folder': "No folder selected",
        'DSLAM_Profile': "",
        
        # DUT Configuration
        'DUT_Name': "DMS12",
        'DUT_Serialnumber': "021018000000 DMS12_CPE 65430",
        'DUT_Port': "0",
        'DUT_Busmap': "DMS12_192_168_254_042",
        'DUT_Driver': "No driver selected",
        'DUT_Profile': "",
        'DUT_Testplan': "No Noise",
        'DUT_Host': "",
        'DUT_Testscenario': "FS",
        'DUT_Folder': "No folder selected",
        
        # Platform Configuration
        'Platform_name': "other",
        'Linesimulator': "",
        'NoiseGenerator': "",
        'Switch': "",
        'trafficgenerator': "",
        
        # Layout Configuration
        'Layout': "default"
    }
        
        # Update all variables with default values
        for key, value in default_values.items():
            if key in dropdowns and dropdowns[key]['var']:
                dropdowns[key]['var'].set(value)
        
        # Clear the tables in page 4
        if 'clear_preview_tables' in page4_functions:
            page4_functions['clear_preview_tables'](config_inner_frame, test_inner_frame)
        
        # Update status messages
        update_status_2("Reset to defaults")
        update_status_3("Reset to defaults")
        update_status_4("Reset to defaults")
        
        # Update the file status
        update_file_status(filename="New Test Plan", source_page='both')
        
        # Show temporary message instead of requiring a click
        show_temp_message("Reset Complete", "  All settings have been reset to default values.")
        
        # If you want to immediately switch to page2 afterward
        root.after(2100, lambda: switch_tab(1))


# Function to create Page 2
def create_page2():
    global dslam_name_var, dslam_busmap_var,dslam_serialnumber_var, dslam_port_var, dslam_slot_var, dslam_frame_var, dslam_folder_var, dslam_profile_var
    global layout_var, output_folder_var, output_filename_var, startup_time_var
    global test_plan_entry,current_tp_file
    global dropdowns,dslam_driver_entry
    global dslam_folder_entry,dut_folder_entry,dslam_folder_entry,dslam_driver_entry,dslam_driver_var,operator_var

    global df, notebook  # Make sure notebook is global
        # Create and configure page2_frame
    page2_frame = ttk.Frame(notebook, padding="5 5 5 5")
    
    # Add the frame to the notebook with a tab name
     # Add the frame to the notebook with a tab name
    notebook.add(page2_frame, text="Test Plan Configurator")
    
    # Create header frame for Test Plan title and buttons
    header_frame = ttk.Frame(page2_frame)
    header_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
    
    # Configure page2_frame grid - this is the key change
    page2_frame.rowconfigure(1, weight=1)  # Make row 1 (config frame) expandable
    page2_frame.columnconfigure(0, weight=1)  # Make column 0 expandable
    
    
    test_plan_entry = ttk.Entry(page2_frame)  # Use your actual parent frame

    
    # Add status label for page2
    global status_label_2, status_var_2
    status_var_2 = tk.StringVar(value="No file opened")
    status_label_2 = ttk.Label(
        header_frame, 
        textvariable=status_var_2,
        font=("Helvetica", 10, "bold"),
        foreground='black'
    )
    status_label_2.pack(side="right", padx=10)
    # Create a left frame for title and buttons
    left_frame = ttk.Frame(header_frame)
    left_frame.pack(side="left", fill="x")
    
    # Add Test Plan title and buttons in left frame
    ttk.Label(left_frame, text="Test Plan Configurator", font=("Helvetica", 10, "bold")).pack(side="left", padx=10)

    # "New" button - use lambda for passing parameters
    ttk.Button(left_frame, 
        text="New", 
        cursor="hand2", 
        command=lambda: root.after(100, lambda: reset_to_defaults())
    ).pack(side="left", padx=5)

    # "Open" button 
    edit_button_page2 = ttk.Button(left_frame, 
        text="Open", 
        cursor="hand2", 
        command=lambda: root.after(100, coordinated_edit_test_plan)
    )
    edit_button_page2.pack(side="left", padx=5)

    # "Save" button
    save_button_page2 = ttk.Button(left_frame, 
        text="Save", 
        cursor="hand2", 
        command=lambda: root.after(100, coordinated_save_test_plan)
    )
    save_button_page2.pack(side="left", padx=5)

    # "Save As" button
    ttk.Button(left_frame, 
        text="Save As", 
        cursor="hand2", 
        command=lambda: root.after(100, coordinated_save_as_test_plan)
    ).pack(side="left", padx=5)

    # "Run" button
    ttk.Button(left_frame, 
        text="Run", 
        cursor="hand2", 
        command=lambda: root.after(100, open_execute_window)
    ).pack(side="left", padx=5)



    global response3_tp, dslam_name_dropdown_tp

    # Create combined config frame with all three sections
    combined_config_frame = ttk.LabelFrame(page2_frame, text="Configuration Settings", padding="5 5 5 5")
    combined_config_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

    # Create the six section frames
    dslam_frame = ttk.LabelFrame(combined_config_frame, text="DSLAM Configuration", padding="5 5 5 5")
    dslam_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

    dut_frame = ttk.LabelFrame(combined_config_frame, text="DUT Configuration", padding="5 5 5 5")
    dut_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

    validation_frame = ttk.LabelFrame(combined_config_frame, text="Validation Settings", padding="5 5 5 5")
    validation_frame.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")

    platform_frame = ttk.LabelFrame(combined_config_frame, text="Platform Configuration", padding="5 5 5 5")
    platform_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

    # Create a container frame for Output and Results
    output_results_container = ttk.Frame(combined_config_frame)
    output_results_container.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
    output_results_container.rowconfigure(0, weight=1)
    output_results_container.rowconfigure(1, weight=1)
    output_results_container.columnconfigure(0, weight=1)

    # Put Output and Results frames in the container
    output_frame = ttk.LabelFrame(output_results_container, text="Output", padding="5 5 5 5")
    output_frame.grid(row=0, column=0, padx=0, pady=(0,2), sticky="nsew")

    results_frame = ttk.LabelFrame(output_results_container, text="Results", padding="5 5 5 5")
    results_frame.grid(row=1, column=0, padx=0, pady=(2,0), sticky="nsew")

    scenario_frame = ttk.LabelFrame(combined_config_frame, text="Test Scenario", padding="5 5 5 5")
    scenario_frame.grid(row=1, column=2, padx=5, pady=5, sticky="nsew")

    # Configure the grid weights
    combined_config_frame.columnconfigure(0, weight=1)
    combined_config_frame.columnconfigure(1, weight=1)
    combined_config_frame.columnconfigure(2, weight=1)
    combined_config_frame.rowconfigure(0, weight=1)
    combined_config_frame.rowconfigure(1, weight=1)

    # ----- DSLAM Configuration Section -----
    # DSLAM Type
    ttk.Label(dslam_frame, text="DSLAM Type:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    dslam_name_dropdown = ttk.Combobox(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                  textvariable=dslam_name_var, state='readonly', width=67)
    dslam_name_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DSLAM_Name']['widget'] = dslam_name_dropdown

    
    # DSLAM Serial Number
    ttk.Label(dslam_frame, text="Serial Number:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    dslam_serialnumber_dropdown = ttk.Combobox(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                          textvariable=dslam_serialnumber_var, state='readonly', width=67)
    dslam_serialnumber_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DSLAM_Serialnumber']['widget'] = dslam_serialnumber_dropdown
    dslam_serialnumber_dropdown.bind('<<ComboboxSelected>>', on_dslam_serialnumber_selected)

    # DSLAM Busmap
    ttk.Label(dslam_frame, text="Busmap:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky="e")
    dslam_busmap_dropdown = ttk.Combobox(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                    textvariable=dslam_busmap_var, state='readonly', width=67)
    dslam_busmap_dropdown.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DSLAM_Busmap']['widget'] = dslam_busmap_dropdown

    
    # DUT Driver
    # DSLAM Driver Frame with Browse Button
    ttk.Label(dslam_frame, text="DSLAM Driver:", font=("Helvetica", 10, "bold")).grid(row=8, column=0, padx=5, pady=5, sticky="e")
    dslam_driver_frame = ttk.Frame(dslam_frame)
    dslam_driver_frame.grid(row=8, column=1, padx=5, pady=5, sticky="ew")
    
    dslam_driver_entry = ttk.Entry(dslam_driver_frame, textvariable=dslam_driver_var, width=67,
                                  foreground='black', font=("Helvetica", 10, "bold"))
    dslam_driver_entry.pack(side="left", fill="x", expand=True)
    
    dslam_driver_button = ttk.Button(dslam_driver_frame, text="Browse", cursor="hand2",
                                    command=lambda: browse_driver(dslam_driver_entry))
    dslam_driver_button.pack(side="right")
    
    dropdowns['DSLAM_Driver'] = {'widget': dslam_driver_entry, 'var': dslam_driver_var}
    


    # DSLAM Port
    ttk.Label(dslam_frame, text="Port:", font=("Helvetica", 10, "bold")).grid(row=4, column=0, padx=5, pady=5, sticky="e")
    dslam_port_dropdown = ttk.Combobox(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                  textvariable=dslam_port_var, width=67)
    dslam_port_dropdown.grid(row=4, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DSLAM_Port'] = {'widget': dslam_port_dropdown, 'var': dslam_port_var}

    # DSLAM Slot
    ttk.Label(dslam_frame, text="Slot:", font=("Helvetica", 10, "bold")).grid(row=5, column=0, padx=5, pady=5, sticky="e")
    dslam_slot_entry = ttk.Entry(dslam_frame, width=67, foreground='black', font=("Helvetica", 10, "bold"), 
                            textvariable=dslam_slot_var)
    dslam_slot_entry.grid(row=5, column=1, padx=5, pady=5, sticky="ew")
    #dslam_slot_entry.grid(row=5, column=1, padx=5, pady=5)

    # DSLAM Frame
    ttk.Label(dslam_frame, text="Frame:", font=("Helvetica", 10, "bold")).grid(row=6, column=0, padx=5, pady=5, sticky="e")
    dslam_frame_entry = ttk.Entry(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                width=67, textvariable=dslam_frame_var)
    dslam_frame_entry.grid(row=6, column=1, padx=5, pady=5, sticky="ew")
    

        # DUT Host
    ttk.Label(dslam_frame, text="Host:", font=("Helvetica", 10, "bold")).grid(row=7, column=0, padx=5, pady=5, sticky="e")
    dslam_host_dropdown = ttk.Combobox(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                               textvariable=dslam_host_var, state='readonly', width=67)
    dslam_host_dropdown.grid(row=7, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DSLAM_Host']['widget'] = dslam_host_dropdown

    #dslam_frame_entry.grid(row=6, column=1, padx=5, pady=5)
    # Initialize dslam_driver_var if not already defined
    if dslam_driver_var is None or not isinstance(dslam_driver_var, tk.StringVar):
        dslam_driver_var = tk.StringVar()
        # Set default value - bcm_file should be defined earlier in your code
        try:
            default_value = bcm_file if 'bcm_file' in globals() else "../DSLAM/BCM6400x/BCM6400x_100.bin"
            dslam_driver_var.set(default_value)
            print(f"Initialize dslam_driver_var with: {default_value}")
        except Exception as e:
            print(f"Error setting default driver path: {str(e)}")
            dslam_driver_var.set("../DSLAM/BCM6400x/BCM6400x_100.bin")
    

    # Profile Folder section - CORRECT POSITION (row 8)
    # Profile Folder section
    ttk.Label(dslam_frame, text="Profile Folder:", font=("Helvetica", 10, "bold")).grid(row=9, column=0, padx=5, pady=5, sticky="e")
    
        # Create frame for folder entry and browse button
    dslam_folder_frame = ttk.Frame(dslam_frame)
    dslam_folder_frame.grid(row=9, column=1, padx=5, pady=5, sticky="ew")
    
    # Create a container frame to hold both entry and button
    entry_button_container = ttk.Frame(dslam_folder_frame)
    entry_button_container.pack(fill="x", expand=True)
    
    # Create entry widget
    dslam_folder_entry = ttk.Entry(entry_button_container, 
                                  textvariable=dslam_folder_var,
                                  width=67,
                                  foreground='black', 
                                  font=("Helvetica", 10, "bold"))
    dslam_folder_entry.pack(side="left", fill="x", expand=True)
    
    # Create browse button
    dslam_browse_button_2 = ttk.Button(entry_button_container, 
                                      text="Browse", 
                                      cursor="hand2",
                                      command=lambda: browse_with_default_path(dslam_folder_entry,
                                                                            dslam_folder_var,
                                                                            is_profile=True,
                                                                            profile_dropdown=dslam_profile_dropdown))
    dslam_browse_button_2.pack(side="left")  # Reduced padding between entry and button # Profile Folder section (combining both declarations into one correct version)
   
    # Add trace to the variable for debugging
    def on_folder_var_change(*args):
        print(f"DEBUG: dslam_folder_var changed to: {dslam_folder_var.get()}")
    dslam_folder_var.trace_add("write", on_folder_var_change)
    # Initialize dslam_folder_var if not already done
    if 'dslam_folder_var' not in globals() or not isinstance(dslam_folder_var, tk.StringVar):
        dslam_folder_var = tk.StringVar()
    
    
    # DSLAM Profile
    dslam_profile_dropdown = ttk.Combobox(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                         textvariable=dslam_profile_var, state='readonly', width=65)
    #dslam_profile_dropdown.grid(row=8, column=1, padx=5, pady=5)
    dropdowns['DSLAM_Profile']['widget'] = dslam_profile_dropdown

    # Get Files Button
    #dslam_get_files_button = ttk.Button(combined_config_frame, text="Get DSLAM Profiles", 
                                       #command=lambda: get_dslam_folder_files())

    #dslam_driver_frame.grid(row=8, column=1, padx=5, pady=5, sticky="ew")

    #dslam_driver_frame.grid(row=9, column=1, padx=5, pady=5, sticky="w")


    # For the profile dropdown - use ComboboxSelected event
    dslam_profile_dropdown.bind('<<ComboboxSelected>>', on_dslam_file_select)
    dslam_folder_entry.config(width=67)



    # Add bindings for DSLAM dropdowns
    dslam_name_dropdown.bind("<<ComboboxSelected>>", 
        lambda event: update_dslam_name_dropdown(dropdown_data, dslam_name_var.get(), 
                                           dslam_name_dropdown, dslam_serialnumber_dropdown, dslam_busmap_var,libname_var))
    dslam_serialnumber_dropdown.bind('<<ComboboxSelected>>', on_dslam_serialnumber_selected) 

    # For the profile dropdown - use ComboboxSelected event
    dslam_profile_dropdown.bind('<<ComboboxSelected>>', on_dslam_file_select)


        
    # Initialize dslam_driver_var if not already defined
    if dslam_driver_var is None or not isinstance(dslam_driver_var, tk.StringVar):
        dslam_driver_var = tk.StringVar()
        # Set default value - bcm_file should be defined earlier in your code
        try:
            default_value = bcm_file if 'bcm_file' in globals() else "../DSLAM/BCM6400x/BCM6400x_100.bin"
            dslam_driver_var.set(default_value)
            print(f"Initialize dslam_driver_var with: {default_value}")
        except Exception as e:
            print(f"Error setting default driver path: {str(e)}")
            dslam_driver_var.set("../DSLAM/BCM6400x/BCM6400x_100.bin")
    




    dropdowns['DSLAM_Driver'] = {'widget': dslam_driver_entry, 'var': dslam_driver_var}
    dslam_driver_button = ttk.Button(dslam_driver_frame, text="Browse", cursor="hand2",
                                command=lambda: browse_driver(dslam_driver_entry))
    dslam_driver_button.place(relx=1.0, rely=0.5, anchor="e")
    dslam_driver_entry.config(width=67)



    # Profile Name section
    ttk.Label(dslam_frame, text="Profile Name:", font=("Helvetica", 10, "bold")).grid(row=10, column=0, padx=5, pady=5, sticky="e")
    dslam_profile_dropdown = ttk.Combobox(dslam_frame, foreground='black', font=("Helvetica", 10, "bold"),
                                         textvariable=dslam_profile_var, state='readonly', width=67)
    dslam_profile_dropdown.grid(row=10, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DSLAM_Profile']['widget'] = dslam_profile_dropdown
    

    # CORRECT EVENT BINDING (ONE binding only)
    dslam_profile_dropdown.bind('<<ComboboxSelected>>', on_dslam_file_select)
    
    # Other DSLAM dropdown bindings
    dslam_name_dropdown.bind("<<ComboboxSelected>>", 
        lambda event: update_dslam_name_dropdown(dropdown_data, dslam_name_var.get(), 
                                               dslam_name_dropdown, dslam_serialnumber_dropdown, dslam_busmap_var,libname_var))
    dslam_serialnumber_dropdown.bind('<<ComboboxSelected>>', on_dslam_serialnumber_selected)


    # DUT Type
    ttk.Label(dut_frame, text="DUT Type:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    dut_name_dropdown = ttk.Combobox(dut_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                               textvariable=dut_name_var, state='readonly', width=67)
    dut_name_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Name']['widget'] = dut_name_dropdown
    #dut_libname_var = 
    #dropdowns['DUT_libName']['widget'] = dut_libname_var

    # DUT Serial Number
    ttk.Label(dut_frame, text="Serial Number:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    dut_serialnumber_dropdown = ttk.Combobox(dut_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                       textvariable=dut_serialnumber_var, state='readonly', width=67)
    dut_serialnumber_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Serialnumber']['widget'] = dut_serialnumber_dropdown

    # DUT Busmap
    ttk.Label(dut_frame, text="Busmap:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky="e")
    dut_busmap_dropdown = ttk.Combobox(dut_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                 textvariable=dut_busmap_var, state='readonly', width=67)
    dut_busmap_dropdown.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Busmap']['widget'] = dut_busmap_dropdown



    # DUT Driver
    ttk.Label(dut_frame, text="DUT Driver:", font=("Helvetica", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky="e") 
    dut_driver_frame = ttk.Frame(dut_frame)
    dut_driver_frame.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

    dut_driver_entry = ttk.Entry(dut_driver_frame, textvariable=dut_driver_var, width=67,
                                foreground='black', font=("Helvetica", 10, "bold"))
    dut_driver_entry.pack(side="left", fill="x", expand=True)
    dut_driver_button = ttk.Button(dut_driver_frame, text="Browse", cursor="hand2",
                             command=lambda: browse_driver(dut_driver_entry))
    dut_driver_button.pack(side="right")

    # DUT Port
    ttk.Label(dut_frame, text="Port:", font=("Helvetica", 10, "bold")).grid(row=4, column=0, padx=5, pady=5, sticky="e")
    dut_port_dropdown = ttk.Combobox(dut_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                               textvariable=dut_port_var, width=67, state='readonly')
    dut_port_dropdown.grid(row=4, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Port']['widget'] = dut_port_dropdown

    # DUT Host
    ttk.Label(dut_frame, text="Host:", font=("Helvetica", 10, "bold")).grid(row=5, column=0, padx=5, pady=5, sticky="e")
    dut_host_dropdown = ttk.Combobox(dut_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                               textvariable=dut_host_var, state='readonly', width=67)
    dut_host_dropdown.grid(row=5, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Host']['widget'] = dut_host_dropdown


    # DUT Folder
    ttk.Label(dut_frame, text="Profile Folder:", font=("Helvetica", 10, "bold")).grid(row=6, column=0, padx=5, pady=5, sticky="e")
    dut_folder_frame = ttk.Frame(dut_frame)
    dut_folder_frame.grid(row=6, column=1, padx=5, pady=5, sticky="ew")

    dut_folder_entry = ttk.Entry(dut_folder_frame, textvariable=dut_folder_var, width=67, 
                                foreground='black', font=("Helvetica", 10, "bold"))
    dut_folder_entry.pack(side="left", fill="x", expand=True)

    dut_browse_button = ttk.Button(
        dut_folder_frame, 
        text="Browse", 
        cursor="hand2",
        command=lambda: browse_with_default_path(
        None,
        dut_folder_var,
        is_profile=True,
        profile_dropdown=dut_profile_dropdown
        )
    )
    dut_browse_button.pack(side="right")



    # DUT Profile
    #ttk.Label(combined_config_frame, text="Profile:", font=("Helvetica", 10, "bold")).grid(row=8, column=3, padx=5, pady=5, sticky="e")
   # DUT Profile should be on DUT frame and gridded
    ttk.Label(dut_frame, text="Profile Name:", font=("Helvetica", 10, "bold")).grid(row=8, column=0, padx=5, pady=5, sticky="e")
    dut_profile_dropdown = ttk.Combobox(dut_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                  textvariable=dut_profile_var, state='readonly', width=67)
    dut_profile_dropdown.grid(row=8, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Profile']['widget'] = dut_profile_dropdown



    # dut_driver_entry = ttk.Entry(dut_driver_frame, textvariable=dut_driver_var, width=67,
    #                             foreground='black', font=("Helvetica", 10, "bold"))
    # dut_driver_entry.pack(side="left", fill="x", expand=True)
    
    # dut_driver_button = ttk.Button(dut_driver_frame, text="Browse", cursor="hand2",
    #                               command=lambda: browse_driver(dut_driver_entry))
    # dut_driver_button.place(relx=1.0, rely=0.5, anchor="e")
    # dut_driver_entry.config(width=67)

    # Initialize the port dropdown with full range
    full_port_range = [str(i) for i in range(0, 12)]  # Assuming max is 11
    full_port_range = ['None'] + full_port_range
    dslam_port_dropdown['values'] = full_port_range
    dslam_port_var.set(full_port_range[0])


    # And update the listbox binding
    dslam_profile_dropdown.bind('<<ListboxSelect>>', on_dslam_file_select)
    # Add bindings for DSLAM dropdowns
    dslam_name_dropdown.bind("<<ComboboxSelected>>", 
        lambda event: update_dslam_name_dropdown(dropdown_data, dslam_name_var.get(), 
                                               dslam_name_dropdown, dslam_serialnumber_dropdown, dslam_busmap_var,libname_var))
    dslam_serialnumber_dropdown.bind('<<ComboboxSelected>>', on_dslam_serialnumber_selected)

    # Update DSLAM Details Button
    #ttk.Button(combined_config_frame, text="Update DSLAM Details", cursor="hand2", 
               #command=lambda: update_dslamdrivers_thread(dslam_name_dropdown, dslam_name_var, 
                                                        #dslam_busmap_var, dslam_busmap_dropdown, 
                                                        #dslam_folder_var, dslam_profile_dropdown, 
                                                        #dslam_port_var, message_var, 
                                                        #dslam_slot_var, dslam_frame_var,dslam_profile_var)
   	#).grid(row=11, column=1, padx=5, pady=5)

    # Message Label for status/feedback
    message_var = tk.StringVar()
    message_label = ttk.Label(dslam_frame, textvariable=message_var, foreground='black', font=("Helvetica", 10, "bold"))
    #message_label.grid(row=10, column=1, padx=5, pady=5,columnspan=2)

    # Create a new validation frame directly in the combined_config_frame instead of settings_container
    validation_frame = ttk.LabelFrame(combined_config_frame, text="Validation Settings", padding="5 5 5 5")
    validation_frame.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")

    # Create a new platform frame directly in the combined_config_frame
    platform_frame = ttk.LabelFrame(combined_config_frame, text="Platform Configuration", padding="5 5 5 5")
    platform_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")


    # Configure row weights for the frames
    combined_config_frame.rowconfigure(1, weight=1)  # Row for output frame
    combined_config_frame.rowconfigure(2, weight=1)  # Row for results frame

    scenario_frame = ttk.LabelFrame(combined_config_frame, text="Test Scenario", padding="5 5 5 5")
    scenario_frame.grid(row=1, column=2, padx=5, pady=5, sticky="nsew")

   
    # === Validation Frame Setup ===
    # Configure columns for wider spacing in validation frame
    validation_frame.grid_columnconfigure(1, minsize=80)  # Value column
    validation_frame.grid_columnconfigure(2, minsize=80)  # Min column
    validation_frame.grid_columnconfigure(3, minsize=80)  # Max column

    # Value/Min/Max column headers
    ttk.Label(validation_frame, text="Value", font=("Helvetica", 10, "bold")).grid(row=0, column=1)
    ttk.Label(validation_frame, text="Min", font=("Helvetica", 10, "bold")).grid(row=0, column=2)
    ttk.Label(validation_frame, text="Max", font=("Helvetica", 10, "bold")).grid(row=0, column=3)

    # Startup Time row
    startup_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                 textvariable=startup_time_var, width=8)
    startup_time_entry.grid(row=2, column=1, padx=5, pady=5)
    min_startup_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                     textvariable=min_startup_time_var, width=8, state='readonly')
    min_startup_time_entry.grid(row=2, column=2, padx=5, pady=5)
    max_startup_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                     textvariable=max_startup_time_var, width=8, state='readonly')
    max_startup_time_entry.grid(row=2, column=3, padx=5, pady=5)

    # Stabilize Time row
    stabilize_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                   textvariable=stabilize_time, width=8)
    stabilize_time_entry.grid(row=3, column=1, padx=5, pady=5)
    min_stabilize_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                       textvariable=min_stabilize_time_var, width=8, state='readonly')
    min_stabilize_time_entry.grid(row=3, column=2, padx=5, pady=5)
    max_stabilize_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                       textvariable=max_stabilize_time_var, width=8, state='readonly')
    max_stabilize_time_entry.grid(row=3, column=3, padx=5, pady=5)

    # Error Free Time row
    error_free_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                    textvariable=error_free_time, width=8)
    error_free_time_entry.grid(row=4, column=1, padx=5, pady=5)
    min_error_free_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                        textvariable=min_error_free_time_var, width=8, state='readonly')
    min_error_free_time_entry.grid(row=4, column=2, padx=5, pady=5)
    max_error_free_time_entry = ttk.Entry(validation_frame, font=("Helvetica", 10, "bold"), 
                                        textvariable=max_error_free_time_var, width=8, state='readonly')
    max_error_free_time_entry.grid(row=4, column=3, padx=5, pady=5)

    # Move validation widgets to the validation frame
    # Startup Time row
    ttk.Label(validation_frame, text="Startup Time:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    startup_time_entry.grid(row=1, column=1, padx=5, pady=5)
    min_startup_time_entry.grid(row=1, column=2, padx=5, pady=5)
    max_startup_time_entry.grid(row=1, column=3, padx=5, pady=5)

    # Stabilize Time row
    ttk.Label(validation_frame, text="Stabilize Time:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky="e")
    stabilize_time_entry.grid(row=2, column=1, padx=5, pady=5)
    min_stabilize_time_entry.grid(row=2, column=2, padx=5, pady=5)
    max_stabilize_time_entry.grid(row=2, column=3, padx=5, pady=5)

    # Error Free Time row
    ttk.Label(validation_frame, text="Error Free Time:", font=("Helvetica", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky="e")
    error_free_time_entry.grid(row=3, column=1, padx=5, pady=5)
    min_error_free_time_entry.grid(row=3, column=2, padx=5, pady=5)
    max_error_free_time_entry.grid(row=3, column=3, padx=5, pady=5)

    
    # === Test Scenario Section ===
    # Move test scenario and layout to scenario_frame
    ttk.Label(scenario_frame, text="Test Scenario:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    testscenario_dropdown = ttk.Combobox(scenario_frame, foreground='black', font=("Helvetica", 10, "bold"),
                                   textvariable=testscenario_var, values=["FS", "UR"], width=30, state='readonly')
    testscenario_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Testscenario']['widget'] = testscenario_dropdown

    ttk.Label(scenario_frame, text="Layout:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    layout_dropdown = ttk.Combobox(scenario_frame, textvariable=layout_var, 
                             values=["default", "custom1"], width=30,
                             foreground='black', font=("Helvetica", 10, "bold"),
                             state='readonly')
    layout_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['Layout'] = {'var': layout_var, 'widget': layout_dropdown}


    #  Test Plan
    ttk.Label(scenario_frame, text="Test Plan:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky="e")
    dut_testplan_dropdown = ttk.Combobox(scenario_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                   textvariable=dut_testplan_var, state='readonly', width=30)
    dut_testplan_dropdown.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['DUT_Testplan']['widget'] = dut_testplan_dropdown

    #operator field
    ttk.Label(scenario_frame, text="Operator:", font=("Helvetica", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky="e")
    operator_dropdown = ttk.Combobox(scenario_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                   textvariable=operator_var, width=30)
    operator_dropdown.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
    dropdowns['Operator']['widget'] = operator_dropdown


    # Create platform-related dropdowns before placing them
    platform_name_dropdown = ttk.Combobox(platform_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                    textvariable=platform_name_var, width=30, state='readonly')
    dropdowns['Platform_name'] = {'var': platform_name_var, 'widget': platform_name_dropdown}

    platform_noise_generator_dropdown = ttk.Combobox(platform_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                               textvariable=platform_noise_generator_var, width=30, state='readonly')
    dropdowns['NoiseGenerator'] = {'var': platform_noise_generator_var, 'widget': platform_noise_generator_dropdown}

    platform_traffic_generator_dropdown = ttk.Combobox(platform_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                                 textvariable=platform_traffic_generator_var, width=30, state='readonly')
    dropdowns['trafficgenerator'] = {'var': platform_traffic_generator_var, 'widget': platform_traffic_generator_dropdown}

    platform_linesimulator_dropdown = ttk.Combobox(platform_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                             textvariable=platform_linesimulator_var, width=30, state='readonly')
    dropdowns['Linesimulator'] = {'var': platform_linesimulator_var, 'widget': platform_linesimulator_dropdown}

    platform_switch_dropdown = ttk.Combobox(platform_frame, foreground='black', font=("Helvetica", 10, "bold"), 
                                      textvariable=platform_switch_var, width=30, state='readonly')
    dropdowns['Switch'] = {'var': platform_switch_var, 'widget': platform_switch_dropdown}

    # === Platform Configuration Section ===
    # Simplify platform section headers and move widgets
    ttk.Label(platform_frame, text="Name:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    platform_name_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    platform_name_dropdown.config(width=30)  # Adjust width for the new layout

    ttk.Label(platform_frame, text="Noise Generator:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    platform_noise_generator_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    platform_noise_generator_dropdown.config(width=30)

    ttk.Label(platform_frame, text="Traffic Generator:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky="e")
    platform_traffic_generator_dropdown.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
    platform_traffic_generator_dropdown.config(width=30)

    ttk.Label(platform_frame, text="Line Simulator:", font=("Helvetica", 10, "bold")).grid(row=3, column=0, padx=5, pady=5, sticky="e")
    platform_linesimulator_dropdown.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
    platform_linesimulator_dropdown.config(width=30)

    ttk.Label(platform_frame, text="Switch:", font=("Helvetica", 10, "bold")).grid(row=4, column=0, padx=5, pady=5, sticky="e")
    platform_switch_dropdown.grid(row=4, column=1, padx=5, pady=5, sticky="ew")
    platform_switch_dropdown.config(width=30)

    dropdowns['Platform_name']['widget'] = platform_name_dropdown

    dropdowns['NoiseGenerator']['widget'] = platform_noise_generator_dropdown

    dropdowns['trafficgenerator']['widget'] = platform_traffic_generator_dropdown

    dropdowns['Linesimulator']['widget'] = platform_linesimulator_dropdown

    dropdowns['Switch']['widget'] = platform_switch_dropdown
    # === Output Section ===
    # Create frames
    database_frame = ttk.Frame(output_frame)
    database_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    database_frame.columnconfigure(0, weight=1)  # Make entry expand

    output_folder_frame = ttk.Frame(output_frame)
    output_folder_frame.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    output_folder_frame.columnconfigure(0, weight=1)  # Make entry expand
    
    # # Excel file entry and button
    # database_file_entry = ttk.Entry(database_frame, textvariable=database_file_var, width=67,
    #                               foreground='black', font=("Helvetica", 10, "bold"))
    # database_file_entry.grid(row=0, column=0, sticky="ew")

    # browse_button = ttk.Button(database_frame, text="Browse", cursor="hand2", 
    #                          command=lambda: browse_with_default_path(database_file_entry))
    # browse_button.grid(row=0, column=1, padx=(5, 0))

    # Output folder entry and button
    output_folder_entry = ttk.Entry(output_folder_frame, textvariable=output_folder_var, width=67,
                                  foreground='black', font=("Helvetica", 10, "bold"), state='readonly')
    output_folder_entry.grid(row=0, column=0, sticky="ew")

    output_browse_button = ttk.Button(output_folder_frame, text="Browse", cursor="hand2",
                                    command=lambda: browse_with_default_path(output_folder_entry, dslam_name_dropdown))
    output_browse_button.grid(row=0, column=1, padx=(5, 0))

    # Output frame labels and layout
    #ttk.Label(output_frame, text="Excel File:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    ttk.Label(output_frame, text="Folder:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    ttk.Label(output_frame, text="Filename:", font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=5, pady=5, sticky="e")

    output_filename_entry = ttk.Entry(output_frame, textvariable=output_filename_var, width=67,
                                    foreground='black', font=("Helvetica", 10, "bold"))
    output_filename_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

    # Configure output frame columns
    output_frame.columnconfigure(1, weight=1)

    # === Results Section ===
    # Create frames
    results_excel_frame = ttk.Frame(results_frame)
    results_excel_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    results_excel_frame.columnconfigure(0, weight=1)  # Make entry expand

    # Results folder section
    results_folder_frame = ttk.Frame(results_frame)
    results_folder_frame.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    results_folder_frame.columnconfigure(0, weight=1)
   

    
    # Results excel file entry and button
    results_excel_entry = ttk.Entry(results_excel_frame, textvariable=database_file_var, width=67,
                                  foreground='black', font=("Helvetica", 10, "bold"))
    results_excel_entry.grid(row=0, column=0, sticky="ew")

    results_excel_button = ttk.Button(results_excel_frame, text="Browse", cursor="hand2", 
                                    command=lambda: browse_file(results_excel_entry, database_file_var))
    results_excel_button.grid(row=0, column=1, padx=(5, 0))
    
    # Results folder entry and buttons
    results_folder_entry = ttk.Entry(results_folder_frame, textvariable=results_folder_var, width=67,
                               foreground='black', font=("Helvetica", 10, "bold"))
    results_folder_entry.grid(row=0, column=0, sticky="ew")

    # Button frame for Browse and Open
    results_folder_buttons = ttk.Frame(results_folder_frame)
    results_folder_buttons.grid(row=0, column=1, padx=(5, 0))

    results_folder_browse = ttk.Button(results_folder_buttons, text="Browse", cursor="hand2",
                                 command=lambda: browse_with_default_path(results_folder_entry,
                                                                       results_folder_var))
    results_folder_browse.grid(row=0, column=0, padx=(0, 2))

    results_folder_open = ttk.Button(results_folder_buttons, text="Open", cursor="hand2",
                               command=lambda: open_in_explorer(results_folder_entry,
                                                             results_folder_var))
    results_folder_open.grid(row=0, column=1)

    # Results frame labels
    ttk.Label(results_frame, text="Excel File:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    ttk.Label(results_frame, text="Results Folder:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")

    # Configure results frame columns
    results_frame.columnconfigure(1, weight=1)

    # Message label at bottom
    message_var = tk.StringVar()
    message_label = ttk.Label(combined_config_frame, textvariable=message_var, foreground='black', font=("Helvetica", 10, "bold"))
    #message_label.grid(row=3, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

    
    

   # Function to scroll entry field to the end when focused or when End key is pressed
    def scroll_to_end(event):
        entry = event.widget
        try:
            entry.icursor(tk.END)  # Move insertion cursor to the end
            entry.xview_moveto(1)  # Scroll view to the end
            print(f"Scrolled {entry} to end")  # Debug print
        except Exception as e:
            print(f"Error scrolling: {e}")
        return "break"  # Prevent default handling

    # Apply scrolling directly to specific widgets rather than in a loop
    results_excel_entry.bind("<FocusIn>", scroll_to_end)
    results_excel_entry.bind("<End>", scroll_to_end)
    results_excel_entry.bind("<Double-1>", lambda e: e.widget.select_range(0, tk.END))

    output_folder_entry.bind("<FocusIn>", scroll_to_end)
    output_folder_entry.bind("<End>", scroll_to_end)
    output_folder_entry.bind("<Double-1>", lambda e: e.widget.select_range(0, tk.END))

    dslam_driver_entry.bind("<FocusIn>", scroll_to_end)
    dslam_driver_entry.bind("<End>", scroll_to_end)
    dslam_driver_entry.bind("<Double-1>", lambda e: e.widget.select_range(0, tk.END))

    dut_driver_entry.bind("<FocusIn>", scroll_to_end)
    dut_driver_entry.bind("<End>", scroll_to_end)
    dut_driver_entry.bind("<Double-1>", lambda e: e.widget.select_range(0, tk.END))

    output_filename_entry.bind("<FocusIn>", scroll_to_end)
    output_filename_entry.bind("<End>", scroll_to_end)
    output_filename_entry.bind("<Double-1>", lambda e: e.widget.select_range(0, tk.END))

    # Alternative: Add a right-click menu with Copy All functionality
    def add_right_click_menu(entry_widget):
        popup_menu = tk.Menu(root, tearoff=0)
        popup_menu.add_command(label="Copy All", command=lambda: copy_all_text(entry_widget))
    
        # Function to copy all text to clipboard
        def copy_all_text(widget):
            content = widget.get()
            if content:
                root.clipboard_clear()
                root.clipboard_append(content)
                # Optional: Show brief feedback that text was copied
                print(f"Copied to clipboard: {content}")
                
                # Optional: Show visual feedback
                original_bg = widget.cget("background")
                widget.config(background="#e0f0ff")  # Light blue flash
                widget.after(200, lambda: widget.config(background=original_bg))
    
        # Show the popup menu on right click
        entry_widget.bind("<Button-3>", lambda e: popup_menu.post(e.x_root, e.y_root))

    # Add right-click menu to problematic entries
    for entry in [results_excel_entry, output_folder_entry, dslam_driver_entry, dut_driver_entry]:
        add_right_click_menu(entry)

   
    
    def add_tooltip_to_entry(entry_widget, parent_frame):
        """Add tooltip to entry widget that shows the full text without requiring scrolling"""
        # Flag to track if tooltip should stay visible
        entry_widget.tooltip_active = False
    
        # Function to show tooltip with full path on hover or click
        def show_tooltip(event):
            full_path = entry_widget.get()
            if full_path and not hasattr(entry_widget, 'tooltip'):
            # Create tooltip window
                tip = tk.Toplevel(parent_frame)
                tip.wm_overrideredirect(True)  # No window decorations
                tip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            
            # Calculate how many lines are needed based on text length
            # Assume about 50 chars per line with reasonable font
                text_length = len(full_path)
                line_count = max(1, min(5, (text_length // 50) + 1))  # At least 1, max 5 lines
            
            # Create a frame for the tooltip
                frame = ttk.Frame(tip)
                frame.pack(fill="both", expand=True)
            
            # Add text widget with word wrapping - no scrollbar needed
                text = tk.Text(frame, wrap="word", height=line_count, width=50,
                      background="#ffffe0", relief="solid", borderwidth=1)
                text.pack(fill="both", expand=True)
            
            # Insert the full path
                text.insert("1.0", full_path)
            
            # Make text selectable but not editable
                text.config(state="disabled")
            # Re-enable selection capability despite being disabled
                text.bind("<B1-Motion>", lambda e: "break")
                text.bind("<ButtonRelease-1>", lambda e: text.tag_add("sel", "1.0", "end"))
            
            # Store reference to destroy later
                entry_widget.tooltip = tip
            
            # Set auto-destruction after 2 seconds
                parent_frame.after(2000, lambda: destroy_tooltip())
    
    # Function to destroy tooltip
        def destroy_tooltip():
            if hasattr(entry_widget, 'tooltip'):
                entry_widget.tooltip.destroy()
                delattr(entry_widget, 'tooltip')
        # Show tooltip on hover (mouse enter)
        entry_widget.bind("<Enter>", show_tooltip)
    
        # Also show tooltip on single click
        entry_widget.bind("<Button-1>", show_tooltip, add="+")  # Use add='+' to not override existing click handlers
    
        # Keep double-click binding for good measure
        entry_widget.bind("<Double-1>", show_tooltip, add="+")
    
        # Add to right-click menu if it exists
        if "<Button-3>" in [binding[0] for binding in entry_widget.bind()]:
        # Entry already has a right-click menu - add our option to it
            orig_binding = entry_widget.bind("<Button-3>")
            entry_widget.unbind("<Button-3>")
        
            def combined_menu(event):
                menu = tk.Menu(root, tearoff=0)
                menu.add_command(label="Show Full Path", command=lambda: show_tooltip(event))
                menu.add_separator()
                # Call original menu handler
                orig_binding(event)
                menu.post(event.x_root, event.y_root)
            
            entry_widget.bind("<Button-3>", combined_menu)
    # Apply tooltip functionality to ALL entry widgets with browse buttons
    add_tooltip_to_entry(results_excel_entry, database_frame)
    add_tooltip_to_entry(output_folder_entry, output_folder_frame)
    add_tooltip_to_entry(dslam_driver_entry, dslam_frame) 
    add_tooltip_to_entry(dut_driver_entry, dut_frame)
    add_tooltip_to_entry(output_filename_entry, output_frame)
    add_tooltip_to_entry(dslam_folder_entry, dslam_frame)
    add_tooltip_to_entry(dut_folder_entry, dut_frame)
    
    # # Also add to any other entry widgets that need path viewing
    # for key, item in dropdowns.items():
    #     if 'widget' in item and item['widget'] is not None:
    #         widget = item['widget']
    #         if isinstance(widget, ttk.Entry) and widget not in [database_file_entry, output_folder_entry, 
    #                                                            dslam_driver_entry, dut_driver_entry, 
    #                                                            output_filename_entry, dslam_folder_entry,
    #                                                            dut_folder_entry]:
    #             add_tooltip_to_entry(widget, parent_frame=combined_config_frame)



    # Set default values explicitly for DSLAM/DPU fields
    dslam_name_var.set("Select a driver")
    dslam_name_dropdown.set(dslam_name_var.get())
    dslam_busmap_var.set("Select an Address")
    dslam_busmap_dropdown.set(dslam_busmap_var.get())
    dslam_port_dropdown.set(dslam_port_var.get())
    dslam_serialnumber_dropdown.set(dslam_serialnumber_var.get())
    dslam_slot_entry.insert(0, dslam_slot_var.get())
    dslam_frame_entry.insert(0, dslam_frame_var.get())
    dslam_driver_entry.insert(0, dslam_driver_var.get())
    dslam_folder_entry.insert(0, dslam_folder_var.get())
    dslam_profile_dropdown.set(dslam_profile_var.get())

   # DUT Configuration fields (middle section of combined_config_frame)
    
    # Add spacing between DUT and Validation sections
    ttk.Label(combined_config_frame, text="").grid(row=0, column=5, padx=20)
    # Bind the dropdown selection event to the update function
    dut_name_dropdown.bind("<<ComboboxSelected>>", lambda event: update_dut_name_dropdown(dropdown_data, dut_name_var.get(), dut_name_dropdown, dut_serialnumber_dropdown, dut_busmap_var))
    dut_serialnumber_dropdown.bind("<<ComboboxSelected>>", lambda event: update_dut_serialnumber_dropdown(dropdown_data, dut_serialnumber_var.get(), dut_serialnumber_dropdown, dut_name_var, dut_busmap_var,dut_libname_var))

    def browse_driver(entry_widget):
        try:
            # First check if there's an existing path in the entry
            current_path = entry_widget.get()
            initial_dir = None
            
            if current_path:
                if current_path.startswith('\\\\'): # Network path
                    # Convert network path to a format that tkinter can understand
                    initial_dir = os.path.dirname(current_path).replace('\\', '/')
                elif current_path.startswith('..'): # Relative path
                    # Try to resolve relative path
                    pwd_process = subprocess.Popen('cd', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                    current_dir, _ = pwd_process.communicate()
                    if current_dir:
                        full_path = os.path.abspath(os.path.join(current_dir.strip(), current_path))
                        initial_dir = os.path.dirname(full_path)
                else: # Local path
                    if os.path.exists(os.path.dirname(current_path)):
                        initial_dir = os.path.dirname(current_path)
            
            print(f"Opening file dialog at: {initial_dir}")  # Debug print
            
            # Open file dialog with initial directory if available
            file_path = filedialog.askopenfilename(
                title="Select Driver File",
                initialdir=initial_dir,
                filetypes=[("Binary Files", "*.bin"), ("All Files", "*.*")]
            )
            
            if file_path:
                # Convert back to Windows path format if needed
                if '/' in file_path:
                    file_path = file_path.replace('/', '\\')
                
                # Normalize the path before setting it
                normalized_path = normalize_path(file_path)

                # Update both the entry widget and the StringVar
                if entry_widget == dslam_driver_entry:
                    dslam_driver_var.set(normalized_path)
                    entry_widget.delete(0, tk.END)
                    entry_widget.insert(0, normalized_path)
                elif entry_widget == dut_driver_entry:
                    dut_driver_var.set(normalized_path)
                    entry_widget.delete(0, tk.END)
                    entry_widget.insert(0, normalized_path)
                
        except Exception as e:
            print(f"Browse driver error: {str(e)}")  # Debug print
            messagebox.showerror("Error", f"Error browsing for driver: {str(e)}") 
    # Set default values explicitly
    dut_name_dropdown.set(dut_name_var.get())
    dut_busmap_dropdown.set(dut_busmap_var.get())
    dut_driver_entry.insert(0, dut_driver_var.get())
    dut_serialnumber_dropdown.set(dut_serialnumber_var.get())
    dut_port_dropdown.set(dut_port_var.get())
    dut_profile_dropdown.set(dut_profile_var.get())
    dut_testplan_dropdown.set(dut_testplan_var.get())
    #file_list_dropdown.set(dut_profile_var.get())

    # Set default values explicitly for DSLAM dropdowns
    dslam_name_dropdown.set(dslam_name_var.get())
    dslam_busmap_dropdown.set(dslam_busmap_var.get())
    dslam_serialnumber_dropdown.set(dslam_serialnumber_var.get())
    dslam_port_dropdown.set(dslam_port_var.get())
    dslam_profile_dropdown.set(dslam_profile_var.get())

    # Set widget references in dropdowns dictionary for DSLAM
    dropdowns['DSLAM_Name']['widget'] = dslam_name_dropdown
    dropdowns['DSLAM_Busmap']['widget'] = dslam_busmap_dropdown
    dropdowns['DSLAM_Serialnumber']['widget'] = dslam_serialnumber_dropdown
    dropdowns['DSLAM_Port']['widget'] = dslam_port_dropdown
    dropdowns['DSLAM_Profile']['widget'] = dslam_profile_dropdown
    dropdowns['DSLAM_Driver']['widget'] = dslam_driver_entry
    # Set widget references in dropdowns dictionary for DUT
    dropdowns['DUT_Name']['widget'] = dut_name_dropdown
    dropdowns['DUT_Serialnumber']['widget'] = dut_serialnumber_dropdown
    dropdowns['DUT_Port']['widget'] = dut_port_dropdown
    dropdowns['DUT_Busmap']['widget'] = dut_busmap_dropdown
    dropdowns['DUT_Testplan']['widget'] = dut_testplan_dropdown
    dropdowns['DUT_Host']['widget'] = dut_host_dropdown
    dropdowns['DUT_Driver']['widget'] = dut_driver_entry


    # Output Results Section
    #ttk.Label(combined_config_frame, text="Output Results", font=("Helvetica", 10, "bold")).grid(row=8, column=10, columnspan=2, padx=5, pady=5)
    
 

    # Bind the platform dropdown selection event
    platform_name_dropdown.bind("<<ComboboxSelected>>", lambda event: on_platform_selected(
        platform_name_dropdown, dropdown_data, platform_name_var, 
        platform_linesimulator_var, platform_noise_generator_var,
        platform_traffic_generator_dropdown, platform_switch_var,
        platform_switch_dropdown, platform_traffic_generator_var,
        platform_linesimulator_dropdown, platform_noise_generator_dropdown))
  

    dropdowns['Platform_name']['widget'] = platform_name_dropdown
    dropdowns['Linesimulator']['widget'] = platform_linesimulator_dropdown
    dropdowns['NoiseGenerator']['widget'] = platform_noise_generator_dropdown
    dropdowns['Switch']['widget'] = platform_switch_dropdown
    dropdowns['trafficgenerator']['widget'] = platform_traffic_generator_dropdown

   


    # Update the dropdowns dictionary with the widget references
    # Update the dropdowns dictionary with the widget references
    dropdowns = {
        # DSLAM Configuration
        'DSLAM_Name': {'var': dslam_name_var, 'widget': dslam_name_dropdown},
        'DSLAM_Serialnumber': {'var': dslam_serialnumber_var, 'widget': dslam_serialnumber_dropdown},  # New entry
        'DSLAM_Busmap': {'var': dslam_busmap_var, 'widget': dslam_busmap_dropdown},
        'DSLAM_Driver': {'var': dslam_driver_var, 'widget': None},
        'DSLAM_Port': {'var': dslam_port_var, 'widget': dslam_port_dropdown},
        'DSLAM_Slot': {'var': dslam_slot_var, 'widget': None},
        'DSLAM_Frame': {'var': dslam_frame_var, 'widget': None},
        'DSLAM_Folder': {'var': dslam_folder_var, 'widget': None},
        'DSLAM_Host': {'var': dslam_host_var, 'widget':dslam_host_dropdown},
        'DSLAM_Profile': {'var': dslam_profile_var, 'widget': dslam_profile_dropdown},
        
        # DUT Configuration
        'DUT_Name': {'var': dut_name_var, 'widget': dut_name_dropdown},
        'DUT_Serialnumber': {'var': dut_serialnumber_var, 'widget': dut_serialnumber_dropdown},
        'DUT_Port': {'var': dut_port_var, 'widget': dut_port_dropdown},
        'DUT_Busmap': {'var': dut_busmap_var, 'widget': dut_busmap_dropdown},
        'DUT_Driver': {'var': dut_driver_var, 'widget': None},
        'DUT_Profile': {'var': dut_profile_var, 'widget': dut_profile_dropdown},
        'DUT_Testplan': {'var': dut_testplan_var, 'widget': dut_testplan_dropdown},
        'DUT_Host': {'var': dut_host_var, 'widget': dut_host_dropdown},
        'DUT_Testscenario': {'var': testscenario_var, 'widget': testscenario_dropdown},
        'DUT_Folder': {'var': dut_folder_var, 'widget': None},
        
        # Platform Configuration
        'Platform_name': {'var': platform_name_var, 'widget': platform_name_dropdown},
        'Linesimulator': {'var': platform_noise_generator_var, 'widget': platform_linesimulator_dropdown},
        'NoiseGenerator': {'var': platform_linesimulator_var, 'widget': platform_noise_generator_dropdown},
        'Switch': {'var': platform_switch_var, 'widget': platform_switch_dropdown},
        'trafficgenerator': {'var': platform_traffic_generator_var, 'widget': platform_traffic_generator_dropdown},
        
        # Layout Configuration
        'Layout': {'var': layout_var, 'widget': layout_dropdown},
        'Operator': {'var': operator_var,'widget': operator_dropdown}
    }
    # # Save button for layout configuration
    # save_layout_button = ttk.Button(page2_frame, text="Save Layout Configuration", command=save_layout_config)
    # save_layout_button.grid(row=9, column=7, padx=5, pady=5)

    #df = load_excel_data(file_path)
   
    # Populate the dropdowns
    #populate_dropdowns(df, dropdowns)


    # Make sure DSLAM driver entry shows the correct value
    if 'DSLAM_Driver' in dropdowns and 'widget' in dropdowns['DSLAM_Driver']:
        driver_widget = dropdowns['DSLAM_Driver']['widget']
        if hasattr(driver_widget, 'delete') and hasattr(driver_widget, 'insert'):
            driver_widget.delete(0, 'end')
            driver_widget.insert(0, dslam_driver_var.get())
            print(f"Set DSLAM driver entry to: {dslam_driver_var.get()}")


    try:
        # Get the file path
        file_path = get_excel_file_path()
        
        # Define sheet names
        sheet_name = "DUT_serialnumber"
        sheet_name1 = "DUT_serialnumber"
        sheet_name2 = "DUT_testplan"
        sheet_name3 = "DUT_host"
        sheet_name4 = 'Validation_Fields'
        sheet_name5 = 'Platform_data'
        sheet_name6 = 'Platform_Fields'
        sheet_name7 = 'DSLAM_serialnumber'
        sheet_name8 = 'operator'

        # Load and populate dropdowns
        dropdown_data = load_and_populate_dropdowns(
            file_path, 
            sheet_name, 
            dropdowns, 
            dut_name_var, 
            dut_testplan_var, 
            dut_host_var, 
            dut_busmap_var, 
            dut_serialnumber_var,
            dslam_name_var,
            dslam_serialnumber_var,
            dslam_busmap_var,
            dslam_host_var,
            testscenario_var, 
            min_startup_time_var, 
            max_startup_time_var, 
            min_error_free_time_var, 
            max_error_free_time_var, 
            min_stabilize_time_var, 
            max_stabilize_time_var, 
            platform_name_var, 
            platform_linesimulator_var, 
            platform_noise_generator_var, 
            platform_switch_var, 
            platform_traffic_generator_var,
            operator_var
        )

        # Update with other sheets
        sheets_to_load = [
            sheet_name1, sheet_name2, sheet_name3, sheet_name4, 
            sheet_name5, sheet_name6, sheet_name7,sheet_name8
        ]
        
        for sheet in sheets_to_load:
            dropdown_data.update(load_and_populate_dropdowns(
                file_path, sheet, dropdowns,
                dut_name_var, dut_testplan_var, dut_host_var, dut_busmap_var, 
                dut_serialnumber_var,
                dslam_name_var, dslam_serialnumber_var, dslam_busmap_var,dslam_host_var,
                testscenario_var, min_startup_time_var, max_startup_time_var,
                min_error_free_time_var, max_error_free_time_var,
                min_stabilize_time_var, max_stabilize_time_var,
                platform_name_var, platform_linesimulator_var,
                platform_noise_generator_var, platform_switch_var,
                platform_traffic_generator_var,operator_var
            ))

    except FileNotFoundError as e:
        messagebox.showerror("Error", str(e))
        # Initialize dropdown_data as empty if file loading fails
        dropdown_data = {}
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred:\n{str(e)}")
        dropdown_data = {}

    return page2_frame

# page3_frame functions

# Function to delete the selected row
def add_checkboxes_to_treeview(tree, frame):
    checkboxes = {}
   
    # Add header checkbox
    header_var = tk.BooleanVar()
    header_checkbox = ttk.Checkbutton(frame, variable=header_var, command=lambda: select_all(tree, checkboxes, header_var))
    tree.update_idletasks()
   
    # Place header checkbox
    bbox = tree.bbox("", column="#0")  # Get bbox of first visible column
    if bbox:
        x, y, width, height = bbox
        header_checkbox.place(x=x, y=0, width=width, height=height)
   
    # Add checkboxes for each row
    for item in tree.get_children():
        var = tk.BooleanVar()
        checkboxes[item] = var
        checkbox = ttk.Checkbutton(frame, variable=var, command=lambda item=item: toggle_row(tree, item))
        tree.update_idletasks()
        bbox = tree.bbox(item, column="#0")
        if bbox:
            x, y, width, height = bbox
            checkbox.place(x=x, y=y, width=width, height=height)
   
    return checkboxes

def select_all(tree, checkboxes, header_var):
    for item, var in checkboxes.items():
        var.set(header_var.get())
        tree.item(item, tags=('checked' if header_var.get() else '',))

def toggle_row(tree, item):
    current_tags = tree.item(item, "tags")
    if "checked" in current_tags:
        tree.item(item, tags=())
    else:
        tree.item(item, tags=("checked",))



def delete_selected_rows(tree, checkboxes):
    for item in list(checkboxes.keys()):  # Use list() to avoid modifying dict during iteration
        if "checked" in tree.item(item, "tags"):
            tree.delete(item)
            del checkboxes[item]
    # Reset header checkbox
    header_checkbox = tree.master.children.get('!checkbutton')
    if header_checkbox:
        header_checkbox.state(['!selected'])





def add_header_checkbox(tree, frame, checkboxes):
    header_var = tk.BooleanVar()
    header_checkbox = ttk.Checkbutton(frame, variable=header_var, command=lambda: select_all(tree, checkboxes, header_var))
    tree.update_idletasks()
    bbox = tree.bbox("heading")
    if bbox:
        x, y, width, height = bbox
        header_checkbox.place(x=x + tree.winfo_rootx(), y=y + tree.winfo_rooty(), width=width, height=height)

def toggle_checkbox(event, checkbox):
    checkbox.invoke()

def select_all(tree, checkboxes, header_var):
    for var in checkboxes.values():
        var.set(header_var.get())



# Function to switch pages
def show_frame(frame):
    frame.tkraise()


def use_preferences():
    global driver_path_changed
    try:
        display_message("Starting to load preferences...", "black")
        # Start the progress bar
        #progress_bar.start()
       
        file_path = filedialog.askopenfilename(
            title="Select Preferences File",
            filetypes=[("INI files", "*.ini"), ("All files", "*.*")]
        )
        if file_path:
            config.read(file_path)
            if 'PREFERENCES' in config:
                display_message("Preferences file loaded successfully.", "black")
               
                ip_entry.delete(0, tk.END)
                ip_entry.insert(0, config['PREFERENCES']['IP'])
               
                port_entry.delete(0, tk.END)
                port_entry.insert(0, config['PREFERENCES']['Port'])
               
                # Load the driver path and reset the flag
                driver_path = config['PREFERENCES']['Driver']
                driver_path_var.set(driver_path)
                driver_path_changed = False  # Reset the flag since we're loading preferences
                display_message(f"Driver path set to: {driver_path}", "black")
               
                profile_path_var.set(config['PREFERENCES']['ProfilePath'])
                selected_port = config['PREFERENCES']['SelectedPort']
                port_var.set(selected_port)
               
                # Load bcm_address
                bcm_address = config['PREFERENCES']['BCMAddress']
               
                # Populate the profile listbox with the default file name profile
                default_file_name_profile = config['PREFERENCES']['FileNameProfile']
                profile_listbox.delete(0, tk.END)
                profile_listbox.insert(tk.END, default_file_name_profile)
                profile_listbox.selection_set(0)
                profile_listbox.activate(0)
                profile_listbox.event_generate("<<ListboxSelect>>")
               
                display_message("Preferences loaded successfully.", "black")
            else:
                display_message("No preferences found in the selected file.", "red")
        else:
            display_message("No file selected.", "red")
    except Exception as e:
        display_message(f"Failed to load preferences: {str(e)}", "red")
    finally:
        # Stop the progress bar
        #progress_bar.stop()
        display_message("Finished loading preferences.", "black")


profile_path_var = tk.StringVar(value="..\\ICL_Server\\DSLAM\\DMS4\\Profiles")

# Flag to track if DSLAM drivers have been retrieved
drivers_retrieved = False



# Validate IP address
def validate_ip(ip):
    import ipaddress
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False

# Validate Port number
def validate_port(port):
    return port.isdigit() and 1 <= int(port) <= 65535

# Function to log commands and responses to an Excel file
def log_command_to_excel(ip, port, cmd, response, command_status, connection_status, end_time):
    try:
        # Check if the Excel file exists
        file_exists = os.path.isfile("server_log.xlsx")
        # Create a new workbook if the file does not exist
        if not file_exists:
            wb = Workbook()
            ws = wb.active
            ws.title = "Server Log"

            # Add headers
            ws.append(["IP", "Port", "Date", "Time", "Command", "Response", "Command Status", "Connection Status", "End Time"])
        else:
            # Load the existing workbook
            wb = openpyxl.load_workbook("server_log.xlsx")
            ws = wb.active

        # Get the current date and time
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")

        # Clean the response to remove non-printable characters
        clean_response = ''.join(c for c in response if c.isprintable())

        # Append the log entry to the worksheet
        ws.append([ip, port, date_str, time_str, cmd, clean_response, command_status, connection_status, end_time])

        # Save the workbook and close it properly
        wb.save("server_log.xlsx")
        wb.close()

    except openpyxl.utils.exceptions.InvalidFileException as e:
        print(f"Failed to log command due to file error: {str(e)}")

    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")

def stop_server_thread():
    threading.Thread(target=stop_server).start()

def start_server_thread():
    threading.Thread(target=start_server).start()

def is_admin():
    """Check if the program has admin rights"""
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False
def check_permissions(computer_name):
    """Check if current user has admin rights on the remote computer"""
    try:
        # Try to query services (requires admin rights)
        test_cmd = f'sc.exe \\\\{computer_name} query'
        result = subprocess.run(test_cmd, shell=True, capture_output=True, text=True)
        
        # Return True if no "Access denied" error
        return "FAILED 5" not in result.stderr
        
    except Exception as e:
        message_queue.put((f"Error checking permissions: {str(e)}", "red"))
        return False

def stop_server():
    """Stop ICL Server on remote computer"""
    try:
        remote_computer = ip_entry.get()
        
        # Step 1: Test basic connectivity
        ping_cmd = f'ping -n 1 {remote_computer}'
        ping_result = subprocess.run(ping_cmd, shell=True, capture_output=True, text=True)
        
        if "Reply from" not in ping_result.stdout:
            message_queue.put((f"Cannot reach {remote_computer}. Check if it's online.", "red"))
            return
            
        # Step 2: Check permissions
        if not check_permissions(remote_computer):
            message_queue.put((
                f"Access denied on {remote_computer}.\n"
                f"You need administrative rights to manage services.", "red"))
            return
            
        # Step 3: Check if service exists and get its status
        query_cmd = f'sc.exe \\\\{remote_computer} query ICLServerService'
        service_result = subprocess.run(query_cmd, shell=True, capture_output=True, text=True)
        
        if "1060" in service_result.stderr:
            message_queue.put((f"ICLServerService is not installed on {remote_computer}", "red"))
            return
            
        # Step 4: Stop the service if it's running
        if "RUNNING" in service_result.stdout:
            message_queue.put((f"Attempting to stop service on {remote_computer}...", "black"))
            
            stop_cmd = f'sc.exe \\\\{remote_computer} stop ICLServerService'
            stop_result = subprocess.run(stop_cmd, shell=True, capture_output=True, text=True)
            
            if stop_result.returncode == 0:
                message_queue.put((f"Successfully stopped server on {remote_computer}", "black"))
            else:
                message_queue.put((f"Failed to stop server: {stop_result.stderr}", "red"))
                
        elif "STOPPED" in service_result.stdout:
            message_queue.put((f"Service is already stopped on {remote_computer}", "black"))
        else:
            message_queue.put((f"Unexpected service status on {remote_computer}", "red"))
            
    except Exception as e:
        message_queue.put((f"Error stopping remote server: {str(e)}", "red"))

def get_remote_netstat(remote_computer):
    """Get netstat output from remote computer"""
    try:
        # Test basic connectivity
        ping_cmd = f'ping -n 1 {remote_computer}'
        ping_result = subprocess.run(ping_cmd, shell=True, capture_output=True, text=True)
        
        if "Reply from" not in ping_result.stdout:
            message_queue.put((f"Cannot reach {remote_computer}. Check if it's online.", "red"))
            return None
            
        # Run basic netstat (no admin needed)
        netstat_cmd = f'netstat -n | findstr ":12005"'  # Filter for ICL Server port
        result = subprocess.run(netstat_cmd, shell=True, capture_output=True, text=True)
        
        if result.returncode == 0:
            return result.stdout
        else:
            message_queue.put((f"Failed to get network info: {result.stderr}", "red"))
            return None
            
    except Exception as e:
        message_queue.put((f"Error getting network info: {str(e)}", "red"))
        return None
            
def start_server():
    """Start ICL Server on remote computer"""
    try:
        remote_computer = ip_entry.get()
        
        # Step 1: Test basic connectivity
        ping_cmd = f'ping -n 1 {remote_computer}'
        ping_result = subprocess.run(ping_cmd, shell=True, capture_output=True, text=True)
        
        if "Reply from" not in ping_result.stdout:
            message_queue.put((f"Cannot reach {remote_computer}. Check if it's online.", "red"))
            return
            
        # Step 2: Check permissions
        if not check_permissions(remote_computer):
            message_queue.put((
                f"Access denied on {remote_computer}.\n"
                f"You need administrative rights to manage services.", "red"))
            return
            
        # Step 3: Check if service exists and get its status
        query_cmd = f'sc.exe \\\\{remote_computer} query ICLServerService'
        service_result = subprocess.run(query_cmd, shell=True, capture_output=True, text=True)
        
        if "1060" in service_result.stderr:
            message_queue.put((f"ICLServerService is not installed on {remote_computer}", "red"))
            return
            
        # Step 4: Start the service if it's not running
        if "STOPPED" in service_result.stdout:
            message_queue.put((f"Attempting to start service on {remote_computer}...", "black"))
            
            start_cmd = f'sc.exe \\\\{remote_computer} start ICLServerService'
            start_result = subprocess.run(start_cmd, shell=True, capture_output=True, text=True)
            
            if start_result.returncode == 0:
                message_queue.put((f"Successfully started server on {remote_computer}", "black"))
            else:
                message_queue.put((f"Failed to start server: {start_result.stderr}", "red"))
                
        elif "RUNNING" in service_result.stdout:
            message_queue.put((f"Service is already running on {remote_computer}", "black"))
        else:
            message_queue.put((f"Unexpected service status on {remote_computer}", "red"))
            
    except Exception as e:
        message_queue.put((f"Error starting remote server: {str(e)}", "red"))

# Thread-safe connect function
def connect_thread():
    threading.Thread(target=connect).start()
def disconnect_thread():
    threading.Thread(target=disconnect).start()

def reconnect_thread():
    threading.Thread(target=disconnect).start()
    threading.Thread(target=connect).start()


# Add the disconnect function
def disconnect():
    """Disconnect from ICL Server by closing socket"""
    try:
        global sock , current_log_file
        
        if sock is not None:
            try:
                disconnect_button.config(state='disabled')
                connect_button.config(state='disabled')
                # Properly shutdown the socket
                sock.shutdown(socket.SHUT_RDWR)
                sock.close()
                message_queue.put(("Successfully disconnected from server", "black"))
                
                
                # Add a disconnection message to the log file
                if current_log_file and os.path.exists(current_log_file):
                    with open(current_log_file, 'a', encoding='utf-8') as f:
                        f.write("\n=== Disconnected ===\n")
                        f.write(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                        f.write("="*40 + "\n\n")
                
                # Reset current_log_file to None to indicate no active connection log
                current_log_file = None
                # Update GUI elements
                status_label.config(text="Disconnected")
                connect_button.config(state='normal')
                disconnect_button.config(state='normal')
                
            except socket.error as e:
                message_queue.put((f"Socket error during disconnect: {str(e)}", "red"))
                disconnect_button.config(state='normal')
                connect_button.config(state='normal')
            finally:
                # Always reset the socket variable
                sock = None
        else:
            message_queue.put(("No active connection to disconnect", "black"))
            disconnect_button.config(state='normal')
            connect_button.config(state='normal')
            

    except Exception as e:
        message_queue.put((f"Error during disconnect: {str(e)}", "red"))

# Thread-safe send commands function
def send_commands_thread():
    threading.Thread(target=send_commands).start()

def get_icl_log_path():
    """Get ICL Server log file path"""
    try:
        # Get current working directory
        current_path = os.getcwd()
        
        # Try to find ICL_Server/Logs relative to current path
        possible_relative_paths = [
            os.path.join(current_path, "..", "..", "ICL_Server", "Logs"),  # Up two levels
            os.path.join(current_path, "..", "ICL_Server", "Logs"),        # Up one level
            os.path.join(current_path, "ICL_Server", "Logs")               # Same level
        ]
        
        # Check if any path exists and contains log files
        for path in possible_relative_paths:
            abs_path = os.path.abspath(path)  # Convert to absolute path
            if os.path.exists(abs_path):
                log_files = glob.glob(os.path.join(abs_path, "ICLServerLog_*.txt"))
                if log_files:
                    return abs_path
        
        # If no existing ICL logs found, create our own logs directory
        local_log_path = os.path.join(current_path, "ICL_Logs")
        os.makedirs(local_log_path, exist_ok=True)
        return local_log_path
        
    except Exception as e:
        message_queue.put((f"Error finding log path: {str(e)}", "red"))
        return None
    
def refresh_network_connections():
    """Refresh network connections table"""
    try:
        network_refresh_button.config(state='disabled')
        if sock is None:
            message_queue.put(("Not connected to the server. Please connect first.", "red"))
            return
        # Check for admin rights first
        if is_admin():
            output = run_as_admin('netstat -b -n')
            if output:
                process_netstat_output(output, netstat_content_frame, netstat_canvas)
            else:
                raise Exception("No output from admin netstat")
        else:
            # Fallback to basic netstat
            remote_computer = ip_entry.get()
            output = get_remote_netstat(remote_computer)
            if output:
                process_netstat_output(output, netstat_content_frame, netstat_canvas)
            else:
                message_queue.put(("Failed to get netstat output", "red"))
        network_refresh_button.config(state='normal')
    except Exception as e:
        message_queue.put((f"Error refreshing network connections: {str(e)}", "red"))
        network_refresh_button.config(state='normal')

def refresh_network_connections_thread():
    threading.Thread(target=refresh_network_connections, daemon=True).start()

def get_command_queue_thread():
    threading.Thread(target=get_command_queue, daemon=True).start()

def create_connections_table(server_details_frame):
    """Create table to show connected clients"""
    global connection_content_frame, connection_canvas
    print("Creating connections table...")
    
    # Add Command Queue label (already exists in parent frame)
    #ttk.Label(server_details_frame, 
    #         text="Command Queue:", 
    #         font=("Helvetica", 10, "bold")).grid(row=10, column=0, padx=5, pady=5, sticky="w")
             
    # Create table frame
    table_frame = ttk.Frame(server_details_frame)
    table_frame.grid(row=11, column=0, columnspan=2, sticky="ew", padx=5)

    # Create header frame
    header_frame = ttk.Frame(table_frame)
    header_frame.grid(row=0, column=0, columnspan=2, sticky="ew")

    # Table headers with specific widths - match netstat table
    headers = [
        ("Client", 15),
        ("Command", 40)
    ]

    for i, (text, width) in enumerate(headers):
        ttk.Label(header_frame, 
                 text=text, 
                 width=width, 
                 font=("Helvetica", 10, "bold"),
                 anchor="w").grid(row=0, column=i, padx=2)

    # Create scrollable content area
    canvas = tk.Canvas(table_frame, height=125)  # Same height as netstat table
    canvas.grid(row=1, column=0, sticky="ew")
    
    scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=canvas.yview)
    scrollbar.grid(row=1, column=1, sticky="ns")
    
    content_frame = ttk.Frame(canvas)

    # Configure scrolling
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas_frame = canvas.create_window((0, 0), window=content_frame, anchor="nw", width=canvas.winfo_width())

    # Configure canvas scrolling
    def configure_scroll_region(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def configure_canvas(event):
        canvas.itemconfig(canvas_frame, width=event.width)

    content_frame.bind("<Configure>", configure_scroll_region)
    canvas.bind("<Configure>", configure_canvas)

    connection_content_frame = content_frame
    connection_canvas = canvas
    
    print("Table created successfully")
    return content_frame, canvas
 
#Update add_connection_row function signature and implementation
def add_connection_row(content_frame, client, command):
    """Add a row to the connections table"""
    try:
        print(f"Adding row: {client} - {command}")
        
        # Get the next row number
        row_num = len(content_frame.winfo_children()) // 2  # Account for separators
        
        row_frame = ttk.Frame(content_frame)
        row_frame.grid(row=row_num*2, column=0, sticky="ew")  # Multiply by 2 to leave space for separators

        # Client text widget
        client_text = tk.Text(row_frame, width=15, height=3)
        client_text.grid(row=0, column=0, padx=2, sticky='w')
        client_text.insert('1.0', client)
        client_text.configure(state='disabled')

        # Command text widget
        command_text = tk.Text(row_frame, width=40, height=3)
        command_text.grid(row=0, column=1, padx=2, sticky='w')
        command_text.insert('1.0', command)
        command_text.configure(state='disabled')

        # Configure text widget appearance - Remove background setting
        for widget in (client_text, command_text):
            widget.configure(
                relief='flat',
                borderwidth=0,
                highlightthickness=0
            )

        # Add separator
        sep = ttk.Separator(content_frame)
        sep.grid(row=row_num*2+1, column=0, columnspan=2, sticky="ew", pady=1)

        print("Row added successfully")
        return row_frame

    except Exception as e:
        print(f"Error adding connection row: {str(e)}")
        return None

    
def create_netstat_table(server_details_frame):
    """Create table to show connected clients from netstat"""
    # Table frame - start at row 13 (after command queue)
    #ttk.Label(server_details_frame, 
    #         text="Network Connections:", 
    #         font=("Helvetica", 10, "bold")).grid(row=13, column=0, padx=5, pady=5, sticky="w")
             
    table_frame = ttk.Frame(server_details_frame)
    table_frame.grid(row=14, column=0, columnspan=2, sticky="ew", padx=5)  # Changed row number

    # Create header frame
    header_frame = ttk.Frame(table_frame)
    header_frame.grid(row=0, column=0, columnspan=2, sticky="ew")  # Changed to grid

    # Table headers with specific widths
    headers = [
        ("Client", 15),
        ("Connection Details", 40)
    ]

    for i, (text, width) in enumerate(headers):
        ttk.Label(header_frame, 
                 text=text, 
                 width=width, 
                 font=("Helvetica", 10, "bold"),
                 anchor="w").grid(row=0, column=i, padx=2)

    # Create scrollable content area
    canvas = tk.Canvas(table_frame, height=125)
    canvas.grid(row=1, column=0, sticky="ew")
    
    scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=canvas.yview)
    scrollbar.grid(row=1, column=1, sticky="ns")
    
    netstat_content_frame = ttk.Frame(canvas)

    # Configure scrolling
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas_frame = canvas.create_window((0, 0), window=netstat_content_frame, anchor="nw", width=canvas.winfo_width())

    # Configure canvas scrolling
    def configure_scroll_region(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def configure_canvas(event):
        canvas.itemconfig(canvas_frame, width=event.width)

    netstat_content_frame.bind("<Configure>", configure_scroll_region)
    canvas.bind("<Configure>", configure_canvas)

    return netstat_content_frame, canvas


def update_command_queue_table(lines):
    """Update command queue table in main thread"""
    try:
        print(f"Updating table with {len(lines)} lines")  # Debug print
        
        if not connection_content_frame:
            print("Error: connection_content_frame is None")
            return
            
        # Clear existing content
        for widget in connection_content_frame.winfo_children():
            widget.destroy()
            
        # Process each command line
        for line in lines:
            print(f"Processing line: {line}")  # Debug print
            if '<' in line and '>' in line:
                try:
                    parts = line.split('->')
                    if len(parts) == 2:
                        client = parts[0].strip()[1:-1]  # Remove < >
                        command = parts[1].strip()
                        add_connection_row(connection_content_frame, client, command)
                except Exception as e:
                    print(f"Error processing line: {str(e)}")
                    continue
        
        # Update scroll region
        if connection_canvas:
            connection_content_frame.update_idletasks()
            connection_canvas.configure(scrollregion=connection_canvas.bbox("all"))
            print("Scroll region updated")  # Debug print
        
    except Exception as e:
        print(f"Error updating command queue table: {str(e)}")
        
def get_command_queue():
    """Get current command queue status"""
    global connection_content_frame, connection_canvas, server_details_frame  # Add reference to server_details_frame
    try:
        command_refresh_button.config(state='disabled')
        if sock is None:
            message_queue.put(("Not connected to the server. Please connect first.", "red"))
            return
        print("Getting command queue...")
        response = socket_query(sock, "commandqueue", timeout=10)
        
        if response:
            print(f"Response received: {response}")
            lines = [line.strip() for line in response.split('\n') 
                    if line.strip() and not line.strip() in ['>', '! Invalid characters found:']]
            
            print(f"Filtered lines: {lines}")
            
            # Ensure table exists before updating
            if connection_content_frame is None or connection_canvas is None:
                print("Table not initialized, creating now...")
                # Use the global server_details_frame instead of finding by path
                connection_content_frame, connection_canvas = create_connections_table(server_details_frame)
                
            # Now that we're sure the table exists, update it
            def update():
                update_command_queue_table(lines)
            root.after(100, update)
            
            if lines:
                message_queue.put(("Command Queue:\n" + "\n".join(lines), "black"))
                command_refresh_button.config(state='normal')
            else:
                message_queue.put(("No active commands in queue", "black"))
                command_refresh_button.config(state='normal')

    except Exception as e:
        print(f"Error in get_command_queue: {str(e)}")
        message_queue.put((f"Error getting command queue: {str(e)}", "red"))        
        command_refresh_button.config(state='normal')

def add_netstat_row(netstat_content_frame, client, executable, ip_port, status):
    """Add a row to the netstat connections table"""
    try:
        # Get the next row number
        row_num = len(netstat_content_frame.winfo_children())
        
        row_frame = ttk.Frame(netstat_content_frame)
        row_frame.grid(row=row_num, column=0, sticky="ew")

        # Client text widget
        client_text = tk.Text(row_frame, width=15, height=3)
        client_text.grid(row=0, column=0, padx=2, sticky='w')
        client_text.insert('1.0', client)
        client_text.configure(state='disabled')

        # Details text widget with executable, IP:port and status
        details = f"{executable}\n{ip_port}\n{status}"
        details_text = tk.Text(row_frame, width=40, height=3)
        details_text.grid(row=0, column=1, padx=2, sticky='w')
        details_text.insert('1.0', details)
        details_text.configure(state='disabled')

        # Configure text widget appearance
        for widget in (client_text, details_text):
            widget.configure(
                relief='flat',
                borderwidth=0,
                highlightthickness=0,
                background=row_frame.cget('background')
            )

        # Add separator
        sep = ttk.Separator(netstat_content_frame)
        sep.grid(row=row_num+1, column=0, columnspan=2, sticky="ew", pady=1)

        return row_frame

    except Exception as e:
        print(f"Error adding netstat row: {str(e)}")
        return None
    
def run_as_admin(cmd):
    """Run command with admin privileges"""
    try:
        # Create a VBS script to elevate privileges
        vbs_content = f'''
        Set UAC = CreateObject("Shell.Application")
        UAC.ShellExecute "cmd.exe", "/c {cmd} > %temp%\\netstat_output.txt", "", "runas", 1
        '''
        
        with open('%temp%\\elevate.vbs', 'w') as f:
            f.write(vbs_content)
            
        # Run the VBS script
        subprocess.run('cscript //nologo %temp%\\elevate.vbs', shell=True)
        
        # Wait briefly for the command to complete
        time.sleep(2)
        
        # Read the output
        with open('%temp%\\netstat_output.txt', 'r') as f:
            return f.read()
            
    except Exception as e:
        print(f"Admin elevation error: {str(e)}")
        return None
    finally:
        # Cleanup temporary files
        try:
            os.remove('%temp%\\elevate.vbs')
            os.remove('%temp%\\netstat_output.txt')
        except:
            pass


def process_netstat_output(output, netstat_content_frame, netstat_canvas):
    """Process netstat output and update table"""
    try:
        # Clear existing content
        if netstat_content_frame:
            for widget in netstat_content_frame.winfo_children():
                widget.destroy()
        
        # Parse the output
        lines = output.splitlines()
        i = 0
        
        # Check if this is full netstat output (with Active Connections header)
        is_full_output = any("Active Connections" in line for line in lines)
        
        if is_full_output:
            # Skip header lines until we find "Active Connections"
            while i < len(lines) and "Active Connections" not in lines[i]:
                i += 1
            i += 2  # Skip "Active Connections" and the header line
            
            # Process connections with executable info
            while i < len(lines):
                line = lines[i].strip()
                if line.startswith('TCP') and ':12005' in line:
                    try:
                        parts = line.split()
                        proto = parts[0]
                        local_addr = parts[1]
                        remote_addr = parts[2]
                        state = parts[3]
                        
                        # Get executable name from next line
                        exe_name = "Unknown"
                        if i + 1 < len(lines):
                            next_line = lines[i + 1].strip()
                            if next_line.startswith('['):
                                exe_name = next_line[1:-1]  # Remove [ ]
                            elif "WpnService" in next_line:  # Handle special case
                                i += 1
                                if i + 1 < len(lines):
                                    next_line = lines[i + 1].strip()
                                    if next_line.startswith('['):
                                        exe_name = next_line[1:-1]
                        
                        # Extract client name from remote address
                        client_ip = remote_addr.split(':')[0]
                        
                        # Add to table
                        if netstat_content_frame and netstat_canvas:
                            add_netstat_row(
                                netstat_content_frame,
                                client_ip,
                                exe_name,
                                f"{local_addr} -> {remote_addr}\n{state}",
                                state
                            )
                    except Exception as e:
                        print(f"Error processing line: {str(e)}")
                i += 1
        else:
            # Process filtered output (no headers, no executable info)
            for line in lines:
                line = line.strip()
                if line.startswith('TCP'):
                    try:
                        parts = line.split()
                        proto = parts[0]
                        local_addr = parts[1]
                        remote_addr = parts[2]
                        state = parts[3] if len(parts) > 3 else "UNKNOWN"
                        
                        # Extract client name from remote address
                        client_ip = remote_addr.split(':')[0]
                        
                        # Add to table
                        if netstat_content_frame and netstat_canvas:
                            add_netstat_row(
                                netstat_content_frame,
                                client_ip,
                                "ICL",  # Default for filtered output
                                f"{local_addr} ->{remote_addr}",
                                state
                            )
                    except Exception as e:
                        print(f"Error processing line: {str(e)}")
        
        # Update scroll region
        if netstat_content_frame and netstat_canvas:
            netstat_content_frame.update_idletasks()
            netstat_canvas.configure(scrollregion=netstat_canvas.bbox("all"))
            
    except Exception as e:
        message_queue.put((f"Error processing netstat output: {str(e)}", "red"))
    
def update_connections_table(content_frame, canvas):
    """Update connections table using ICL log file"""
    try:
        # Clear existing content
        for widget in content_frame.winfo_children():
            widget.destroy()

        # Get latest log file
        log_path = get_icl_log_path()
        if not log_path:
            message_queue.put(("No ICL log path found", "red"))
            return

        log_files = glob.glob(os.path.join(log_path, "ICLServerLog_*.txt"))
        if not log_files:
            message_queue.put(("No log files found", "red"))
            return

        latest_log = max(log_files, key=os.path.getmtime)

        # Read and filter last 10 lines
        with open(latest_log, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        # Filter out Signal and WAITING messages
        filtered_lines = [
            line.strip() for line in lines 
            if line.strip() and 
            not "Signal" in line and 
            not "WAITING" in line
        ]


        # Add rows for last 10 entries
        for line in filtered_lines[-10:]:
            try:
                if '[' in line and ']' in line:
                    time_part = line[line.find('[')+1:line.find(']')]
                    rest = line[line.find(']')+1:].strip()

                    if ':' in rest:
                        client, details = rest.split(':', 1)
                    else:
                        client, details = rest, ""

                    # Update to match new function signature
                    add_connection_row(content_frame, client.strip(), details.strip())
            except Exception as e:
                continue

        # Update scroll region
        content_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

        # Schedule next update
        #root.after(1000, lambda: update_connections_table(content_frame, canvas))

    except Exception as e:
        message_queue.put((f"Error updating connections table: {str(e)}", "red"))


def check_license_expiry(init_response):
    try:
        import re
        from datetime import datetime, timedelta
        
        expiry_match = re.search(r'expire on (\d{2}/\d{2}/\d{4})', init_response)
        if expiry_match:
            expiry_date_str = expiry_match.group(1)
            expiry_date = datetime.strptime(expiry_date_str, '%d/%m/%Y')
            current_date = datetime.now()
            days_until_expiry = (expiry_date - current_date).days
            
            if days_until_expiry <= 30:
                messagebox.showwarning(
                    "License Expiry Warning",
                    f"Your license will expire on {expiry_date_str}\n"
                    f"({days_until_expiry} days remaining)"
                )
                return True
        return False
    except Exception as e:
        print(f"Error checking license expiry: {str(e)}")
        return False
    
def check_socket(silent=False):
    global _startup_complete
    if not sock:
        if not _startup_complete:
            return False
        if not silent:
            root.after(0, lambda: messagebox.showerror("Error", "Please connect to server first"))
        return False
    return True    


# Connect to the server (Non-blocking)
def connect(retries=3):
    global sock ,connection_content_frame, connection_canvas, netstat_content_frame, netstat_canvas
    global root_dir, workspace_dir, out_dir, dslam_profiles_dir, icl_scripts_dir , current_log_file
    _startup_complete = True

    try:
        # Check if we're shutting down before updating widgets
        if is_shutting_down:
            return
        # Use root.after to schedule widget updates on the main thread
        root.after(0, lambda: connect_button.config(state='disabled'))
        root.after(0, lambda: disconnect_button.config(state='disabled'))
        #root.after(0, lambda: #progress_bar.start())
        
        # If already connected, disconnect first
        if sock is not None:
            message_queue.put(("Already connected. Disconnecting first...", "black"))
            disconnect()

        host = ip_entry.get()
        port = port_entry.get()
                
        # Create new socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(10)  # 10 second timeout
        
        message_queue.put((f"Attempting to connect to {host}:{port}...", "black"))
        
        # Validate port first
        if not validate_port(port):
            message_queue.put(("Invalid Port: " + port, "red"))
            update_window_title("Invalid Port")
            show_notification("Connection Error", "Invalid Port")
            return

        # Convert IP to hostname if an IP is provided
        try:
            if validate_ip(host):
                try:
                    # Get hostname from IP
                    hostname = socket.gethostbyaddr(host)[0]
                    # Update the IP entry with hostname
                    ip_entry.delete(0, tk.END)
                    ip_entry.insert(0, hostname)
                    message_queue.put((f"Converted IP {host} to hostname: {hostname}", "black"))
                    host = hostname  # Use hostname for connection
                except socket.herror as e:
                    message_queue.put((f"Could not resolve hostname for IP {host}, using IP address", "black"))
            # If hostname provided, use it directly
            ip = host
        except Exception as e:
            message_queue.put((f"Error resolving address: {str(e)}", "red"))
            return

        # Store the hostname in config
        config['SERVER'] = {
            'Host': host,
            'Port': str(port)
        }
        
        # Connection attempts
        attempt = 0
        while attempt < retries:
            try:
                status_label.config(text="Connecting", foreground='black')
                message_queue.put((f"Connecting to {host} ({port})... (Attempt {attempt + 1})", "black"))
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.setblocking(0)
                result = sock.connect_ex((host, int(port)))  # Use hostname directly
                
                # Use select to wait for the socket to be ready
                ready = select.select([sock], [], [], 10)  # 10 seconds timeout
                if ready:
                    # Get initialization response
                    init_response = ""
                    for _ in range(1):
                        response = sock.recv(4096).decode()
                        init_response = response

                        create_connection_log(host)

                        #message_queue.put((f"Initialization response {response}", "black"))
                        # Add this right after capturing the initialization response in the connect function
                        message_queue.put((f"Initialization response: {init_response}", "black"))
                    
                    # Check license expiry
                    license_warning_shown = check_license_expiry(init_response)

                    
                    if not license_warning_shown:
                        message_queue.put(("Connected successfully!", "black"))
                        show_notification("Connection Status", "Connected successfully!")
                    disconnect_button.config(state='normal')
                    connect_button.config(state='normal')
                    update_window_title(f"Connected to {host}")

                    get_command_queue()
                    # Create tables
                    connection_content_frame, connection_canvas = create_connections_table(server_details_frame)
                    netstat_content_frame, netstat_canvas = create_netstat_table(server_details_frame)
                    # Add a small delay to ensure tables are created
                    root.after(100, lambda: initialize_after_connect(host))
        
        
                    
                    # Update status to "ready"
                    status_label.config(text="Ready", foreground='black')
                    
                    # Save successful connection details
                    config['SERVER'] = {
                        'Host': host,
                        'Port': str(port)
                    }
                    with open('user_config.ini', 'w') as configfile:
                        config.write(configfile)
                        
                    return  # Exit the loop after successful connection and initialization
                else:
                    message_queue.put(("Connection timed out", "red"))
                    update_window_title("Connection failed")
                    disconnect_button.config(state='normal')
                    connect_button.config(state='normal')
                    #show_notification("Connection Status", "Connection timed out")
            except socket.error as e:
                if e.errno == 10035:
                    message_queue.put(("Non-blocking socket operation could not be completed immediately. Retrying...", "red"))
                    time.sleep(1)
                    disconnect_button.config(state='normal')
                    connect_button.config(state='normal')
                    continue
                else:
                    message_queue.put((f"Socket error: {str(e)}", "red"))
                    update_window_title("Connection failed")
                    disconnect_button.config(state='normal')
                    connect_button.config(state='normal')
                    #show_notification("Connection Error", f"Socket error: {str(e)}")
            except Exception as e:
                message_queue.put((f"Failed to connect: {str(e)}", "red"))
                update_window_title("Connection failed")
                disconnect_button.config(state='normal')
                connect_button.config(state='normal')
                #show_notification("Connection Error", f"Failed to connect: {str(e)}")
            attempt += 1
            time.sleep(5)
        
        if attempt == retries:
            message_queue.put(("Max retries reached. Connection failed.", "red"))
            update_window_title("Connection failed")
            show_notification("Connection Status", "Max retries reached. Connection failed.")
            disconnect_button.config(state='normal')
            connect_button.config(state='normal')
        

        status_label.config(text="Ready", foreground='black')
    finally:
        # Safely stop progress bar
        if not is_shutting_down:
            #root.after(0, lambda: #progress_bar.stop())
            print(" ")
        else:
            print("Skipping widget updates during shutdown")

def find_dir_case_insensitive(parent_path, target_name):
    """
    Find a subdirectory inside parent_path, ignoring case.
    Returns Path object if found, else None.
    """
    target_name_lower = target_name.lower()
    for entry in os.listdir(parent_path):
        full_path = Path(parent_path) / entry
        if full_path.is_dir() and entry.lower() == target_name_lower:
            return full_path
    return None


def initialize_after_connect(host):
    """Handle initialization after successful connection"""
    # Create a thread to handle initialization to prevent GUI freezing
    threading.Thread(target=_initialize_after_connect_thread, args=(host,), daemon=True).start()
def _initialize_after_connect_thread(host):
    """Thread function to handle initialization after successful connection"""
    try:
        global root_dir, workspace_dir, out_dir, dslam_profiles_dir, icl_scripts_dir
        # Get command queue first
        get_command_queue()
        
        # Update connections table
        update_connections_table(connection_content_frame, connection_canvas)

        # Check for admin rights and handle netstat
        try:
            if is_admin():
                message_queue.put(("Admin rights detected, trying advanced netstat...", "black"))
                output = run_as_admin('netstat -b -n')
                if output:
                    process_netstat_output(output, netstat_content_frame, netstat_canvas)
                else:
                    raise Exception("No output from admin netstat")
            else:
                message_queue.put(("No admin rights, using basic netstat...", "black"))
                raise Exception("No admin rights")
        except Exception as e:
            message_queue.put((f"Using basic netstat: {str(e)}", "black"))
            remote_computer = ip_entry.get()
            output = get_remote_netstat(remote_computer)
            if output:
                process_netstat_output(output, netstat_content_frame, netstat_canvas)
            else:
                message_queue.put(("Failed to get any netstat output", "red"))

        # Check if command queue has data
        has_active_commands = False
        try:
            response = socket_query(sock, "commandqueue", timeout=5)
            if response and any(line.strip() for line in response.split('\n') if '->' in line):
                has_active_commands = True
                message_queue.put(("Active commands found in queue. Skipping additional commands.", "black"))
                # Refresh command queue display
                get_command_queue()
        except Exception as e:
            message_queue.put((f"Error checking command queue: {str(e)}", "red"))

        if not has_active_commands:
            # Only send additional commands if no active commands in queue
            message_queue.put(("No active commands in queue. Proceeding with initialization...", "black"))
            
            # Get the remote current working directory
            message_queue.put((f"Retrieving directory structure on remote host {host}...", "black"))
            
            # First, get the current working directory on the remote machine
            pwd_cmd = "cmd = 'cd'"
            res_pwd = socket_query(sock, pwd_cmd)
            callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
            res_dir = socket_query(sock, callos_cmd)
        
            current_dir = None
            if res_dir:
                # Extract the current directory path
                lines = res_dir.strip().split('\n')
                for line in lines:
                    if ':\\' in line:  # Look for any drive path, not just C:
                        current_dir = line.strip()
                        break
        
            if current_dir:
                message_queue.put((f"Remote current directory: {current_dir}", "black"))
                #local_root_dir = os.path.join(local_root_dir, "ICL_Server")
                # If we found a root directory, try to locate key subdirectories
                if 1:
                    # Convert local paths to UNC paths with the format \\hostname\drive\path
                    # Only add hostname prefix if it's not localhost/127.0.0.1
                    is_local = host.lower() in ['localhost', '127.0.0.1']
                    # Convert to UNC if needed
                    def to_unc(path: Path, host: str):
                        if host.lower() in ["localhost", "127.0.0.1"]:
                            return str(path)
                        drive = path.drive[0]
                        rest = path.relative_to(path.anchor).as_posix().replace("/", "\\")
                        return f"\\\\{host}\\{drive}$\\{rest}"
                    # # Function to convert path to UNC format if needed
                    # def to_unc_path(path):
                    #     if is_local:
                    #         return path  # Keep local paths as is
                    #     else:
                    #         drive = path[0]
                    #         path_without_colon = path[2:]  # Skip "C:"
                    #         cleaned_path = path_without_colon.lstrip('\\')  # do this first
                    #         return f"\\\\{host}\\{drive}$\\{cleaned_path}"
                    def to_unc_path(path):
                        if not path:
                            return ""
                        if path.startswith("\\\\"):  # already UNC
                            return path
                        if is_local:
                            return path
                        drive = path[0]
                        path_without_colon = path[2:].lstrip("\\")
                        return f"\\\\{host}\\{drive}$\\{path_without_colon}"
                    
                    # Try to find ICL_Server
                    #local_icl_server_dir = os.path.join(local_root_dir, "ICL_Server")

                    
                    current_dir_path = Path(current_dir)  # e.g., D:\Sparnex.testplan.mlt25\ICL Server\ICL_Scripts
                    root_dir_path = current_dir_path.parents[1]  # one level above 'ICL Server', i.e., Sparnex root
                    #root_dir_path = to_unc(root_dir_path, host)
                    #root_dir_path = Path(to_unc(root_dir_path, host))
                    icl_server_dir_path = root_dir_path / "ICL Server"
                    icl_scripts_dir_path = icl_server_dir_path / "ICL_Scripts"
                    workspace_dir_path = root_dir_path / "MyWorkspace"
                    if workspace_dir_path :
                        out_dir_path = workspace_dir_path / "Out"
                    else:
                        workspace_dir_path = root_dir_path / "myworkspace"    
                    if workspace_dir_path :
                        out_dir_path = workspace_dir_path / "Out"
                    else:
                        workspace_dir_path = root_dir_path / "Myworkspace"   

                    out_dir_path = workspace_dir_path / "Out"
                    dslam_profiles_dir_path = icl_server_dir_path / "dslam" / "dms4" / "profiles"
                    if dslam_profiles_dir_path :
                        dslam_profiles_dir_path = dslam_profiles_dir_path
                    else:
                        dslam_profiles_dir_path = icl_server_dir_path / "DSLAM" / "DMS4" / "Profiles"  
                    if dslam_profiles_dir_path :
                        dslam_profiles_dir_path = dslam_profiles_dir_path
                    else:
                        dslam_profiles_dir_path = icl_server_dir_path / "Dslam" / "DMS4" / "Profiles"  
                    root_dir = to_unc(root_dir_path, host)
                    icl_scripts_dir = to_unc(icl_scripts_dir_path, host)
                    workspace_dir = to_unc(workspace_dir_path, host)
                    out_dir = to_unc(out_dir_path, host)
                    dslam_profiles_dir = to_unc(dslam_profiles_dir_path, host)

                    # Store the discovered paths in config with UNC format
                    config['DIRECTORIES'] = {
                        'RemoteHost': host,
                        'RootDir': root_dir,
                        'WorkspaceDir': workspace_dir,
                        'OutDir': out_dir,
                        'DSLAMProfilesDir': dslam_profiles_dir,
                        'ICLScriptsDir': icl_scripts_dir
                    }
                
                    with open('user_config.ini', 'w') as configfile:
                        config.write(configfile)
                
                    # Display the paths that were found
                    message_queue.put(("Remote directory structure discovery complete.", "black"))
                    if not is_local:
                        message_queue.put((f"Using UNC paths with format \\\\{host}\\drive\\path", "black"))
                        message_queue.put((f"Root Dir: {root_dir}", "black"))
                        message_queue.put((f"Workspace Dir: {workspace_dir}", "black"))
                        message_queue.put((f"Out Dir: {out_dir}", "black"))
                        message_queue.put((f"DSLAM Profiles Dir: {dslam_profiles_dir}", "black"))
                        message_queue.put((f"ICL Scripts Dir: {icl_scripts_dir}", "black"))
                        
                        # Add a message to indicate the global variables have been updated
                        message_queue.put(("Global path variables updated to use UNC paths", "black"))
                else:
                    message_queue.put(("Could not determine remote root directory structure", "red"))
            else:
                message_queue.put(("Could not determine current remote directory", "red"))
            
            command = f"option('_debuginfo','off')"
            message_queue.put((f"Command: {command}", "black"))
            try:
                response = socket_query(sock, command, timeout=2)
                if response:
                    message_queue.put((f"Response: {response}", "black"))
                else:
                    message_queue.put(("No debug info response, continuing...", "black"))
            except Exception as e:
                message_queue.put(("Debug info command timed out, continuing...", "black"))

            setup_validation_traces()
            
            # Automatically get DSLAM drivers after successful connection
            #message_queue.put(("Automatically retrieving DSLAM drivers...", "black"))
            #show_dslamdrivers_thread()
        
        # Update status to "ready"
        status_label.config(text="Ready", foreground='black')
        
    except Exception as e:
        message_queue.put((f"Error in initialization: {str(e)}", "red"))
        # Add traceback for debugging
        message_queue.put((f"Traceback: {traceback.format_exc()}", "red"))

def thread_safe_display_message(message, color):
    root.after(0, display_message, message, color)
def create_connection_log(host):
    """Create a new log file for the current connection"""
    global current_log_file
    
    # Get the correct base directory whether running as script or frozen executable
    if getattr(sys, 'frozen', False):
        # If the application is frozen (PyInstaller)
        base_dir = os.path.dirname(sys.executable)
    else:
        # Regular Python script
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Create logs directory if it doesn't exist
    logs_dir = os.path.join(base_dir, "icl_logs")
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)
    
    # Create a new log file with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"icl_log_{timestamp}.txt"
    current_log_file = os.path.join(logs_dir, log_filename)
    
    # Write header to the log file
    with open(current_log_file, 'w', encoding='utf-8') as f:
        f.write(f"=== ICL Server Connection Log ===\n")
        f.write(f"Host: {host}\n")
        f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("="*40 + "\n\n")
    
    message_queue.put((f"Created new log file: {log_filename}", "black"))
    return current_log_file
# Display message in the response text area with specified color
def display_message(message, color):
    global current_log_file
    response_text.tag_configure("red", foreground="red")
    response_text.tag_configure("black", foreground="black")
    response_text.config(state=tk.NORMAL)  # Enable editing to insert message
    response_text.insert(tk.END, message + "\n", color)
    response_text.insert(tk.END, "-"*70 + "\n")
    response_text.see(tk.END)  # Scroll to the end
    response_text.config(state=tk.DISABLED)  # Disable editing after inserting message

    # Also log to the current log file if it exists
    if current_log_file and os.path.exists(current_log_file):
        try:
            with open(current_log_file, 'a', encoding='utf-8') as f:
                # Add timestamp to log entry
                timestamp = datetime.now().strftime("%H:%M:%S")
                f.write(f"[{timestamp}] {message}\n")
                f.write("-"*70 + "\n")
        except Exception as e:
            print(f"Error writing to log file: {str(e)}")

   
# Function to handle the button click and perform selected action on driver in a separate thread
def dslam_actions_thread():
    if sock is None:
        message_queue.put(("Not connected to the server. Please connect first.", "red"))
        dslam_actions_button.config(state='normal')
        return
    if not drivers_retrieved:
        display_message("Get DSLAM drivers details before sending any actions", "red")
        dslam_actions_button.config(state='normal')
        return
    threading.Thread(target=dslam_actions).start()

# Function to handle the button click and perform selected action on driver
def dslam_actions():
    global bcm_address, bcm_file, selected_port, selected_driver_name
    
    try:
        if sock is None:
            message_queue.put(("Not connected to the server. Please connect first.", "red"))
            dslam_actions_button.config(state='normal')
            return
        dslam_actions_button.config(state='disabled')
        # Initialize variables
        selected_index = None
        selected_driver = None
        command = None
        
        # Get the currently selected driver
        selected_index = devices_listbox.curselection()
        if not selected_index:
            display_message("Please select a driver first", "red")
            dslam_actions_button.config(state='normal')
            return
            
        selected_driver = devices_listbox.get(selected_index)
        if selected_driver in ["Name Address"]:  # Skip header row
            #display_message("Please select a valid driver", "red")
            dslam_actions_button.config(state='normal')
            return
            
        # Use the stored selected_driver_name instead of parsing again
        if not selected_driver_name:
            display_message("No driver selected", "red")
            dslam_actions_button.config(state='normal')
            return
            
        # Get driver details using the stored name
        new_address, _, new_file = get_all_drivers_details(selected_driver_name)
        if new_address and new_file:
            bcm_address = new_address
            bcm_file = new_file
            display_message(f"Using selected driver: {selected_driver_name}", "black")
        else:
            display_message(f"Could not get details for driver: {selected_driver_name}", "red")
            dslam_actions_button.config(state='normal')
            return
        
        # Load  driver
        load_driver(bcm_file)
        print(f"Using bcm_file: {bcm_file}")
        print(f"Using bcm_address: {bcm_address}")
        
        selected_action = dslam_var.get()
        thread_safe_display_message(f"Sending DSLAM Action - {selected_action}", "black")
        thread_safe_display_message(f"Address : {bcm_address}, Port : {selected_port}", "black")
        
        # Create command based on selected action
        if selected_action == "Read Status":
            command = f"(e, r) = readstatus('{bcm_address}', '{selected_port}') <<< e <<< r"
        elif selected_action == "Performance Parameters":
            command = f"(e, r) = PerformanceParameters('{bcm_address}', '{selected_port}') <<< e <<< r"
        elif selected_action == "Get Configuration":
            command = f"(e, r) = GetConfiguration('{bcm_address}', '{selected_port}') <<< e <<< r"
        elif selected_action == "Inventory":  # Add Inventory action
            command = f"(e, r) = Inventory('{bcm_address}', '{selected_port}') <<< e <<< r"
        else:
            display_message("Invalid action selected", "red")
            dslam_actions_button.config(state='normal')
            return

        if command:
            try:
                thread_safe_display_message(f"Command: {command}", "black")
                thread_safe_display_message("Loading... Please wait for response", "blue")
                
                # Wait for response with a maximum timeout of 10 seconds
                sock.settimeout(10)  # Set a 10-second timeout

                try:
                    response = socket_query(sock, command)
                    if response:
                        thread_safe_display_message("Response received:", "green")
                        thread_safe_display_message(response, "black")
                        dslam_actions_button.config(state='normal')
                    else:
                        thread_safe_display_message("No response received", "red")
                        dslam_actions_button.config(state='normal')
                except socket.timeout:
                    thread_safe_display_message("Response timeout after 10 seconds", "red")
                    dslam_actions_button.config(state='normal')
                except socket.error as e:
                    thread_safe_display_message(f"Socket error while receiving: {str(e)}", "red")
                    dslam_actions_button.config(state='normal')
            finally:
                # Reset the timeout to None (blocking mode)
                sock.settimeout(None)
                dslam_actions_button.config(state='normal')

    except Exception as e:
        display_message(f"Error in DSLAM action: {str(e)}", "red")
        dslam_actions_button.config(state='normal')

    finally:
        # Maintain selection after action
        if selected_index:
            devices_listbox.selection_clear(0, tk.END)
            devices_listbox.selection_set(selected_index)
            devices_listbox.see(selected_index)
        dslam_actions_button.config(state='normal')

# Function to handle the button click and display the drivers in a separate thread
def update_dslamdrivers_thread(dslam_name_dropdown,dslam_name_var,dslam_serialnumber_dropdown,dslam_serialnumber_var, dslam_busmap_var, dslam_busmap_dropdown, dslam_folder_var,dslam_profile_var, dslam_profile_dropdown, dslam_port_var,message_var,dslam_slot_var,dslam_frame_var):
    # Start a new thread to update DSLAM drivers
    threading.Thread(target=update_dslamdrivers, args=(dslam_name_dropdown,dslam_name_var,dslam_serialnumber_dropdown,dslam_serialnumber_var, dslam_busmap_var, dslam_busmap_dropdown, dslam_folder_var,dslam_profile_var, dslam_profile_dropdown, dslam_port_var,message_var,dslam_slot_var,dslam_frame_var)).start()
   


   
# Function to handle the button click and display the drivers
def update_dslamdrivers(dslam_name_dropdown=None, dslam_name_var=None,dslam_serialnumber_dropdown=None,dslam_serialnumber_var=None, dslam_busmap_var=None, dslam_busmap_dropdown=None, dslam_folder_var=None,dslam_profile_var=None, dslam_profile_dropdown=None, dslam_port_var=0, message_var=None, dslam_slot_var=0, dslam_frame_var=0):
    try:
        message_var.set("Loading DSLAM details...")
        response_setting = socket_query(sock, "option _hidden show")
        response2 = socket_query(sock, "<<<_DSLAMdrivers")

        if isinstance(response2, set):
            response2 = "\n".join(response2)
        elif not isinstance(response2, str):
            response2 = str(response2)

        response2 = response2.replace('\r', '').replace('\t', ' ')

        global drivers_list
        drivers_list = extract_name_and_address(response2)

        global drivers_details
        drivers_details = {}
        name_count = {}
        for driver in drivers_list:
            name = driver['Name']
            if name not in name_count:
                name_count[name] = 0
            name_count[name] += 1
            key = f"{name}_{name_count[name]}"
            drivers_details[key] = driver

        response3 = socket_query(sock, "for(idx,1:!_DSLAMdrivers)<<<_dslamdrivers[idx].name")
        # Directly populate the dropdown with the response
        dslam_names_list_tp = [driver_name.strip() for driver_name in response3.split('\n') if driver_name.strip()]

        dslam_name_dropdown_tp['values'] = []
        name_count = {}
        driver_display_list = []    # For display in dropdown
        driver_value_list = []      # For actual values (addresses only)


        for driver_name in response3.split('\n'):
            driver_name = driver_name.strip()
            if not driver_name:
                continue
                
            if driver_name not in name_count:
                name_count[driver_name] = 0
            name_count[driver_name] += 1
            
            key = f"{driver_name}_{name_count[driver_name]}"
            address = drivers_details.get(key, {}).get('Address', 'N/A')
            
            # Create display string (name + address) and value string (address only)
            display_string = f"{driver_name:<10} {address:<20}"
            value_string = address
            
            driver_display_list.append(display_string)
            driver_value_list.append(value_string)

        if dslam_name_dropdown:
            dslam_name_dropdown['values'] = ["Select a driver"] + dslam_names_list_tp
        if dslam_busmap_dropdown:
            # Store the value list for later use
            dslam_busmap_dropdown.value_list = driver_value_list
            # Show the display list in dropdown
            dslam_busmap_dropdown['values'] = ["Select an Address"] + driver_display_list
            # Bind selection event to handle the conversion
            def on_busmap_select(event):
                selected_idx = dslam_busmap_dropdown.current()
                if selected_idx > 0:  # If not "Select an Address"
                    # Set the actual value (address only) to the variable
                    dslam_busmap_var.set(driver_value_list[selected_idx - 1])
                    
            dslam_busmap_dropdown.bind('<<ComboboxSelected>>', on_busmap_select)            


        profile_listbox.delete(0, tk.END)
        folder = profile_path_var.get()

        cmd = f"folder = '{folder}'"
        response = socket_query(sock, cmd)
        cmd = f"(e,r)=callos('dir '++folder) <<< e <<< r"

        response = socket_query(sock, cmd)
        time.sleep(1)

        if response:
            file_names = extract_file_names(response)
            populate_profile_listbox(file_names)
        else:
            display_message("No response received", "red")

        dslam_folder_var.set(profile_path_var.get())
        default_value = 0
        dslam_frame_var.set(default_value)
        dslam_slot_var.set(default_value)

        bcm_profile_var = str(bcm_profile)
       
        if dslam_profile_dropdown:
            dslam_profile_dropdown['values'] = [bcm_profile_var] + ["Select other Profiles"] + [profile_listbox.get(i) for i in range(1, profile_listbox.size())]
            dslam_profile_dropdown.set(bcm_profile_var)

        dslam_port_var.set(port_var.get())


        response_setting = socket_query(sock, "option _hidden hide")
       
        message_var.set("Load successful")
    except Exception as e:
        thread_safe_display_message(f"Failed to retrieve DSLAM details: {str(e)}", "red")






# Function to handle the button click and display the drivers in a separate thread
def show_dslamdrivers_thread():
    global drivers_retrieved
    threading.Thread(target=show_dslamdrivers).start()
    drivers_retrieved = True


# Function to handle the button click and display the drivers
def show_dslamdrivers():
    try:
        # Start the progress bar
        #progress_bar.start()
       
        # Update status to "executing"
        status_label.config(text="Executing", foreground='black')
        response_setting = socket_query(sock, "option _hidden show")
        thread_safe_display_message(f"Response setting : option _hidden show {response_setting}", "black")
       
        # Send the first command to get the number of DSLAM drivers
        response1 = socket_query(sock, "<<<!_DSLAMdrivers")
        if response1:
            thread_safe_display_message(f"No. of DSLAM drivers: {response1}", "black")
        else:
            thread_safe_display_message("Failed to get number of DSLAM drivers.", "red")
       
        # Send the second command to get the details of DSLAM drivers
        response2 = socket_query(sock, "<<<_DSLAMdrivers")
        thread_safe_display_message(f"Details of DSLAM drivers: {response2}", "black")
       
        # Debugging: Print the full content of response2
        print(f"Full content of response2:\n{response2}")
       
        # Ensure response2 is a string
        if isinstance(response2, set):
            response2 = "\n".join(response2)
        elif not isinstance(response2, str):
            response2 = str(response2)
       
        # Sanitize the response
        response2 = response2.replace('\r', '').replace('\t', ' ')
       
        # Parse the response data into a list of dictionaries
        global drivers_list
        drivers_list = extract_name_and_address(response2)
        print(f"Drivers list: {drivers_list}")
       
        # Check the structure of drivers_list
        if drivers_list and isinstance(drivers_list, list) and all(isinstance(driver, dict) for driver in drivers_list):
            print("Drivers list is correctly formatted.")
        else:
            print("Drivers list is not correctly formatted.")
            thread_safe_display_message("Failed to parse drivers list.", "red")
            return
       
        # Ensure each dictionary has a 'Name' key
        for driver in drivers_list:
            if 'Name' not in driver:
                print(f"Missing 'Name' key in driver: {driver}")
                thread_safe_display_message("Driver entry missing 'Name' key.", "red")
                return
       
        # Create a dictionary for quick lookup of addresses by driver name
        global drivers_details
        drivers_details = {}
        name_count = {}
        for driver in drivers_list:
            name = driver['Name']
            if name not in name_count:
                name_count[name] = 0
            name_count[name] += 1
            key = f"{name}_{name_count[name]}"
            drivers_details[key] = driver
        print(f"driver details - {drivers_details}")
       
        # Retrieve DMS4 details and assign global variables
        dms4_details_list = get_all_dms4_details(drivers_list)
        print(dms4_details_list)
       
        # Display DMS4 details
        if dms4_details_list:
            bcm_address, FN_profile, bcm_file = dms4_details_list
            thread_safe_display_message("DMS4 Entry:", "black")
            thread_safe_display_message(f"  Address of DMS4 : {bcm_address}", "black")
            thread_safe_display_message(f"  Profile of DMS4 : {FN_profile}", "black")
            thread_safe_display_message(f"  Bin file location of DMS4 bcm: {bcm_file}", "black")
        else:
            thread_safe_display_message("BCM details not found.", "red")
        if bcm_file:
            thread_safe_display_message("", "red")
            dslam_driver_var.set("../DSLAM/BCM6400x/BCM6400x_100.bin")
            #dslam_path_var.set("../DSLAM/BCM6400x/BCM6400x_100.bin")

            #driver_path_var.set(bcm_file)
            #dslam_driver_var.set(bcm_file)
        else:
            driver_path_var.set("No driver is selected")

        # Send the third command to get the list of DSLAM drivers
        response3 = socket_query(sock, "for(idx,1:!_DSLAMdrivers)<<<_dslamdrivers[idx].name")
        print(f"Response3: {response3}") # Debugging print
        if response3:
            thread_safe_display_message(f"List of DSLAM drivers:\n{response3}", "black")
        else:
            thread_safe_display_message("No list received for DSLAM drivers.", "red")
       
        # Clear the listbox before adding new items
        devices_listbox.delete(0, tk.END)
       
        # Add "Name" and "Address" as the first item
        devices_listbox.insert(tk.END, f"{'Name':<30} {'Address':<20}")
        devices_listbox.itemconfig(0, {'bg':'lightgrey'})
       
        # Add each driver name and address to the listbox
        name_count = {}
        driver_names = []
        for driver_name in response3.split('\n'):
            driver_name = driver_name.strip() # Remove any leading/trailing whitespace
            if driver_name not in name_count:
                name_count[driver_name] = 0
            name_count[driver_name] += 1
            key = f"{driver_name}_{name_count[driver_name]}"
            print(f"Processing driver name: {driver_name} with key: {key}") # Debugging print
            # Find the corresponding address from drivers_details
            address = drivers_details.get(key, {}).get('Address', 'N/A')
            print(f"Address for {driver_name}: {address}") # Debugging print
            devices_listbox.insert(tk.END, f"{driver_name:<20} {address:<20}")
            driver_names.append(driver_name)
       
        # Remove the last item from the listbox
        if devices_listbox.size() > 0:
            last_item = devices_listbox.get(devices_listbox.size() - 1)
            if ">" in last_item:
                devices_listbox.delete(devices_listbox.size() - 1)
       
        response_setting = socket_query(sock, "option _hidden hide")
        thread_safe_display_message(f"Response setting : option _hidden hide {response_setting}", "black")
        # Update status to "ready"
        status_label.config(text="Ready", foreground='black')
    except Exception as e:
        thread_safe_display_message(f"Failed to retrieve DSLAM drivers: {str(e)}", "red")
        # Update status to "ready" in case of failure
        status_label.config(text="Ready", foreground='black')
       
    finally:
        # Stop the progress bar
        #progress_bar.stop()
        thread_safe_display_message(f" ", "red")



# Function send inventory_dslam of a DSLAM port
def inventory_dslam(bcm_address, port):
    try:
         # Start the progress bar
        #progress_bar.start()
        command = f"(e, r) = Inventory('{bcm_address}', '{port}') <<< e <<< r"
        display_message(f"Inventory Action for : {bcm_address}, Port: {port}", "black")
        response = socket_query(sock, command)
        if response:
            display_message(f"Response :\n{response}", "black")
        else:
            display_message("No response received", "red")
    except Exception as e:
        display_message(f"Failed to read DSLAM status: {str(e)}", "red")
    finally:
        # Stop the progress bar
        #progress_bar.stop()
        display_message(f" ", "red")

#Parse the response data into a dictionary
def extract_name_and_address(response):
    sections = response.strip().split("\n\n")
    results = []
    
    # Clear existing dictionary before processing new data
    dslam_drivers_dict.clear()
    
    current_name = None
    current_address = None
    current_file = None
    current_profile = None
    
    for line in response.split('\n'):
        line = line.strip()
        if not line:
            continue
            
        if line.startswith("Name: '"):
            # Save previous entry if exists
            if current_name:
                # Create details dictionary
                details_dict = {
                    'name': current_name,
                    'address': current_address,
                    'file': current_file.strip('"') if current_file else None,
                    'profile': current_profile,
                    'dms4': ''
                }
                
                # Special handling for SFP addresses
                if 'sfp' in current_name.lower():
                    if current_address:
                        mac_parts = current_address.split()
                        if mac_parts:
                            details_dict['address'] = mac_parts[-1]  # Keep only MAC address
                
                # Store in both results list and global dictionary
                results.append({
                    'Name': current_name,
                    'Address': details_dict['address'],
                    'File': details_dict['file'],
                    'Profile': details_dict['profile']
                })
                dslam_drivers_dict[current_name] = details_dict
                
            current_name = line.split("'")[1]
            
        elif line.startswith("Address: '"):
            current_address = line.split("'")[1]
            
        elif line.startswith("File: '"):
            current_file = line.split("'")[1]
            
        elif line.startswith("profile: '"):
            current_profile = line.split("'")[1]
            
        elif 'OnSet:' in line and 'DSLAM::SetDefault' in line:
            try:
                parts = line.split('"')
                for i, part in enumerate(parts):
                    if part.strip() == "address":
                        current_address = parts[i+2]
                        break
            except Exception as e:
                print(f"Error parsing OnSet line: {e}")
                
    if current_name:
        # Create details dictionary for last entry
        details_dict = {
            'name': current_name,
            'address': current_address,
            'file': current_file.strip('"') if current_file else None,
            'profile': current_profile,
            'dms4': ''
        }
        
        # Special handling for SFP addresses
        if 'sfp' in current_name.lower():
            if current_address:
                mac_parts = current_address.split()
                if mac_parts:
                    details_dict['address'] = mac_parts[-1]
        
        # Store in both results list and global dictionary
        results.append({
            'Name': current_name,
            'Address': details_dict['address'],
            'File': details_dict['file'],
            'Profile': details_dict['profile']
        })
        dslam_drivers_dict[current_name] = details_dict
    
    # Debug print
    print(f"Populated drivers dictionary: {dslam_drivers_dict}")
    
    return results


#extract the required details
def get_all_dms4_details(drivers_dict):
    """
    Get DMS4 details from the drivers list, prioritizing the correct path
    Returns: tuple (address, profile, file) or (None, None, None) if not found
    """
    dms4_entries = []
    
    # Collect all DMS4 entries
    for driver in drivers_dict:
        if driver['Name'] == 'DMS4':
            dms4_entries.append(driver)
    
    # If no DMS4 entries found, return None
    if not dms4_entries:
        return None, None, None
        
    # Find the entry with the correct path pattern
    for entry in dms4_entries:
        file_path = entry.get('File', '')
        if r'..\DSLAM\DMS4' in file_path:
            return (
                entry.get('Address'),
                entry.get('Profile'),
                entry.get('File')
            )
    
    # If no entry with correct path found, return the first DMS4 entry
    return (
        dms4_entries[0].get('Address'),
        dms4_entries[0].get('Profile'),
        dms4_entries[0].get('File')
    )

def get_all_drivers_details(name):
    """
    Get driver details from the drivers list, prioritizing the correct path
    Args:
        name: Driver name and address string (format: "name address")
    Returns:
        tuple: (address, profile, file) or (None, None, None) if not found
    """
    try:
        print(f"Searching for driver: {name}")  # Debug print
        print(f"Available drivers: {drivers_list}")  # Debug print
        
        # Split the input into name and address
        parts = name.split()
        if len(parts) < 2:
            print(f"Invalid driver format: {name}")
            return None, None, None
            
        driver_name = parts[0].strip()
        driver_address = parts[-1].strip()  # Take the last part as address
        
        matching_entries = []
        
        # Collect all matching entries
        for driver in drivers_list:
            if driver['Name'].strip() == driver_name:
                matching_entries.append(driver)
                
        # If no matching entries found, return None
        if not matching_entries:
            print(f"No matching entries found for: {driver_name}")
            return None, None, None
            
        # First try to find exact match with both name and address
        for entry in matching_entries:
            if entry['Address'].strip() == driver_address:
                print(f"Found exact match: {entry}")
                return (
                    entry.get('Address'),
                    entry.get('Profile'),
                    entry.get('File')
                )
                
        # If no exact match, find entry with correct path pattern
        for entry in matching_entries:
            file_path = entry.get('File', '')
            if rf'..\DSLAM\{driver_name}' in file_path:
                print(f"Found path match: {entry}")
                return (
                    entry.get('Address'),
                    entry.get('Profile'),
                    entry.get('File')
                )
        
        # If still no match, return the first matching entry
        print(f"Using first matching entry: {matching_entries[0]}")
        return (
            matching_entries[0].get('Address'),
            matching_entries[0].get('Profile'),
            matching_entries[0].get('File')
        )
        
    except Exception as e:
        print(f"Error in get_all_drivers_details: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return None, None, None
    
# def get_all_drivers_details(name):
#     """
#     Get driver details from the drivers list
#     Args:
#         name: Driver name to search for
#     Returns:
#         tuple: (address, profile, file) or (None, None, None) if not found
#     """
#     try:
#         print(f"Searching for driver: {name}")  # Debug print
#         print(f"Available drivers: {drivers_list}")  # Debug print
        
#         for driver in drivers_list:
#             print(f"Checking driver: {driver}")  # Debug print
#             if driver['Name'].strip() == name.strip():
#                 print(f"Found matching driver: {driver}")  # Debug print
#                 return (
#                     driver.get('Address'),
#                     driver.get('Profile'),
#                     driver.get('File')
#                 )
                
#         print(f"No matching driver found for: {name}")  # Debug print
#         return None, None, None
        
#     except Exception as e:
#         print(f"Error in get_all_drivers_details: {str(e)}")
#         return None, None, None

# #function to get bcm details
# def get_bcm_details(response):
#     print("Full Response:\n", response)  # Print the full response for debugging
#     sections = response.strip().split("Name: '")
#     for section in sections[1:]:
#         print("Section:\n", section)  # Print each section for debugging
#         lines = section.split('\n')
#         driver_data = {}
#         for line in lines:
#             if line.strip() == '':
#                 continue
#             if ':' in line:
#                 key, value = line.split(':', 1)
#                 key = key.strip().lower()  # Convert key to lowercase for case-insensitive comparison
#                 value = value.strip().strip("'")
#                 driver_data[key] = value
#         # Debugging information
#         print(f"Driver Data: {driver_data}")
#         if 'dms4' in driver_data.get('name', '').lower():
#             bcm_address = driver_data.get('address')
#             FN_profile = driver_data.get('profile')
#             bcm_file = driver_data.get('file')
#             return bcm_address, FN_profile, bcm_file
#     return None, None, None


# Function to send commands to the server
def send_commands(cmds=None):
    if cmds is None:
        cmds = []

    try:
        send_button.config(state='disabled')
        if sock is None:
            thread_safe_display_message("Not connected to the server. Please connect first.", "red")
            update_window_title("Not connected")
            show_temp_message("Connection Error", "Not connected to the server. Please connect first.")
            send_button.config(state='normal')
            return
        # Get the command from the combobox
        cmd = command_var.get()
        if cmd:
            cmds.append(cmd)

        # Validate commands
        if not cmds:
            thread_safe_display_message("Command is empty", "red")
            update_window_title("Command is empty")
            show_temp_message("Command Error", "Command is empty")
            send_button.config(state='normal')
            return

        # Check connection
        if sock is None:
            thread_safe_display_message("Not connected to the server. Please connect first.", "red")
            update_window_title("Not connected")
            show_temp_message("Connection Error", "Not connected to the server. Please connect first.")
            send_button.config(state='normal')
            return

        # Update status and start progress bar
        status_label.config(text="Execution in progress", foreground='black')
        #progress_bar.start()

        # Process commands
        for cmd in cmds:
            if not cmd.strip():
                continue  # Skip empty commands
                
            try:
                # Send command and get response
                response = socket_query(sock, cmd)
                
                if response is not None:
                    send_button.config(state='normal')
                    # Display successful response
                    thread_safe_display_message(f"Command: {cmd}", "black")
                    thread_safe_display_message(f"Response:\n{response}", "black")
                    thread_safe_display_message("-" * 60, "black")
                    
                    try:
                        # Log successful command
                        log_command(cmd, response)
                        add_to_history(cmd)
                    except Exception as log_error:
                        print(f"Logging error: {str(log_error)}")
                        # Continue execution even if logging fails
                    
                    # Clear command entry if it's a single command
                    if len(cmds) == 1:
                        root.after(0, lambda: command_combobox.set(''))
                else:
                    thread_safe_display_message(f"No response received for command: {cmd}", "red")
                    update_window_title("No response received")
                    show_temp_message("Command Error", f"No response received for command: {cmd}")
                    send_button.config(state='normal')

            except socket.timeout:
                thread_safe_display_message(f"Command timed out: {cmd}", "red")
                update_window_title("Command timed out")
                show_temp_message("Command Error", f"Command timed out: {cmd}")
                send_button.config(state='normal')
            except Exception as e:
                thread_safe_display_message(f"Error executing command '{cmd}': {str(e)}", "red")
                update_window_title("Command failed")
                show_notification("Command Error", f"Failed to execute command: {cmd}")
                print(f"Command error details: {traceback.format_exc()}")
                send_button.config(state='normal')

    except Exception as e:
        thread_safe_display_message(f"Error in command processing: {str(e)}", "red")
        update_window_title("Command processing failed")
        show_notification("Error", "Command processing failed")
        print(f"General error details: {traceback.format_exc()}")
        send_button.config(state='normal')

    finally:
        # Always reset status and progress bar
        status_label.config(text="Ready", foreground='black')
        #progress_bar.stop()
        send_button.config(state='normal')
        update_window_title("Connected to Server")

def validate_output_files(files):
        """Validate that all required file types are present"""
        required_extensions = {'.tp.txt', '.inv.txt', '.action.txt', 
                             '.cfg.txt', '.setup.txt', '.results.csv'}
    
        file_extensions = {os.path.splitext(os.path.splitext(f)[0])[1] + 
                      os.path.splitext(f)[1] for f in files}
    
        return required_extensions == file_extensions   

def organize_output_files(output_path, timestamp, files):
        """Create folder and move files into it"""
        try:
            # Create folder name (e.g., "TC.0.FS.20241118_124411")
            base_name = files[0].split('.', 3)[:3]
            folder_name = '.'.join(base_name) + '.' + timestamp
            
            # Create full folder path
            folder_path = os.path.join(output_path, folder_name)
            
            # Create folder if it doesn't exist
            os.makedirs(folder_path, exist_ok=True)
            
            # Move files to folder
            for file in files:
                src = os.path.join(output_path, file)
                dst = os.path.join(folder_path, file)
                shutil.move(src, dst)
            
            return folder_path
        
        except Exception as e:
            print(f"Error organizing files: {str(e)}")
            return None 


def process_output_files(display_func, is_timeout=False):
    """Main function to process output files after test plan execution"""
    global current_results_csv, timestamp_start, direct_filename_result
    try:
        display_func("\nStarting file processing...")

        # Get the output path and ensure it exists
        output_path = output_folder_var.get()
        if not output_path:
            output_path = "..\\..\\MyWorkspace\\Out"  # Default path
        
        display_func(f"Processing output files in: {output_path}")

        # Check if we have a direct filename result
        if not is_timeout and 'direct_filename_result' in globals() and direct_filename_result:
            base_name = direct_filename_result
            display_func(f"Using direct filename: {base_name}")
            
            # Get all files with the same timestamp pattern
            timestamp_pattern = base_name.split('.')[-1]
            cmd1 = f'dir /b /a-d /tc "{output_path}\\*.{timestamp_pattern}*"'
            res1 = socket_query(sock, f"cmd = '{cmd1}'")
            if res1:
                display_func(f"Directory listing command: {cmd1}")
                socket_query(sock, "(e, r) = callos(cmd) <<< e <<< r")
                files_response = socket_query(sock, '<<< r')
                
                if files_response:
                    files = [f.strip() for f in files_response.strip().split('\n') 
                            if f.strip() and "TC." in f]
                    
                    if files:
                        display_func(f"Found {len(files)} matching files using direct filename")
                        
                        # Create new directory in HERMES (with quotes to handle spaces)
                        hermes_path = f'\\\\HERMES\\4_Engineering\\7-test campaign\\Results\\{base_name}'
                        mkdir_cmd = f"cmd = 'mkdir \"{hermes_path}\"'"
                        
                        try:
                            # Update results path if possible
                            if 'update_results_path' in page4_functions and base_name:
                                page4_functions['update_results_path'](base_name)
                            elif 'results_path_var' in globals() and results_path_var and base_name:
                                # Direct fallback if we have the variable
                                base_path = r"\\HERMES\4_Engineering\7-test campaign\Results"
                                full_path = os.path.join(base_path, base_name)
                                results_path_var.set(full_path)
                        except Exception as e:
                            # Log the error but continue processing
                            display_func(f"Note: Could not update results path display: {str(e)}", "black")
                        
                        # Execute mkdir command
                        res_mkdir = socket_query(sock, mkdir_cmd)
                        display_func(f"Creating directory: {hermes_path}")
                        
                        # Execute callos for mkdir
                        callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
                        res_callos = socket_query(sock, callos_cmd)
                        
                        # Copy files using timestamp pattern
                        try:
                            # Source path with wildcard
                            src = f'..\\..\\MyWorkspace\\Out\\{base_name}*'
                            display_func(f"\nCopying files matching pattern: {base_name}*")
                            
                            # Copy all matching files to HERMES
                            copy_cmd = f"cmd = 'copy \"{src}\" \"{hermes_path}\"'"
                            socket_query(sock, copy_cmd)
                            socket_query(sock, "<<< cmd")
                            socket_query(sock, "(e, r) = callos(cmd) <<< e <<< r")
                            
                            # Display success summary
                            display_func("\nFile Copy Summary:", "green")
                            display_func(f"Successfully copied {len(files)} files to:", "black")
                            display_func(f"  {hermes_path}", "black")
                            for file in files:
                                display_func(f"  - {file}", "black")
                            display_func("\nOperation completed successfully!", "green")
                            
                            # Update the measurement database
                            try:
                                # Get the current test plan file name
                                tp_filename = os.path.basename(current_tp_file) if current_tp_file else ""
                                
                                # Start database update in a separate thread
                                threading.Thread(
                                    target=lambda: update_measurement_database(base_name, tp_filename, display_func),
                                    daemon=True
                                ).start()
                                
                                display_func("Database update initiated", "black")
                                
                            except Exception as e:
                                display_func(f"Warning: Failed to update measurement database: {str(e)}", "black")
                                print(f"Database update error: {traceback.format_exc()}")
                                
                            # Clear the direct filename result for next run
                            direct_filename_result = None
                            return
                            
                        except Exception as e:
                            display_func(f"Error during file copy: {str(e)}", "red")
                            # Continue to fall back to time-based approach
                
            # If we get here, something went wrong with the direct filename approach
            display_func("Direct filename approach unsuccessful, falling back to time-based approach.")
            direct_filename_result = None
            
        # Access the global timestamp
        if not timestamp_start:
            display_func("Error: No test start time available", "black")
            return
        
        # Get year and month from timestamp_start
        year_month = timestamp_start.strftime("%Y%m")
        timestamp_str = timestamp_start.strftime("%Y%m%d_%H%M%S")
        display_func(f"Using test start time: {timestamp_str}")
        time.sleep(1)
        # Define ref_time here so it's available for both timeout and non-timeout cases
        ref_time = time.mktime(timestamp_start.timetuple())
        display_func(f"Reference time set to: {datetime.fromtimestamp(ref_time).strftime('%Y-%m-%d %H:%M:%S')}")
        
        
        display_func(f"Processing output files in: {output_path}")
        
        if is_timeout:
            try:
                # Get remote host and construct remote output path
                remote_host = ip_entry.get()
                driver_path = dslam_driver_var.get()
                if driver_path:
                    base_parts = driver_path.split('ICL_Server')[0]
                    remote_output_path = base_parts.replace('C:', f'\\\\{remote_host}\\C') + 'MyWorkspace\\Out'
                    display_func(f"Using remote output path: {remote_output_path}")
                else:
                    display_func("Error: No DSLAM driver path available", "red")
                    return

                # Use remote_output_path for file operations
                search_pattern = os.path.join(remote_output_path, f"TC.*.{year_month}*.*")
                files = glob.glob(search_pattern)
                
                if not files:
                    display_func(f"No files found matching pattern: {search_pattern}", "red")
                    return None
                
                # Group files by test case and timestamp
                file_groups = {}
                for file_path in files:
                    file = os.path.basename(file_path)
                    match = re.search(r'(TC\.\d+\.[\w]+)\.(\d{8}_\d{6})', file)
                    if match:
                        base = match.group(1)
                        timestamp = match.group(2)
                        key = f"{base}.{timestamp}"
                        if key not in file_groups:
                            file_groups[key] = []
                        file_groups[key].append(file)
                
                # Find the most recent complete set of files
                best_group = None
                latest_timestamp = None
                for key, group_files in file_groups.items():
                    timestamp = key.split('.')[-1]
                    if latest_timestamp is None or timestamp > latest_timestamp:
                        latest_timestamp = timestamp
                        best_group = (key, group_files)

                if best_group:
                    base_name, files = best_group
                    display_func(f"\nFound file group: {base_name}")
                    display_func(f"Files found: {len(files)}")
                    
                    # Process the files (copy to HERMES etc.)
                    hermes_path = f'\\\\HERMES\\4_Engineering\\7-test campaign\\Results\\{base_name}'
                    
                    try:
                        # Create HERMES directory
                        os.makedirs(hermes_path, exist_ok=True)
                        display_func(f"Created directory: {hermes_path}")
                        
                        # Copy files using shutil with remote_output_path
                        copied_files = []
                        for file in files:
                            src = os.path.join(remote_output_path, file)
                            dst = os.path.join(hermes_path, file)
                            try:
                                shutil.copy2(src, dst)
                                copied_files.append(file)
                                #display_func(f"Copied: {file}", "black")
                                
                                # Special handling for Results.csv
                                #if file.endswith('.results.csv'):
                                    #ptu_dest = r'\\HERMES\4_Engineering\7-test campaign\PTU\Results.csv'
                                    #shutil.copy2(src, ptu_dest)
                                    #current_results_csv = src
                                    #display_func(f"Copied Results.csv to PTU folder", "green")
                            except Exception as copy_error:
                                display_func(f"", "red")
        
                        # Display success summary
                        if copied_files:
                            display_func("\nFile Copy Summary:", "green")
                            display_func(f"Successfully copied {len(copied_files)} files to:", "black")
                            display_func(f"  {hermes_path}", "black")
                            for file in copied_files:
                                display_func(f"  - {file}", "black")
                            display_func("\nOperation completed successfully!", "green")
        
                        # Update the measurement database
                        try:
                            # Get the current test plan file name
                            tp_filename = os.path.basename(current_tp_file) if current_tp_file else ""
            
                            # Start database update in a separate thread
                            threading.Thread(
                                target=lambda: update_measurement_database(base_name, tp_filename, display_func),
                                daemon=True
                            ).start()
                            
                            display_func("Database update initiated", "black")
                            
                        except Exception as e:
                            display_func(f"Warning: Failed to update measurement database: {str(e)}", "black")
                            print(f"Database update error: {traceback.format_exc()}")
        
                        return base_name, files
                        
                    except Exception as e:
                        display_func(f"Error processing files: {str(e)}", "red")
                        return None
                        
                else:
                    display_func("No complete file groups found", "red")
                    return None
                    
            except Exception as e:
                display_func(f"Error processing files after timeout: {str(e)}", "red")
                traceback.print_exc()
                return None
        else:
            # Get directory listing
            cmd1 = f'dir /b /a-d /tc "{output_path}\\TC.*.{year_month}*"'
            res1 = socket_query(sock, f"cmd = '{cmd1}'")
            if res1 :
                display_func(f"Directory listing command: {cmd1}")
                all_files_cmd = f'dir /b "{output_path}\\TC.*.{year_month}*"'
                all_files_res = socket_query(sock, f"cmd = '{all_files_cmd}'")
                socket_query(sock, "(e, r) = callos(cmd) <<< e <<< r")
                display_func(f"All files command: {socket_query(sock, '<<< r')}")
            else:
                display_func(f"No directory listing command: {cmd1}")
            # Execute callos for directory listing
            cmd2 = "(e, r) = callos(cmd) <<< e <<< r"
            res2 = socket_query(sock, cmd2)
            if res2:
                display_func(f"File list obtained")
            else:
                display_func(f"No file list obtained")

            # Find matching files using current time
            result = find_latest_output_files(ref_time, display_func, res2 or "", is_timeout)
            if not result:
                display_func("No output files found", "black")
                return
          
            base_name, files = result
            if not base_name or not files:
                display_func("Error: Invalid base name or no files found", "red")
                return
            
            # Create new directory in HERMES (with quotes to handle spaces)
            hermes_path = f'\\\\HERMES\\4_Engineering\\7-test campaign\\Results\\{base_name}'
            mkdir_cmd = f"cmd = 'mkdir \"{hermes_path}\"'"
            
            try:
                # Update results path if possible
                if 'update_results_path' in page4_functions and base_name:
                    page4_functions['update_results_path'](base_name)
                elif 'results_path_var' in globals() and results_path_var and base_name:
                    # Direct fallback if we have the variable
                    base_path = r"\\HERMES\4_Engineering\7-test campaign\Results"
                    full_path = os.path.join(base_path, base_name)
                    results_path_var.set(full_path)
            except Exception as e:
                # Log the error but continue processing
                display_func(f"Note: Could not update results path display: {str(e)}", "black")
            # Execute mkdir command
            res_mkdir = socket_query(sock, mkdir_cmd)
            display_func(f"Creating directory : {hermes_path}")
            #display_func(f"Response: {res_mkdir}")
            
            # Execute callos for mkdir
            callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
            res_callos = socket_query(sock, callos_cmd)
            #display_func(f"Directory creation result: {res_callos}")
        
            # Copy each file with the specified command sequence
            copied_files = []
            # Copy files using timestamp pattern
            try:
                # Extract timestamp from base_name
                timestamp_pattern = base_name.split('.')[-1]
                copy_pattern = f"TC.*.{timestamp_pattern}.*"
                
                # Source path with wildcard
                src = f'..\\..\\MyWorkspace\\Out\\{copy_pattern}'
                display_func(f"\nCopying files matching pattern: {copy_pattern}")
                
                # Copy all matching files to HERMES
                copy_cmd = f"cmd = 'copy \"{src}\" \"{hermes_path}\"'"
                socket_query(sock, copy_cmd)
                socket_query(sock, "<<< cmd")
                socket_query(sock, "(e, r) = callos(cmd) <<< e <<< r")
                
                # # Handle Results.csv separately
                # csv_file = next((f for f in files if f.endswith('.results.csv')), None)
                # if csv_file:
                #     #current_results_csv = csv_file
                #     src_csv = f'..\\..\\MyWorkspace\\Out\\{csv_file}'
                #     ptu_des = r'\\HERMES\4_Engineering\7-test campaign\PTU\Results.csv'
                #     copy_cmd = f"cmd = 'copy \"{src_csv}\" \"{ptu_des}\"'"
                #     #socket_query(sock, copy_cmd)
                #     #socket_query(sock, "<<< cmd")
                #     #socket_query(sock, "(e, r) = callos(cmd) <<< e <<< r")
                #     #display_func(f"Copied Results.csv to PTU folder", "green")
                #     current_results_csv = src_csv
            
                # Display success summary
                display_func("\nFile Copy Summary:", "green")
                display_func(f"Successfully copied {len(files)} files to:", "black")
                display_func(f"  {hermes_path}", "black")
                for file in files:
                    display_func(f"  - {file}", "black")
                #if csv_file:
                    #display_func("\nResults.csv also copied to:", "black")
                    #display_func(f"  {ptu_des}", "black")
                display_func("\nOperation completed successfully!", "green")    
            except Exception as e:
                display_func(f"Error during file copy: {str(e)}", "red")
                return
                  
            # Update the measurement database
            try:
                # Get the current test plan file name
                tp_filename = os.path.basename(current_tp_file) if current_tp_file else ""
                
                # Start database update in a separate thread
                threading.Thread(
                    target=lambda: update_measurement_database(base_name, tp_filename, display_func),
                    daemon=True
                ).start()
            
                display_func("Database update initiated", "black")
            
            except Exception as e:
                display_func(f"Warning: Failed to update measurement database: {str(e)}", "black")
                print(f"Database update error: {traceback.format_exc()}")
        
    except Exception as e:
            error_msg = f"Error processing output files: {str(e)}"
            display_func(error_msg, "black")
            # Print full traceback for debugging
            traceback.print_exc()



def find_latest_output_files(ref_time_seconds, display_func, response_text, is_timeout=False):
    """Process file list to find matching files, with different logic for timeout cases"""
    try:
        if not response_text or not response_text.strip():
            display_func("Error: No response text provided")
            return None, None
            
        files_dict = {}
        timestamps_dict = {}
        
        # Parse filenames first to avoid duplicate code
        filenames = [line.strip().split()[-1] for line in response_text.splitlines() 
                    if line.strip() and "TC." in line]
                    
        if not filenames:
            display_func("No TC files found in response")
            return None, None
            
        if is_timeout:
            # For timeout case - find latest files
            filenames = [line.strip().split()[-1] for line in response_text.splitlines() 
                        if line.strip() and "TC." in line]
            
            # Extract unique timestamps
            timestamps = set()
            for f in filenames:
                match = re.search(r'\d{8}_\d{6}', f)
                if match:
                    timestamps.add(match.group())
            
            if not timestamps:
                display_func("No files with timestamps found")
                return None, None
            
            # Get the latest timestamp
            latest_timestamp = sorted(timestamps)[-1]
            display_func(f"Latest timestamp found: {latest_timestamp}")
            
            # Get all files matching the latest timestamp
            matching_files = [f for f in filenames if latest_timestamp in f]
            
            # Get the base pattern from first matching file
            base_pattern = matching_files[0].split('.' + latest_timestamp)[0]
            display_func(f"Found matching pattern: {base_pattern}")
            display_func(f"Files found: {len(matching_files)}")
            
            return f"{base_pattern}.{latest_timestamp}", matching_files
            
        else:
            # Normal execution - use time window matching
            ref_time = datetime.fromtimestamp(ref_time_seconds)
            ref_stamp = ref_time.strftime("%Y%m%d_%H%M%S")
            display_func(f"Looking for files with timestamps between 60s before and 60s after: {ref_stamp}")
            
            filenames = [line.strip().split()[-1] for line in response_text.splitlines() 
                        if line.strip() and "TC." in line]
            
            for filename in filenames:
                try:
                    match = re.search(r'((TC\.\d+\.[\w]+)\.(\d{8}_\d{6}))', filename)
                    if not match:
                        continue
                        
                    base_pattern = match.group(2)
                    file_timestamp_str = match.group(3)
                    
                    # Validate timestamp format
                    try:
                        hour = int(file_timestamp_str[9:11])
                        minute = int(file_timestamp_str[11:13])
                        second = int(file_timestamp_str[13:15])
                        if not (0 <= hour <= 23 and 0 <= minute <= 59 and 0 <= second <= 59):
                            continue
                    except ValueError:
                        continue
                    
                    file_time = datetime.strptime(file_timestamp_str, "%Y%m%d_%H%M%S")
                    time_diff = (file_time - ref_time).total_seconds()
                    
                    if -60 <= time_diff <= 60:
                        if base_pattern not in files_dict:
                            files_dict[base_pattern] = []
                            timestamps_dict[base_pattern] = {
                                'time': file_time,
                                'diff': abs(time_diff)
                            }
                        files_dict[base_pattern].append(filename)
                    
                except Exception as e:
                    continue
            
            if not timestamps_dict:
                display_func("No files found within the time window (-60s to +60s)")
                return None, None
                
            best_pattern = min(timestamps_dict.items(), 
                             key=lambda x: x[1]['diff'])[0]
            
            display_func(f"Found matching pattern: {best_pattern}")
            display_func(f"Files found: {len(files_dict[best_pattern])}")
            
            return (f"{best_pattern}.{timestamps_dict[best_pattern]['time'].strftime('%Y%m%d_%H%M%S')}", 
                    files_dict[best_pattern])
    
    except Exception as e:
        display_func(f"Error processing files: {str(e)}")
        display_func(f"Traceback: {traceback.format_exc()}")
        return None, None

def clear_page4_tables_safely(config_frame, test_frame):
    """Safely clear all tables in page 4 with proper error handling"""
    try:
        def clear_widgets():
            try:
                # Clear the preview tree first
                if preview_tree and preview_tree.winfo_exists():
                    preview_tree.delete(*preview_tree.get_children())
                
                # Then clear the inner frames
                if config_frame and config_frame.winfo_exists():
                    for widget in config_frame.winfo_children():
                        widget.destroy()
                
                if test_frame and test_frame.winfo_exists():
                    for widget in test_frame.winfo_children():
                        widget.destroy()
                
                print("All preview sections cleared successfully")
                update_status_4("All preview sections cleared")
                
            except Exception as e:
                print(f"Error in widget clearing: {str(e)}")
        
        # Schedule the clearing operation on the main thread
        root.after(0, clear_widgets)
        
    except Exception as e:
        print(f"Error in clear_page4_tables_safely: {str(e)}")

# At global scope (outside any function)
def socket_query_continuous(sock, cmd, display_func, update_status=None, config_frame=None, test_frame=None):
    global timestamp_start, preview_tree , current_log_file
    """Send command once and continuously receive/display responses"""
    print('command:', cmd)
    if isinstance(cmd, str):
        cmd = cmd.encode()
    
    try:
        # Capture timestamp before running commands
        timestamp_start = datetime.now()
        timestamp_str = timestamp_start.strftime("%Y%m%d_%H%M%S")
        display_func(f"Test start time: {timestamp_str}")
        is_timeout = False
        # Send command once
        sock.sendall(cmd + b'\n')
        print(f"Sent command: {cmd.decode()}")
        display_func(f"Sent command: {cmd.decode()}")
        # Log the command to the current log file
        if current_log_file and os.path.exists(current_log_file):
            with open(current_log_file, 'a', encoding='utf-8') as f:
                timestamp = datetime.now().strftime("%H:%M:%S")
                f.write(f"[{timestamp}] COMMAND: {cmd.decode()}\n")
                f.write("-"*70 + "\n")
        
        # Set a reasonable timeout for initial response
        sock.settimeout(10)
        display_func("\nCommand sent to server, receiving responses...\n")
        
        # Initialize for receiving responses
        full_response = ""
        last_response_time = time.time()
        interval_count = 0
        end_sequence_seen = False
        buffer = ""
        
        # Continuous receiving loop
        while True:
            try:
                sock.settimeout(2.0)
                data = sock.recv(65536)
                
                if data:
                    part = data.decode()
                    full_response += part
                    display_func(part)
                    last_response_time = time.time()
                    interval_count = 0
                    
                    # Add to buffer and check for patterns
                    buffer += part
                    lines = buffer.split('\n')
                    buffer = lines[-1]  # Keep the last incomplete line
                    
                    # Check for end sequence in complete lines
                    for line in lines[:-1]:  # Process all complete lines
                        if "end of Run_TP" in line:
                            end_sequence_seen = True
                            display_func("\nEnd of Run_TP\n")
                          # Clear preview tree in thread-safe way
                            root.after(1000, lambda:clear_page4_tables_safely(config_frame, test_frame))
                            
                            #display_func("\nCleared preview tree after execution...\n")
                    
                        # If we've seen end sequence and get a prompt alone on a line
                        if end_sequence_seen and line.strip() == '>':
                            display_func("\n=== Response ended from server ===\n")
                            root.after(1000, lambda:clear_page4_tables_safely(config_frame, test_frame))
                            if update_status:
                                update_status("Execution completed.")
                            
                            # New approach: Query for output filename directly
                            display_func("\nQuerying server for output filename...\n")
                            filename_cmd = "(e, I) = environment() <<< i.output.filename"
                            filename_response = socket_query(sock, filename_cmd)
                            
                            direct_filename_found = False
                            base_name = None
                            if filename_response and "TC." in filename_response:
                                # Extract the filename from the response
                                lines = filename_response.strip().split('\n')
                                for resp_line in lines:
                                    if "TC." in resp_line:
                                        # Extract the filename (last word in the line)
                                        filename = resp_line.strip().split()[-1]
                                        if "TC." in filename:
                                            display_func(f"Filename obtained directly from server: {filename}")
                                            
                                            # Get the base pattern from the filename
                                            match = re.search(r'(TC\.\d+\.[\w]+)\.(\d{8}_\d{6})', filename)
                                            if match:
                                                base_pattern = match.group(1)
                                                timestamp = match.group(2)
                                                base_name = f"{base_pattern}.{timestamp}"
                                                direct_filename_found = True
                                                break
                            
                            # Process output files with display function
                            display_func("\nProcessing output files...\n")
                            if direct_filename_found and base_name:
                                display_func(f"Using directly obtained filename: {base_name}")
                                # Store the base_name for process_output_files to use
                                global direct_filename_result
                                direct_filename_result = base_name
                            else:
                                display_func("Falling back to time-based file matching...")
                                if 'direct_filename_result' in globals():
                                    direct_filename_result = None
                                    
                            root.after(3000, lambda: process_output_files(display_func, is_timeout=False))
                            return full_response
            except socket.timeout:
                current_time = time.time()
                if current_time - last_response_time > 60:
                    interval_count += 1
                    display_func(f"\nWaiting: No new data for {interval_count} minute(s)...\n")
                    last_response_time = current_time
                    
                    # If no data for 5 minutes, proceed with processing
                    if interval_count >= 5:
                        display_func("\n=== No response for 5 minutes, proceeding with file processing ===\n")
                        root.after(1000, lambda:clear_page4_tables_safely(config_frame, test_frame))
                        if update_status:
                            update_status("Execution completed (timeout).")
                        display_func("\nProcessing output files...\n")
                        process_output_files(display_func,is_timeout=True)
                        return full_response
                
                # If we've seen end sequence but no prompt for a while, check buffer
                if end_sequence_seen and current_time - last_response_time > 5:
                    if buffer.strip().endswith('>'):
                        display_func("\n=== Response ended after 'end of Run_TP' ===\n")
                        root.after(1000, lambda:clear_page4_tables_safely(config_frame, test_frame))
                        if update_status:
                            update_status("Execution completed.")
                        display_func("\nProcessing output files...\n")
                        process_output_files(display_func,is_timeout=False)
                        return full_response
                continue
                
        return full_response
        
    except Exception as e:
        error_msg = f"Error in socket communication: {str(e)}"
        root.after(1000, lambda:clear_page4_tables_safely(config_frame, test_frame))
        display_func(f"\n{error_msg}\n")
        if update_status:
            update_status(f"Socket error: {str(e)}", True)
        
        # Proceed with file processing even after socket error
        display_func("\n=== Proceeding with file processing despite socket error ===\n")
        root.after(1000, lambda:clear_page4_tables_safely(config_frame, test_frame))
        process_output_files(display_func,is_timeout=True)
        return full_response  # Return whatever response we got before the error
    
def socket_query(sock, cmd, timeout=35):
    print('command:', cmd)
    if isinstance(cmd, str):
        cmd = cmd.encode()
    print(cmd, type(cmd))
    try:
        # Flush the socket buffer
        sock.settimeout(0.5)
        while True:
            try:
                sock.recv(65536)  # Increased buffer size to 64 KB
            except socket.timeout:
                break
        # Send the command
        sock.sendall(cmd + b'\n')  # Append newline character
        print(f"Sent command: {cmd.decode()}")
    except socket.error as reason:
        print('SocketQuery failed!', str(reason))
        return None
   
    # Read the response
    sock.settimeout(timeout)  # Set a timeout for receiving data
    response = b""
    try:
        while True:
            reply = sock.recv(65536)  # Increased buffer size to 64 KB
            if not reply:
                break
            response += reply
            if b">" in reply:  # Assuming '>' is the end of response marker
                break
        response = response.decode().strip()
        print(f"Received response: {response}")
        return response
    except socket.timeout:
        print("Socket receive timed out")
        return None

def check_server_connection():
    """Check if server is responsive"""
    try:
        if not sock:
            return False
            
        # Simple math command to verify server response
        test_cmd = "2+2"
        response = socket_query(sock, test_cmd)
        
        # Check if response contains '4'
        return response is not None and '4' in str(response)
        
    except Exception as e:
        print(f"Connection check failed: {str(e)}")
        return False
    
def execute_profile_line_by_line(sock,bcm_address,selected_port_temp, full_path):
    try:
        # Check if 'bcm' is part of the profile text file
        with open(full_path, 'r') as file:
            profile_content = file.read()
       
        if 'bcm' in profile_content:
            display_message(f"'bcm cfg' found in {FN_profile} : 'api apply config' is not implemented", "black")
            display_message(f"Executing...Please wait ","black")
            with open(full_path, 'r') as file:
               for line in file:
                  line = line.strip()
                  if line:
                    # Replace $port$ with selected_port_temp if found
                    if '$port$' in line:
                        line = line.replace('$port$', selected_port_temp)
                   
                    cmd = f"(e,r) = sendto('{bcm_address}', '{line}') <<< e <<< r"
                    response = socket_query(sock, cmd)
                    if response:
                        display_message(f"Response for '{cmd}': {response}", "black")
                    else:
                        display_message(f"No response for '{cmd}'", "black")
                    time.sleep(0.5)  # Optional: Add a delay between commands
        else:
            display_message(f"Executing...Please wait ","black")
            with open(full_path, 'r') as file:
             for line in file:
                line = line.strip()
                if line:
                    cmd = f"(e,r) = sendto('{bcm_address}', '{line}') <<< e <<< r"
                    response = socket_query(sock, cmd)
                    if response:
                        display_message(f"Response for '{cmd}': {response}","black")
                    else:
                        display_message(f"No response for '{cmd}'","black")
                    time.sleep(0.5)  # Optional: Add a delay between commands
    except Exception as e:
        print(f"Failed to execute profile: {str(e)}")


# Log commands and responses to a file
def log_command(cmd, response):
    """Log commands and responses to Excel file"""
    command_status = "Pass" if response else "Fail"
    connection_status = "Connected" if sock else "Disconnected"
    end_time = datetime.now().strftime("%H:%M:%S")  # Fixed datetime usage
    log_command_to_excel(ip_entry.get(), port_entry.get(), cmd, response, 
                        command_status, connection_status, end_time)
def process_queue():
    while not message_queue.empty():
        message, color = message_queue.get()
        display_message(message, color)
    root.after(100, process_queue)  # Check the queue every 100ms


# Upload file to the server (non-blocking)
def upload_file():
    threading.Thread(target=upload_file_dialog).start()

# Modified upload_file_dialog function
def upload_file_dialog():
    file_path = filedialog.askopenfilename()
    if file_path:
        try:
            # Only upload the file, process_file will be called after upload completes
            threading.Thread(target=process_file, args=(file_path,)).start()
        except Exception as e:
            display_message(f"Failed while processing file: {str(e)}", "red")

# Modified upload_file_thread function
# def upload_file_thread(file_path):
#     try:
#         #progress_bar.start()
#         status_label.config(text="Uploading", foreground='black')

#         display_message(f"Uploading file '{file_path}'...", "black")
#         with open(file_path, 'rb') as file:
#             sock.sendall(file.read())
#             display_message(f"Loading...", "black")
#             display_message(f"File upload successful", "black")

#         # After successful upload, process the file
#         threading.Thread(target=process_file, args=(file_path,)).start()
        
#     except Exception as e:
#         display_message(f"File upload failed: {str(e)}", "red")
#         status_label.config(text="Ready", foreground='black')
#         #progress_bar.stop()

# def execute_file_thread(file_path):
#     try:
#         status_label.config(text="Executing", foreground='black')
#         display_message(f"File execution started", "black")
#         print("File execution started")
        
#         # Prepare log file
#         filename = os.path.join(os.path.dirname(file_path), f"log_{os.path.basename(file_path)}.txt")
#         log_file = open(filename, 'w')
#         log_file.write(f"Log created at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
#         log_file.write("-" * 50 + "\n\n")

#         # Read all lines from the file
#         with open(file_path, 'r') as file:
#             lines = file.readlines()
        
#         # Process each line sequentially
#         for line_index, line in enumerate(lines):
#             original_line = line.strip()
#             print(f"Processing line {line_index + 1}: {original_line}")  # Debug print
            
#             # Skip empty lines
#             if not original_line:
#                 log_file.write("\n")
#                 log_file.flush()
#                 continue
            
#             # Handle comments
#             if original_line.startswith('//') or original_line.startswith('#'):
#                 log_file.write(f"COMMENT: {original_line}\n")
#                 log_file.flush()
#                 continue
            
#             # Log the command BEFORE execution
#             log_file.write(f"COMMAND: {original_line}\n")
#             log_file.flush()  # Flush immediately
            
#             # Execute the command
#             response = execute_single_command_adaptive(original_line)
#             print(f"Received response for command {line_index + 1}: {response}")  # Debug print
            
#             # Log the response AFTER execution
#             log_file.write(f"RESPONSE: {response}\n")
#             log_file.write("-" * 50 + "\n\n")
#             log_file.flush()  # Ensure response is written immediately
            
#             # Optional: Add a delay or wait for user input before proceeding
#             time.sleep(1)  # Adjust as necessary

#         log_file.close()
#         display_message(f"Responses saved to '{filename}'", "black")
#         display_message("File execution successful", "black")
#         status_label.config(text="Ready", foreground='black')
        
#     except Exception as e:
#         display_message(f"File execution failed: {str(e)}", "red")
#         status_label.config(text="Ready", foreground='black')
#         try:
#             if 'log_file' in locals() and not log_file.closed:
#                 log_file.close()
#         except:
#             pass

# def execute_file_thread(file_path):
#     try:
#         #progress_bar.start()
#         status_label.config(text="Executing", foreground='black')
#         display_message(f"File execution started", "black")
#         print("File execution started")
#         # Prepare log file
#         filename = os.path.join(os.path.dirname(file_path), f"log_{os.path.basename(file_path)}.txt")
#         log_file = open(filename, 'w')
#         log_file.write(f"Log created at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
#         log_file.write("-" * 50 + "\n\n")

#         # Read all lines from the file
#         with open(file_path, 'r') as file:
#             lines = file.readlines()
        
#         # First, completely drain any existing data in the socket
#         drain_socket_completely()
        
#         # Process each line sequentially
#         line_index = 0
#         total_lines = len(lines)
        
#         while line_index < total_lines:
#             line = lines[line_index]
#             original_line = line.strip()
#             line_index += 1  # Move to next line for next iteration
            
#             # Skip empty lines but log them
#             if not original_line:
#                 log_file.write("\n")
#                 log_file.flush()
#                 continue
                
#             # Handle comments - log them but don't send to server
#             if original_line.startswith('//') or original_line.startswith('#'):
#                 log_file.write(f"COMMENT: {original_line}\n")
#                 log_file.flush()
#                 continue
                
#             # Check for termination commands - more precise matching
#             if (original_line.lower() in ["end", "eof"] or
#                 original_line.lower() == "end of run_tp" or
#                 any(re.search(r'\b' + re.escape(term) + r'\b', original_line.lower()) for term in ["end", "eof"])):
#                 display_message(f"Termination command '{original_line}' received. Stopping execution.", "black")
#                 log_file.write(f"COMMAND: {original_line}\n")
#                 log_file.write(f"RESPONSE: (Command not sent - execution terminated)\n")
#                 log_file.write("-" * 50 + "\n\n")
#                 log_file.flush()
#                 break    

#             # Log the command BEFORE execution
#             display_message(f"Executing command {line_index}/{total_lines}: {original_line}", "black")
#             log_file.write(f"COMMAND: {original_line}\n")
#             log_file.flush()  # Flush immediately so we see the command in the log
            
#             # Determine if this is a potentially long-running command
#             is_long_command = is_long_running_command(original_line)
#             is_long_command = 1
#             # Execute the command with appropriate timeout
#             timeout = 40 if is_long_command else 20
#             response = execute_single_command_adaptive(original_line, timeout=timeout)
            
#             print("Received response for command", line_index, "/", total_lines)
#             print("Response:", response)

#             # Log the response AFTER execution
#             display_message(f"Received response for command {line_index}/{total_lines}", "black")
#             log_file.write(f"RESPONSE: {response}\n")
#             log_file.write("-" * 50 + "\n\n")
#             log_file.flush()  # Ensure response is written immediately
            
#             # Adaptive delay based on command type and response size
#             delay = calculate_adaptive_delay(original_line, response)
#             time.sleep(delay)

#         log_file.close()
#         display_message(f"Responses saved to '{filename}'", "black")
#         display_message("File execution successful", "black")
#         status_label.config(text="Ready", foreground='black')
#         #progress_bar.stop()

#     except Exception as e:
#         display_message(f"File execution failed: {str(e)}", "red")
#         status_label.config(text="Ready", foreground='black')
#         #progress_bar.stop()
#         # Try to close the log file if it's open
#         try:
#             if 'log_file' in locals() and not log_file.closed:
#                 log_file.close()
#         except:
#             pass

def execute_file_thread(file_path):
    try:
        #progress_bar.start()
        status_label.config(text="Executing", foreground='black')
        display_message(f"File execution started", "black")
        print("File execution started")
        # Prepare log file
        filename = os.path.join(os.path.dirname(file_path), f"log_{os.path.basename(file_path)}.txt")
        log_file = open(filename, 'w')
        log_file.write(f"Log created at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        log_file.write("-" * 50 + "\n\n")

        # Read all lines from the file
        with open(file_path, 'r') as file:
            lines = file.readlines()
        
        # First, completely drain any existing data in the socket
        drain_socket_completely()
        
        # Process each line sequentially
        line_index = 0
        total_lines = len(lines)
        
        while line_index < total_lines:
            line = lines[line_index]
            original_line = line.strip()
            line_index += 1  # Move to next line for next iteration
            
            # Skip empty lines but log them
            if not original_line:
                log_file.write("\n")
                log_file.flush()
                continue
                
            # Handle comments - log them but don't send to server
            if original_line.startswith('//') or original_line.startswith('#'):
                log_file.write(f"COMMENT: {original_line}\n")
                log_file.flush()
                continue
                
            # Check for termination commands - more precise matching
            if (original_line.lower() in ["end", "eof"] or
                original_line.lower() == "end of run_tp" or
                any(re.search(r'\b' + re.escape(term) + r'\b', original_line.lower()) for term in ["end", "eof"])):
                display_message(f"Termination command '{original_line}' received. Stopping execution.", "black")
                log_file.write(f"COMMAND: {original_line}\n")
                log_file.write(f"RESPONSE: (Command not sent - execution terminated)\n")
                log_file.write("-" * 50 + "\n\n")
                log_file.flush()
                break    

            # Log the command BEFORE execution
            display_message(f"Executing command {line_index}/{total_lines}: {original_line}", "black")
            log_file.write(f"COMMAND: {original_line}\n")
            log_file.flush()  # Flush immediately so we see the command in the log
            
            start_time = time.time()
            # Determine if this is a potentially long-running command
            is_long_command = is_long_running_command(original_line)
            #is_long_command = 1
            # Execute the command with appropriate timeout
            timeout = 40 if is_long_command else 20
            response = execute_single_command_adaptive(original_line, timeout=timeout)
            end_time = time.time()
            elapsed = max(0, end_time - start_time - 6)

            print("Received response for command", line_index, "/", total_lines)
            print("Response:", response)
            #time.sleep(5)
            # Log the response AFTER execution
            display_message(f"Received response for command {line_index}/{total_lines}", "black")
            log_file.write(f"RESPONSE: \n{response}\n")
            log_file.write(f"Execution time : {elapsed:.2f} seconds\n")
            log_file.write("-" * 50 + "\n\n")
            log_file.flush()  # Ensure response is written immediately
            
            
            # Wait for additional responses
            time.sleep(1)  # Wait for 5 seconds to check for more output
            

            
            
            # Optional: Add a delay based on command type and response size
            delay = calculate_adaptive_delay(original_line, response)
            time.sleep(delay)
            


        log_file.close()
        display_message(f"Responses saved to '{filename}'", "black")
        display_message("File execution successful", "black")
        status_label.config(text="Ready", foreground='black')
        #progress_bar.stop()

    except Exception as e:
        display_message(f"File execution failed: {str(e)}", "red")
        status_label.config(text="Ready", foreground='black')
        #progress_bar.stop()
        # Try to close the log file if it's open
        try:
            if 'log_file' in locals() and not log_file.closed:
                log_file.close()
        except:
            pass


def is_long_running_command(command):
    """Determine if a command is likely to be long-running"""
    long_running_patterns = [
        'load', 'upload', 'pause', 'test', 'scan','readstatus','performanceparameters','inventory','getstatus'
        'start', 'initialize', 'reboot', 'reset', 'callos'
    ]
    return any(pattern in command.lower() for pattern in long_running_patterns)

def calculate_adaptive_delay(command, response):
    """Calculate an appropriate delay based on command type and response size"""
    base_delay = 0.5  # Minimum delay
    
    # Add delay based on response size
    response_length = len(response)
    size_factor = min(response_length / 1000, 2.0)  # Cap at 2 seconds
    
    # Add delay for specific command types
    if is_long_running_command(command):
        command_factor = 1.0
    else:
        command_factor = 0.2
        
    # If response indicates command is still processing, add more delay
    if "processing" in response.lower() or "**" in response.lower() or "please wait" in response.lower():
        processing_factor = 2.0
    else:
        processing_factor = 0.0
        
    total_delay = base_delay + size_factor + command_factor + processing_factor
    
    # Cap the maximum delay
    return min(total_delay, 5.0)

# def execute_single_command_adaptive(command, timeout=300, max_retries=3):
#     """Execute a command with adaptive response handling"""
#     for attempt in range(max_retries):
#         try:
#             # First, completely drain the socket to ensure no leftover data
#             drain_socket_completely()
            
#             # Send the command with a clear terminating newline
#             print(f"Sending command: {command}")
#             sock.sendall(command.encode() + b'\n')
            
#             # Wait for a complete response with adaptive timeout
#             response = wait_for_complete_response_enhanced(timeout=timeout)
            
#             # Verify we got a valid response
#             if response:
#                 # Simply return the response as is, without checking for error patterns
#                 return response
                
#             # If no response, try again
#             print(f"No response received on attempt {attempt+1}, retrying...")
#             time.sleep(1.0)
            
#         except Exception as e:
#             print(f"Attempt {attempt+1} failed: {str(e)}")
#             time.sleep(1.0)
    
#     # If we get here, all attempts failed
#     return ""

def wait_for_complete_response_adaptive(timeout=20):
    """Wait for a complete response with adaptive handling"""
    response = ""
    start_time = time.time()
    last_data_time = time.time()
    prompt_seen = False
    data_chunks = 0
    
    while time.time() - start_time < timeout:
        try:
            ready = select.select([sock], [], [], 0.5)
            
            if ready[0]:
                sock.settimeout(0.5)
                data = sock.recv(8192)
                
                if data:
                    last_data_time = time.time()
                    data_chunks += 1
                    chunk = data.decode(errors='replace')
                    response += chunk
                    
                    # Check for prompt character
                    if '>' in chunk:
                        prompt_seen = True
                        if response.rstrip().endswith('>'):
                            if len(response) < 1000:
                                break
                            time.sleep(0.5)
                            try:
                                sock.settimeout(0.5)
                                more_data = sock.recv(8192)
                                if more_data:
                                    response += more_data.decode(errors='replace')
                            except (socket.timeout, BlockingIOError):
                                pass
                            break
                else:
                    if data_chunks > 0:
                        break
            else:
                if prompt_seen and time.time() - last_data_time > 1.0:
                    break
                if data_chunks > 0 and time.time() - last_data_time > 2.0:
                    if "processing" in response.lower() or "please wait" in response.lower():
                        pass
                    else:
                        break
                if data_chunks == 0 and time.time() - start_time > timeout / 2:
                    print("Warning: No data received for an extended period")
                    continue
                    # Send a newline to try to get a response

        except socket.timeout:
            if prompt_seen and time.time() - last_data_time > 1.0:
                continue
        except Exception as e:
            print(f"Error receiving data: {str(e)}")
            break
    
    return response.strip()

# def wait_for_complete_response_adaptive(timeout=20):
#     """Wait for a complete response with adaptive handling"""
#     response = ""
#     start_time = time.time()
#     last_data_time = time.time()
#     prompt_seen = False
#     data_chunks = 0
    
#     while time.time() - start_time < timeout:
#         try:
#             # Use select to check if data is available without blocking
#             ready = select.select([sock], [], [], 0.5)
            
#             if ready[0]:
#                 # Data is available to read
#                 sock.settimeout(0.5)
#                 data = sock.recv(8192)
                
#                 if data:
#                     # Reset the last data time
#                     last_data_time = time.time()
#                     data_chunks += 1
                    
#                     # Decode and append to response
#                     chunk = data.decode(errors='replace')
#                     response += chunk
                    
#                     # Check for prompt character
#                     if '>' in chunk:
#                         prompt_seen = True
                        
#                         # If the prompt is at the end of the response or followed by whitespace only,
#                         # we're likely done
#                         if response.rstrip().endswith('>'):
#                             # For short responses with prompt, we can return immediately
#                             if len(response) < 1000:
#                                 break
                            
#                             # For longer responses, wait a bit more to ensure we get any final data
#                             time.sleep(0.5)
                            
#                             # Try to read any remaining data
#                             try:
#                                 sock.settimeout(0.5)
#                                 more_data = sock.recv(8192)
#                                 if more_data:
#                                     response += more_data.decode(errors='replace')
#                             except (socket.timeout, BlockingIOError):
#                                 pass  # No more data
                            
#                             break
#                 else:
#                     # Connection closed or no data
#                     if data_chunks > 0:  # Only break if we've received some data
#                         break
#             else:
#                 # No data available in this check interval
                
#                 # If we've seen a prompt and there's been no new data for a while,
#                 # we can consider the response complete
#                 if prompt_seen and time.time() - last_data_time > 1.0:
#                     break
                
#                 # If we've received some data but no new data for a while, check if we need to wait longer
#                 if data_chunks > 0 and time.time() - last_data_time > 2.0:
#                     # Check if response contains indicators that more data might be coming
#                     if "processing" in response.lower() or "please wait" in response.lower():
#                         # Keep waiting
#                         pass
#                     else:
#                         # Probably done
#                         break
                
#                 # If we haven't seen any data for too long, something might be wrong
#                 if data_chunks == 0 and time.time() - start_time > timeout / 2:
#                     print("Warning: No data received for an extended period")
#                     continue
#         except socket.timeout:
#             # No data available in this check interval
#             if prompt_seen and time.time() - last_data_time > 1.0:
#                 continue
#         except Exception as e:
#             print(f"Error receiving data: {str(e)}")
#             break
    
#     # Clean up the response
#     return response.strip()


def execute_single_command_adaptive(command, timeout=300, max_retries=3):
    """Execute a command with adaptive response handling and clean up leftovers."""
    for attempt in range(max_retries):
        try:
            # Drain any leftover output from previous commands
            drain_socket_completely()

            print(f"Sending command: {command}")
            sock.sendall(command.encode() + b'\n')



            # Wait for the command response
            response = wait_for_complete_response_enhanced(idle_timeout=15)

            if response:
                # Ensure no trailing data from slow prompt arrival
                leftover = drain_socket_completely(return_data=True)
                if leftover:
                    print(f"Extra data received after response: {leftover[:100]}...")
                    response += leftover
                return response.strip()

            print(f"No response received on attempt {attempt+1}, retrying...")
            time.sleep(1.0)

        except Exception as e:
            print(f"Attempt {attempt+1} failed: {str(e)}")
            time.sleep(1.0)

    # If all retries fail, still drain so next command starts clean
    drain_socket_completely()
    return ""


def drain_socket_completely(return_data=False):
    """Completely drain the socket of any pending data.
       If return_data=True, returns all drained data as a decoded string."""
    try:
        original_blocking = sock.getblocking()
        original_timeout = sock.gettimeout()

        sock.setblocking(False)

        buffer_data = b""
        start_time = time.time()

        while time.time() - start_time < 3.0:
            try:
                data = sock.recv(8192)
                if data:
                    buffer_data += data
                    if len(data) > 0:
                        start_time = time.time()  # reset timer
                else:
                    break
            except BlockingIOError:
                time.sleep(0.1)
            except Exception as e:
                print(f"Error while draining buffer: {str(e)}")
                break

    finally:
        try:
            sock.setblocking(original_blocking)
            if original_timeout is not None:
                sock.settimeout(original_timeout)
        except:
            sock.setblocking(True)
            sock.settimeout(30.0)

    if return_data:
        return buffer_data.decode(errors='replace')
    else:
        if buffer_data:
            print(f"Drained {len(buffer_data)} bytes leftover")
        return None

def execute_single_command_with_retry(command, max_retries=3):
    """Execute a single command with retries and ensure a proper response"""
    for attempt in range(max_retries):
        try:
            # First, completely drain the socket
            drain_socket_completely()
            
            # Send the command
            sock.sendall(command.encode() + b'\n')
            
            #time.sleep(5)
            # Wait for a complete response
            response = wait_for_complete_response_enhanced(idle_timeout=15)
            
            # If we got a valid response, return it
            if response and '>' in response:
                return response
                
            # If response doesn't contain a prompt, wait a bit more and try again
            time.sleep(1.0)
            
        except Exception as e:
            print(f"Attempt {attempt+1} failed: {str(e)}")
            time.sleep(1.0)
    
    # If we get here, all attempts failed
    return " "

# def drain_socket_completely():
#     """Completely drain the socket of any pending data"""
#     try:
#         # Save original socket settings
#         original_blocking = sock.getblocking()
#         original_timeout = sock.gettimeout()
        
#         # Set socket to non-blocking mode
#         sock.setblocking(False)
        
#         buffer_data = b""
#         start_time = time.time()
        
#         # Keep reading until there's no more data or timeout
#         while time.time() - start_time < 3.0:  # Try for up to 3 seconds
#             try:
#                 data = sock.recv(8192)
#                 if data:
#                     buffer_data += data
#                     print(f"Drained {len(data)} bytes from buffer")
#                     # Reset timer if we're still getting data
#                     if len(data) > 0:
#                         start_time = time.time()
#                 else:
#                     # Connection closed or no more data
#                     break
#             except BlockingIOError:
#                 # No more data to read right now, wait a bit and try again
#                 time.sleep(0.1)
#             except Exception as e:
#                 print(f"Error while draining buffer: {str(e)}")
#                 break
        
#         if buffer_data:
#             print(f"Total drained: {len(buffer_data)} bytes from buffer")
#             # Try to decode and log what was drained (for debugging)
#             try:
#                 print(f"Drained content: {buffer_data.decode(errors='replace')[:100]}...")
#             except:
#                 pass
            
#     except Exception as e:
#         print(f"Error draining socket: {str(e)}")
#     finally:
#         # Restore original socket settings
#         try:
#             sock.setblocking(original_blocking)
#             if original_timeout is not None:
#                 sock.settimeout(original_timeout)
#         except:
#             # In case of any issues, set reasonable defaults
#             sock.setblocking(True)
#             sock.settimeout(30.0)


def wait_for_prompt(timeout=5):
    """
    Wait specifically for the prompt character '>'.
    Returns all received data up to and including the prompt or timeout.
    """
    response = ""
    start_time = time.time()

    def has_prompt(text):
        # Same prompt detection logic for consistency
        if re.search(r"(^|\n)\s*>\s*$", text):
            return True
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        return bool(lines and lines[-1] == ">")

    while time.time() - start_time < timeout:
        try:
            ready, _, _ = select.select([sock], [], [], 0.5)
            if ready:
                data = sock.recv(8192)
                if data:
                    chunk = data.decode(errors='replace')
                    response += chunk
                    if has_prompt(response):
                        break
                else:
                    # Socket closed by remote end
                    break
        except Exception:
            break

    return response


def wait_for_complete_response_enhanced(idle_timeout=15):
    """
    Wait until the prompt '>' is received.
    Uses regex check + last-line check inside main loop.
    Never exits on timeout — only on confirmed prompt or socket close.
    Before final return, also runs wait_for_prompt() as a third check.
    """
    def has_prompt(text):
        # Method 1: Regex at end of line
        if re.search(r"(^|\n)\s*>\s*$", text):
            return True
        # Method 2: Last non-empty line is '>'
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        return bool(lines and lines[-1] == ">")

    response = ""
    last_data_time = time.time()
    #time.sleep(5)
    while True:
        try:
            ready, _, _ = select.select([sock], [], [], 0.5)
            if ready:
                data = sock.recv(8192)
                if data:
                    chunk = data.decode(errors='replace')
                    response += chunk
                    last_data_time = time.time()

                    if has_prompt(response):
                        # Third check before leaving
                        extra = wait_for_prompt(timeout=5)
                        if extra:
                            response += extra
                        if has_prompt(response):
                            print("Final prompt received — leaving.")
                            return response.strip()
                else:
                    # Socket closed
                    return response.strip()
            else:
                # Idle timeout hit — still just keep looping
                if time.time() - last_data_time >= idle_timeout:
                    # Reset idle timer but don’t exit
                    last_data_time = time.time()
        except socket.timeout:
            continue
        except Exception:
            return response.strip()


# def wait_for_complete_response_enhanced(timeout=300):
#     """
#     Waits until a full response is received, ending with a prompt '>' on a line by itself.
#     Never returns early due to a pause in data; only returns when the prompt is seen or timeout is reached.
#     """
#     response_lines = []
#     start_time = time.time()
#     buffer = ""
#     while time.time() - start_time < timeout:
#         try:
#             ready = select.select([sock], [], [], 0.5)
#             if ready[0]:
#                 sock.settimeout(0.5)
#                 data = sock.recv(8192)
#                 if data:
#                     chunk = data.decode(errors='replace')
#                     buffer += chunk
#                     start_time = time.time()
#                     print(f"Received data: {chunk}")
#                     # Process complete lines
#                     while '\n' in buffer or '\r' in buffer:
#                         if '\n' in buffer:
#                             line, buffer = buffer.split('\n', 1)
#                         else:
#                             line, buffer = buffer.split('\r', 1)
#                         line = line.rstrip('\r')
#                         response_lines.append(line)
#                         # Only break if the prompt is the last line received
#                         if line.strip() == '>':
#                             return '\n'.join(response_lines).strip()
#                 else:
#                     time.sleep(0.1)
#                     continue  # Connection closed
#             # If no data, just continue waiting until timeout or prompt
#         except socket.timeout:
#             start_time = time.time()
#             continue
#         except Exception as e:
#             print(f"Error receiving data: {str(e)}")
#             start_time = time.time()
#             continue
#         #print("Warning: Timeout reached before prompt '>' was received!")
#         #
#     return '\n'.join(response_lines).strip()
# # Thread to handle file execution and response processing
# def execute_file_thread(file_path):
#     try:
#         # Start the progress bar
#         #progress_bar.start()
#         # Update status to "executing"
#         status_label.config(text="Executing", foreground='black')

#         time.sleep(1)
#         display_message(f"File execution started", "black")
#         responses = []
        
#         # Read commands from the file
#         with open(file_path, 'r') as file:
#             commands = [line.strip() for line in file if line.strip() and not line.startswith('//') and not line.startswith('#')]

#         for cmd in commands:
#             # Check for termination commands
#             if any(term in cmd.lower() for term in ["end", "eof", "exit", "quit", "bye", "stop", "close", "end of run_tp"]):
#                 display_message(f"Termination command '{cmd}' received. Stopping execution.", "black")
#                 break
#             print(f"about to send command :{cmd}")
#             display_message(f"Executing command: {cmd}", "black")
#             sock.sendall(cmd.encode() + b'\n')
#             display_message(f"Sent command: {cmd}", "black")

#             response = ""
#             start_time = time.time()
#             while time.time() - start_time < 60:  # 100 seconds timeout
#                 ready = select.select([sock], [], [], 50)  # Check for data every 50 seconds

#                 if ready[0]:
#                     data = sock.recv(4096)
#                     if data:
#                         resp_text = data.decode().strip()
#                         response += resp_text
#                         display_message(f"Received response from server: {resp_text}", "black")
#                     else:
#                         break
#                 else:
#                     break

#             responses.append((cmd, response))  # Store command and response
#             display_message(f"Final response for command '{cmd}': {response}", "black")

#         # Log all responses with commands after processing
#         log_responses_to_file(responses, file_path)  
#         display_message("File execution successful", "black")

#         # Update status to "ready"
#         status_label.config(text="Ready", foreground='black')
#         #progress_bar.stop()  # Stop the progress bar

#     except Exception as e:
#         display_message(f"File execution failed: {str(e)}", "red")
#         # Update status to "ready" in case of failure
#         status_label.config(text="Ready", foreground='black')

def log_responses_to_file(responses, file_path):
    try:
        # Update status to "executing"
        status_label.config(text="Executing", foreground='black')

        # Create log file in the same directory as the command file
        filename = os.path.join(os.path.dirname(file_path), 
                                 f"log_{os.path.basename(file_path)}.txt")

        with open(filename, 'w') as log_file:
            # Add timestamp at the top of the log file
            log_file.write(f"Log created at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            log_file.write("-" * 50 + "\n\n")

            for cmd, response in responses:
                log_file.write(f"COMMAND: {cmd}\n")
                log_file.write(f"RESPONSE: {response if response else '(No response)'}\n")
                log_file.write("-" * 50 + "\n")

        display_message(f"Responses saved to '{filename}'", "black")

        # Update status to "ready"
        status_label.config(text="Ready", foreground='black')
    except Exception as e:
        display_message(f"Failed to log responses: {str(e)}", "red")
        # Update status to "ready" in case of failure
        status_label.config(text="Ready", foreground='black')


# Function to load the driver for DMS4
def load_driver(driver):
    try:
        # Determine the driver path to load
        driver_path_to_load = driver_path_var.get() if driver_path_changed else driver

        script_dir = os.path.dirname(__file__)
        driver_path = os.path.join(script_dir, driver_path_to_load)
        cmd = f"load('{driver_path_to_load}')"
        display_message(f"Command to load driver in the path: {driver_path_to_load}", "black")
        response = socket_query(sock, cmd)
        if response:
            display_message(f"Response :\n{response}", "black")
            display_message("Driver loaded successfully", "black")
        else:
            display_message("No response received", "red")
         
        # Simulate loading the driver
        time.sleep(1)  # Simulate time delay for loading

    except Exception as e:
        display_message(f"Failed to load driver: {str(e)}", "red")


# Function to show a notification
def show_notification(title, message):
    messagebox.showinfo(title, message)


def update_file_list(dut_folder_var, file_list_var,dut_profile_dropdown):
    folder_path = dut_folder_var.get()
    if os.path.isdir(folder_path):
        files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
        file_list_var.set(files)
        dut_profile_dropdown['values'] = files
    else:
        file_list_var.set([])
        dut_profile_dropdown['values'] = []

# Function to display the default list and select the default field
def get_files_list_thread():
    threading.Thread(target=display_default_list).start()

# Function to display the default list and select the default field
def display_default_list():
    global profile_path_var
    try:
        # Start the progress bar
        #progress_bar.start()
       
        # Clear the listbox first
        profile_listbox.delete(0, tk.END)
       
        # Update status to "executing"
        status_label.config(text="Fetching files list", foreground='black')
       
        # Use default path if no folder is selected
        default_path = "C:\\Sparnex.testplan\\ICL_Server\\DSLAM\\DMS4\\Profiles"
        folder = profile_path_var.get() if profile_path_var.get() else default_path
       
        # Debugging: Display the folder path being used
        display_message(f"Using folder path: {folder}", "black")
       
        # Command to get the list of files from the selected folder
        cmd = f"folder = '{folder}'"
        display_message(f"Command to display profiles in given path: {folder}", "black")
        response = socket_query(sock, cmd)
        if response:
            display_message(f"Response :\n{response}", "black")
        else:
            display_message("No response received", "red")

        cmd = f"(e,r)=callos('dir '++folder) <<< e <<< r"
        display_message(f"Executing command: {cmd}", "black")
       
        response = socket_query(sock, cmd)
        time.sleep(3) # Wait
       
        if response:
            display_message(f"Response :\n{response}", "black")
            file_names = extract_file_names(response)
            time.sleep(1)
            populate_profile_listbox(file_names)
           
            # Select the default field
            default_field = "BCM6400x_Default_VDSL.txt"
            if default_field in file_names:
                index = file_names.index(default_field)
                profile_listbox.selection_set(index)
                profile_listbox.activate(index)
               
                # Trigger the on_profile_select event
                profile_listbox.event_generate("<<ListboxSelect>>")
        else:
            display_message("No response received", "red")
       
        # Update status to "ready"
        status_label.config(text="Ready", foreground='black')
    except Exception as e:
        display_message(f"Failed to fetch files list: {str(e)}", "red")
    finally:
        # Stop the progress bar
        #progress_bar.stop()
        display_message(f" ", "red")


def browse_save_profile():
    try:
        # Start the progress bar
        #progress_bar.start()
       
        # Get current path from the entry, or default if none
        current_path = profile_path_var.get()
        if current_path == "No profile folder selected":
            current_path = os.getcwd()

        # Browse for the folder
        folder_path = filedialog.askdirectory(
            initialdir=current_path,
            title="Select DSLAM Profile Folder"
        )
        
        if folder_path:
            # Convert forward slashes to backward slashes
            folder_path = folder_path.replace('/', '\\')

            # Save the selected folder path
            profile_path_var.set(folder_path)
            display_message(f"Selected folder: {folder_path}", "black")
            
            # Automatically trigger get_files_list after folder selection
            #root.after(100, get_files_list_thread)  # Small delay to ensure UI updates
        else:
            display_message("No folder selected.", "red")
            
    except Exception as e:
        display_message(f"Failed to select folder: {str(e)}", "red")
        profile_path_var.set("No profile folder selected")
    finally:
        # Stop the progress bar
        #progress_bar.stop()
        display_message(f" ", "red")




# Thread-safe function to load default Broadcom configuration
def load_profile():
    try:
        # Start the progress bar
        #progress_bar.start()
       
        if sock is None:
            message_queue.put(("Not connected to the server. Please connect first.", "red"))
            return
        load_profile_button.config(state='disabled')
       
        # Check if a profile is selected in the listbox
        selected_index = profile_listbox.curselection()
        if selected_index:
            selected_profile = profile_listbox.get(selected_index)
        else:
            # Default profile if none is selected
            selected_profile = "BCM6400x_Default_VDSL.txt"
       
        bcm_profile = selected_profile
       
        # Start the Default_Broadcom_Configuration thread with the selected profile
        threading.Thread(target=Default_Broadcom_Configuration, args=(bcm_address, bcm_profile)).start()
        load_profile_button.config(state='normal')
    except Exception as e:
        display_message(f"Failed to load profile: {str(e)}", "red")
        load_profile_button.config(state='normal')
    finally:
        # Stop the progress bar
        ##progress_bar.stop()
        load_profile_button.config(state='normal')

def Default_Broadcom_Configuration(bcm_address, FN_profile):
    try:
        # Update status to "executing"
        status_label.config(text="Executing", foreground='black')
       
        # Simulate Default Broadcom Configuration
        display_message(f"Loading DSL Configuration: {FN_profile}", "black")

        full_path = os.path.join(profile_path_var.get(), FN_profile)
       
        display_message(f"Executing command - api stop", "black")
        cmd = f"(e, r) = sendto('{bcm_address}', 'api stop ' ++ {selected_port}) <<< e <<< r"
        response = socket_query(sock, cmd)
        if response:
            display_message(f"Response :\n{response}", "black")
        else:
            display_message("No response received", "red")


        display_message(f"Reading status", "black")
   
        selected_port_temp = port_var.get()
        # Execute the profile line by line
        execute_profile_line_by_line(sock,bcm_address, selected_port_temp,full_path)

        # Additional commands after profile execution
        time.sleep(1)

        # Check if 'bcm' is part of the profile text file
        with open(full_path, 'r') as file:
            profile_content = file.read()
       
        if 'bcm' in profile_content:
            display_message(f"'bcm cfg' found in {FN_profile} : 'api apply config' is not implemented", "black")
        else:
            display_message(f"Executing command - api cfg apply", "black")
            cmd = f"(e, r) = sendto('{bcm_address}', 'api cfg apply ' ++ {selected_port}) <<< e <<< r"
            response = socket_query(sock, cmd)
            if response:
                display_message(f"Response :\n{response}", "black")
            else:
                display_message("No response received", "red")
            display_message(f"'bcm cfg' not found in {FN_profile}. Proceeding with normal load.", "black")

        display_message(f"Executing command - api start", "black")
        cmd = f"(e, r) = sendto('{bcm_address}', 'api start ' ++ {selected_port}) <<< e <<< r"
        response = socket_query(sock, cmd)
        if response:
            display_message(f"Response :\n{response}", "black")
        else:
            display_message("No response received", "red")
    except Exception as e:
        display_message(f"Failed to load Default DSL Configuration: {str(e)}", "red")
   
    # Update status to "ready"
    status_label.config(text="Ready", foreground='black')

        # display_message(f"Executing command - api getstate all", "black")
        # cmd = f"(e, r) = sendto('{bcm6400x_address}', 'api api getstate all') <<< e <<< r"

        # time.sleep(1)  # Wait
        # response = socket_query(sock, cmd)
     
        # if response:
        #     display_message(f"Response :\n{response}", "black")
        # else:
        #     display_message("No response received", "red")




# Add a variable to store the driver path
driver_path_var = tk.StringVar(value=bcm_file)
driver_path_changed = False


# Function to browse and save the new driver path
def browse_save_driver():
    global driver_path_changed
    try:
        # Get current driver path
        current_path = driver_path_var.get()
        initial_dir = None
        
        if current_path:
            if current_path.startswith('\\\\'): # Network path
                # Convert network path to a format that tkinter can understand
                initial_dir = os.path.dirname(current_path).replace('\\', '/')
                display_message(f"Opening network path: {initial_dir}", "black")
            elif current_path.startswith('..'): # Relative path
                # Try to resolve relative path
                pwd_process = subprocess.Popen('cd', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                current_dir, _ = pwd_process.communicate()
                if current_dir:
                    full_path = os.path.abspath(os.path.join(current_dir.strip(), current_path))
                    initial_dir = os.path.dirname(full_path)
                    display_message(f"Resolved relative path to: {initial_dir}", "black")
            else: # Local path
                initial_dir = os.path.dirname(current_path)
                display_message(f"Opening local path: {initial_dir}", "black")
        
        file_path = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="Select DSLAM driver",
            filetypes=[("All files", "*.*")]
        )
        
        if file_path:
            # Convert back to Windows path format if needed
            if '/' in file_path:
                file_path = file_path.replace('/', '\\')
            
            driver_path_var.set(file_path)
            driver_path_changed = True
            display_message(f"Selected file: {file_path}", "black")
        else:
            driver_path_var.set(bcm_file)
            display_message(f"Using default driver: {bcm_file}", "black")
            
    except Exception as e:
        display_message(f"Failed to select driver: {str(e)}", "red")
        print(f"Browse save driver error: {str(e)}")  # Debug print


def on_driver_select(event):
    try:
        # Initialize variables
        selected_driver = None
        global bcm_address, FN_profile, bcm_file, selected_driver_name  # Declare globals
        
        selected_index = devices_listbox.curselection()
        if not selected_index:
            return
            
        selected_driver = devices_listbox.get(selected_index[0])  # Get the first selected item
        print(f"Selected driver: {selected_driver}")  # Debug print
        
        # Enhanced header row check - check if it's the header row or contains header text
        if any(header in selected_driver.upper() for header in ["NAME", "ADDRESS"]) or selected_index[0] == 0:
            devices_listbox.selection_clear(0, tk.END)
            return
            
        # Split the selected driver text and check if it has enough parts
        driver_parts = selected_driver.split()
        if len(driver_parts) < 2:  # Must have at least name and address
            devices_listbox.selection_clear(0, tk.END)
            return
            
        # For SFP drivers, keep full name except last part (address)
        if "SFP" in selected_driver:
            driver_name = ' '.join(selected_driver.split()[:-1])  # Keep the full SFP name
            selected_driver_name = driver_name  # Store the full name
        else:
            # For other drivers, store both name and address
            driver_name = driver_parts[0]
            driver_address = driver_parts[-1]  # Get the address part
            selected_driver_name = f"{driver_name} {driver_address}"  # Store full string
            
        print(f"Looking up driver: {selected_driver_name}")  # Debug print
        
        # Get driver details with full string
        details = get_all_drivers_details(selected_driver_name)
        print(f"Got details: {details}")  # Debug print
        
        if details:
            bcm_address, FN_profile, bcm_file = details
            
            # Debug prints
            print(f"Address: {bcm_address}")
            print(f"Profile: {FN_profile}")
            print(f"File: {bcm_file}")
            
            display_message(f"Selected DSLAM driver: {driver_name}", "black")
            display_message(f"Selected Port: {selected_port}", "black")
            
            if bcm_address:
                display_message(f"Selected Driver Address: {bcm_address}", "black")
            if FN_profile:
                display_message(f"Selected Driver Profile: {FN_profile}", "black")
            if bcm_file:
                display_message(f"Selected Driver File: {bcm_file}", "black")
                driver_path_var.set(bcm_file)  # Update the driver path display
                
                # Only start thread if we have all required details
                threading.Thread(target=load_and_read_driver, 
                               args=(selected_port, bcm_file, bcm_address)).start()
            else:
                display_message("Error: Missing driver file path", "red")
        else:
            thread_safe_display_message(f"Could not find details for driver: {selected_driver_name}", "red")
            
    except Exception as e:
        thread_safe_display_message(f"Error selecting DSLAM driver: {str(e)}", "red")
        import traceback
        print(traceback.format_exc())  # Print full error trace
    finally:
        # Restore selection after any action
        if selected_index:
            devices_listbox.selection_clear(0, tk.END)
            devices_listbox.selection_set(selected_index)
            devices_listbox.see(selected_index)

# # Function to handle the selection of a DSLAM driver
# def on_driver_select(event):
#     try:
#         # Initialize variables
#         selected_driver = None
#         global bcm_address, FN_profile, bcm_file, selected_driver_name  # Declare globals
        
#         selected_index = devices_listbox.curselection()
#         if not selected_index:
#             return
            
#         selected_driver = devices_listbox.get(selected_index[0])  # Get the first selected item
#         print(f"Selected driver: {selected_driver}")  # Debug print
        
#         # Enhanced header row check - check if it's the header row or contains header text
#         if any(header in selected_driver.upper() for header in ["NAME", "ADDRESS"]) or selected_index[0] == 0:
#             devices_listbox.selection_clear(0, tk.END)
#             return
            
#         # Split the selected driver text and check if it has enough parts
#         driver_parts = selected_driver.split()
#         if len(driver_parts) < 2:  # Must have at least name and address
#             devices_listbox.selection_clear(0, tk.END)
#             return
            
#         # For SFP drivers, keep full name except last part (address)
#         if "SFP" in selected_driver:
#             driver_name = ' '.join(selected_driver.split()[:-1])  # Keep the full SFP name
#             selected_driver_name = driver_name  # Store the full name
#         else:
#             # For other drivers, take first part as name
#             driver_name = selected_driver.split()[0]
#             selected_driver_name = driver_name
            
#         print(f"Looking up driver name: {driver_name}")  # Debug print
        
#         # Get driver details
#         details = get_all_drivers_details(driver_name)
#         print(f"Got details: {details}")  # Debug print
        
#         if details:
#             bcm_address, FN_profile, bcm_file = details
            
#             # Debug prints
#             print(f"Address: {bcm_address}")
#             print(f"Profile: {FN_profile}")
#             print(f"File: {bcm_file}")
            
#             display_message(f"Selected DSLAM driver: {driver_name}", "black")
#             display_message(f"Selected Port: {selected_port}", "black")
            
#             if bcm_address:
#                 display_message(f"Selected Driver Address: {bcm_address}", "black")
#             if FN_profile:
#                 display_message(f"Selected Driver Profile: {FN_profile}", "black")
#             if bcm_file:
#                 display_message(f"Selected Driver File: {bcm_file}", "black")
                
#                 # Only start thread if we have all required details
#                 threading.Thread(target=load_and_read_driver, 
#                                args=(selected_port, bcm_file, bcm_address)).start()
#             else:
#                 display_message("Error: Missing driver file path", "red")
#         else:
#             thread_safe_display_message(f"Could not find details for driver: {driver_name}", "red")
            
#     except Exception as e:
#         thread_safe_display_message(f"Error selecting DSLAM driver: {str(e)}", "red")
#         import traceback
#         print(traceback.format_exc())  # Print full error trace
#     finally:
#         # Restore selection after any action
#         if selected_index:
#             devices_listbox.selection_clear(0, tk.END)
#             devices_listbox.selection_set(selected_index)
#             devices_listbox.see(selected_index)



# Function to read the status of a DSLAM port
def load_and_read_driver(port,bcm_file,bcm_address):
    try:
        load_driver(bcm_file)
        inventory_dslam(bcm_address, port)
    except Exception as e:
        thread_safe_display_message(f"Error loading driver or reading status: {str(e)}", "red")

# Function to load selected command into the command entry field
# Function to load selected command into the command combobox
def load_command_from_history(event):
    try:
        selected_index = history_listbox.curselection()
        if selected_index:
            selected_command = history_listbox.get(selected_index)
            command_combobox.set(selected_command)
            
            # Add the command to the dropdown list if it's not already there
            current_values = list(command_combobox['values'])
            if selected_command not in current_values:
                command_combobox['values'] = tuple([selected_command] + current_values)
                
    except Exception as e:
        display_message(f"Error loading command from history: {str(e)}", "red")

# Function to add command to history
def add_to_history(command):
    history_listbox.insert(tk.END, command)

# Function to add a new line to the file
def add_new_line_to_file(file_path):
    try:
        with open(file_path, 'a') as file:
            file.write('\n')  # Add a new line at the end of the file
        display_message(f"New line added to '{file_path}'", "black")
    except Exception as e:
        display_message(f"Failed to add new line to file: {str(e)}", "red")

# Function to remove the last line from the file
def remove_last_line_from_file(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
        if lines:
            with open(file_path, 'w') as file:
                file.writelines(lines[:-1])  # Write all lines except the last one
            display_message(f"Last line removed from '{file_path}'", "black")
    except Exception as e:
        display_message(f"Failed to remove last line from file: {str(e)}", "red")

# Function to process the file: add a new line, execute it, and then remove the added line
def process_file(file_path):
    add_new_line_to_file(file_path)
    #process_file_part2(file_path)
    execute_file_thread(file_path)
    remove_last_line_from_file(file_path)


#Add a function to parse the directory listing and extract file names.
def extract_file_names(output):
    file_names = []
    lines = output.split('\n')
    for line in lines:
        if line.strip() and line.startswith((' Volume', ' Directory', '0', '1', '2', '7', 'Processing line:')):
            parts = line.split()
            print(f"Line parts: {parts}")  # Debugging statement
            if len(parts) > 0 and parts[-1].endswith('.txt'):
                file_names.append(parts[-1])
               # print(f"Added file name: {parts[-1]}")  # Debugging statement
               # display_message(f"Added file name: {parts[-1]}", "black")
    print(f"Extracted file names: {file_names}")  # Debugging statement
   # display_message(f"Extracted file names: {file_names}", "black")
    return file_names


#Add a function to populate the profile list box with the extracted file names.
def populate_profile_listbox(file_names):
    global profile_listbox
    try:
        profile_listbox.delete(0, tk.END)
        for file_name in file_names:
            profile_listbox.insert(tk.END, file_name)
         # # Trigger the on_profile_select event
                # devices_listbox.event_generate("<<ListboxSelect>>")
        profile_listbox.bind('<<ListboxSelect>>', on_profile_select)
    except Exception as e:
        print(f"Error populating profile listbox: {str(e)}")


#Add a function to handle the selection of a profile from the list box.
def on_profile_select(event):
    try:
        selected_index = profile_listbox.curselection()
        if selected_index:
            selected_profile = profile_listbox.get(selected_index)
            display_message(f"Selected Profile: {selected_profile}", "black")
    except Exception as e:
        print(f"Error selecting profile: {str(e)}")

def update_preferences_mid_session():
    if toggle_var.get():
        threading.Thread(target=use_preferences).start()


# Create a frame for the navigation buttons
#nav_frame = ttk.Frame(root, padding="5 5 5 5")
#nav_frame.grid(row=0, column=0, sticky="w")

# # Add buttons to navigate to Page 1, Page 2, and Page 3
# ttk.Button(nav_frame, text="Device Configurator",cursor="hand2", command=lambda: show_frame(main_frame)).pack(side="left", padx=5, pady=5)
# ttk.Button(nav_frame, text="Test Plan Configurator", cursor="hand2",command=lambda: show_frame(page2_frame)).pack(side="left", padx=5, pady=5)
# ttk.Button(nav_frame, text="Test Plan Selection", cursor="hand2",command=lambda: show_frame(page3_frame)).pack(side="left", padx=5, pady=5)
# ttk.Button(nav_frame, text="Go to Page 4", cursor="hand2",command=lambda: show_frame(page4_frame)).pack(side="left", padx=5, pady=5)

# Main Frame for padding and organization
#main_frame = ttk.Frame(root, padding="5 5 5 5")
#main_frame.grid(row=1, column=0, sticky="nsew")
# Create notebook
# Create notebook
# Create notebook below the logo
notebook = ttk.Notebook(main_container)
notebook.pack(fill="both", expand=True, padx=5, pady=2)

# Create frames with notebook as parent
main_frame = ttk.Frame(notebook, padding="5 5 5 5")
main_frame.columnconfigure(0, weight=1)
main_frame.columnconfigure(1, weight=1)

# Create the server frame (left side)
combined_server_frame = ttk.LabelFrame(main_frame, text="Device Configuration", padding="5 5 5 5")
combined_server_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

# Create page2_frame for Test Plan Configurator
page2_frame = ttk.Frame(notebook, padding="5 5 5 5")
page2_frame.columnconfigure(0, weight=1)
page2_frame.columnconfigure(1, weight=1)
page2_frame.columnconfigure(2, weight=1)
page2_frame.rowconfigure(0, weight=1)

# Create other pages
page3_frame = ttk.Frame(notebook, padding="5 5 5 5")
page4_frame = ttk.Frame(notebook, padding="5 5 5 5")

# Add frames to notebook in order
notebook.add(main_frame, text='Device Configurator')
#notebook.add(page2_frame, text='Test Plan Configurator')
#notebook.add(page3_frame, text='Test Plan Selection')
#notebook.add(page4_frame, text='Results')
# Now call create_page2() to populate page2_frame
# Function to switch notebook tabs
def switch_tab(tab_index):
    notebook.select(tab_index)




# Page 3 specific status update function
def update_status_3(message=None, is_error=False):
    """Updates Page 3 status label with either custom message or current file status"""
    try:
        global current_tp_file, status_label_3, status_var_3
        if status_label_3 is None or status_var_3 is None:
            return  # Exit if labels aren't created yet        
        # If a specific message is provided, display it
        if message:
            status_var_3.set(message)
            status_label_3.configure(
                font=("Helvetica", 10, "bold"),
                foreground="red" if is_error else "black"
            )
        # Otherwise, show current file status
        else:
            if current_tp_file and os.path.exists(current_tp_file):
                file_name = os.path.basename(current_tp_file)
                status_var_3.set(f"Selected Test Plan: {file_name}")
                status_label_3.configure(
                    font=("Helvetica", 10, "bold"),
                    foreground="black"
                )
            else:
                status_var_3.set("No Test Plan selected")
                status_label_3.configure(
                    font=("Helvetica", 10, "bold"),
                    foreground="black"
                )
        
        # Force update
        status_label_3.update()
        
    except Exception as e:
        print(f"Error updating page 3 status: {str(e)}")
        # Set default error state
        status_var_3.set("Error updating status")
        status_label_3.configure(foreground="red")






    # Update the filter_items function to use the new variable
def filter_items(*args):
        search_text = listbox_search_var.get().lower()  # Use listbox_search_var instead
        listbox1.delete(0, tk.END)  # Clear current items
        
        if not search_text:
            for item in all_items:
                listbox1.insert(tk.END, item)
            return
    
        for item in all_items:
            if search_text in item.lower():
                listbox1.insert(tk.END, item)

def clear_search():
        listbox_search_var.set("")  # Update to use listbox_search_var


            
def transfer_items(from_list, to_list):
        """Transfer from listbox1 to listbox2 - prevent duplicates"""
        selections = from_list.curselection()
        if not selections:
            update_status_3("Please select items to transfer", True)
            return
        
        added_count = 0
        skipped_count = 0
        
        for index in selections:
            test_name = from_list.get(index)
            # Check if already in listbox2
            if test_name in to_list.get(0, tk.END):
                skipped_count += 1
                update_status_3(f"Skipping duplicate in selected plans: {test_name}")
                continue
            
            to_list.insert(tk.END, test_name)
            added_count += 1
    
        update_status_3(f"Added {added_count} plans to selection, skipped {skipped_count} duplicates")
    
def load_test_inputs():
        """Load test inputs for selected test plan into preview tree"""
        selection = listbox2.curselection()
        if not selection:
            update_status_3("Please select a test from the list first", True)
            return
    
        selected_test = listbox2.get(selection[0])
        update_status_3(f"Loading test inputs for {selected_test}...")

        # Clear previous test inputs
        preview_tree.delete(*preview_tree.get_children())

        try:
            # Get test inputs from server
            cmd1 = f'spt = "{selected_test}"'
            socket_query(sock, cmd1)
            
            # Get test inputs using PerfTable_properties
            cmd2 = 'myproperties=ST::PerfTable_properties(spt)'
            socket_query(sock, cmd2)

            cmd3 = '<<< myproperties.arg1'
            response = socket_query(sock, cmd3)
    
            if response:
                test_inputs = response.strip().split('\n')
                valid_inputs = []
                no_param_inputs = []
                
                # First separate valid inputs and "No Parameter" inputs
                for test_input in test_inputs:
                    test_input = test_input.strip()
                    if test_input:
                        if '>' in test_input:
                            no_param_inputs.append(test_input)
                        else:
                            valid_inputs.append(test_input)
            
                # If we have valid inputs, only show those
                if valid_inputs:
                    for test_input in valid_inputs:
                        preview_tree.insert("", "end", values=(selected_test, test_input))
                    update_status_3(f"Loaded {len(valid_inputs)} test inputs for {selected_test}")
            
                # If no valid inputs exist, show non-selectable "No Parameter" entries
                elif no_param_inputs:
                    for _ in no_param_inputs:
                        item = preview_tree.insert("", "end", 
                                                values=(selected_test, "No Parameter"),
                                                tags=('no_parameter',))
                        # Make item non-selectable
                        preview_tree.tag_bind('no_parameter', '<1>', lambda e: 'break')
                
                    # Configure the "No Parameter" style
                    preview_tree.tag_configure('no_parameter', 
                                             font=('Helvetica', 9, 'italic'),
                                         foreground='gray')
                    update_status_3(f"Test {selected_test} has no valid parameters", True)
            
                else:
                    update_status_3(f"No test inputs found for {selected_test}")
        
            else:
                update_status_3(f"No test inputs found for {selected_test}")
        
        except Exception as e:
            update_status_3(f"Error loading test inputs: {str(e)}", True)


def load_selected_plans():
        """Load all plans from listbox2 to main table"""
        try:
            print("Load Selected Plans button clicked")
        
            # Get all items from listbox2
            all_tests = listbox2.get(0, tk.END)
            if not all_tests:
                update_status_3("No plans in the selection list", True)
                return
        
            total_tests = len(all_tests)
            added_count = 0
            skipped_count = 0
        
            update_status_3(f"Starting to load {total_tests} test plans...")
            root.update_idletasks()

            def process_next_test(index=0):
                """Process tests one by one using recursion"""
                nonlocal added_count, skipped_count
            
                if index >= len(all_tests):
                    # All tests processed
                    update_status_3(f"Completed: Added {added_count} plans.")
                    #skipped {skipped_count} duplicates")
                    canvas.configure(scrollregion=canvas.bbox("all"))
                    return

                test_name = all_tests[index]
                print(f"Processing test: {test_name}")  # Debug print

                # Check for duplicates
                duplicate = False
                for widget in content_frame.winfo_children():
                    if isinstance(widget, ttk.Frame):
                        try:
                            name_entry = widget.grid_slaves(row=0, column=0)[0]
                            input_entry = widget.grid_slaves(row=0, column=1)[0]
                            existing_name = name_entry.get()
                            existing_input = input_entry.get()
                            if existing_name == test_name and existing_input.strip():
                                duplicate = False
                                skipped_count += 1
                                #update_status_3(f"Skipping duplicate: {test_name} ({index + 1}/{total_tests})")

                                root.update_idletasks()
                                # Process next test after a short delay
                                root.after(100, lambda: process_next_test(index + 1))
                                return
                        except IndexError:
                            continue

                if not duplicate:
                    try:
                        # Get test details from server
                        cmd1 = f'spt = "{test_name}"'
                        socket_query(sock, cmd1)
                    
                        cmd2 = '(err, info) = test_plan::Get_TPlist(spt) <<< info'
                        response = socket_query(sock, cmd2)

                        profile = ""
                        if response and isinstance(response, str):
                            info_lines = response.strip().split('\n')
                            if len(info_lines) >= 3:
                                profile = info_lines[2].strip()

                        # Add row to main table
                        add_row_with_details(
                            name=test_name,
                            test_input="",
                            profile=profile
                        )
                        added_count += 1
                    
                        remaining = total_tests - (index + 1)
                        if profile:
                            update_status_3(f"Added test ({index + 1}/{total_tests}, {remaining} remaining)")
                        else:
                            update_status_3(f"Added test (no profile) ({index + 1}/{total_tests}, {remaining} remaining)")
                    
                    except Exception as e:
                        print(f"Error processing test {test_name}: {str(e)}")
                        update_status_3(f"Error with {test_name}: {str(e)} ({index + 1}/{total_tests})", True)
                    
                        # Still add the row with available information
                        add_row_with_details(
                            name=test_name,
                            test_input="",
                            profile=""
                        )
                        added_count += 1

                    content_frame.update_idletasks()
                    canvas.configure(scrollregion=canvas.bbox("all"))
                    root.update_idletasks()

                # Process next test after a short delay
                root.after(100, lambda: process_next_test(index + 1))

            # Start processing the first test
            process_next_test(0)
            update_status_3("Completed: Added plans")
            
            


        except Exception as e:
            print(f"Error in load_selected_plans: {str(e)}")
            update_status_3(f"Error loading plans: {str(e)}", True)



table_search_var = tk.StringVar()


    
    # Clear search button
def clear_search():
            table_search_var.set("")
    
def clear_main_table():
        """Clear all rows from the main table"""
        try:
            for widget in content_frame.winfo_children():
                widget.destroy()
            canvas.configure(scrollregion=canvas.bbox("all"))
            update_status_3("Table cleared")
        except Exception as e:
            update_status_3(f"Error clearing table: {str(e)}", True)

def filter_table_rows(*args):
        """Filter the main table based on search text"""
        search_term = table_search_var.get().lower()
        
        try:
            # Get all row frames
            rows = [widget for widget in content_frame.winfo_children() 
                   if isinstance(widget, ttk.Frame) and hasattr(widget, 'searchable_text')]
        
            for row_frame in rows:
                if not search_term:
                    # Show all rows if search is empty
                    row_frame.pack(fill="x", padx=2, pady=1)
                    if hasattr(row_frame, 'separator'):
                        row_frame.separator.pack(fill="x")
                else:
                    # Show/hide based on search term
                    if search_term in row_frame.searchable_text:
                        row_frame.pack(fill="x", padx=2, pady=1)
                        if hasattr(row_frame, 'separator'):
                            row_frame.separator.pack(fill="x")
                    else:
                        row_frame.pack_forget()
                        if hasattr(row_frame, 'separator'):
                            row_frame.separator.pack_forget()
        
            # Update canvas scroll region
            content_frame.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
                
        except Exception as e:
            print(f"Error filtering table: {str(e)}")



def create_test_plan():
        global current_tp_file
        """Create a new test plan"""
        try:
            # # Check server connection
            # if not sock:
            #     messagebox.showerror("Error", "Not connected to server. Please connect first.")
            #     return

            # Get number of rows in main table
            rows = [widget for widget in content_frame.winfo_children() if isinstance(widget, ttk.Frame)]
            row_count = len(rows)

            # Prompt user to save the file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".tp",
                filetypes=[("Test Plan Files", "*.tp"), ("All Files", "*.*")],
                title="Save New Test Plan As"
            )
        
            if not file_path:  # User cancelled
                return
            
            # Update current file reference
            current_tp_file = file_path

            tests_saved = 0
            with open(file_path, 'w') as tp_file:
                # Write config section with newlines
                tp_file.write("[config]\n")
                tp_file.write("[\\config]\n\n")  # Added newline
            
                # Write test section with newline
                tp_file.write("[test]\n")  # Added newline
            
                # If there are rows in the main table
                if row_count > 0:
                    for row_frame in rows:
                        try:
                            # Find all entries in the row
                            entries = [widget for widget in row_frame.winfo_children() 
                                     if isinstance(widget, ttk.Entry)]
                        
                            # Get test name from first entry
                            test_name = entries[0].get() if entries else ""
                        
                            # Find profile frame and get profile value
                            profile_frames = [widget for widget in row_frame.winfo_children() 
                                        if isinstance(widget, ttk.Frame)]
                            if profile_frames:
                                profile_entry = [widget for widget in profile_frames[0].winfo_children() 
                                               if isinstance(widget, ttk.Entry)]
                                profile_value = profile_entry[0].get() if profile_entry else ""
                            else:
                                profile_value = ""
                        
                            if test_name:  # Save test if name exists
                                # Only add ", use: profile" if profile exists
                                if profile_value:
                                    tp_file.write(f"{test_name}, use: {profile_value}\n")
                                else:
                                    tp_file.write(f"{test_name}\n")
                                tests_saved += 1
                            
                        except Exception as e:
                            print(f"Error processing row: {str(e)}")
                            continue
            
                tp_file.write("[\\test]")
        
            # Show success message
            if tests_saved == 0:
                messagebox.showinfo("Success", 
                    f"Created {os.path.basename(file_path)} file with 0 tests successfully.")
            else:
                messagebox.showinfo("Success", 
                    f"Created {os.path.basename(file_path)} file with {tests_saved} tests successfully.")

            # Update the view without destroying/recreating widgets
            content_frame.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
        
            update_status_3(f"Test plan created: {os.path.basename(file_path)}")
            update_file_status(source_page='3')
            return True
        
        except Exception as e:
            messagebox.showerror("Error", f"Error creating test plan: {str(e)}")
            update_status_3(f"Error creating test plan: {str(e)}", True)
            update_file_status(str(e), True, source_page='3')
            return False

    # Store the function reference in the global dictionary
page3_functions['create_test_plan'] = create_test_plan


    


def edit_page3_test_plan(tp_file):
        """Edit an existing test plan"""
        global current_tp_file
        try:
            # Initial checks
            if not sock:
                if not hasattr(edit_page3_test_plan, 'initialized'):
                    edit_page3_test_plan.initialized = True
                    return
                #messagebox.showerror("Error", "Please connect to server first")
                #return

            # tp_file = filedialog.askopenfilename(
            #     title="Select Test Plan File",
            #     filetypes=[("Test Plan Files", "*.tp"), ("All Files", "*.*")]
            # )
            
            if not tp_file:
                return

            current_tp_file = tp_file
            update_status_3("Loading test plan file...", False)
            thread_safe_display_message("Loading test plan, please wait...", "black")
            
            # Clear existing data
            preview_tree.delete(*preview_tree.get_children())
            for widget in content_frame.winfo_children():
                widget.destroy()

            test_lines = []
            added_count = 0
            syntax_errors = []

            # Read test section
            with open(tp_file, 'r') as f:
                in_test_section = False
                for line in f:
                    line = line.strip()
                    if not line:  # Skip empty lines
                        continue
                    if line == '[test]':
                        in_test_section = True
                        continue
                    if line == '[\\test]':
                        break
                    if in_test_section:
                        test_lines.append(line)

            # Process each line
            for line_num, line in enumerate(test_lines, 1):
                try:
                    line = line.strip()
                    print(f"Processing line {line_num}: {line}")  # Debug print

                    # Case 1: Multiple inputs with profile (handles both formats)
                    # _TR114i3_B_00({0,100,200}), use: profile.txt
                    # _TR114i3_B_00({200}),use:profile.kht.uuh
                    if '({' in line and ('), use:' in line or '),use:' in line or ');use:' in line or ' ;use:' in line):
                        if '), use:' in line:
                            parts = line.split('), use:')
                        elif '),use:' in line:
                            parts = line.split('),use:')
                        elif ');use:' in line:
                            parts = line.split(');use:')
                        elif ' ;use:' in line:
                            parts = line.split(' ;use:')
                        else:
                            parts = line.split('),use:')
                            
                        test_part = parts[0]
                        profile = parts[1].strip()
                        
                        test_name = test_part[:test_part.find('({')].strip()
                        input_str = test_part[test_part.find('({') + 2:test_part.find('})')]
                        inputs = [x.strip() for x in input_str.split(',') if x.strip()]
                        
                        for input_val in inputs:
                            add_row_with_details(
                                name=test_name,
                                test_input=input_val,
                                profile=profile
                            )
                            added_count += 1

                    # Case 2: Multiple inputs without profile
                    # _TR114i3_B_00({0,100,200})
                    # elif '({' in line and '})' in line:
                    #     test_name = line[:line.find('({')].strip()
                    #     input_str = line[line.find('({') + 2:line.find('})')]
                    #     inputs = [x.strip() for x in input_str.split(',') if x.strip()]
                        
                    #     for input_val in inputs:
                    #         add_row_with_details(
                    #             name=test_name,
                    #             test_input=input_val,
                    #             profile=""
                    #         )
                    #         added_count += 1
                    elif '({' in line and '})' in line:
                        # Extract the test name and input string
                        test_name = line[:line.find('({')].strip()
                        input_str = line[line.find('({') + 2:line.find('})')]
                        inputs = [x.strip() for x in input_str.split(',') if x.strip()]
        
                        
                        # Keep it as is, do not split further
                        add_row_with_details(
                            name=line.strip(),
                            test_input="",
                            profile=""
                        )
                    #         else:
                    #             #   Process as normal if not exactly two numbers
                    #             for input_val in inputs:
                    #                 add_row_with_details(
                    #                     name=test_name,
                    #                     test_input=input_val,
                    #                     profile=""
                    #                 )
                        

                    # Case 3: Special function calls
                    # setloop(100)
                    elif ('(' in line and ')' in line and 
                          not '({' in line):
                        add_row_with_details(
                            name=line.strip(),
                            test_input="",
                            profile=""
                        )
                        added_count += 1

                    # Case 4: Simple test with profile (handles both formats)
                    # _TR114i3_B_00,use:profile.txteer
                    # _TR114i3_B_00, use: profile.txt
                    elif ',use:' in line or ', use:' in line or ';use:' in line or ' ;use:' in line:
                        if ', use:' in line:
                            name_part, profile_part = line.split(', use:')
                        else:
                            name_part, profile_part = line.split(',use:')

                        if ';use:' in line:
                            name_part, profile_part = line.split(';use:')
                        elif ' ;use:' in line:
                            name_part, profile_part = line.split(' ;use:')

                        add_row_with_details(
                            name=name_part.strip(),
                            test_input="",
                            profile=profile_part.strip()
                        )
                        added_count += 1

                    # Case 5: Simple test name
                    # _TR114i3_B_00
                    else:
                        add_row_with_details(
                            name=line.strip(),
                            test_input="",
                            profile=""
                        )
                        added_count += 1

                    # Update progress
                    if added_count % 5 == 0:
                        root.update_idletasks()

                except Exception as e:
                    print(f"Error processing line {line_num}: {line}")
                    print(f"Error details: {str(e)}")
                    syntax_errors.append(f"Line {line_num}: {line} - {str(e)}")
                    continue

            # Show syntax errors if any
            if syntax_errors:
                messagebox.showwarning(
                    "Syntax Errors",
                    "Syntax errors in following lines:\n" + "\n".join(syntax_errors)
                )

            # Final updates
            filename = os.path.basename(tp_file)
            update_status_3(f"{filename} loaded successfully - Added {added_count} test plans")
            thread_safe_display_message(f"Successfully loaded: {filename}", "black")
            update_file_status(source_page='3')
            
            content_frame.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        except Exception as e:
            messagebox.showerror("Error", f"Error loading test plan: {str(e)}")
            thread_safe_display_message(f"Error loading test plan: {str(e)}", "red")
            update_status_3("Error loading test plan", True)
            update_file_status(str(e), True, source_page='3')
    
page3_functions['edit_page3_test_plan'] = edit_page3_test_plan

def get_all_test_data():
        """Helper function to get all test data from either source"""
        test_rows = []
        
        try:
            # Get ALL frames in their visual order, including duplicates
            # Important: Use pack_slaves() instead of winfo_children() to get exact visual order
            test_frames = [w for w in content_frame.pack_slaves() 
                      if isinstance(w, ttk.Frame) and not isinstance(w, ttk.Separator)]
        
   
        
            total_frames = len(test_frames)
            print(f"Found {total_frames} test frames in content frame")
            
            for i, frame in enumerate(test_frames):
                try:
                    # Get test name
                    name_entries = frame.grid_slaves(row=0, column=0)
                    if not name_entries or not isinstance(name_entries[0], ttk.Entry):
                        continue
                    test_name = name_entries[0].get().strip()
                    
                    # Skip empty rows but preserve order
                    if not test_name:
                        continue

                    # Get test input
                    test_input = ""
                    input_entries = frame.grid_slaves(row=0, column=1)
                    if input_entries and isinstance(input_entries[0], ttk.Entry):
                        test_input = input_entries[0].get().strip()

                    # Get profile if exists
                    profile = ""
                    profile_frames = [w for w in frame.winfo_children() 
                                    if isinstance(w, ttk.Frame)]
                    if profile_frames:
                        profile_entries = [w for w in profile_frames[0].winfo_children() 
                                         if isinstance(w, ttk.Entry)]
                        if profile_entries:
                            profile = profile_entries[0].get().strip()
                
                
                    # Check if test name already contains parameters
                    already_has_params = '(' in test_name 
                    #and test_name.rstrip().endswith(')')
                    
                    # Format the row based on available data
                    if already_has_params:
                        # If test name already has parameters, don't add more parentheses
                        if profile:
                            test_rows.append(f"{test_name} ;use: {profile}")
                        else:
                            test_rows.append(f"{test_name}")
                    else:
                        # Original logic for test names without parameters
                        if test_input:
                            if profile:
                                test_rows.append(f"{test_name}({test_input}) ;use: {profile}")
                            else:
                                test_rows.append(f"{test_name}({test_input})")
                        else:
                            if profile:
                                test_rows.append(f"{test_name}() ;use: {profile}")
                            else:
                                test_rows.append(f"{test_name}()")
                
                
                    # Progress update
                    if (i + 1) % 10 == 0:
                        progress_msg = f"Processing content... ({i + 1}/{total_frames})"
                        print(progress_msg)
                        root.after(0, lambda msg=progress_msg: update_status_3(msg))
                    
                except Exception as e:
                    print(f"Error processing frame {i}: {str(e)}")
                    continue

            # Final status update
            total_rows = len(test_rows)


            root.after(0, lambda: update_status_3(f"Collected {total_rows} test cases"))
            return test_rows

        except Exception as e:
            error_msg = f"Error collecting test data: {str(e)}"
            print(f"Error in get_all_test_data: {error_msg}")
            print(f"Traceback: {traceback.format_exc()}")
            root.after(0, lambda: messagebox.showerror("Error", error_msg))
            root.after(0, lambda: update_status_3(error_msg, True))
            return []

    # Store function reference
page3_functions['get_all_test_data'] = get_all_test_data
def save_page3_test_plan():
        """Save current test plan"""
        def save_in_thread():
            try:
                # Initial checks
                if not sock:
                    if not hasattr( save_page3_test_plan, 'initialized'):
                        save_page3_test_plan.initialized = True
                        return False
                #     root.after(0, lambda: messagebox.showerror("Error", "Please connect to server first"))
                #     return False
                
                if not current_tp_file:
                    return save_page3_test_plan()

                # Get all test data using the helper function
                test_rows = get_all_test_data()
                if not test_rows:
                    return False

                root.after(0, lambda: update_status_3("Writing test plan file..."))

                # Read existing file to preserve config section
                if os.path.exists(current_tp_file):
                    with open(current_tp_file, 'r') as f:
                        lines = f.readlines()

                    # Find test section
                    test_start = -1
                    test_end = -1
                    for i, line in enumerate(lines):
                        if line.strip() == '[test]':
                            test_start = i
                        elif line.strip() == '[\\test]':
                            test_end = i
                            break

                    # Replace test section content
                    new_lines = lines[:test_start + 1]  # Keep everything before [test]
                    
                    # Write test lines
                    total_rows = len(test_rows)
                    for i, row in enumerate(test_rows, 1):
                        new_lines.append(f"{row}\n")
                        if i % 100 == 0:
                            progress_msg = f"Writing test plan... ({i}/{total_rows})"
                            root.after(0, lambda msg=progress_msg: update_status_3(msg))

                    new_lines.append("[\\test]")
                    if test_end + 1 < len(lines):
                        new_lines.extend(lines[test_end + 1:])

                    # Write back to file
                    with open(current_tp_file, 'w') as f:
                        f.writelines(new_lines)
                else:
                    # Create new file
                    with open(current_tp_file, 'w') as f:
                        f.write("[config]\n")
                        f.write("[\\config]\n")
                        f.write("\n[test]\n")
                        for row in test_rows:
                            f.write(f"{row}\n")
                        f.write("[\\test]")

                success_msg = f"Test plan saved successfully in {os.path.basename(current_tp_file)}"
                root.after(0, lambda: update_status_3(success_msg))
                root.after(0, lambda: update_file_status(source_page='3'))
                return True

            except Exception as e:
                error_msg = f"Error saving test plan: {str(e)}"
                root.after(0, lambda: messagebox.showerror("Error", error_msg))
                root.after(0, lambda: update_status_3(error_msg, True))
                root.after(0, lambda msg=str(e): update_file_status(msg, True, source_page='3'))
                return False

        # Start save operation in a separate thread
        threading.Thread(target=save_in_thread, daemon=True).start()
page3_functions['save_page3_test_plan'] = save_page3_test_plan

def save_as_page3_test_plan(source_page='3'):
        """Save test plan with a new name"""
        # Create a queue for thread communication
        file_path_queue = queue.Queue()
        
        def get_file_path_from_main_thread():
            """Get file path using dialog in main thread"""
            try:
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".tp",
                    filetypes=[("Test Plan Files", "*.tp")],
                    title="Save Test Plan As"
                )
                file_path_queue.put(file_path)
            except Exception as e:
                print(f"Error in file dialog: {e}")
                file_path_queue.put(None)

        def save_as_in_thread_3():
            global current_tp_file
            try:
                # if not check_socket(silent=True):
                #     return False

                # Get all test data first
                test_rows = get_all_test_data()
                if not test_rows:
                    return False

                # Store original file info
                original_file = current_tp_file
                original_filename = os.path.basename(original_file) if original_file else None

                # Request file path from main thread
                root.after(0, get_file_path_from_main_thread)
                
                # Wait for file path from dialog
                new_file_path = file_path_queue.get()
                if not new_file_path:  # User cancelled
                    return False

                # Update current file reference
                current_tp_file = new_file_path
                root.after(0, lambda: update_status_3(f"Saving to {os.path.basename(new_file_path)}..."))

                # Copy config from original file if it exists
                config_section = ""
                if original_file and os.path.exists(original_file):
                    try:
                        with open(original_file, 'r') as f:
                            content = f.read()
                            config_start = content.find("[config]")
                            config_end = content.find("[\\config]")
                            if config_start != -1 and config_end != -1:
                                config_section = content[config_start:config_end + 9] + '\n'
                                print(f"Copied config section from {original_filename}")
                    except Exception as e:
                        print(f"Error copying config: {str(e)}")
                        config_section = "[config]\n[\\config]\n"

                # Write new file with progress updates
                try:
                    with open(new_file_path, 'w') as f:
                        # Write config section
                        f.write(config_section or "[config]\n[\\config]\n")
                        
                        # Write test section
                        f.write("\n[test]\n")
                        total_rows = len(test_rows)
                        
                        # Write in chunks for progress updates
                        chunk_size = max(1, min(100, total_rows // 10))  # Update every 100 rows or less
                        for i in range(0, total_rows, chunk_size):
                            chunk = test_rows[i:i + chunk_size]
                            f.write('\n'.join(chunk) + '\n')
                            
                            # Update progress
                            progress = min(i + chunk_size, total_rows)
                            progress_msg = f"Writing test plan... ({progress}/{total_rows})"
                            root.after(0, lambda msg=progress_msg: update_status_3(msg))

                        f.write("[\\test]")

                    # Show success message
                    success_msg = (f"Test Plan saved to {os.path.basename(new_file_path)}\n"
                                 f"Test data and config copied from {original_filename}" if original_filename
                                 else f"Test Plan saved to {os.path.basename(new_file_path)}")
                    
                    root.after(0, lambda msg=success_msg: messagebox.showinfo("Success", msg))
                    root.after(0, lambda: update_status_3(f"Saved: {os.path.basename(new_file_path)}"))
                    root.after(0, lambda: update_file_status(source_page='3'))
                    return True

                except Exception as e:
                    error_msg = f"Error writing file: {str(e)}"
                    root.after(0, lambda msg=error_msg: messagebox.showerror("Error", msg))
                    root.after(0, lambda msg=error_msg: update_status_3(msg, True))
                    return False

            except Exception as e:
                error_msg = f"Failed to save test plan: {str(e)}"
                root.after(0, lambda msg=error_msg: messagebox.showerror("Error", msg))
                root.after(0, lambda msg=error_msg: update_status_3(msg, True))
                root.after(0, lambda msg=str(e): update_file_status(msg, True, source_page='3'))
                return False

        # Start save operation in a separate thread
        threading.Thread(target=save_as_in_thread_3, daemon=True).start()

    # Store function reference
page3_functions['save_as_page3_test_plan'] = lambda: save_as_page3_test_plan('3')



def add_empty_row():
        """Add an empty row to the table"""
        try:
            add_row_with_details(
                name="",
                test_input="",
                profile=""
            )
            canvas.configure(scrollregion=canvas.bbox("all"))
            update_status_3("Added empty row")
            
            # Scroll to the bottom to show the new row
            canvas.yview_moveto(1.0)
            
        except Exception as e:
            update_status_3(f"Error adding empty row: {str(e)}", True)


def load_selected_parameters():
        """Load selected test inputs from preview to main table"""
        selections = preview_tree.selection()
        if not selections:
            update_status_3("Please select test inputs to load", True)
            return

        try:
            added_count = 0
            skipped_count = 0

            for item in selections:
                values = preview_tree.item(item)['values']
                if not values or len(values) < 2:
                    continue
        
                test_name = str(values[0])  # Convert to string explicitly
                test_input = str(values[1])  # Convert to string explicitly
            
                # Skip rows with "No Parameter"
                if test_input == "No Parameter":
                    skipped_count += 1
                    update_status_3(f"Skipping {test_name} - No valid parameter")
                    continue
                
                try:
                    # Combine test name and input in format: testname(testinput)
                    combined_name = f"{test_name}({test_input})"
                    
                    cmd1 = f'spt = "{test_name}"'
                    socket_query(sock, cmd1)
                    cmd2 = f'(err, info) = test_plan::Get_TPlist(spt) <<< info'
                    response = socket_query(sock, cmd2)
        
                    if response:
                        info_lines = response.strip().split('\n')
                        profile = info_lines[2] if len(info_lines) > 2 else ""
            
                        add_row_with_details(
                            name=test_name,
                            test_input=test_input,
                            profile=profile
                        )
                        added_count += 1
                    else:
                        # Add row even if no profile is found
                        add_row_with_details(
                            name=test_name,
                            test_input=test_input,
                            profile=""
                        )
                        added_count += 1
                    
                except Exception as e:
                    update_status_3(f"Error getting details for {test_name}: {str(e)}", True)
                    # Still add the row even if profile fetch fails
                    add_row_with_details(
                        name=test_name,
                        test_input=test_input,
                        profile=""
                    )
                added_count += 1

            canvas.configure(scrollregion=canvas.bbox("all"))
            update_status_3(f"Added {added_count} test inputs, skipped {skipped_count} invalid parameters")

        except Exception as e:
            update_status_3(f"Error loading test inputs: {str(e)}", True)
                   
    # # Add Load Parameters button
    # ttk.Button(preview_frame, 
    #            text="Load Selected Parameters", 
    #            command=load_selected_parameters).grid(row=2, column=0, columnspan=2, pady=10)        


def update_test_plans():
        if not sock:
            messagebox.showerror("Error", "Not connected to server. Please connect first.")
            return
            
        try:
            # Show loading indicator
            refresh_button.configure(state='disabled')
            
            cmd = "(e, r) = get_tplist() <<< e <<< r"
            print(f"Command - {cmd}")
            response = socket_query(sock, cmd)
            if response:
                print(f"Response:\n{response}")
                
                cmd = "For(idx, 1:!r) <<< r[idx][1]"
                print(f"Command - {cmd}")
                response = socket_query(sock, cmd)
                if response:
                    # Clear existing items
                    all_items.clear()
                    listbox1.delete(0, tk.END)
                    
                    # Update with new items
                    for line in response.splitlines():
                        item = line.strip()
                        if item and item != '>':  # Only add non-empty items and skip '>' lines
                            all_items.append(item)
                            listbox1.insert(tk.END, item)
                    
                    messagebox.showinfo("Success", f"Loaded {len(all_items)} test plans")
                else:
                    messagebox.showwarning("Warning", "No test plans received from server")
            else:
                messagebox.showwarning("Warning", "No response received from server")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error updating test plans: {str(e)}")
        finally:
            refresh_button.configure(state='normal')
    
    




def add_row_with_details(name, test_input="", profile=""):
        """Add a row to the table with selectable and copyable values"""
        try:
            row_frame = ttk.Frame(content_frame, width=sum(width for _, width in headers))
            row_frame.pack(fill="x", padx=2, pady=1)
            row_frame.pack_propagate(False)

            # Store searchable text in the frame for filtering
            row_frame.searchable_text = name.lower()

            # Test Name Entry
            name_entry = ttk.Entry(row_frame, width=60, justify='left')
            name_entry.grid(row=0, column=0, padx=(2,0), sticky='w')
            
            # Only clean name if it's not empty
            if name:
                #clean_name = name.split('({')[0].strip() if '({' in name else name
                #clean_name = clean_name.split(')}')[0].strip() if ')}' in clean_name else clean_name
                clean_name = name.strip()
                name_entry.insert(0, clean_name)
            
            # Always set state to normal for empty rows
            name_entry.configure(state='normal')

            def update_searchable_text(event=None):
                """Update searchable text when name changes"""
                new_name = name_entry.get().strip()
                row_frame.searchable_text = new_name.lower()
                filter_table_rows()

            # Bind to both keyup and focus out
            name_entry.bind('<KeyRelease>', update_searchable_text)
            name_entry.bind('<FocusOut>', update_searchable_text)

            # Test Input Entry
            input_entry = ttk.Entry(row_frame, width=15)
            input_entry.grid(row=0, column=1, padx=(0,2), sticky='w')
            
            # Only clean and set readonly for non-empty inputs
            if test_input:
                clean_input = test_input.strip('{}() ')
                input_entry.insert(0, clean_input)
                input_entry.configure(state='readonly')
            else:
                input_entry.configure(state='normal')  # Keep editable for empty rows

            # Profile Frame
            profile_frame = ttk.Frame(row_frame)
            profile_frame.grid(row=0, column=2)

            # Profile Entry
            profile_entry = ttk.Entry(profile_frame, width=40)
            profile_entry.pack(side="left")
            if profile:
                profile_entry.delete(0, tk.END)
                profile_entry.full_path = profile
                display_name = profile.replace('.txt', '')
                profile_entry.insert(0, display_name)

            def browse_profile():
                file_path = filedialog.askopenfilename(
                    title="Select Profile",
                    filetypes=[("All Files", "*.*")]
                )
                if file_path:
                    profile_entry.delete(0, tk.END)
                    profile_entry.full_path = file_path
                    display_name = os.path.basename(file_path).replace('.txt', '')
                    profile_entry.insert(0, display_name)

            # Browse button
            ttk.Button(profile_frame, 
                    text="Browse", 
                    width=10, 
                    command=browse_profile).pack(side="left")

            def on_double_click(event):
                widget = event.widget
                if widget.cget('state') == 'readonly':
                    widget.configure(state='normal')
                    widget.select_range(0, tk.END)
                    widget.configure(state='readonly')
                    widget.focus_set()

            # Only bind double-click for non-empty rows
            if name or test_input:
                name_entry.bind('<Double-1>', on_double_click)
                input_entry.bind('<Double-1>', on_double_click)

            def delete_row():
                try:
                    if hasattr(row_frame, 'separator'):
                        row_frame.separator.destroy()
                    row_frame.destroy()
                    content_frame.update_idletasks()
                    canvas.configure(scrollregion=canvas.bbox("all"))
                    filter_table_rows()
                except Exception as e:
                    print(f"Error deleting row: {str(e)}")


            def duplicate_row():
                try:
                    current_name = name_entry.get()
                    current_input = input_entry.get()
                    current_profile = getattr(profile_entry, 'full_path', '')
                    
                    # Get all row frames
                    all_rows = [w for w in content_frame.winfo_children() 
                              if isinstance(w, ttk.Frame) and hasattr(w, 'searchable_text')]
                    
                    # Find the index of current row
                    current_index = all_rows.index(row_frame)
                    
                    # Store all rows after current index
                    following_rows = all_rows[current_index + 1:]
                
                    def create_duplicate_row(parent_row, name, input_val, profile_val):
                        new_row = ttk.Frame(content_frame, width=sum(width for _, width in headers))
                        new_row.searchable_text = name.lower()
                        
                        # Add all entries and buttons
                        name_entry_new = ttk.Entry(new_row, width=60, justify='left')
                        name_entry_new.grid(row=0, column=0, padx=(2,0), sticky='w')
                        name_entry_new.insert(0, name)
                        name_entry_new.configure(state='normal')
                        
                        input_entry_new = ttk.Entry(new_row, width=15)
                        input_entry_new.grid(row=0, column=1, padx=(0,2), sticky='w')
                        input_entry_new.insert(0, input_val)
                        input_entry_new.configure(state='readonly')
                        
                        profile_frame_new = ttk.Frame(new_row)
                        profile_frame_new.grid(row=0, column=2)
                        profile_entry_new = ttk.Entry(profile_frame_new, width=40)
                        profile_entry_new.pack(side="left")
                        if profile_val:
                            profile_entry_new.delete(0, tk.END)
                            profile_entry_new.full_path = profile_val
                            display_name = os.path.basename(profile_val).replace('.txt', '')
                            profile_entry_new.insert(0, display_name)
                    

                    
                        # Create row-specific delete function
                        def delete_this_row():
                            try:
                                if hasattr(new_row, 'separator'):
                                    new_row.separator.destroy()
                                new_row.destroy()
                                content_frame.update_idletasks()
                                canvas.configure(scrollregion=canvas.bbox("all"))
                                filter_table_rows()
                            except Exception as e:
                                print(f"Error deleting row: {str(e)}")
                        def browse_profile_new():
                            file_path = filedialog.askopenfilename(
                                title="Select Profile",
                                filetypes=[("All Files", "*.*")]
                            )
                            if file_path:
                                profile_entry_new.delete(0, tk.END)
                                profile_entry_new.full_path = file_path
                                display_name = os.path.basename(file_path).replace('.txt', '')
                                profile_entry_new.insert(0, display_name)

                        ttk.Button(profile_frame_new, text="Browse", width=10,
                        command=browse_profile_new).pack(side="left")  # Use the new function

                        def update_searchable_text_new(event=None):
                            new_name = name_entry_new.get().strip()
                            new_row.searchable_text = new_name.lower()
                            filter_table_rows()

                        name_entry_new.bind('<KeyRelease>', update_searchable_text_new)
                        name_entry_new.bind('<FocusOut>', update_searchable_text_new)

                        def on_double_click_new(event):
                            widget = event.widget
                            if widget.cget('state') == 'readonly':
                                widget.configure(state='normal')
                                widget.select_range(0, tk.END)
                                widget.configure(state='readonly')
                                widget.focus_set()

                        name_entry_new.bind('<Double-1>', on_double_click_new)
                        input_entry_new.bind('<Double-1>', on_double_click_new)

                        # Create row-specific duplicate function
                        def duplicate_this_row():
                            try:
                                create_duplicate_row(new_row, 
                                              name_entry_new.get(),
                                              input_entry_new.get(),
                                              getattr(profile_entry_new, 'full_path', ''))
                            except Exception as e:
                                print(f"Error duplicating row: {str(e)}")
                    
                        ttk.Button(new_row, text="Delete", width=10,
                             command=delete_this_row).grid(row=0, column=4)
                        ttk.Button(new_row, text="Copy", width=10,
                             command=duplicate_this_row).grid(row=0, column=5)
                    
                        # Add separator
                        sep_frame = ttk.Frame(content_frame)
                        new_row.separator = sep_frame
                    
                        # Pack after parent row
                        new_row.pack(fill="x", padx=2, pady=1, after=parent_row)
                        sep_frame.pack(fill="x", after=new_row)
                        ttk.Separator(sep_frame, orient="horizontal").pack(fill="x")
                    
                        return new_row
                
                # Create the duplicate row
                    new_row = create_duplicate_row(row_frame, current_name, current_input, current_profile)
                
                    # Repack following rows
                    for row in following_rows:
                        row.pack(fill="x", padx=2, pady=1)
                        if hasattr(row, 'separator'):
                            row.separator.pack(fill="x")
                
                    content_frame.update_idletasks()
                    canvas.configure(scrollregion=canvas.bbox("all"))
                
                except Exception as e:
                    print(f"Error duplicating row: {str(e)}")
                    traceback.print_exc()
                    

            # Action buttons (moved outside and aligned with duplicate_row)
            ttk.Button(row_frame, text="Delete", width=10, 
                      command=delete_row).grid(row=0, column=4)
            ttk.Button(row_frame, text="Copy", width=10,
                      command=duplicate_row).grid(row=0, column=5)

            # Row visibility toggle (moved outside and aligned with duplicate_row)
            def toggle_visibility(show):
                if show:
                    row_frame.pack(fill="x", padx=1, pady=1)
                    if hasattr(row_frame, 'separator'):
                        row_frame.separator.pack(fill="x")
                else:
                    row_frame.pack_forget()
                    if hasattr(row_frame, 'separator'):
                        row_frame.separator.pack_forget()
            
            row_frame.toggle_visibility = toggle_visibility

            # Add separator (moved outside and aligned with duplicate_row)
            sep_frame = ttk.Frame(content_frame, width=sum(width for _, width in headers))
            sep_frame.pack(fill="x")
            ttk.Separator(sep_frame, orient="horizontal").pack(fill="x")
            sep_frame.pack_propagate(False)
            row_frame.separator = sep_frame

            # Update UI periodically (moved outside and aligned with duplicate_row)
            if hasattr(add_row_with_details, 'row_count'):
                add_row_with_details.row_count += 1
            else:
                add_row_with_details.row_count = 1

            if add_row_with_details.row_count % 50 == 0:
                content_frame.update_idletasks()

            return row_frame

        except Exception as e:
            print(f"Error adding row: {str(e)}")
            return None
  
    

def check_duplicate(test_name):
        """
        Check for duplicates in main table
        Args:
            test_name: Test name
        Returns:
            bool: True if duplicate found
        """
        for widget in content_frame.winfo_children():
            if isinstance(widget, ttk.Frame):
                name_entry = widget.grid_slaves(row=0, column=0)[0]
                if name_entry.get() == test_name:
                    return True
        return False

def load_test_data_for_page3(file_path):
        threading.Thread(target=load_data, args=(file_path,), daemon=True).start()




def load_data(file_path):
        """Load test data directly from file path"""
        try:
            # Initial checks
            if not sock and not hasattr(load_test_data_for_page3, 'initialized'):
                load_test_data_for_page3.initialized = True
                return

            # Clear existing data
            preview_tree.delete(*preview_tree.get_children())
            root.after(0, lambda: [widget.destroy() for widget in content_frame.winfo_children()])
            root.after(0, lambda: update_status_3("Loading test plan file...", False))
            root.after(0, lambda: thread_safe_display_message("Loading test plan, please wait...", "black"))

            test_lines = []
            added_count = 0
            
            # Read test section
            with open(file_path, 'r') as f:
                in_test_section = False
                for line in f:
                    line = line.strip()
                    if not line:  # Skip empty lines
                        continue
                    if line == '[test]':
                        in_test_section = True
                        continue
                    if line == '[\\test]':
                        break
                    if in_test_section:
                        test_lines.append(line)

            # Process each line
            for line_num, line in enumerate(test_lines, 1):
                try:
                    line = line.strip()
                    print(f"Processing line {line_num}: {line}")  # Debug print

                    # Case 1: Multiple inputs with profile (handles both formats)
                    # _TR114i3_B_00({0,100,200}), use: profile.txt
                    # _TR114i3_B_00({200}),use:profile.kht.uuh
                    if '({' in line and ('), use:' in line or '),use:' in line or ');use:' in line or ' ;use:' in line):
                        if '), use:' in line:
                            parts = line.split('), use:')
                        elif '),use:' in line:
                            parts = line.split('),use:')
                        elif ');use:' in line:
                            parts = line.split(');use:')
                        elif ' ;use:' in line:
                            parts = line.split(' ;use:')
                        else:
                            parts = line.split('),use:')
                            
                        test_part = parts[0]
                        profile = parts[1].strip()
                        test_name = test_part[:test_part.find('(')].strip()
                        input_str = test_part[test_part.find('(') + 1:].strip()
                        inputs = [x.strip() for x in input_str.split(',') if x.strip()]
                        
                        for input_val in inputs:
                            add_row_with_details(
                                name=test_name,
                                test_input=input_val,
                                profile=profile
                            )
                            added_count += 1

                    # Case 2: Multiple inputs without profile
                    # _TR114i3_B_00({0,100,200})
                    elif '({' in line and '})' in line:
                        test_name = line[:line.find('({')].strip()
                        input_str = line[line.find('({') + 2:line.find('})')]
                        inputs = [x.strip() for x in input_str.split(',') if x.strip()]
                        
                        for input_val in inputs:
                            add_row_with_details(
                                name=test_name,
                                test_input=input_val,
                                profile=""
                            )
                            added_count += 1

                    # Case 3: Special function calls
                    # setloop(100)
                    elif ('(' in line and ')' in line and 
                          not '({' in line):
                        add_row_with_details(
                            name=line.strip(),
                            test_input="",
                            profile=""
                        )
                        added_count += 1

                    # Case 4: Simple test with profile (handles both formats)
                    # _TR114i3_B_00,use:profile.txteer
                    # _TR114i3_B_00, use: profile.txt
                    elif ',use:' in line or ', use:' in line or ';use:' in line or ' ;use:' in line:
                        if ', use:' in line:
                            name_part, profile_part = line.split(', use:')
                        else:
                            name_part, profile_part = line.split(',use:')

                        if ';use:' in line:
                            name_part, profile_part = line.split(';use:')
                        elif ' ;use:' in line:
                            name_part, profile_part = line.split(' ;use:')

                        add_row_with_details(
                            name=name_part.strip(),
                            test_input="",
                            profile=profile_part.strip()
                        )
                        added_count += 1

                    # Case 5: Simple test name
                    # _TR114i3_B_00
                    else:
                        add_row_with_details(
                            name=line.strip(),
                            test_input="",
                            profile=""
                        )
                        added_count += 1

                    # Update progress every few rows
                    if added_count % 10 == 0:
                        root.after(0, lambda c=added_count: update_status_3(
                            f"Loading... ({c} tests added)"
                        ))
                        root.update_idletasks()

                    # Update progress every few rows
                    if added_count % 10 == 0:
                        root.after(0, lambda c=added_count: update_status_3(
                            f"Loading... ({c} tests added)"
                        ))
                        root.update_idletasks()

                except Exception as e:
                    print(f"Error processing line {line_num}: {line}")
                    print(f"Error details: {str(e)}")
                    continue

            # Final update
            filename = os.path.basename(file_path)
            root.after(0, lambda: update_status_3(
                f"{filename} loaded successfully - Added {added_count} test plans"
            ))
            root.after(0, lambda: canvas.configure(scrollregion=canvas.bbox("all")))

        except Exception as e:
            root.after(0, lambda: messagebox.showerror("Error", f"Error loading test plan: {str(e)}"))
            root.after(0, lambda: thread_safe_display_message(f"Error loading test plan: {str(e)}", "red"))
            root.after(0, lambda: update_status_3("Error loading test plan", True))



    # Store the function reference in the global dictionary
page3_functions['load_test_data'] = load_test_data_for_page3

def create_page3():
    
    global page3_frame, df, notebook, content_frame, canvas, current_tp_file
    global test_plan_entry, status_label_3, status_var_3
    global listbox1, listbox2, preview_tree, refresh_button
    global all_items, page3_functions, table_search_var, listbox_search_var
    global headers, sock, root
    page3_frame = ttk.Frame(root, padding="5 5 5 5")
   
    content_frame = None
    canvas = None
    current_tp_file = None
    test_plan_entry = None
    status_label_3 = None
    status_var_3 = None
    listbox1 = None
    listbox2 = None
    preview_tree = None
    page3_functions = {}
    notebook.add(page3_frame, text='Test Plan Selection')
    # Style for headers and cells
    header_style = {"font": ("Helvetica", 10, "bold"), "anchor": "center"}
    cell_style = {"font": ("Helvetica", 10), "anchor": "center"}


    #global status_label_3, status_var_3
    # Create table header frame with status
    table_header_frame = ttk.Frame(page3_frame)
    table_header_frame.grid(row=11, column=0, columnspan=26, padx=5, pady=5, sticky="w")

    # Initialize Page 3 specific status
    status_var_3 = tk.StringVar()

    #status_label_3.pack(side="right")
    test_plan_tree = ttk.Treeview(page3_frame)  # Use your actual parent frame

    # Lists frame configuration
    lists_frame = ttk.Frame(page3_frame)
    lists_frame.grid(row=1, column=0, columnspan=26, rowspan=10, padx=10, pady=5, sticky="nsew")
    
    # Configure column weights
    lists_frame.grid_columnconfigure(5, weight=1)  # For preview table

    list_width = 60
    list_height = 8
    
    # Store all items for filtering
    all_items = []

    



    # Test Plan Details label
    ttk.Label(table_header_frame, 
         text="Test Plan Details:", 
         font=("Helvetica", 10, "bold")).pack(side="left", padx=(0,10))



    search_entry = ttk.Entry(table_header_frame, 
                        textvariable=table_search_var,
                        width=20)
    search_entry.pack(side="left", padx=(0,5))

    # Clear search button
    ttk.Button(table_header_frame, 
              text="Clear Search", 
          command=lambda: table_search_var.set("")).pack(side="left", padx=(0,5))

    # Clear table button
    ttk.Button(table_header_frame, 
              text="Clear Table", 
              command=clear_main_table).pack(side="left", padx=(0,5))


  
    # Add new "Add Line" button
    ttk.Button(table_header_frame,
              text="Add Empty Line",
              command=lambda: add_empty_row()).pack(side="left", padx=(0,5))



    # Create frame for list boxes
    listbox_frame = ttk.Frame(lists_frame)
    listbox_frame.grid(row=0, column=3, columnspan=5, padx=10, pady=5, sticky="nsew")

    listbox2 = tk.Listbox(listbox_frame, height=list_height, width=list_width, 
                      font=("Helvetica", 10, "bold"), exportselection=False)
    listbox2.grid(row=1, column=3, padx=10, pady=5, sticky="nsew")
    scrollbar2 = ttk.Scrollbar(listbox_frame, orient="vertical", command=listbox2.yview)
    scrollbar2.grid(row=1, column=4, sticky="ns")
    listbox2.configure(yscrollcommand=scrollbar2.set)

    # Test Inputs Preview
    preview_header_frame = ttk.Frame(listbox_frame)
    preview_header_frame.grid(row=0, column=6, columnspan=2, padx=5, pady=5, sticky="w")

    ttk.Label(preview_header_frame, 
         text="Test Inputs Preview:", 
         font=("Helvetica", 10, "bold")).pack(side="left", padx=(0,10))
         
    clear_preview_button = ttk.Button(preview_header_frame, 
                                 text="Clear Table", 
                                 command=lambda: preview_tree.delete(*preview_tree.get_children()))
    clear_preview_button.pack(side="left", padx=5)

    # Header frame for listbox2 (Selected Test Plans)
    selected_header_frame = ttk.Frame(listbox_frame)
    selected_header_frame.grid(row=0, column=3, columnspan=2, padx=5, pady=5, sticky="w")

    ttk.Label(selected_header_frame, 
             text="Selected Test Plans:", 
             font=("Helvetica", 10, "bold")).pack(side="left", padx=(0,10))
         
    clear_listbox2_button = ttk.Button(selected_header_frame, 
                                  text="Clear Table", 
                                  command=lambda: listbox2.delete(0, tk.END))
    clear_listbox2_button.pack(side="left", padx=5)

    update_status_3("Ready to load test plans")  

    # Table frame
    table_frame = ttk.Frame(page3_frame)
    table_frame.grid(row=12, column=0, columnspan=26,  padx=5, pady=5, sticky="w")

    # Create header frame
    # Create header frame with fixed width
    header_frame = ttk.Frame(table_frame)
    header_frame.pack(fill="none", padx=2)

    # Table headers with specific widths
    headers = [
        ("Test Name", 35),
        ("Test Input", 45),
        ("Profile", 30),
        ("Action", 33)
    ]

    for i, (text, width) in enumerate(headers):
        ttk.Label(header_frame, 
                 text=text, 
                 width=width, 
                 font=("Helvetica", 10, "bold"),
                 anchor="center").grid(row=0, column=i, padx=2)

    total_width = sum(width for _, width in headers)

    # Create scrollable content area
    canvas = tk.Canvas(table_frame, width=total_width, height=400)
    scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=canvas.yview)
    content_frame = ttk.Frame(canvas)

    # Configure scrolling
    canvas.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    canvas_frame = canvas.create_window((0, 0), window=content_frame, anchor="nw")  



    # Configure canvas scrolling
    def configure_scroll_region(event):
        """Update scroll region when content changes"""
        canvas.configure(scrollregion=canvas.bbox("all"))

    def configure_canvas(event):
        """Update canvas width when window resizes"""
        canvas = event.widget
        canvas.itemconfig(canvas_frame, width=event.width)

    # Bind configuration events
    content_frame.bind("<Configure>", configure_scroll_region)
    canvas.bind("<Configure>", configure_canvas)

    # Update scroll region initially
    content_frame.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))


    # Header frame for search
    header_frame = ttk.Frame(listbox_frame)
    header_frame.grid(row=0, column=0, columnspan=1, padx=5, pady=5, sticky="w")

    # Available Test Plans Label and Search
    ttk.Label(header_frame, text="Test Plans:", font=("Helvetica", 10, "bold")).pack(side="left", padx=(0,10))
    table_framesearch_var = tk.StringVar()
    listbox_search_var = tk.StringVar()

    listbox_search_var.trace('w', filter_items)
    listbox_search_entry = ttk.Entry(header_frame, textvariable=listbox_search_var, width=20)
    listbox_search_entry.pack(side="left", padx=5)
    clear_button = ttk.Button(header_frame, text="Clear", command=clear_search)
    clear_button.pack(side="left", padx=5)

    # Listbox1 - Available Test Plans
    listbox1 = tk.Listbox(listbox_frame, height=list_height, width=list_width, 
                      font=("Helvetica", 10, "bold"), selectmode=tk.MULTIPLE, 
                      exportselection=False)
    listbox1.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
    scrollbar1 = ttk.Scrollbar(listbox_frame, orient="vertical", command=listbox1.yview)
    scrollbar1.grid(row=1, column=1, sticky="ns")
    listbox1.configure(yscrollcommand=scrollbar1.set)

    # Transfer Button
    ttk.Button(listbox_frame, text=">>", 
           command=lambda: transfer_items(listbox1, listbox2), 
           width=5).grid(row=1, column=2, padx=5, pady=5)


    # Preview Tree
    preview_tree = ttk.Treeview(listbox_frame, 
                           columns=("test", "test_input"), 
                           show="headings", 
                           selectmode="extended", 
                           height=15)
    preview_tree.heading("test", text="Test Name")
    preview_tree.heading("test_input", text="Test Input")
    preview_tree.column("test", width=350)
    preview_tree.column("test_input", width=100)
    preview_tree.grid(row=1, column=6, sticky="nsew")

    # Preview Scrollbar
    preview_scroll = ttk.Scrollbar(listbox_frame, orient="vertical", command=preview_tree.yview)
    preview_scroll.grid(row=1, column=7, sticky="ns")
    preview_tree.configure(yscrollcommand=preview_scroll.set)

    # Load Test Inputs Button
    ttk.Button(listbox_frame, text=">>", 
           command=load_test_inputs, 
           width=5).grid(row=1, column=5, padx=5, pady=5)

    # Action Buttons
    load_button = ttk.Button(listbox_frame, 
                            text="Load Selected Plans to Table", 
                            command=load_selected_plans)
    load_button.grid(row=2, column=3, columnspan=2, pady=5)

    load_inputs_button = ttk.Button(listbox_frame, 
                               text="Load Selected Test Inputs", 
                               command=load_selected_parameters)
    load_inputs_button.grid(row=2, column=6, columnspan=2, pady=5)

    refresh_button = ttk.Button(listbox_frame, 
                           text="Load Test Plans", 
                           command=update_test_plans)
    refresh_button.grid(row=2, column=0, columnspan=2, pady=5)


    # When you reach the header_frame_buttons creation part, modify it like this:
    header_frame_buttons = ttk.Frame(page3_frame)
    header_frame_buttons.grid(row=0, column=0, columnspan=26, padx=5, pady=5, sticky="ew")

    # Add title and buttons to header
    ttk.Label(header_frame_buttons, 
        text="Test Plan Selection", 
        font=("Helvetica", 10, "bold")
    ).pack(side="left", padx=10)

    # "New" button - use lambda for passing parameters
    ttk.Button(header_frame_buttons, 
        text="New",
        cursor="hand2",
        command=lambda: root.after(100, lambda: reset_to_defaults())
    ).pack(side="left", padx=5)

    # "Open" button
    ttk.Button(header_frame_buttons, 
        text="Open",
        cursor="hand2",
        command=lambda: root.after(100, coordinated_edit_test_plan)
    ).pack(side="left", padx=5)

    # "Save" button
    ttk.Button(header_frame_buttons, 
        text="Save",
        cursor="hand2",
        command=lambda: root.after(100, coordinated_save_test_plan)
    ).pack(side="left", padx=5)

    # "Save As" button
    ttk.Button(header_frame_buttons, 
        text="Save As",
        cursor="hand2",
        command=lambda: root.after(100, coordinated_save_as_test_plan)
    ).pack(side="left", padx=5)

    # "Run" button
    ttk.Button(header_frame_buttons, 
        text="Run",
        cursor="hand2",
        command=lambda: root.after(100, open_execute_window)
    ).pack(side="left", padx=5)

    # Add status label to the right of the buttons
    status_label_3 = ttk.Label(
        header_frame_buttons,
        textvariable=status_var_3,
        font=("Helvetica", 10, "bold"),
        foreground="black"
    )
    status_label_3.pack(side="right", padx=10)


    return page3_frame



def update_status_4(message=None, is_error=False):
    """Update status message in page 4"""
    global status_label_4, status_var_4
    
    if not message:
        message = f"Current file: {os.path.basename(current_tp_file) if current_tp_file else 'None'}"
    
    # Print for debug
    print(f"Status 4: {message}")
    
    try:
        # Safely update status variable if it exists
        if status_var_4 and hasattr(status_var_4, 'get'):
            status_var_4.set(message)
        
        # Safely update label color based on error status
        if status_label_4:
            # Check if the widget still exists and is valid
            if hasattr(status_label_4, 'winfo_exists') and status_label_4.winfo_exists():
                try:
                    status_label_4.config(
                        foreground='red' if is_error else 'black'
                    )
                except Exception as e:
                    print(f"Could not update status label config: {str(e)}")
            else:
                print("Status label widget no longer exists")
    except Exception as e:
        print(f"Error in update_status_4: {str(e)}")

def update_file_status(filename=None, is_error=False, source_page=None):
    """Update status labels and load test data when file changes"""
    global current_tp_file
    
    try:
        if filename is None and current_tp_file:
            filename = os.path.basename(current_tp_file)
        
        if not filename:
            if source_page == '2':
                update_status_3("No file opened")
                update_status_4("No file selected")
            elif source_page == '3':
                update_status_2("No file opened")
                update_status_4("No file selected")
            elif source_page == '4':
                update_status_2("No file opened")
                update_status_3("No file opened")
            return

        # Always update page 4's status regardless of source page
        if current_tp_file:
            update_status_4(f"Current file: {os.path.basename(current_tp_file)}")
        
        # Update based on source page
        if source_page == '3':
            if is_error:
                update_status_2(f"Error: {filename}", True)
            else:
                update_status_2(f"Current file: {filename}")
                # Load config in page 2
                threading.Thread(
                    target=lambda: load_config_data_page2(current_tp_file),
                    daemon=True
                ).start()
        
        elif source_page == '2':
            if is_error:
                update_status_3(f"Error: {filename}", True)
            else:
                update_status_3(f"Current file: {filename}")
                # Load test data in page 3
                if 'load_test_data' in page3_functions:
                    threading.Thread(
                        target=lambda: page3_functions['load_test_data'](current_tp_file), 
                        daemon=True
                    ).start()
        
        elif source_page == '4':
            if is_error:
                update_status_2(f"Error: {filename}", True)
                update_status_3(f"Error: {filename}", True)
            else:
                update_status_2(f"Current file: {filename}")
                update_status_3(f"Current file: {filename}")
                # Load both config and test data
                threading.Thread(
                    target=lambda: load_config_data_page2(current_tp_file),
                    daemon=True
                ).start()
                if 'load_test_data' in page3_functions:
                    threading.Thread(
                        target=lambda: page3_functions['load_test_data'](current_tp_file), 
                        daemon=True
                    ).start()
            
    except Exception as e:
        print(f"Error updating status: {str(e)}")

def count_tp_test_lines():
    """
    Counts non-empty lines between [test] and [/test] sections in the TP file.
    """
    try:
        if not current_tp_file:
            print("No TP file currently set")
            return 0
        
        count = 0
        in_test_section = False
        
        with open(current_tp_file, 'r') as file:
            for line in file:
                line = line.strip()
                
                # Check for section markers
                if line.lower() == '[test]':
                    in_test_section = True
                    continue
                elif line.lower() == '[/test]':
                    in_test_section = False
                    continue
                
                # Count non-empty lines within test section
                if in_test_section and line:
                    count += 1
                    
        return count
    except Exception as e:
        print(f"Error counting TP test lines: {str(e)}")
        return 0

def count_csv_data_rows():
    """
    Counts the number of data rows in the current Results.csv file (excluding header).
    """
    try:
        if not current_results_csv:
            print("No Results.csv file currently set")
            return 0
            
        # Wait briefly for file operations to complete
        time.sleep(1)
        
        # Try to find the file in possible locations
        base_filename = os.path.basename(current_results_csv)
        possible_paths = [
            current_results_csv,  # Original path
            os.path.join('..', '..', 'MyWorkspace', 'Out', base_filename),  # Relative workspace path
            os.path.join(os.getcwd(), '..', '..', 'MyWorkspace', 'Out', base_filename),  # From current directory
            os.path.join(output_folder_var.get(), base_filename)  # From GUI output folder
        ]
        
        for path in possible_paths:
            try:
                if os.path.exists(path):
                    with open(path, 'r') as file:
                        next(file)  # Skip header
                        count = sum(1 for line in file if line.strip())
                        print(f"Successfully read CSV from: {path}")
                        return count
            except Exception as file_error:
                print(f"Could not read from {path}: {str(file_error)}")
                continue
                
        # If no file found, log the error
        print(f"Could not find CSV file: {base_filename}")
        print(f"Tried paths: {possible_paths}")
        return 0
            
    except Exception as e:
        print(f"Error counting CSV rows: {str(e)}")
        return 0

def update_measurement_database(folder_name, tp_filename, display_func):
    """
    Updates the measurement database Excel file with new measurement details.
    Uses existing temp file when main database is locked.
    """
    try:
        excel_path = '//HERMES/4_Engineering/7-test campaign/Results/database.Measurement.xlsx'
        temp_file = '//HERMES/4_Engineering/7-test campaign/Results/Python Temp/temp_measurements.xlsx'
        base_results_path = '//HERMES/4_Engineering/7-test campaign/Results'

        # Get the line count from TP file
        tp_line_count = count_tp_test_lines()
        # Get CSV row count
        csv_row_count = count_csv_data_rows()

        def write_data_to_sheet(sheet, row_num):
            """Helper function to write data to a sheet"""
            for col in range(1, sheet.max_column + 1):
                header = str(sheet.cell(row=1, column=col).value or '')
                
                # Measurement name [Raw Results]
                if "Measurement name [Raw Results]" in header:
                    sheet.cell(row=row_num, column=col, value=os.path.basename(folder_name))
                # Filename TP
                elif "Filename TP" in header:
                    sheet.cell(row=row_num, column=col, value=tp_filename)
                # Platform
                elif "platform" in header:
                    sheet.cell(row=row_num, column=col, value=ip_entry.get())
                # DUT Name
                elif "DUT Name" in header:
                    sheet.cell(row=row_num, column=col, value=dut_name_var.get())
                # DUT SN
                elif "DUT SN" in header:
                    sheet.cell(row=row_num, column=col, value=dut_serialnumber_var.get())
                # Host
                elif "DUT Host" in header:
                    sheet.cell(row=row_num, column=col, value=dut_host_var.get())

                elif "DPU Host" in header:
                    sheet.cell(row=row_num, column=col, value=dslam_host_var.get())    
                # DPU Name
                elif "DPU Name" in header:
                    sheet.cell(row=row_num, column=col, value=dslam_name_var.get())
                # DPU SN
                elif "DPU SN" in header:
                    sheet.cell(row=row_num, column=col, value=dslam_serialnumber_var.get())
                # Port (capitalized)
                elif header == "Port":
                    sheet.cell(row=row_num, column=col, value=dut_port_var.get())
                # port (lowercase)
                elif header == "port":
                    sheet.cell(row=row_num, column=col, value=dslam_port_var.get())
                # Profile
                elif "Profile" in header:
                    sheet.cell(row=row_num, column=col, value=dslam_profile_var.get())
                # Linesimulator
                elif "Linesimulator" in header:
                    sheet.cell(row=row_num, column=col, value=platform_linesimulator_var.get())
                # Noise generator
                elif "Noise generator" in header:
                    sheet.cell(row=row_num, column=col, value=platform_noise_generator_var.get())
                # Traffic Generator
                elif "Traffic Generator" in header:
                    sheet.cell(row=row_num, column=col, value=platform_traffic_generator_var.get())

                elif "operator" in header:
                    sheet.cell(row=row_num, column=col, value=operator_var.get())    
                elif "Test Plan" in header:
                    sheet.cell(row=row_num, column=col,  value=dut_testplan_var.get())
                # Switch
                elif "Switch" in header:
                    sheet.cell(row=row_num, column=col, value=platform_switch_var.get())
                # Link
                elif header == "Link":
                    measurement_folder = os.path.basename(folder_name)
                    network_path = f"\\\\HERMES\\4_Engineering\\7-test campaign\\Results\\{measurement_folder}"
                    hyperlink_formula = f'=HYPERLINK("{network_path}","Click to open")'
                    sheet.cell(row=row_num, column=col, value=hyperlink_formula)
                # Expected
                elif header == "Expected":
                    sheet.cell(row=row_num, column=col, value=tp_line_count)
                # Actual
                elif header == "Actual":
                    sheet.cell(row=row_num, column=col, value=csv_row_count)

        try:
            # Try to write to main database first
            wb = openpyxl.load_workbook(excel_path)
            
            # Get appropriate sheet (2024 or 2025)
            current_date = datetime.now()
            sheet_name = "2025" if current_date >= datetime(2025, 1, 1) else "2024"
            
            if sheet_name == "2025" and "2025" not in wb.sheetnames:
                wb.copy_worksheet(wb["2024"]).title = "2025"
                sheet_2025 = wb["2025"]
                for row in range(2, sheet_2025.max_row + 1):
                    for col in range(1, sheet_2025.max_column + 1):
                        sheet_2025.cell(row=row, column=col).value = None
            
            sheet = wb[sheet_name]
            
            # Find last row in main sheet
            last_row = 1
            while sheet.cell(row=last_row, column=1).value is not None:
                last_row += 1

            # Write current data
            write_data_to_sheet(sheet, last_row)
            
            # Save main database
            wb.save(excel_path)
            
            # Process temp file if it has data
            temp_wb = openpyxl.load_workbook(temp_file)
            temp_sheet = temp_wb['Temp']
            
            if temp_sheet.max_row > 1:  # If there's data beyond headers
                # Copy all data from temp to main
                for row in range(2, temp_sheet.max_row + 1):
                    for col in range(1, temp_sheet.max_column + 1):
                        value = temp_sheet.cell(row=row, column=col).value
                        if value is not None:  # Only copy non-empty cells
                            sheet.cell(row=last_row + row - 1, column=col).value = value
                
                # Save main database with temp data
                wb.save(excel_path)
                
                # Clear temp sheet (keep headers)
                for row in range(2, temp_sheet.max_row + 1):
                    for col in range(1, temp_sheet.max_column + 1):
                        temp_sheet.cell(row=row, column=col).value = None
                
                # Save cleared temp file
                temp_wb.save(temp_file)
                display_func("Successfully transferred temp data to main database", "green")
            
            display_func(f"Database updated successfully with measurement: {folder_name}", "green")
            
        except PermissionError:
            # If main database is locked, save to temp file
            temp_wb = openpyxl.load_workbook(temp_file)
            temp_sheet = temp_wb['Temp']
            
            # Find last row in temp sheet
            last_row = 1
            while temp_sheet.cell(row=last_row, column=1).value is not None:
                last_row += 1
            
            # Write current data to temp sheet
            write_data_to_sheet(temp_sheet, last_row)
            
            # Save temp file
            temp_wb.save(temp_file)
            display_func(f"Data saved to temporary file. Will update database when available.", "yellow")

    except Exception as e:
        error_msg = f"Error updating measurement database: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        messagebox.showerror("Database Update Error", error_msg)

def thread_safe_display_message_4(message, color="black"):
    """Thread-safe way to update the display text"""
    # Don't use lambda here to avoid closure issues with widget references
    def safe_update():
        if response_text_4 and hasattr(response_text_4, 'winfo_exists') and response_text_4.winfo_exists():
            try:
                display_message_4(message, color)
            except Exception as e:
                print(f"Error in thread_safe_display_message_4: {str(e)}")
        else:
            print(f"Warning: Cannot display message - text widget doesn't exist: {message}")
            
    # Schedule the update on the main thread
    if 'root' in globals() and root and root.winfo_exists():
        root.after(0, safe_update)
def display_message_4(message, color="black"):
    """Update the response text with the message and log to file"""
    global response_text_4, current_log_file
    
    # Safety check - make sure widget exists before trying to use it
    if not response_text_4 or not hasattr(response_text_4, 'winfo_exists') or not response_text_4.winfo_exists():
        print(f"Warning: Cannot display message - text widget doesn't exist: {message}")
        return
        
    try:
        # Configure basic black color
        response_text_4.tag_configure("black", foreground="black")
        
        # Enable text widget for editing
        response_text_4.config(state=tk.NORMAL)

        current_view = response_text_4.yview()
        is_scrolled_up = current_view[1] < 1.0

        # Insert message
        response_text_4.insert(tk.END, message)
        
        # Add newline if needed
        if not message.endswith('\n'):
            response_text_4.insert(tk.END, "\n")

        # Only auto-scroll if user hasn't scrolled up
        if not is_scrolled_up:
            response_text_4.see(tk.END)
        else:
            response_text_4.yview_moveto(current_view[0])    

        # Disable text widget
        response_text_4.config(state=tk.DISABLED)
        
        # Also log to the current log file if it exists
        if current_log_file and os.path.exists(current_log_file):
            try:
                with open(current_log_file, 'a', encoding='utf-8') as f:
                    # Add timestamp to log entry
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    
                    # Only add timestamp for new lines, not for continuations
                    if message.strip() and (message.startswith('\n') or not message[0].isalnum()):
                        f.write(f"[{timestamp}] {message}")
                    else:
                        f.write(message)
                    
                    # Add newline if needed
                    if not message.endswith('\n'):
                        f.write("\n")
            except Exception as e:
                print(f"Error writing to log file: {str(e)}")
                
    except Exception as e:
        print(f"Error displaying message in text widget: {str(e)}")

def run_pretest():
        """
        Runs pretest checks before executing the test plan.
        Returns True if pretest passes, False otherwise.
        """
        try:

            # Clear the response text area first
            if response_text_4:
                response_text_4.config(state=tk.NORMAL)  # Enable editing temporarily
                response_text_4.delete(1.0, tk.END)      # Clear all text
                response_text_4.config(state=tk.DISABLED) # Disable editing again
            else:
                print(f"Error clearing response text: {str(e)}")
            thread_safe_display_message_4("\n===Pre Test Initiated===\n", "black")
            # Get the driver path from the entry
            driver_path = dslam_driver_var.get()
            if not driver_path:
                thread_safe_display_message_4("Error: No driver selected\n", "red")
                return False
                
            # Load the driver
            cmd_load = f"load('{driver_path}')"
            res_load = socket_query(sock, cmd_load)
            thread_safe_display_message_4(f"Command sent : {cmd_load}\n", "black")            
            # Get DSLAM busmap and port values
            dslam_busmap = dropdowns['DSLAM_Busmap']['widget'].get()  # Get current value from combobox
            dslam_port = dslam_port_var.get() or "0"  # Default to 0 if empty
            
            # Command 1: Set DSLAM busmap
            cmd1 = f"dslam = '{dslam_busmap}'"
            res1 = socket_query(sock, cmd1)  # No need to check res1
            
            # Command 2: Set port number (without quotes)
            cmd2 = f"port = {dslam_port}"
            res2 = socket_query(sock, cmd2)
            thread_safe_display_message_4(f"Commands sent : {cmd1}\n{cmd2}\n", "black")
                
            # Command 3: Read status
            cmd3 = "(e, r) = readstatus(dslam, port) <<< e"
            res3 = socket_query(sock, cmd3)
            thread_safe_display_message_4(f"Command sent : {cmd3}\n", "black")
                
            # Check if res3 contains '?' or '!' or if it's in red color
            if ('?' in str(res3) or '!' in str(res3) or 
                "\033[31m" in str(res3) or  # ANSI escape code for red
                "\033[91m" in str(res3)):   # ANSI escape code for bright red
                error_msg = f"Error: Verify the setup, error in 'Read Status' function.\nResponse: {res3}\n"
                thread_safe_display_message_4(error_msg, "red")
                return False
            # Check if line simulator is DLA type
            line_simulator = dropdowns['Linesimulator']['widget'].get().upper()
            if 'DLA' in line_simulator or 'LSX 2030VX' in line_simulator.upper():                
                cmd4 = "setloop(0)"
                res4 = socket_query(sock, cmd4)
                thread_safe_display_message_4(f"Command sent : {cmd4}\n", "black")

                # Command 5: Get status after setting loop to 0
                if 'DLA' in line_simulator:
                    cmd5 = "(e, i) = dla35x::getstatus() <<< e <<< i"
                elif 'LSX' in line_simulator:
                    cmd5 = "(e, i) = lsx2030vx::getstatus() <<< e <<< i"
                else:
                    cmd5 = "(err, reply) = sendto([dslam], "") <<< e"
                    res5 = socket_query(sock,cmd5)
                    thread_safe_display_message_4(f"Command sent : {cmd5}\n", "black")
                    if '?' in str(res5) or '!' in str(res5) :
                        error_msg = f"Error: Verify the setup, error in 'sendto()' function.\nResponse:\n{res5}\n"
                        thread_safe_display_message_4(error_msg, "red")
                        return False

                res5 = socket_query(sock, cmd5)
                thread_safe_display_message_4(f"Command sent : {cmd5}\n", "black")

                # Check if linelength is 0
                if "length: 0" not in str(res5) and "length:0" not in str(res5) and "length" not in str(res5):
                    error_msg = f"Error: Verify the setup, error in 'getstatus()' function.\nResponse for length 0:\n{res5}\n"
                    thread_safe_display_message_4(error_msg, "red")
                    return False
                

                # Command 6: Set loop to 1000
                cmd6 = "setloop(100)"
                res6 = socket_query(sock, cmd6)
                thread_safe_display_message_4(f"Command sent : {cmd6}\n", "black")

                # Command 7: Get status after setting loop to 1000
                if 'DLA' in line_simulator:
                    cmd7 = "(e, i) = dla35x::getstatus() <<< e <<< i"
                else:
                    cmd7 = "(e, i) = lsx2030vx::getstatus() <<< e <<< i"
                res7 = socket_query(sock, cmd7)
                thread_safe_display_message_4(f"Command sent : {cmd7}\n", "black")

                # Check if linelength is 1000
                if "length: 100" not in str(res7) and "length:100" not in str(res7) and "length" not in str(res7):
                    error_msg = (
                        f"Error: Verify the setup, error in 'getstatus()' function.\n"
                        f"Response for length 0:\n{res5}\n"
                        f"Response for length 100:\n{res7}\n"
                    )
                    thread_safe_display_message_4(error_msg, "red")
                    return False
            else:
                thread_safe_display_message_4("Line simulator is not DLA/LSX 2030VX type, skipping loop commands.\n", "black")
                
            thread_safe_display_message_4("\n===Pre Test completed===\n", "black")
            return True                           

            
        except Exception as e:
            thread_safe_display_message_4(f"Error in Pre Test: {str(e)}\n", "red")
            return False


def run_server_commands(tp_file):
        try:
            # Run pretest first
            if not run_pretest():
                update_status_4("Pre Test failed - Test plan execution aborted", True)
                return            
            server_path = tp_file.replace('/', '\\')
            if ' ' in tp_file:
                directory = os.path.dirname(tp_file)
                filename = os.path.basename(tp_file)

                new_name =filename.replace(' ','_')
                new_tp_name = os.path.join(directory,new_name)
                try:
                    shutil.copy2(tp_file,new_tp_name)
                    thread_safe_display_message_4(f"Renamed file to avoid spaces in filename :{new_tp_name}\n", "black")
                    tp_file = new_tp_name
                
                except Exception as e:
                    thread_safe_display_message_4(f"Error: Failed to copy file: {str(e)}\n", "red")
                    return
            server_path = tp_file.replace('/', '\\')    
            global current_tp_file
            current_tp_file = tp_file
            # Set file path command
            cmd1 = f'tp_file="{server_path}"'
            response1 = socket_query(sock, cmd1)
            
            if response1 is None:
                thread_safe_display_message_4("Error: No response from server for file path command\n", "red")
                update_status_4("Error: Server not responding", True)
                return
            
   
            # Run test plan command with progress monitoring
            cmd2 = '(err) = Run_tp(tp_file) <<< err'
            thread_safe_display_message_4("\nStarting test plan execution...\n")
            update_status_4(f"Executing: {os.path.basename(tp_file)}")
            socket_query_continuous(sock, cmd2, thread_safe_display_message_4, update_status_4, config_inner_frame, test_inner_frame)            
            # Use the new handle_long_running_command
            # handle_long_running_command(cmd2, thread_safe_display_message_4)
            
        except Exception as e:
            thread_safe_display_message_4(f"Error in server communication: {str(e)}\n", "red")
            update_status_4("Error: Server not responding", True)
    
def run_current_or_new_tp():
        """Run current test plan or select new one"""
        if not sock:
            messagebox.showerror("Error", "Please connect to server first")
            update_status_4("Error: Server not connected", True)
            return
        # Check if server is actually responsive
        if not check_server_connection():
            messagebox.showerror("Connection Error", 
                           "Server connection lost!\nPlease reconnect to the server.")
            update_status_4("Error: Server connection lost", True)
            return            
        if current_tp_file:
            try:
                thread_safe_display_message_4("\nAuto-saving changes before execution...\n")
                
                def save_changes():
                    try:
                        # Save Page 2 changes (config)
                        thread_safe_display_message_4("Saving configuration changes...\n")
                        save_test_plan.current_tp_file = current_tp_file
                        save_test_plan()
                        thread_safe_display_message_4("Configuration saved successfully.\n")
                        
                        # Save Page 3 changes (test data)
                        thread_safe_display_message_4("Saving test data changes...\n")

                        # Now show the run dialog in the main thread
                        #root.after(0, lambda: show_run_dialog())
                        execute_tp(current_tp_file)
                    except Exception as e:
                        error_msg = f"Error during auto-save: {str(e)}"
                        thread_safe_display_message_4(f"{error_msg}\n", "red")
                        root.after(0, lambda: messagebox.showerror("Auto-save Error", error_msg))
                        print(f"Save error: {traceback.format_exc()}")
                
                # Start save operations in background
                save_thread = threading.Thread(target=save_changes, daemon=True)
                save_thread.start()
                
            except Exception as e:
                error_msg = f"Error initiating auto-save: {str(e)}"
                thread_safe_display_message_4(f"{error_msg}\n", "red")
                messagebox.showerror("Error", error_msg)
                print(f"Init error: {traceback.format_exc()}")
                return
        else:
            update_status_4("No file selected")
            select_and_run_tp()

def show_run_dialog():
        """Show dialog to run current or select new test plan"""
        response = messagebox.askyesnocancel(
            "Run Test Plan",
            f"Use current test plan?\n{os.path.basename(current_tp_file)}\n\n"
            "Yes: Use current file\n"
            "No: Select different file\n"
            "Cancel: Abort"
        )

        if response is None:  # Cancel
            return
        elif response:  # Yes - use current file
            execute_tp(current_tp_file)
        else:  # No - select new file
            select_and_run_tp()

def select_and_run_tp():
        """Select and run new test plan"""
        global current_tp_file
        tp_file = filedialog.askopenfilename(
            title="Select Test Plan to Run",
            filetypes=[("Test Plan Files", "*.tp")]
        )
        
        if tp_file:
            current_tp_file = tp_file
            update_status_4(f"Current file: {os.path.basename(tp_file)}")
            update_file_status(source_page='4')
            execute_tp(tp_file)
        else:
            update_status_4("No file selected", True)
        
    
def execute_tp(tp_file):

        """Execute the test plan file"""
        try:
            update_status_4(f"Executing: {os.path.basename(tp_file)}")
            # Clear previous content
            for widget in config_inner_frame.winfo_children():
                widget.destroy()
            for widget in test_inner_frame.winfo_children():
                widget.destroy()
            
            # Load and parse file
            with open(tp_file, 'r') as f:
                lines = f.readlines()
        
            in_config = False
            in_test = False
            test_count = 0
            current_test = ""
        
            for line in lines:
                line = line.strip()
                if not line:  # Skip empty lines
                    continue
                
                if line == '[config]':
                    in_config = True
                    in_test = False
                elif line == '[\\config]':
                    in_config = False
                elif line == '[test]':
                    in_test = True
                    in_config = False
                elif line == '[\\test]':
                    in_test = False
                     # Add any remaining test
                    if current_test:
                        test_count += 1
                        if ", use:" in current_test or ";use:" in current_test or " ;use:" in current_test:
                            test_name, profile = current_test.split(", use:")
                            if ";use:" in current_test:
                                test_name, profile = current_test.split(";use:")
                            elif " ;use:" in current_test:
                                test_name, profile = current_test.split(" ;use:")
                            elif ",use:" in current_test:
                                test_name, profile = current_test.split(",use:")
                            add_test_row(test_count, test_name.strip(), profile.strip())
                        else:
                            add_test_row(test_count, current_test.strip(), "")
                        current_test = ""
                elif in_config and '=' in line:
                    key, value = line.split('=', 1)
                    value = value.strip().strip("'")
                    add_config_row(key.strip(), value)
                elif in_test:
                    # Handle test cases with parameters
                    if '{' in line:
                        # Add previous test if exists
                        if current_test:
                            test_count += 1
                            if ", use:" in current_test:
                                test_name, profile = current_test.split(", use:")
                                add_test_row(test_count, test_name.strip(), profile.strip())
                            else:
                                add_test_row(test_count, current_test.strip(), "")
                        current_test = line
                    else:
                        # If line contains "use:" profile, it's a new test
                        if ", use:" in line and not current_test:
                            test_count += 1
                            test_name, profile = line.split(", use:")
                            add_test_row(test_count, test_name.strip(), profile.strip())
                        # Otherwise, it might be a continuation of current test
                        elif current_test:
                            current_test = (current_test + " " + line).strip()
                        else:
                            test_count += 1
                            add_test_row(test_count, line.strip(), "")
        
        # Start execution in background
            update_status_4(f"Starting: {os.path.basename(tp_file)}")
            threading.Thread(
                target=lambda: run_server_commands(tp_file), 
                daemon=True
            ).start()
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to execute test plan: {str(e)}")
            update_status_4(f"Error: {str(e)}", True)


def add_config_row(setting, value):
        row_frame = ttk.Frame(config_inner_frame)
        row_frame.grid(sticky="ew")  # Changed from pack to grid
        
        setting_entry = ttk.Entry(row_frame, width=42)
        setting_entry.grid(row=0, column=0, sticky="ew")
        setting_entry.insert(0, setting)
        setting_entry.configure(state='readonly')
        
        value_entry = ttk.Entry(row_frame, width=65)
        value_entry.grid(row=0, column=1, sticky="ew")
        value_entry.insert(0, value)
        value_entry.configure(state='readonly')
        
        row_frame.grid_columnconfigure(0, weight=1)
        row_frame.grid_columnconfigure(1, weight=2)
        
        
    
        def on_double_click(event):
            widget = event.widget
            if widget.cget('state') == 'readonly':
                widget.configure(state='normal')
                widget.select_range(0, tk.END)
                widget.configure(state='readonly')
    
        setting_entry.bind('<Double-1>', on_double_click)
        value_entry.bind('<Double-1>', on_double_click)


        # Add after creating canvas windows
        def on_canvas_configure(event, canvas, window):
            """Update canvas window width when canvas is resized"""
            canvas.itemconfig(window, width=event.width)


        # Bind canvas configure events
        config_canvas.bind('<Configure>', 
                          lambda e: on_canvas_configure(e, config_canvas, config_window))
        test_canvas.bind('<Configure>', 
                          lambda e: on_canvas_configure(e, test_canvas, test_window))

        # Mouse wheel scrolling
        def on_mousewheel(event, canvas):
            """Handle mouse wheel scrolling"""
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        # Bind mouse wheel events - add these after canvas creation
        def bind_mousewheel(canvas):
            """Bind mouse wheel to canvas and its children"""
            if canvas and hasattr(canvas, 'bind_all'):
                canvas.bind_all("<MouseWheel>", lambda e: on_mousewheel(e, canvas))

        def unbind_mousewheel(canvas):
            """Unbind mouse wheel when mouse leaves the canvas"""
            if canvas and hasattr(canvas, 'unbind_all'):
                canvas.unbind_all("<MouseWheel>")

        # Bind mouse enter/leave events for each canvas
        config_canvas.bind('<Enter>', lambda e: bind_mousewheel(config_canvas))
        config_canvas.bind('<Leave>', lambda e: unbind_mousewheel(config_canvas))
        test_canvas.bind('<Enter>', lambda e: bind_mousewheel(test_canvas))
        test_canvas.bind('<Leave>', lambda e: unbind_mousewheel(test_canvas))    
        return row_frame
    



def add_test_row(num, name, profile=""):
        row_frame = ttk.Frame(test_inner_frame)
        row_frame.grid(sticky="ew")
        
        num_entry = ttk.Entry(row_frame, width=5)
        num_entry.grid(row=0, column=0, sticky="w")
        num_entry.insert(0, num)
        num_entry.configure(state='readonly')
        
        name_entry = ttk.Entry(row_frame, width=60)
        name_entry.grid(row=0, column=1,  sticky="ew")
        name_entry.insert(0, name)
        name_entry.configure(state='readonly')
        
        profile_entry = ttk.Entry(row_frame, width=40)
        profile_entry.grid(row=0, column=2, sticky="ew")
        profile_entry.insert(0, profile)
        profile_entry.configure(state='readonly')
        
        row_frame.grid_columnconfigure(1, weight=2)
        row_frame.grid_columnconfigure(2, weight=1)
        
        def on_double_click(event):
            widget = event.widget
            if widget.cget('state') == 'readonly':
                widget.configure(state='normal')
                widget.select_range(0, tk.END)
                widget.configure(state='readonly')
    
        for entry in [num_entry, name_entry, profile_entry]:
            entry.bind('<Double-1>', on_double_click)
    
        return row_frame

    # Configure scroll regions
def configure_scroll_region(canvas, inner_frame):
    """Configure the scroll region for a canvas"""
    # Safety check to prevent NoneType errors
    if not canvas or not inner_frame:
        print("Warning: Cannot configure scroll region - canvas or frame is None")
        return
        
    def _configure_scroll_region(event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))
    
    # Bind the configuration event
    inner_frame.bind("<Configure>", _configure_scroll_region)


def refresh_database():
        """Start refresh database operation in a separate thread"""
        threading.Thread(target=refresh_database_thread, daemon=True).start()

def refresh_database_thread():
        """Thread function to handle database refresh operation"""
        try:
            excel_path = '//HERMES/4_Engineering/7-test campaign/Results/database.Measurement.xlsx'
            temp_file = '//HERMES/4_Engineering/7-test campaign/Results/Python Temp/temp_measurements.xlsx'
            
            # Load both workbooks
            wb = openpyxl.load_workbook(excel_path)
            temp_wb = openpyxl.load_workbook(temp_file)
            temp_sheet = temp_wb['Temp']
            
            if temp_sheet.max_row > 1:  # If there's data beyond headers
                # Get appropriate sheet (2024 or 2025)
                current_date = datetime.now()
                sheet_name = "2025" if current_date >= datetime(2025, 1, 1) else "2024"
                
                if sheet_name == "2025" and "2025" not in wb.sheetnames:
                    wb.copy_worksheet(wb["2024"]).title = "2025"
                    sheet_2025 = wb["2025"]
                    for row in range(2, sheet_2025.max_row + 1):
                        for col in range(1, sheet_2025.max_column + 1):
                            sheet_2025.cell(row=row, column=col).value = None
                
                sheet = wb[sheet_name]
                
                # Find last row in main sheet
                last_row = 1
                while sheet.cell(row=last_row, column=1).value is not None:
                    last_row += 1

                # Copy all data from temp to main
                for row in range(2, temp_sheet.max_row + 1):
                    for col in range(1, temp_sheet.max_column + 1):
                        value = temp_sheet.cell(row=row, column=col).value
                        if value is not None:  # Only copy non-empty cells
                            sheet.cell(row=last_row + row - 1, column=col).value = value
                
                # Save main database with temp data
                wb.save(excel_path)
                
                # Clear temp sheet (keep headers)
                for row in range(2, temp_sheet.max_row + 1):
                    for col in range(1, temp_sheet.max_column + 1):
                        temp_sheet.cell(row=row, column=col).value = None
                
                # Save cleared temp file
                temp_wb.save(temp_file)
                thread_safe_display_message_4("Successfully transferred temp data to main database", "green")
            else:
                thread_safe_display_message_4("No temporary data found to transfer", "black")
                
        except Exception as e:
            error_msg = f"Error refreshing database: {str(e)}"
            thread_safe_display_message_4(error_msg, "red")
            messagebox.showerror("Database Refresh Error", error_msg)


def filter_config_rows(*args):
        search_text = config_search_var.get().lower()
        for child in config_inner_frame.winfo_children():
            show = True
            for entry in child.winfo_children():
                if isinstance(entry, ttk.Entry):
                    if search_text and search_text not in entry.get().lower():
                        show = False
                        break
            if show:
                child.grid()
            else:
                child.grid_remove()

def filter_test_rows(*args):
        search_text = test_search_var.get().lower()
        for child in test_inner_frame.winfo_children():
            show = True
            for entry in child.winfo_children():
                if isinstance(entry, ttk.Entry):
                    if search_text and search_text not in entry.get().lower():
                        show = False
                        break
            if show:
                child.grid()
            else:
                child.grid_remove()

# # Function to create Page 4
# def create_page4(parent=None, as_tab=True):
#     """
#     Create the Test Plan Execution page
    
#     Args:
#         parent: Optional parent widget. If None, uses notebook.
#         as_tab: Whether to add as a notebook tab or just create a frame
#     """

#     global page4_frame, config_inner_frame, test_inner_frame, config_canvas, test_canvas
#     global response_text_4, status_label_4, status_var_4, config_search_var, test_search_var
#     global config_window, test_window
    
    
#     # Reset canvas references first to avoid binding to old references
#     config_canvas = None
#     test_canvas = None
    
#     print(f"Creating page4: parent={parent}, as_tab={as_tab}")
    
#     try:
#         # Use provided parent or default to notebook
#         if parent is None:
#             parent = notebook
        
#         # Create main frame
#         page4_frame = ttk.Frame(parent, padding="5 5 5 5")
        
#         # Add as tab if needed
#         if as_tab:
#             notebook.add(page4_frame, text="Test Plan Execution")
#             # Bind tab change event
#             notebook.bind('<<NotebookTabChanged>>', lambda e: update_status_4())
#         else:
#             # Just pack the frame in the parent
#             page4_frame.pack(fill="both", expand=True)
        
#         # Initialize StringVars - only if they don't already exist
#         if not config_search_var:
#             config_search_var = tk.StringVar()
#         if not test_search_var:
#             test_search_var = tk.StringVar()
#         if not status_var_4:
#             status_var_4 = tk.StringVar(value="No file loaded")
        
#         print("Successfully created page4 frame and variables")
    
#         # Create main frame
#         #page4_frame = ttk.Frame(parent, padding="5 5 5 5")
    
    
#         global df, current_tp_file, status_label_4
    
#         #notebook.add(page4_frame, text="Test Plan  Execution")
#         # After creating notebook and all pages
#         #notebook.bind('<<NotebookTabChanged>>', lambda e: update_status_4())
#         # Create two main vertical frames
#         left_frame = ttk.Frame(page4_frame ,relief="solid")
#         right_frame = ttk.Frame(page4_frame, relief="solid")
    
#         left_frame.grid(row=0, column=0, sticky="nsew")
#         right_frame.grid(row=0, column=1, sticky="nsew")
    
#         # Configure grid weights
#         page4_frame.grid_columnconfigure(0, weight=2)    # Left frame smaller
#         page4_frame.grid_columnconfigure(1, weight=4)    # Right frame larger
#         page4_frame.grid_rowconfigure(0, weight=1)       # Full height

#         # Create a container frame for the entire config section
#         config_section_frame = ttk.Frame(left_frame)
#         config_section_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
#         # Config Header Section
#         config_header_frame = ttk.Frame(config_section_frame)
#         config_header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0,5))

#         # Config Search
#         config_search_var = tk.StringVar()
#         config_search_entry = ttk.Entry(config_header_frame, textvariable=config_search_var, width=30)
#         config_search_entry.grid(row=0, column=1, padx=5)
#         ttk.Button(config_header_frame, 
#                text="Clear", 
#                command=lambda: config_search_var.set("")).grid(row=0, column=2, padx=5)

#         # Config Content Frame
#         config_content_frame = ttk.Frame(config_section_frame)
#         config_content_frame.grid(row=1, column=0, sticky="nsew")

#         # Create canvas and scrollbar in content frame
#         config_canvas = tk.Canvas(config_content_frame, height=200)
#         config_canvas.grid(row=0, column=0, sticky="nsew")

#         config_scrollbar = ttk.Scrollbar(config_content_frame, orient="vertical", command=config_canvas.yview)
#         config_scrollbar.grid(row=0, column=1, sticky="ns")

#         config_inner_frame = ttk.Frame(config_canvas)
#         config_window = config_canvas.create_window((0, 0), window=config_inner_frame, anchor="nw")
#         config_canvas.configure(yscrollcommand=config_scrollbar.set)

#         # Configure weights for proper expansion
#         config_section_frame.grid_columnconfigure(0, weight=1)
#         config_section_frame.grid_rowconfigure(1, weight=1)

#         config_content_frame.grid_columnconfigure(0, weight=1)
#         config_content_frame.grid_rowconfigure(0, weight=1)

#         # Test Section Frame (similar structure)
#         test_section_frame = ttk.Frame(left_frame)
#         test_section_frame.grid(row=2, column=0, sticky="nsew", padx=5, pady=(10,5))

#         # Test Header Section
#         test_header_frame = ttk.Frame(test_section_frame)
#         test_header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0,5))


#         configure_scroll_region(config_canvas, config_inner_frame)
#         configure_scroll_region(test_canvas, test_inner_frame)

#         # Bind search variables after they're initialized
#         #config_search_var.trace('w', filter_config_rows)
#         #test_search_var.trace('w', filter_test_rows)

#         ttk.Label(test_header_frame, 
#              text="Test Plans", 
#              font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0,10))

#         test_search_var = tk.StringVar()
#         test_search_entry = ttk.Entry(test_header_frame, textvariable=test_search_var, width=30)
#         test_search_entry.grid(row=0, column=1, padx=5)
#         ttk.Button(test_header_frame, 
#           text="Clear", 
#           command=lambda: test_search_var.set("")).grid(row=0, column=2, padx=5)

#         # Test Content Frame
#         test_content_frame = ttk.Frame(test_section_frame)
#         test_content_frame.grid(row=1, column=0, sticky="nsew")

#         # Create canvas and scrollbar in content frame
#         test_canvas = tk.Canvas(test_content_frame, height=300)
#         test_canvas.grid(row=0, column=0, sticky="nsew")

#         test_scrollbar = ttk.Scrollbar(test_content_frame, orient="vertical", command=test_canvas.yview)
#         test_scrollbar.grid(row=0, column=1, sticky="ns")

#         test_inner_frame = ttk.Frame(test_canvas)
#         test_window = test_canvas.create_window((0, 0), window=test_inner_frame, anchor="nw")
#         test_canvas.configure(yscrollcommand=test_scrollbar.set)

#         # Configure weights for proper expansion
#         test_section_frame.grid_columnconfigure(0, weight=1)
#         test_section_frame.grid_rowconfigure(1, weight=1)

#         test_content_frame.grid_columnconfigure(0, weight=1)
#         test_content_frame.grid_rowconfigure(0, weight=1)

#         # Header label and search
#         ttk.Label(config_header_frame, 
#              text="Configuration", 
#              font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0,10))


#         # Add a status message label below the table
#         status_var_4 = tk.StringVar()
#         table_search_var = tk.StringVar()



#         # Store the function in the global dictionary
#         page4_functions['clear_preview_tables'] = clear_page4_tables_safely
#         page4_functions['clear_preview_tables'] = clear_page4_tables_safely
#         page4_functions['display_message'] = thread_safe_display_message_4
#         page4_functions['run_test_plan'] = run_current_or_new_tp
#         # Preview Section - just a label instead of LabelFrame
#         #ttk.Label(left_frame, 
#         #          text="Test Plan Preview", 
#         #          font=("Helvetica", 10, "bold")).grid(row=1, column=0, sticky="w", padx=5)


   
#         # Header frame in left_frame - at row 0
#         header_frame = ttk.Frame(left_frame)
#         header_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
    
#         # Configure header frame columns
#         header_frame.grid_columnconfigure(0, weight=1)
#         header_frame.grid_columnconfigure(1, weight=0)

#         # Status Label - left aligned
#         status_label_4 = ttk.Label(header_frame, 
#                             text="No file loaded", 
#                             font=("Helvetica", 10, "bold"))
#         status_label_4.grid(row=0, column=0, sticky="w", padx=(5,10))

#         # Run Button - right aligned
#         ttk.Button(header_frame, 
#               text="Run Test Plan", 
#               width=20,
#               command=run_current_or_new_tp).grid(row=0, column=1, sticky="e", padx=5)
#         # Configure left frame grid weights for new implementation
#         # Left frame configuration
#         left_frame.grid_columnconfigure(0, weight=1)
#         left_frame.grid_rowconfigure(1, weight=1)
#         left_frame.grid_rowconfigure(2, weight=2)

#         # Configure right frame for server response
#         # Right frame server response setup
#         right_frame.grid_columnconfigure(0, weight=1)
#         right_frame.grid_rowconfigure(0, weight=0)  # Header row
#         right_frame.grid_rowconfigure(1, weight=1)  # Response text area

#         # Server Response Frame
#         response_frame_4 = ttk.Frame(right_frame, padding="5 5 5 5")
#         response_frame_4.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
#         response_frame_4.grid_columnconfigure(0, weight=1)
#         response_frame_4.grid_rowconfigure(1, weight=1)

#         header_response_frame = ttk.Frame(response_frame_4)
#         header_response_frame.grid(row=0, column=0, sticky="ew")
#         header_response_frame.grid_columnconfigure(0, weight=0)
#         header_response_frame.grid_columnconfigure(1, weight=0)
#         header_response_frame.grid_rowconfigure(2, weight=1)



#         # Server Response Header
#         ttk.Label(header_response_frame, 
#              text="Server Response:", 
#              foreground='black',
#              font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0,10), sticky="w")

#         # Refresh Database Button
#         ttk.Button(header_response_frame,
#               text="Refresh Database",
#               width=20,
#               command=refresh_database).grid(row=0, column=1, padx=(0,5), sticky="w")
  
#         # Add Close Page4 Button
#         ttk.Button(header_response_frame,
#               text="Close Execution",
#               width=15,
#               command=lambda: close_page4()).pack(side="right", padx=5)
        
#         # Response Text Area
#         response_text_4 = scrolledtext.ScrolledText(response_frame_4, 
#                                             wrap=tk.WORD,
#                                             foreground='black',
#                                             font=("Helvetica", 10, "bold"))
#         response_text_4.grid(row=1, column=0, columnspan=2, sticky="nsew")
#         response_text_4.config(state=tk.DISABLED)

#         # Add Ctrl+A binding
#         def on_select_all_4(event=None):
#             response_text_4.tag_add("sel", "1.0", "end")
#             return "break"
#         response_text_4.bind("<Control-a>", on_select_all_4)
#         try:
#             if config_search_var:
#                config_search_var.trace('w', filter_config_rows)
#             if test_search_var:
#                 test_search_var.trace('w', filter_test_rows)
#             print("Added trace functions to search variables")
#         except Exception as e:
#             print(f"Warning: Could not add trace functions: {str(e)}")
        
#             # Safely configure scroll regions
#         try:
#             if config_canvas and config_inner_frame:
#                 configure_scroll_region(config_canvas, config_inner_frame)
#             if test_canvas and test_inner_frame:
#                 configure_scroll_region(test_canvas, test_inner_frame)
#             print("Configured scroll regions")
#         except Exception as e:
#             print(f"Warning: Could not configure scroll regions: {str(e)}")
        
#         print("Page4 creation completed successfully")
        
#         # Return the frame
#         return page4_frame
        
#     except Exception as e:
#         print(f"Error creating page4: {str(e)}")
#         traceback.print_exc()  # Print full traceback for debugging
#         # Create a minimal frame to prevent further errors
#         if parent:
#             error_frame = ttk.Frame(parent)
#             ttk.Label(error_frame, text=f"Error creating execution page: {str(e)}").pack(pady=20)
#             error_frame.pack(fill="both", expand=True)
#             return error_frame
#         return None    
    
# Function to update results path when base_name changes

# Update the update_results_path function
def update_results_path(base_name=None):
    """Update the results path variable with the current base_name"""
    global results_path_var
    
    if not results_path_var:
        results_path_var = tk.StringVar(value=r"\\HERMES\4_Engineering\7-test campaign\Results")
    
    base_path = r"\\HERMES\4_Engineering\7-test campaign\Results"
    if base_name:
        full_path = os.path.join(base_path, base_name)
    else:
        full_path = base_path
        
    results_path_var.set(full_path)
    print(f"Updated results path to: {full_path}")

# Store the update function in page4_functions for access from other parts of code
page4_functions['update_results_path'] = update_results_path


# Add this function to handle both reset and navigation
def reset_and_go_to_page2():
    """Reset all fields to defaults and then navigate to page2"""
    # First reset everything
    reset_to_defaults()
    
    # Then switch to page2
    switch_tab(1)  # Index 1 corresponds to page2

def update_button_function(folder_path_var, timestamp_var, display_func, results_path_var=None, current_tp_file=None):
    """
    Function to handle Update button click - processes files and updates database
    
    Arguments:
        folder_path_var: StringVar for the folder path
        timestamp_var: StringVar for the timestamp
        display_func: Function to display messages in the UI
        results_path_var: StringVar for the results path (optional)
        current_tp_file: String path to current test plan file (optional)
    """

    global root_dir, workspace_dir, out_dir  
    # Get values from GUI variables
    folder_path = folder_path_var.get()
    timestamp = timestamp_var.get()
    current_test_plan_file = current_tp_file
    
    # Run in a separate thread to prevent UI freezing
    def process_files_thread():
        # Validate inputs
        if not folder_path or not timestamp:
            display_func("Please enter a folder path and timestamp")
            return
            
        # Validate timestamp format and check if future date
        if len(timestamp) >= 8:
            try:
                year = int(timestamp[0:4])
                month = int(timestamp[4:6])
                day = int(timestamp[6:8])
                
                # Check if date is in the future
                current_year = datetime.now().year
                if year > current_year:
                    display_func(f"Warning: Timestamp date {year}-{month:02d}-{day:02d} is in the future. Current year is {current_year}.")
            except ValueError:
                display_func("Note: Couldn't parse timestamp as YYYYMMDD_HHMMSS format")
        
        # Retrieve remote host
        remote_host = ip_entry.get() if 'ip_entry' in globals() else "lab-11"
        display_func(f"Using remote host: {remote_host}")
        
        # Get the correct OUT folder path without hardcoding
        try:
            if folder_path.startswith('\\\\'):
                # Case 1: Already a UNC path
                out_path = folder_path.rstrip("\\")
                display_func(f"Using specified network path: {out_path}")

            elif re.match(r'^[A-Za-z]:\\', folder_path):
                # Case 2: Local absolute path with drive letter (e.g. D:\MyWorkspace\Out)
                drive = folder_path[0].upper()
                path_without_drive = folder_path[2:].lstrip("\\")
        
                if remote_host.lower() in ("localhost", "127.0.0.1", socket.gethostname().lower()):
                    # Same machine → keep local path
                    out_path = folder_path.rstrip("\\")
                    display_func(f"Using local drive path: {out_path}")
                else:
                    # Remote machine → convert to UNC
                    out_path = f"\\\\{remote_host}\\{drive}$\\{path_without_drive}"
                    #unc_path = f"\\\\{remote_host}\\{drive}$\\{path_without_drive}"

                    display_func(f"Converted local drive path to UNC: {out_path}")

            else:
                # Case 3: Relative path (e.g. ..\..\MyWorkspace\Out)
                if 'out_dir' in globals() and out_dir:
                    drive = out_dir[0].upper() if out_dir else 'C'
                    network_out_dir = f"\\\\{remote_host}\\{drive}$\\{out_dir[3:]}"
                    out_path = network_out_dir.rstrip("\\")
                    display_func(f"Using discovered OUT directory: {out_path}")
                else:
                    # Try to get remote working directory via socket
                    pwd_cmd = "cmd = 'cd'"
                    res_pwd = socket_query(sock, pwd_cmd)
                    callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
                    res_dir = socket_query(sock, callos_cmd)
            
                    remote_cwd = None
                    if res_dir:
                        for line in res_dir.strip().splitlines():
                            if ':' in line:  # found a drive path
                                remote_cwd = line.strip()
                                break
            
                    if remote_cwd:
                        drive = remote_cwd[0].upper()
                        # Handle ".." relative paths
                        if folder_path.startswith(".."):
                            up_count = folder_path.count("..")
                            cwd_parts = remote_cwd.split("\\")
                            base_parts = cwd_parts[:-up_count] if len(cwd_parts) > up_count else cwd_parts[:1]
                            target_parts = [p for p in folder_path.split("\\") if p and p != ".."]
                            rel_path = "\\".join(target_parts)
                            #out_path = f"\\\\{remote_host}\\{drive}$\\{'\\'.join(base_parts[1:])}\\{rel_path}"
                            joined_base = "\\".join(base_parts[1:])
                            out_path = f"\\\\{remote_host}\\{drive}$\\{joined_base}\\{rel_path}"
                        else:
                            # Append relative directly
                            out_path = f"\\\\{remote_host}\\{drive}$\\{remote_cwd[3:]}\\{folder_path}"
                            
                
                        out_path = out_path.rstrip("\\")
                        display_func(f"Constructed OUT folder path: {out_path}")
                    else:
                        display_func("Warning: Could not determine remote working directory")
                        out_path = f"\\\\{remote_host}\\C$\\Sparnex.testplan.DLA352\\MyWorkspace\\Out"
                        display_func(f"Using fallback path: {out_path}")

        except Exception as e:
            display_func(f"Error determining OUT folder path: {str(e)}")
            print(f"Error details: {traceback.format_exc()}")
            out_path = f"\\\\{remote_host}\\C$\\Sparnex.testplan.DLA352\\MyWorkspace\\Out"
            display_func(f"Using emergency fallback path: {out_path}")




        # # Get the correct OUT folder path without hardcoding
        # if folder_path.startswith('\\\\'):
        #     # Already a network path - use as is
        #     out_path = folder_path
        #     display_func(f"Using specified network path: {out_path}")
        # else:
        #     try:
        #         # Use the discovered out_dir if available
        #         if 'out_dir' in globals() and out_dir:
        #             # Convert local path to network path
        #             drive = out_dir[0] if out_dir and len(out_dir) > 0 else 'C'
        #             network_out_dir = f"\\\\{remote_host}\\{drive}$\\{out_dir[3:]}"
        #             out_path = network_out_dir
        #             display_func(f"Using discovered OUT directory: {out_path}")
        #         else:
        #             # Get remote working directory using socket_query
        #             pwd_cmd = "cmd = 'cd'"
        #             res_pwd = socket_query(sock, pwd_cmd)
        #             callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
        #             res_dir = socket_query(sock, callos_cmd)
                    
        #             # Clean up the response - important to handle line breaks correctly
        #             remote_cwd = None
        #             if res_dir:
        #                 # Remove any leading numbers and extract just the path
        #                 lines = res_dir.strip().split('\n')
        #                 for line in lines:
        #                     if 'C:' in line:
        #                         remote_cwd = line.strip()
        #                         break
                    
        #             if remote_cwd:
        #                 display_func(f"Remote current directory: {remote_cwd}")
        #                 out_path = None  # Initialize out_path to avoid reference error
                        
        #                 # Handle relative paths
        #                 if folder_path.startswith('..'):
        #                     # For "..\\..\\MyWorkspace\\Out" pattern, go up from ICL_Server
        #                     if "ICL_Server" in remote_cwd:
        #                         # Split path at ICL_Server and keep the base part
        #                         base_path = remote_cwd.split("ICL_Server")[0]
                                
        #                         # Replace relative path directly with full target
        #                         if "myworkspace\\out" in folder_path.lower() or "myworkspace\\Out" in folder_path:
        #                             # Use a direct path to MyWorkspace\\Out
        #                             base_path_fixed = base_path.replace('C:\\', '')
        #                             out_path = f"\\\\{remote_host}\\C$\\{base_path_fixed}MyWorkspace\\Out"
        #                         else:
        #                             # For other relative paths, extract the target folder
        #                             target_parts = [p for p in folder_path.split('\\') if p and p != '..']
        #                             target_path = '\\'.join(target_parts)
        #                             base_path_fixed = base_path.replace('C:\\', '')
        #                             out_path = f"\\\\{remote_host}\\C$\\{base_path_fixed}{target_path}"
        #                     else:
        #                         # Fallback for non-ICL_Server paths
        #                         # Go up levels based on number of ".." segments
        #                         parts = remote_cwd.split('\\')
        #                         up_count = folder_path.count('..')
        #                         if len(parts) > up_count:
        #                             base_path = '\\'.join(parts[:-up_count])
        #                             target_parts = [p for p in folder_path.split('\\') if p and p != '..']
        #                             if target_parts:
        #                                 rel_path = '\\'.join(target_parts)
        #                                 out_path = f"\\\\{remote_host}\\{parts[0][0]}$\\{base_path[3:]}\\{rel_path}"
        #                             else:
        #                                 out_path = f"\\\\{remote_host}\\{parts[0][0]}$\\{base_path[3:]}"
        #                         else:
        #                             # Too many levels up, use root
        #                             target_parts = [p for p in folder_path.split('\\') if p and p != '..']
        #                             rel_path = '\\'.join(target_parts)
        #                             out_path = f"\\\\{remote_host}\\C$\\{rel_path}"
        #                 else:
        #                     # Not a relative path - append to current directory if no drive letter
        #                     if ':' in folder_path:
        #                         # Has drive letter, use as is
        #                         drive = folder_path[0]
        #                         path_without_drive = folder_path[3:] if folder_path[2:3] == '\\' else folder_path[2:]
        #                         out_path = f"\\\\{remote_host}\\{drive}$\\{path_without_drive}"
        #                     else:
        #                         # Append to current directory
        #                         drive = remote_cwd[0]
        #                         out_path = f"\\\\{remote_host}\\{drive}$\\{remote_cwd[3:]}\\{folder_path}"
                    
        #                 # Normalize path to avoid double backslashes
        #                 if out_path:  # Check to avoid error if out_path wasn't set
        #                     out_path = out_path.replace('\\\\\\', '\\\\').replace('\\\\\\', '\\\\')
        #                     display_func(f"Constructed OUT folder path: {out_path}")
        #                 else:
        #                     # If out_path wasn't set, use fallback
        #                     out_path = f"\\\\{remote_host}\\C$\\Sparnex.testplan.DLA352\\MyWorkspace\\Out"
        #                     display_func(f"Using fallback path (out_path not set): {out_path}")
        #             else:
        #                 display_func("Warning: Could not determine remote working directory")
        #                 # Use known path formats based on previous observations
        #                 out_path = f"\\\\{remote_host}\\C$\\Sparnex.testplan.DLA352\\MyWorkspace\\Out"
        #                 display_func(f"Using fallback path: {out_path}")
        #     except Exception as e:
        #         display_func(f"Error determining OUT folder path: {str(e)}")
        #         print(f"Error details: {traceback.format_exc()}")
                
        #         # Emergency fallback - use the path we know worked in socket approach
        #         out_path = f"\\\\{remote_host}\\C$\\Sparnex.testplan.DLA352\\MyWorkspace\\Out"
        #         display_func(f"Using emergency fallback path: {out_path}")
        
        display_func(f"Looking for files in {out_path} with timestamp {timestamp}")
        
        try:
            # Always use socket approach for file listing since it worked
            display_func("Using socket method to list files...")
            
            # Extract date components from timestamp for filtering
            if len(timestamp) >= 8:
                try:
                    year_filter = timestamp[0:4]
                    month_filter = timestamp[4:6]
                    day_filter = timestamp[6:8]
                    date_pattern = f"{year_filter}{month_filter}{day_filter}"
                    
                    # Use more targeted file search with date pattern
                    print(f"Filtering for files from date: {year_filter}-{month_filter}-{day_filter}")
                    dir_cmd = f"cmd = 'dir /b \"{out_path}\\*{date_pattern}*\"'"
                except ValueError:
                    # If we can't parse the timestamp, just use the original format
                    dir_cmd = f"cmd = 'dir /b \"{out_path}\\*{timestamp}*\"'"
            else:
                # If timestamp isn't long enough, use as-is
                dir_cmd = f"cmd = 'dir /b \"{out_path}\\*{timestamp}*\"'"
            
            # Send the command
            res_dir = socket_query(sock, dir_cmd)
            
            if not res_dir:
                print("Failed to send directory command. Trying without filtering...")
                # Try without filtering as fallback
                dir_cmd = f"cmd = 'dir /b \"{out_path}\"'"
                res_dir = socket_query(sock, dir_cmd)
                if not res_dir:
                    print("Failed to list directory. Please check the path.")
                    return
                
            callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
            files_result = socket_query(sock, callos_cmd)
            
            if not files_result:
                print("Failed to get directory listing results")
                return
                
            if "invalid" in files_result.lower() or "error" in files_result.lower():
                print(f"Error from socket command: {files_result}")
                # Try alternative path formats
                #alt_path = out_path.replace("\\\\", "\\").replace("\\\\", "\\")
                alt_path = os.path.normpath(out_path)
                print(f"Trying alternative path format: {alt_path}")
                if alt_path.startswith("\\") and not alt_path.startswith("\\\\"):
                    alt_path = "\\" + alt_path  # restore missing slash
                dir_cmd = f"cmd = 'dir /b \"{alt_path}\"'"
                socket_query(sock, dir_cmd)
                files_result = socket_query(sock, callos_cmd)
                
                if not files_result or "invalid" in files_result.lower():
                    print(f"Still failed with alternative format. Trying direct path.")
                    
                    # Try a direct command with known working path
                    dir_cmd = f"cmd = 'dir /b \"..\\..\\MyWorkspace\\Out\\*{timestamp}*\"'"
                    socket_query(sock, dir_cmd)
                    files_result = socket_query(sock, callos_cmd)
            
            # Parse the file list
            files = [line.strip() for line in files_result.splitlines() if line.strip()]
            
            # Show what files we actually found
            print(f"Found {len(files)} files in directory")
            if files:
                #display_func("First 5 files found:")
                
                # Show available dates in the files
                dates_found = set()
                for file in files:
                    # Look for date patterns (YYYYMMDD_HHMMSS)
                    match = re.search(r'(\d{8}_\d{6})', file)
                    if match:
                        date_str = match.group(1)
                        yyyy = date_str[0:4]
                        mm = date_str[4:6]
                        dd = date_str[6:8]
                        dates_found.add(f"{yyyy}-{mm}-{dd}")
                
                if dates_found:
                    print("Dates available in files:")
                    #for date in sorted(dates_found):
                    print(f"  - {dates_found}")
            
            # Find matching files - use more flexible matching
            matching_files = []
            for file in files:
                # Try different match approaches:
                # 1. Exact match
                if timestamp in file:
                    matching_files.append(file)
                    continue
                    
                # 2. Case-insensitive match
                if timestamp.lower() in file.lower():
                    matching_files.append(file)
                    continue
                    
                # 3. Try to match just the date part if timestamp has date_time format
                if '_' in timestamp and len(timestamp) > 8:
                    date_part = timestamp.split('_')[0]
                    if date_part in file:
                        matching_files.append(file)
                        continue
            
            if not matching_files:
                display_func(f"No files found with timestamp {timestamp}")
                display_func("Please verify the timestamp format matches your files")
                return
            
            display_func(f"Found {len(matching_files)} files matching timestamp {timestamp}")
            print("Matching files:")
            for i, file in enumerate(matching_files[:5]):  # Show first 5
                print(f"  {i+1}. {file}")
            
            # After finding matching files, look specifically for .tp.txt file
            tp_txt_file = None
            tp_config = {}  # Dictionary to store config values from .tp.txt file
        
            for file in matching_files:
                if file.endswith('.tp.txt'):
                    tp_txt_file = file
                    display_func(f"Found TP configuration file: {tp_txt_file}")
                    break
                
            # If we found a .tp.txt file, extract configuration from it
            if tp_txt_file:
                try:
                    # Full path to the tp.txt file
                    tp_txt_path = os.path.join(out_path, tp_txt_file)
                
                    # Read the file content using socket commands
                    read_cmd = f"cmd = 'type \"{tp_txt_path}\"'"
                    socket_query(sock, read_cmd)
                    callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
                    tp_content = socket_query(sock, callos_cmd)
                
                    if tp_content:
                        display_func("Successfully read TP configuration file")
                    
                        # Parse the content to extract config values
                        # The format is typically config.section.key='value'
                        for line in tp_content.splitlines():
                            line = line.strip()
                            if line.startswith('config.'):
                                try:
                                    # Split at the first '=' character
                                    key_part, value_part = line.split('=', 1)
                                    key = key_part.strip()
                                
                                    # Clean up the value (remove quotes)
                                    value = value_part.strip()
                                    if value.startswith("'") and value.endswith("'"):
                                        value = value[1:-1]
                                    elif value.startswith('"') and value.endswith('"'):
                                        value = value[1:-1]
                                
                                    # Store in our config dictionary
                                    tp_config[key] = value
                                except Exception as e:
                                    print(f"Error parsing config line '{line}': {str(e)}")
                    
                        display_func(f"Extracted {len(tp_config)} configuration values from TP file")
                    else:
                        display_func("Could not read TP configuration file content", "red")
                except Exception as e:
                    display_func(f"Error processing TP configuration file: {str(e)}", "red")
                    print(f"Error details: {traceback.format_exc()}")
            else:
                display_func("No .tp.txt configuration file found for this timestamp", "yellow")
        

            # Create directory for results
            hermes_dir = ""
            try:
                # Extract base name from the timestamp or first file
                if len(matching_files) > 0:
                    first_file = matching_files[0]
                    # Try to extract TC.number.name pattern if present
                    match = re.search(r'(TC\.\d+\.[\w]+)', first_file)
                    if match:
                        base_name = match.group(1)
                    else:
                        # Fallback to simpler approach
                        base_name = first_file.split('_')[0]
                else:
                    # If no files found, use generic naming
                    base_name = f"TC.Unknown"
                
                # Add timestamp to the base name
                base_name_with_timestamp = f"{base_name}.{timestamp}"
                
                hermes_dir = f'\\\\HERMES\\4_Engineering\\7-test campaign\\Results\\{base_name_with_timestamp}'
                os.makedirs(hermes_dir, exist_ok=True)
                display_func(f"Created directory: {hermes_dir}")
            
                # OPTIMIZED APPROACH: Copy all files at once using wildcard pattern
                #display_func("Using optimized file copy with wildcard pattern")
            
                # Format wildcard copy command - copy all files with the timestamp at once
                wildcard_pattern = f"*{timestamp}*"
                source_path = os.path.join(out_path, wildcard_pattern)
                #copy_cmd = f"cmd = 'copy \"{os.path.join(out_path, wildcard_pattern)}\" \"{hermes_dir}\"'"
                copy_cmd = f"cmd = 'copy \"{source_path}\" \"{hermes_dir}\"'"
                socket_query(sock, copy_cmd)
                # Execute the command directly with callos
                callos_cmd = f"(e, r) = callos(cmd) <<< e <<< r"
                copy_result = socket_query(sock, callos_cmd)
                
                print(f"Copy command result: {copy_result}")
                if "copied" in copy_result.lower():
                    # Extract number of files copied
                    match = re.search(r'(\d+)\s+file\(s\)\s+copied', copy_result)
                    num_files = match.group(1) if match else "multiple"
                    print(f"Successfully copied {num_files} files with a single command")
                
                    # Get a list of the copied files for display
                    dir_cmd = f"cmd = 'dir /b \"{hermes_dir}\"'"
                    socket_query(sock, dir_cmd)
                    callos_cmd = "(e, r) = callos(cmd) <<< e <<< r"
                    files_result = socket_query(sock, callos_cmd)
                
                    copied_files = [line.strip() for line in files_result.splitlines() if line.strip()]
                    display_func(f"Files copied to {hermes_dir}:")
                    for i, file in enumerate(copied_files[:10]):  # Show first 10
                        display_func(f"  {i+1}. {file}")
                
                    if len(copied_files) > 10:
                        display_func(f"  ... and {len(copied_files) - 10} more files")


                else:
                    #display_func(f"Bulk copy failed: {copy_result}")
                    display_func("Falling back to individual file copy...")
                    # Alternative approach - try using xcopy instead of copy
                    xcopy_cmd = f"cmd = 'xcopy \"{os.path.join(out_path, wildcard_pattern)}\" \"{hermes_dir}\" /I /Y'"
                    socket_query(sock, xcopy_cmd)
                    callos_cmd = f"(e, r) = callos(cmd) <<< e <<< r"
                    xcopy_result = socket_query(sock, callos_cmd)
    
                    if "file(s) copied" in xcopy_result.lower():
                        display_func(f"Successfully used xcopy for bulk copy")
                    else:
                        display_func(f"Alternative bulk copy also failed, falling back to individual files")
                        # Fall back to individual file copy if bulk copy fails
                        copied_files = []
                        for file in matching_files:
                            try:
                                source = os.path.join(out_path, file)
                                destination = hermes_dir
                        
                                # Use direct callos for individual copies
                                copy_cmd = f"copy \"{source}\" \"{destination}\""
                                callos_cmd = f"(e, r) = callos('{copy_cmd}') <<< e <<< r"
                                callos_result = socket_query(sock, callos_cmd)
                        
                                if "copied" in callos_result.lower() or "1 file(s)" in callos_result.lower():
                                    copied_files.append(file)
                                    display_func(f"Copied {file} to {hermes_dir}")
                                else:
                                    display_func(f"Failed to copy {file}: {callos_result}")
                            except Exception as e:
                                display_func(f"Error copying {file}: {str(e)}")
            
                    # Update results path variable and display
                    if results_path_var is not None:
                        results_path_var.set(hermes_dir)
                        display_func(f"Updated results path to: {hermes_dir}")
            
                    # Summary message
                    display_func(f"\nSuccessfully copied files to:", "green")
                    display_func(f"  {hermes_dir}", "green")
            
                    # # Start database update in a separate thread
                    # try:
                    #     threading.Thread(
                    #         target=lambda: update_measurement_database(base_name_with_timestamp, current_test_plan_file, display_func),
                    #         daemon=True
                    #     ).start()
                
                    #     display_func("Database update initiated", "black")
                    # except Exception as e:
                    #     display_func(f"Error starting database update: {str(e)}", "red")
                    #     print(f"Error details: {traceback.format_exc()}")

            except Exception as e:
                display_func(f"Error in copy operation: {str(e)}", "red")
                print(f"Error details: {traceback.format_exc()}")
                
            # After successfully copying files, update the database with the extracted config
            if hermes_dir:
                try:
                    threading.Thread(
                        target=lambda: update_measurement_database_from_tp_config(
                            base_name_with_timestamp, 
                            tp_txt_file, 
                            tp_config, 
                            display_func
                            ),
                            daemon=True
                    ).start()
                
                    display_func("Database update initiated", "black")
                except Exception as e:
                    display_func(f"Error starting database update: {str(e)}", "red")
                    print(f"Error details: {traceback.format_exc()}")

        except Exception as e:
            display_func(f"Error in copy operation: {str(e)}", "red")
            print(f"Error details: {traceback.format_exc()}")
    
    # Start the process in a separate thread to prevent UI freezing
    display_func("Starting file update process...")
    threading.Thread(target=process_files_thread, daemon=True).start()
    
    return "Process started"
        

def update_measurement_database_from_tp_config(folder_name, tp_filename, tp_config, display_func):
    """
    Updates the measurement database Excel file with new measurement details,
    using configuration extracted from a .tp.txt file.
    
    Arguments:
        folder_name: The folder name for the measurement results
        tp_filename: The name of the TP file
        tp_config: Dictionary containing configuration values from the TP file
        display_func: Function to display messages in the UI
    """
    try:
        excel_path = '//HERMES/4_Engineering/7-test campaign/Results/database.Measurement.xlsx'
        temp_file = '//HERMES/4_Engineering/7-test campaign/Results/Python Temp/temp_measurements.xlsx'
        base_results_path = '//HERMES/4_Engineering/7-test campaign/Results'

        # Get the line count from TP file
        tp_line_count = count_tp_test_lines()
        # Get CSV row count
        csv_row_count = count_csv_data_rows()

        def write_data_to_sheet(sheet, row_num):
            """Helper function to write data to a sheet"""
            for col in range(1, sheet.max_column + 1):
                header = str(sheet.cell(row=1, column=col).value or '')
                
                # Measurement name [Raw Results]
                if "Measurement name [Raw Results]" in header:
                    sheet.cell(row=row_num, column=col, value=os.path.basename(folder_name))
                # Filename TP
                elif "Filename TP" in header:
                    sheet.cell(row=row_num, column=col, value=tp_filename)
                # Platform
                elif "platform" in header:
                    # Use the connected server IP or hostname
                    sheet.cell(row=row_num, column=col, value=ip_entry.get())
                # DUT Name
                elif "DUT Name" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dut.name', ''))
                # DUT SN
                elif "DUT SN" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dut.serialnumber', ''))
                # Host
                elif "DUT Host" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dut.host', ''))
                elif "DPU Host" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dpu.host', ''))    
                # DPU Name
                elif "DPU Name" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dpu.name', ''))
                # DPU SN
                elif "DPU SN" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dpu.serialnumber', ''))
                # Port (capitalized)
                elif header == "Port":
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dut.port', ''))
                # port (lowercase)
                elif header == "port":
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dpu.port', ''))
                # Profile
                elif "Profile" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dpu.profile', ''))
                # Linesimulator
                elif "Linesimulator" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.platform.linesimulator', ''))
                # Noise generator
                elif "Noise generator" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.platform.noise_generator', ''))
                # Traffic Generator
                elif "Traffic Generator" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.platform.traffic_generator', ''))
                # Operator
                elif "operator" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.operator', ''))    
                # Switch
                elif "Switch" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.platform.switch', ''))

                # Test Plan
                elif "Test Plan" in header:
                    sheet.cell(row=row_num, column=col, value=tp_config.get('config.dut.testplan', ''))
                # Link
                elif header == "Link":
                    measurement_folder = os.path.basename(folder_name)
                    network_path = f"\\\\HERMES\\4_Engineering\\7-test campaign\\Results\\{measurement_folder}"
                    hyperlink_formula = f'=HYPERLINK("{network_path}","Click to open")'
                    sheet.cell(row=row_num, column=col, value=hyperlink_formula)
                # Expected
                elif header == "Expected":
                    sheet.cell(row=row_num, column=col, value=tp_line_count)
                # Actual
                elif header == "Actual":
                    sheet.cell(row=row_num, column=col, value=csv_row_count)

        # Rest of the function remains the same as update_measurement_database
        try:
            # Try to write to main database first
            wb = openpyxl.load_workbook(excel_path)
            
            # Get appropriate sheet (2024 or 2025)
            current_date = datetime.now()
            sheet_name = "2025" if current_date >= datetime(2025, 1, 1) else "2024"
            
            if sheet_name == "2025" and "2025" not in wb.sheetnames:
                wb.copy_worksheet(wb["2024"]).title = "2025"
                sheet_2025 = wb["2025"]
                for row in range(2, sheet_2025.max_row + 1):
                    for col in range(1, sheet_2025.max_column + 1):
                        sheet_2025.cell(row=row, column=col).value = None
            
            sheet = wb[sheet_name]
            
            # Find last row in main sheet
            last_row = 1
            while sheet.cell(row=last_row, column=1).value is not None:
                last_row += 1

            # Write current data
            write_data_to_sheet(sheet, last_row)
            
            # Save main database
            wb.save(excel_path)
            
            # Process temp file if it has data
            temp_wb = openpyxl.load_workbook(temp_file)
            temp_sheet = temp_wb['Temp']
            
            if temp_sheet.max_row > 1:  # If there's data beyond headers
                # Copy all data from temp to main
                for row in range(2, temp_sheet.max_row + 1):
                    for col in range(1, temp_sheet.max_column + 1):
                        value = temp_sheet.cell(row=row, column=col).value
                        if value is not None:  # Only copy non-empty cells
                            sheet.cell(row=last_row + row - 1, column=col).value = value
                
                # Save main database with temp data
                wb.save(excel_path)
                
                # Clear temp sheet (keep headers)
                for row in range(2, temp_sheet.max_row + 1):
                    for col in range(1, temp_sheet.max_column + 1):
                        temp_sheet.cell(row=row, column=col).value = None
                
                # Save cleared temp file
                temp_wb.save(temp_file)
                display_func("Successfully transferred temp data to main database", "green")
            
            display_func(f"Database updated successfully with measurement: {folder_name}", "green")
            
        except PermissionError:
            # If main database is locked, save to temp file
            temp_wb = openpyxl.load_workbook(temp_file)
            temp_sheet = temp_wb['Temp']
            
            # Find last row in temp sheet
            last_row = 1
            while temp_sheet.cell(row=last_row, column=1).value is not None:
                last_row += 1
            
            # Write current data to temp sheet
            write_data_to_sheet(temp_sheet, last_row)
            
            # Save temp file
            temp_wb.save(temp_file)
            display_func(f"Data saved to temporary file. Will update database when available.", "yellow")

    except Exception as e:
        error_msg = f"Error updating measurement database: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        messagebox.showerror("Database Update Error", error_msg)        

page4_functions['update_button_function'] = update_button_function

def create_page4(parent=None, as_tab=True):
    """
    Create the Test Plan Execution page
    
    Args:
        parent: Optional parent widget. If None, uses notebook.
        as_tab: Whether to add as a notebook tab or just create a frame
    """

    global page4_frame, config_inner_frame, test_inner_frame, config_canvas, test_canvas
    global response_text_4, status_label_4, status_var_4, config_search_var, test_search_var
    global config_window, test_window
    
    
    global results_path_var
    if not results_path_var:
        results_path_var = tk.StringVar(value=r"\\HERMES\4_Engineering\7-test campaign\Results")

    # Reset canvas references first to avoid binding to old references
    config_canvas = None
    test_canvas = None
    
    print(f"Creating page4: parent={parent}, as_tab={as_tab}")
    
    try:
        # Use provided parent or default to notebook
        if parent is None:
            print("No parent provided, using notebook")
            parent = notebook
        
        if not parent:
            print("Error: No parent widget available")
            return None
            
        print("Creating main frame")
        # Create main frame
        page4_frame = ttk.Frame(parent, padding="5 5 5 5")
        
        # Add as tab if needed
        if as_tab:
            print("Adding as notebook tab")
            notebook.add(page4_frame, text="Test Plan Execution")
            print("Added as notebook tab")
            # Bind tab change event
            notebook.bind('<<NotebookTabChanged>>', lambda e: update_status_4())
        else:
            print("Adding as standalone frame")
            page4_frame.pack(fill="both", expand=True)
        
        # Initialize StringVars - only if they don't already exist
        if not config_search_var:
            config_search_var = tk.StringVar()
        if not test_search_var:
            test_search_var = tk.StringVar()
        if not status_var_4:
            status_var_4 = tk.StringVar(value="")
        
        print("Successfully created page4 frame and variables")
    
        global df, current_tp_file, status_label_4
    
        # Create two main vertical frames
        left_frame = ttk.Frame(page4_frame, relief="solid")
        right_frame = ttk.Frame(page4_frame, relief="solid")
    
        left_frame.grid(row=0, column=0, sticky="nsew")
        right_frame.grid(row=0, column=1, sticky="nsew")
    
        # Configure grid weights
        page4_frame.grid_columnconfigure(0, weight=2)    # Left frame smaller
        page4_frame.grid_columnconfigure(1, weight=4)    # Right frame larger
        page4_frame.grid_rowconfigure(0, weight=1)       # Full height

        button_frame = ttk.Frame(left_frame)
        button_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)

        # Test Plan title and buttons
        ttk.Label(button_frame, text="Test Plan Execution", 
          font=("Helvetica", 10, "bold")).pack(side="left", padx=10)

        # "New" button
        ttk.Button(button_frame, 
            text="New", 
            cursor="hand2", 
            command=lambda: root.after(100, lambda: reset_and_go_to_page2())
        ).pack(side="left", padx=5)

        # "Open" button 
        edit_button_page4 = ttk.Button(button_frame, 
            text="Open", 
            cursor="hand2", 
            command=lambda: root.after(100, coordinated_edit_test_plan)
        ).pack(side="left", padx=5)

        # "Save" button
        save_button_page4 = ttk.Button(button_frame, 
            text="Save", 
            cursor="hand2", 
            command=lambda: root.after(100, coordinated_save_test_plan)
        ).pack(side="left", padx=5)

        # "Save As" button
        save_as_button_page4 = ttk.Button(button_frame, 
            text="Save As", 
            cursor="hand2", 
            command=lambda: root.after(100, lambda: coordinated_save_as_test_plan)
        ).pack(side="left", padx=5)

        # "Run" button
        run_button_page4 = ttk.Button(button_frame, 
            text="Run", 
            cursor="hand2", 
            command=lambda: root.after(100, run_current_or_new_tp)
        ).pack(side="left", padx=5)

        # Create a container frame for the entire config section
        config_section_frame = ttk.Frame(left_frame)
        config_section_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        
        
        # 4. Add a separator between config and test sections
        separator = ttk.Separator(left_frame, orient='horizontal')
        separator.grid(row=2, column=0, sticky="ew", padx=10, pady=5)

        # Config Header Section
        config_header_frame = ttk.Frame(config_section_frame)
        config_header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0,5))

        # Config Search
        config_search_var = tk.StringVar()
        config_search_entry = ttk.Entry(config_header_frame, textvariable=config_search_var, width=30)
        config_search_entry.grid(row=0, column=1, padx=5)
        ttk.Button(config_header_frame, 
               text="Clear", 
               command=lambda: config_search_var.set("")).grid(row=0, column=2, padx=5)

        # Config Content Frame
        config_content_frame = ttk.Frame(config_section_frame)
        config_content_frame.grid(row=1, column=0, sticky="nsew")

        # Create canvas and scrollbar in content frame
        config_canvas = tk.Canvas(config_content_frame, height=350)
        config_canvas.grid(row=0, column=0, sticky="nsew")

        config_scrollbar = ttk.Scrollbar(config_content_frame, orient="vertical", command=config_canvas.yview)
        config_scrollbar.grid(row=0, column=1, sticky="ns")

        config_inner_frame = ttk.Frame(config_canvas)
        config_window = config_canvas.create_window((0, 0), window=config_inner_frame, anchor="nw")
        config_canvas.configure(yscrollcommand=config_scrollbar.set)

        # Configure weights for proper expansion
        config_section_frame.grid_columnconfigure(0, weight=1)
        config_section_frame.grid_rowconfigure(1, weight=1)

        config_content_frame.grid_columnconfigure(0, weight=1)
        config_content_frame.grid_rowconfigure(0, weight=1)

        # Test Section Frame (similar structure)
        test_section_frame = ttk.Frame(left_frame)
        test_section_frame.grid(row=3, column=0, sticky="nsew", padx=10, pady=(10,5))

        # Test Header Section
        test_header_frame = ttk.Frame(test_section_frame)
        test_header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0,5))

        # Test Search
        test_search_var = tk.StringVar()
        test_search_entry = ttk.Entry(test_header_frame, textvariable=test_search_var, width=30)
        test_search_entry.grid(row=0, column=1, padx=5)
        ttk.Button(test_header_frame, 
               text="Clear", 
               command=lambda: test_search_var.set("")).grid(row=0, column=2, padx=5)

        # Test Content Frame
        test_content_frame = ttk.Frame(test_section_frame)
        test_content_frame.grid(row=1, column=0, sticky="nsew")

        # Create canvas and scrollbar in content frame
        test_canvas = tk.Canvas(test_content_frame, height=200)
        test_canvas.grid(row=0, column=0, sticky="nsew")

        test_scrollbar = ttk.Scrollbar(test_content_frame, orient="vertical", command=test_canvas.yview)
        test_scrollbar.grid(row=0, column=1, sticky="ns")

        test_inner_frame = ttk.Frame(test_canvas)
        test_window = test_canvas.create_window((0, 0), window=test_inner_frame, anchor="nw")
        test_canvas.configure(yscrollcommand=test_scrollbar.set)

        # Configure weights for proper expansion
        test_section_frame.grid_columnconfigure(0, weight=1)
        test_section_frame.grid_rowconfigure(1, weight=1)

        test_content_frame.grid_columnconfigure(0, weight=1)
        test_content_frame.grid_rowconfigure(0, weight=1)

        # Header label and search
        ttk.Label(config_header_frame, 
             text="Configuration", 
             font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0,10))

        ttk.Label(test_header_frame, 
             text="Test Plans", 
             font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0,10))

        # Configure scroll regions
        configure_scroll_region(config_canvas, config_inner_frame)
        configure_scroll_region(test_canvas, test_inner_frame)

        # Store the function in the global dictionary
        page4_functions['clear_preview_tables'] = clear_page4_tables_safely
        page4_functions['display_message'] = thread_safe_display_message_4
        page4_functions['run_test_plan'] = run_current_or_new_tp

        # # Header frame in left_frame - at row 0
        # header_frame = ttk.Frame(left_frame)
        # header_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)
    
        # # Configure header frame columns
        # header_frame.grid_columnconfigure(0, weight=1)
        # header_frame.grid_columnconfigure(1, weight=0)


        # # Run Button - right aligned
        # ttk.Button(header_frame, 
        #       text="Run Test Plan", 
        #       width=20,
        #       command=run_current_or_new_tp).grid(row=0, column=1, sticky="e", padx=5)
              
        # Configure left frame grid weights for new implementation
        # Left frame configuration
        left_frame.grid_rowconfigure(0, weight=0)    # Button frame - fixed height
        left_frame.grid_rowconfigure(1, weight=2)    # Config section - 2 parts
        left_frame.grid_rowconfigure(2, weight=0)    # Spacing row
        left_frame.grid_rowconfigure(3, weight=3)    # Test section - 3 parts
        left_frame.grid_columnconfigure(0, weight=1)

        # Configure right frame for server response
        right_frame.grid_columnconfigure(0, weight=1)  # Make column expandable
        right_frame.grid_rowconfigure(0, weight=0)  # Header row - fixed
        right_frame.grid_rowconfigure(1, weight=1)  # Response text area - expandable
        right_frame.grid_rowconfigure(2, weight=0)  # Results frame - fixed

        # Server Response Frame (update height)
        response_frame_4 = ttk.Frame(right_frame)
        response_frame_4.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        response_frame_4.grid_columnconfigure(0, weight=1)
        response_frame_4.grid_rowconfigure(1, weight=1)

        # Header response frame (no changes needed)
        header_response_frame = ttk.Frame(response_frame_4)
        header_response_frame.grid(row=0, column=0, sticky="ew")
        header_response_frame.grid_columnconfigure(0, weight=0)
        header_response_frame.grid_columnconfigure(1, weight=0)
        header_response_frame.grid_columnconfigure(2, weight=1)
        
        # Add Results Frame below response text area
        results_frame = ttk.Frame(right_frame)
        results_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        results_frame.grid_columnconfigure(1, weight=1)  # Make entry expand

        header_response_frame = ttk.Frame(response_frame_4)
        header_response_frame.grid(row=0, column=0, sticky="ew")
        header_response_frame.grid_columnconfigure(0, weight=0)
        header_response_frame.grid_columnconfigure(1, weight=0)
        header_response_frame.grid_columnconfigure(2, weight=1)
        header_response_frame.grid_rowconfigure(2, weight=1)


        # Results Label
        ttk.Label(results_frame, 
          text="Results Folder:", 
          font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0,10), sticky="w")

        # Results Entry
        results_path_var = tk.StringVar(value=r"\\HERMES\4_Engineering\7-test campaign\Results")
        results_entry = ttk.Entry(results_frame, 
                         textvariable=results_path_var,
                         font=("Helvetica", 10),
                         state='readonly',
                         width=50)
        results_entry.grid(row=0, column=1, sticky="ew", padx=5)

        # Open Button
        ttk.Button(results_frame, 
           text="Open", 
           cursor="hand2",
           width=10,
           command=lambda: open_in_explorer(None, var=results_path_var)
           ).grid(row=0, column=2, padx=5)

        # Folder Path
        ttk.Label(results_frame, 
            text="Folder Path:", 
            font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=(0,10), sticky="w")

        folder_path_var_4 = tk.StringVar(value="..\..\myworkspace\Out")  # Default to relative 'Out' folder

        folder_entry = ttk.Entry(results_frame, 
                         textvariable=folder_path_var_4,
                         font=("Helvetica", 10),
                         width=50)
        folder_entry.grid(row=1, column=1, sticky="ew", padx=5)

        

        browse_button = ttk.Button(
            results_frame, 
            text="Browse", 
            command=lambda: browse_with_default_path(folder_entry, var=folder_path_var_4)
        )
        browse_button.grid(row=1, column=2, padx=5)

        # Timestamp Field
        ttk.Label(results_frame, 
            text="Timestamp:", 
            font=("Helvetica", 10, "bold")).grid(row=2, column=0, padx=(0,10), sticky="w")

        timestamp_var = tk.StringVar()
        timestamp_entry = ttk.Entry(results_frame, 
                            textvariable=timestamp_var,
                            font=("Helvetica", 10),
                            width=50)
        timestamp_entry.grid(row=2, column=1, sticky="ew", padx=5)

        update_button = ttk.Button(results_frame, text="Update", command=lambda: update_button_function( folder_path_var_4, timestamp_var, display_message_4,  results_path_var, current_tp_file if  'current_tp_file' in globals() else None))
        update_button.grid(row=2, column=2, padx=5)



        # Server Response Header
        ttk.Label(header_response_frame, 
             text="Server Response:", 
             foreground='black',
             font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=(0,10), sticky="w")

        # Refresh Database Button
        ttk.Button(header_response_frame,
              text="Refresh Database",
              width=20,
              command=refresh_database).grid(row=0, column=1, padx=(0,5), sticky="w")
  
       
        # Status Label - right aligned
        status_label_4 = ttk.Label(header_response_frame, 
                    text=f"Current Test Plan: {os.path.basename(current_tp_file) if current_tp_file else 'No Test Plan Selected'}", 
                    font=("Helvetica", 10, "bold"))
        status_label_4.grid(row=0, column=2, padx=10, sticky="e")  # Changed to grid

        # # Add Close Page4 Button
        # close_button = ttk.Button(header_response_frame,
        #     text="Close Execution",
        #     width=15,
        #     command=lambda: close_page4())
        # close_button.grid(row=0, column=2, padx=5, sticky="e")
        
        # Response Text Area
        response_text_4 = scrolledtext.ScrolledText(response_frame_4, 
                                            wrap=tk.WORD,
                                            height=30,
                                            foreground='black',
                                            font=("Helvetica", 10, "bold"))
        response_text_4.grid(row=1, column=0, columnspan=2, sticky="nsew")
        response_text_4.config(state=tk.DISABLED)

        # Add Ctrl+A binding
        def on_select_all_4(event=None):
            response_text_4.tag_add("sel", "1.0", "end")
            return "break"
        response_text_4.bind("<Control-a>", on_select_all_4)
        
        try:
            if config_search_var:
               config_search_var.trace('w', filter_config_rows)
            if test_search_var:
                test_search_var.trace('w', filter_test_rows)
            print("Added trace functions to search variables")
        except Exception as e:
            print(f"Warning: Could not add trace functions: {str(e)}")
        
        # Safely configure scroll regions
        try:
            if config_canvas and config_inner_frame:
                configure_scroll_region(config_canvas, config_inner_frame)
            if test_canvas and test_inner_frame:
                configure_scroll_region(test_canvas, test_inner_frame)
            print("Configured scroll regions")
        except Exception as e:
            print(f"Warning: Could not configure scroll regions: {str(e)}")
        
        print("Page4 creation completed successfully")
        
        # Return the frame
        return page4_frame
        
    except Exception as e:
        print(f"Error creating page4: {str(e)}")
        traceback.print_exc()  # Print full traceback for debugging
        # Create a minimal frame to prevent further errors
        if parent:
            error_frame = ttk.Frame(parent)
            error_label = ttk.Label(error_frame, text=f"Error creating execution page: {str(e)}")
            error_label.grid(row=0, column=0, pady=20)
            error_frame.grid(row=0, column=0, sticky="nsew")
            return error_frame
        return None


   

def close_page4():
    """Close page4 and clean up resources"""
    global page4_frame, config_inner_frame, test_inner_frame, config_canvas, test_canvas
    global response_text_4, status_label_4
    
    try:
        result = True
        
        if result:
            # Clean up all references
            if page4_frame and page4_frame.winfo_exists():
                # Clear tables first
                clear_page4_tables_safely(config_inner_frame, test_inner_frame)
                # Remove from notebook
                notebook.forget(page4_frame)
                # Destroy the frame
                page4_frame.destroy()
            
            # Reset all global variables
            page4_frame = None
            config_inner_frame = None
            test_inner_frame = None
            config_canvas = None
            test_canvas = None
            response_text_4 = None
            status_label_4 = None
            
            print("Page4 closed and resources cleaned up")
            
            # Switch back to page2
            notebook.select(0)
            
    except Exception as e:
        print(f"Error closing page4: {str(e)}")
        messagebox.showerror("Error", f"Could not close execution page:\n{str(e)}")


create_page2()
create_page3()
#create_page4()
# Function to handle Control-a event for selecting all text
def on_select_all(event=None):
    response_text.tag_add("sel", "1.0", "end")
    return "break"

# def on_tab_changed(event):
#     """Handle tab change events"""
#     try:
#         selected_tab = event.widget.select()
#         tab_text = event.widget.tab(selected_tab, "text")
        
#         # Save changes when switching between tabs
#         if tab_text == "Test Plan Configurator":
#             # Coming to page 2, potentially save page 3 changes
#             if 'save_test_plan' in page3_functions:
#                 print("Saving changes from page 3")
#                 page3_functions['save_test_plan']()
#         elif tab_text == "Test Plan Selection":
#             if status_label_3 and status_label_3.winfo_exists():
#                 if current_tp_file:
#                     print(f"Updating execution page status with: {current_tp_file}")
#                     status_label_3.config(text=f"Current Test Plan: {os.path.basename(current_tp_file)}")
#                 else:
#                     print("No current test plan file")
#                     status_label_3.config(text="No Test Plan Selected")
#             # Coming to page 3, potentially save page 2 changes
#             print("Saving changes from page 2")
#             save_test_plan()
#         elif tab_text == "Test Plan Execution":
#             # Update status label when switching to execution tab
#             if status_label_4 and status_label_4.winfo_exists():
#                 if current_tp_file:
#                     print(f"Updating execution page status with: {current_tp_file}")
#                     status_label_4.config(text=f"Current Test Plan: {os.path.basename(current_tp_file)}")
#                 else:
#                     print("No current test plan file")
#                     status_label_4.config(text="No Test Plan Selected")
            
#         # Execute window will handle its own save logic when opened
            
#     except Exception as e:
#         print(f"Error handling tab change: {str(e)}")

def on_tab_changed(event):
    """Handle tab change events"""
    try:
        selected_tab = event.widget.select()
        tab_text = event.widget.tab(selected_tab, "text")
        
        # Save changes when switching between tabs, but only if we have a current test plan file
        if current_tp_file:  # Only attempt to save if we have a file
            if tab_text == "Test Plan Configurator":
                # Coming to page 2, potentially save page 3 changes
                if 'save_test_plan' in page3_functions:
                    print("Saving changes from page 3")
                    page3_functions['save_test_plan']()
            elif tab_text == "Test Plan Selection":
                if status_label_3 and status_label_3.winfo_exists():
                    status_label_3.config(text=f"Current Test Plan: {os.path.basename(current_tp_file)}")
                # Coming to page 3, potentially save page 2 changes
                print("Saving changes from page 2")
                save_test_plan()
            elif tab_text == "Test Plan Execution":
                # Update status label when switching to execution tab
                if status_label_4 and status_label_4.winfo_exists():
                    status_label_4.config(text=f"Current Test Plan: {os.path.basename(current_tp_file)}")
        else:
            # No current test plan file, just update status labels
            print("No current test plan file - skipping automatic save")
            if tab_text == "Test Plan Selection" and status_label_3 and status_label_3.winfo_exists():
                status_label_3.config(text="No Test Plan Selected")
            elif tab_text == "Test Plan Execution" and status_label_4 and status_label_4.winfo_exists():
                status_label_4.config(text="No Test Plan Selected")
            
    except Exception as e:
        print(f"Error handling tab change: {str(e)}")

notebook.bind('<<NotebookTabChanged>>', on_tab_changed)

# Add the selection maintenance code here
def on_dslam_action_select(event):
    selected_index = devices_listbox.curselection()
    event.widget.selection_clear()
    if selected_index:
        devices_listbox.selection_set(selected_index)
        devices_listbox.see(selected_index)

# Create two main vertical frames for page 1
left_frame_p1 = ttk.Frame(main_frame, relief="solid")
right_frame_p1 = ttk.Frame(main_frame, relief="solid")

left_frame_p1.grid(row=0, column=0, sticky="nsew")
right_frame_p1.grid(row=0, column=1, sticky="nsew")
# 1. Combined Server Details and Command Controls Section
server_command_frame = ttk.LabelFrame(left_frame_p1, text="Server & Command Controls", padding="5 5 5 5")
server_command_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

# Server Details
server_details_frame = ttk.Frame(server_command_frame)
server_details_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

ttk.Label(server_details_frame, text="Server IP/Hostname:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="e")
ip_entry = ttk.Entry(server_details_frame, width=21, font=("Helvetica", 10, "bold"))
ip_entry.grid(row=0, column=1, padx=5, pady=5)
ip_entry.insert(0,"PC-KPA")

ttk.Label(server_details_frame, text="Port:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="e")
port_entry = ttk.Entry(server_details_frame, width=21, font=("Helvetica", 10, "bold"))
port_entry.grid(row=1, column=1, padx=5, pady=5)
port_entry.insert(0, "12005")

ttk.Label(server_details_frame, text="Session Status:", font=("Helvetica", 10, "bold")).grid(row=2, column=0,padx=5, pady=5, sticky="e")

connect_button = ttk.Button(server_details_frame, text="Connect", cursor="hand2", width=20, command=connect_thread)
connect_button.grid(row=2, column=1, padx=5, pady=5)

disconnect_button = ttk.Button(server_details_frame, text="Disconnect", cursor="hand2", width=20, command=disconnect_thread)
disconnect_button.grid(row=3, column=1, padx=5, pady=5)

# ttk.Label(server_details_frame, text="Server Status:", font=("Helvetica", 10, "bold")).grid(row=8, column=0,padx=5, pady=5, sticky="e")
# ttk.Button(server_details_frame, text="Stop", cursor="hand2", width=20, command=stop_server_thread).grid(row=9, column=1, padx=5, pady=5)
# ttk.Button(server_details_frame, text="Start", cursor="hand2", width=20, command=start_server_thread).grid(row=8, column=1, padx=5, pady=5)

ttk.Label(server_details_frame, text="Command Queue:", font=("Helvetica", 10, "bold")).grid(row=10, column=0,padx=5, pady=5, sticky="e")
command_refresh_button = ttk.Button(server_details_frame, text="Refresh", cursor="hand2", width=20, command=get_command_queue_thread)
command_refresh_button.grid(row=10, column=1, padx=5, pady=5)

ttk.Label(server_details_frame, text="Network Connections:", font=("Helvetica", 10, "bold")).grid(row=16, column=0,padx=5, pady=5, sticky="e")
network_refresh_button = ttk.Button(server_details_frame, text="Refresh", cursor="hand2", width=20, command=refresh_network_connections_thread)
network_refresh_button.grid(row=16, column=1, padx=5, pady=5)

# Command Controls
command_controls_frame = ttk.Frame(server_command_frame)
command_controls_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

# Command Label and Entry/Dropdown Section
command_entry_frame = ttk.Frame(command_controls_frame)
command_entry_frame.grid(row=0, column=0, columnspan=2, sticky="nsew")

ttk.Label(command_entry_frame, text="Command:", font=("Helvetica", 10, "bold")).grid(row=1, column=0, padx=5, pady=5, sticky="w")

# Combined Command Entry and Dropdown
command_var = tk.StringVar()
command_combobox = ttk.Combobox(command_entry_frame, width=48, font=("Helvetica", 10, "bold"),
                               textvariable=command_var, state='normal', values = ["busmap()"]
                               )
command_combobox.grid(row=1, column=1, sticky="ew")

# Send Button
send_button = ttk.Button(command_entry_frame, cursor="hand2", text="Send", width=15, command=send_commands_thread)
send_button.grid(row=1, column=2)
# Command History (existing)
history_listbox = tk.Listbox(command_controls_frame, height=8, width=40, font=("Helvetica", 10, "bold"))
history_scrollbar = ttk.Scrollbar(command_controls_frame, orient="vertical", command=history_listbox.yview)
history_listbox.configure(yscrollcommand=history_scrollbar.set)
history_listbox.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
history_scrollbar.grid(row=2, column=2,padx=1,pady=5, sticky="ns")
history_listbox.bind('<<ListboxSelect>>', load_command_from_history)
history_listbox.configure(yscrollcommand=history_scrollbar.set)


# Port Controls



#devices_button = ttk.Button(port_controls_frame, cursor="hand2", text="Get DSLAM drivers", 
#                           width=17, command=lambda: show_dslamdrivers_thread())
#devices_button.grid(row=0, column=1, padx=1)

# # Create a combined frame for port and driver controls
# port_driver_frame = ttk.Frame(command_controls_frame)
# port_driver_frame.grid(row=4, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# Driver Label
##ttk.Label(port_driver_frame, text="Driver:", 
 #       font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=1, sticky="w")

# # Driver Entry
# driver_path_var = tk.StringVar(value="No driver selected")
# driver_entry = ttk.Entry(port_driver_frame, textvariable=driver_path_var, width=30, 
#                         font=("Helvetica", 10, "bold"), state='readonly')
# driver_entry.grid(row=0, column=0, padx=1, sticky="ew")

# # Browse Button
# dslam_browse_button = ttk.Button(port_driver_frame, cursor="hand2", text="Browse", 
#                                 width=15, command=browse_save_driver)
# dslam_browse_button.grid(row=0, column=1, padx=5)

# # Configure column weights
# port_driver_frame.grid_columnconfigure(1, weight=1)  # Make driver entry expandable

# # DSLAM Actions
# ttk.Label(command_controls_frame, text="DSLAM Action:", font=("Helvetica", 10, "bold")).grid(row=6, column=0, padx=5, pady=5, sticky="w")

# # Create a frame for DSLAM action controls to keep them together
# dslam_actions_frame = ttk.Frame(command_controls_frame)
# dslam_actions_frame.grid(row=6, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

# dslam_var = tk.StringVar(value="List of DSLAM actions")
# dslam_dropdown = ttk.Combobox(dslam_actions_frame, width=25, font=("Helvetica", 10, "bold"),
#                              textvariable=dslam_var, state='readonly',
#                              values=["Read Status", "Performance Parameters", 
#                                     "Inventory", "Get Configuration"])
# dslam_dropdown.grid(row=0, column=0, padx=1, sticky="ew")

# dslam_actions_button = ttk.Button(dslam_actions_frame, cursor="hand2", text="Send", 
#                                  width=15, command=dslam_actions_thread)
# dslam_actions_button.grid(row=0, column=1, padx=1)

# # Port Label
# ttk.Label(dslam_actions_frame, text="Port:", 
#           font=("Helvetica", 10, "bold")).grid(row=0, column=2, padx=1, sticky="w")

# # Port Dropdown
# port_var = tk.StringVar(value="0")
# port_dropdown = ttk.Combobox(dslam_actions_frame, width=7, font=("Helvetica", 10, "bold"),
#                             textvariable=port_var, 
#                             values=["0", "1", "2", "3"])
# port_dropdown.grid(row=0, column=3, padx=1)


# # DSLAM Devices Listbox
# #ttk.Label(command_controls_frame, text="DSLAM Devices:", font=("Helvetica", 10, "bold")).grid(row=7, column=0, padx=5, pady=5, sticky="w")
# devices_listbox = tk.Listbox(command_controls_frame, height=8, width=40, font=("Helvetica", 10, "bold"))
# devices_listbox.grid(row=8, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
# devices_scrollbar = ttk.Scrollbar(command_controls_frame, orient="vertical", command=devices_listbox.yview)
# devices_scrollbar.grid(row=8, column=2,padx=1,pady=5, sticky="ns")
# devices_listbox.configure(yscrollcommand=devices_scrollbar.set)
# devices_listbox.bind('<<ListboxSelect>>', on_driver_select)

# Profile Controls
#ttk.Label(command_controls_frame, text="Profile:", font=("Helvetica", 10, "bold")).grid(row=9, column=0, padx=5, pady=5, sticky="w")

# Profile Controls
# Create a frame for profile controls to keep them together
# profile_controls_frame = ttk.Frame(command_controls_frame)
# profile_controls_frame.grid(row=9, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# profile_entry_frame = ttk.Frame(profile_controls_frame)
# profile_entry_frame.grid(row=0, column=0, columnspan=2,padx=5,pady=5, sticky="ew")

# ttk.Label(profile_entry_frame, text="Profile:", font=("Helvetica", 10, "bold")).grid(row=0, column=0, padx=5, pady=5, sticky="w")

# load_profile_button = ttk.Button(profile_entry_frame, cursor="hand2", text="Load Profile", 
#                                 width=15, command=load_profile)
# load_profile_button.grid(row=0, column=1, padx=1,pady=5)


# # Add profile path display entry
# profile_path_var = tk.StringVar(value="No profile folder selected")
# profile_entry = ttk.Entry(profile_controls_frame, textvariable=profile_path_var, width=30, 
#                          font=("Helvetica", 10, "bold"), state='readonly')
# profile_entry.grid(row=1, column=0, padx=1, sticky="ew")

# Modify browse function to automatically trigger get_files_list
def browse_and_get_files():
    browse_save_profile()  # Original browse function
    folder_path = profile_path_var.get()
    
    if folder_path != "No profile folder selected" and os.path.exists(folder_path):
        # Clear existing items in profile listbox
        profile_listbox.delete(0, tk.END)
        
        # Get all files in the folder
        files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
        
        # Add header
        profile_listbox.insert(0, f"{'Profile List':<30}")
        profile_listbox.itemconfig(0, {'bg': 'lightgrey'})
        
        # Add files to listbox
        for file in files:
            profile_listbox.insert(tk.END, file)

# profile_browse_button = ttk.Button(profile_controls_frame, cursor="hand2", text="Browse", 
#                                  width=15, command=browse_and_get_files)
# profile_browse_button.grid(row=1, column=1, padx=1)

# # Configure column weights for profile_controls_frame
# profile_controls_frame.grid_columnconfigure(0, weight=1)  # Make entry expand
# profile_controls_frame.grid_columnconfigure(1, weight=0)  # Keep button size fixed



# # Profile Listbox
# #ttk.Label(command_controls_frame, text="Profile List:", font=("Helvetica", 10, "bold")).grid(row=10, column=0, padx=5, pady=5, sticky="w")
# profile_listbox = tk.Listbox(command_controls_frame, height=8, width=40, font=("Helvetica", 10, "bold"))
# profile_listbox.grid(row=11, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
# profile_scrollbar = ttk.Scrollbar(command_controls_frame, orient="vertical", command=profile_listbox.yview)
# profile_scrollbar.grid(row=11, column=2,padx=0,pady=5, sticky="ns")
# profile_listbox.configure(yscrollcommand=profile_scrollbar.set)
# profile_listbox.bind('<<ListboxSelect>>', on_profile_select)
# profile_listbox.configure(yscrollcommand=profile_scrollbar.set)

# File Operations
ttk.Label(command_controls_frame, text="File Operations:", font=("Helvetica", 10, "bold")).grid(row=14, column=0, padx=13, pady=5, sticky="w")

# Create a frame for file operation buttons to keep them together
file_ops_buttons_frame = ttk.Frame(command_controls_frame)
file_ops_buttons_frame.grid(row=14, column=1, columnspan=2, padx=8,pady=5, sticky="ew")

upload_button = ttk.Button(file_ops_buttons_frame, cursor="hand2", text="Upload File", 
                          width=15, command=upload_file)
upload_button.grid(row=0, column=0, padx=8)

# save_button = ttk.Button(file_ops_buttons_frame, cursor="hand2", text="Save Preferences", 
#                         width=15, command=save_preferences)
# save_button.grid(row=0, column=1, padx=8)

# use_preferences_button = ttk.Button(file_ops_buttons_frame, cursor="hand2", text="Load Preferences", 
#                                   width=15, command=use_preferences)
# use_preferences_button.grid(row=0, column=2, padx=8)

# Right Frame Content
# Server Response Section
response_frame = ttk.Frame(right_frame_p1, padding="5 5 5 5")
response_frame.grid(row=0, column=0, sticky="nsew")

# Server Response Header
ttk.Label(response_frame, text="Server Response:", foreground='black', font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)

# Status Label
status_label = ttk.Label(response_frame, text="Ready", foreground="black", font=("Helvetica", 10, "bold"))
status_label.grid(row=0, column=1, sticky="w", padx=5, pady=5)

# Response Text Area
response_text = scrolledtext.ScrolledText(response_frame, 
                                        width=80,  # Reduced width
                                        height=30, # Adjusted height
                                        wrap=tk.WORD,
                                        foreground='black',
                                        font=("Helvetica", 10, "bold"))
response_text.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
response_text.config(state=tk.DISABLED)

response_text.bind("<Control-a>", on_select_all)

# Progress Bar at the bottom
##progress_bar = ttk.Progressbar(main_frame, mode='indeterminate')
##progress_bar.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
# Expandability
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

# Main frame configuration
main_frame.columnconfigure(0, weight=3)  # Left frame larger
main_frame.columnconfigure(1, weight=2)  # Right frame smaller
main_frame.rowconfigure(0, weight=1)

# Server command frame configuration
server_command_frame.grid_columnconfigure(0, weight=1)
server_command_frame.grid_columnconfigure(1, weight=1)
server_command_frame.grid_columnconfigure(2, weight=1)
server_command_frame.grid_columnconfigure(3, weight=1)

# Command controls frame configuration
command_controls_frame.grid_columnconfigure(0, weight=0)  # Label column
command_controls_frame.grid_columnconfigure(1, weight=1)  # Content column
command_controls_frame.grid_columnconfigure(2, weight=1)  # Content column
# command_controls_frame.grid_columnconfigure(3, weight=0)  # Button/scrollbar column
# # Configure column weights
# port_driver_frame.grid_columnconfigure(0, weight=1)  # Make driver entry expand
# port_driver_frame.grid_columnconfigure(1, weight=0)  # Keep button size fixed
# Make response text expandable
response_frame.columnconfigure(0, weight=1)
response_frame.rowconfigure(1, weight=1)

# Left frame configuration
left_frame_p1.columnconfigure(0, weight=1)
left_frame_p1.rowconfigure(0, weight=1)

# Right frame configuration
right_frame_p1.columnconfigure(0, weight=1)
right_frame_p1.rowconfigure(0, weight=1)

if __name__ == "__main__":
    root.after(100, process_queue)  # Start processing the queue
    root.mainloop()